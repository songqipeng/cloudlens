
from fastapi import APIRouter, HTTPException, Query, BackgroundTasks, Body, Depends, Request
from typing import List, Dict, Optional, Any
from datetime import datetime, timedelta
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import sys
import logging

logger = logging.getLogger(__name__)

# åˆ›å»ºé™æµå™¨ï¼ˆä½¿ç”¨ IP åœ°å€ä½œä¸º keyï¼‰
limiter = Limiter(key_func=get_remote_address)
from cloudlens.core.config import ConfigManager, CloudAccount
from web.backend.i18n import get_translation, get_locale_from_request, Locale
from cloudlens.core.context import ContextManager
from cloudlens.core.cost_trend_analyzer import CostTrendAnalyzer
from cloudlens.core.cache import CacheManager  # MySQLç¼“å­˜ç®¡ç†å™¨ï¼ˆç»Ÿä¸€ä½¿ç”¨ï¼‰
from cloudlens.core.rules_manager import RulesManager
from cloudlens.core.services.analysis_service import AnalysisService
from cloudlens.core.virtual_tags import VirtualTagStorage, VirtualTag, TagRule, TagEngine
from cloudlens.core.bill_storage import BillStorageManager
from web.backend.api_cost import _get_billing_overview_totals
from cloudlens.core.dashboard_manager import DashboardStorage, Dashboard, WidgetConfig
from web.backend.api_resources import _get_cost_map, _estimate_monthly_cost, _estimate_monthly_cost_from_spec
from web.backend.error_handler import api_error_handler
from pydantic import BaseModel

def _get_provider_for_account(account: Optional[str] = None):
    """è·å–æŒ‡å®šè´¦å·çš„ Provider (æœ¬åœ°è¾…åŠ©å‡½æ•°)"""
    cm = ConfigManager()
    if not account:
        accounts = cm.list_accounts()
        if not accounts:
            raise HTTPException(status_code=404, detail="No accounts configured")
        account_config = accounts[0]
        account_name = account_config.name
    else:
        account_config = cm.get_account(account)
        if not account_config:
            raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
        account_name = account
    
    from cloudlens.cli.utils import get_provider
    return get_provider(account_config), account_name

router = APIRouter(prefix="/api")

class AccountInfo(BaseModel):
    name: str
    region: str
    access_key_id: str

class AccountUpdateRequest(BaseModel):
    """è´¦å·æ›´æ–°è¯·æ±‚"""
    alias: Optional[str] = None
    provider: Optional[str] = None
    region: Optional[str] = None
    access_key_id: Optional[str] = None
    access_key_secret: Optional[str] = None

class AccountCreateRequest(BaseModel):
    """è´¦å·åˆ›å»ºè¯·æ±‚"""
    name: str
    alias: Optional[str] = None
    provider: str = "aliyun"
    region: str = "cn-hangzhou"
    access_key_id: str
    access_key_secret: str

class DashboardSummary(BaseModel):
    account: str
    total_cost: float
    idle_count: int
    cost_trend: str
    trend_pct: float

class TriggerAnalysisRequest(BaseModel):
    account: str
    days: int = 7
    force: bool = True

@router.get("/accounts")
@limiter.limit("100/minute")
def list_accounts(request: Request) -> List[Dict]:
    """List all configured accounts (é™æµ: 100æ¬¡/åˆ†é’Ÿ)"""
    cm = ConfigManager()
    accounts = cm.list_accounts()
    result = []
    for account in accounts:
        if isinstance(account, CloudAccount):
            result.append({
                "name": account.name,
                "region": account.region,
                "access_key_id": account.access_key_id,
            })
    return result

@router.get("/config/rules")
def get_rules() -> Dict[str, Any]:
    """Get current optimization rules"""
    rm = RulesManager()
    return rm.get_rules()

@router.post("/config/rules")
def set_rules(rules: Dict[str, Any]):
    """Update optimization rules"""
    rm = RulesManager()
    try:
        rm.set_rules(rules)
        return {"status": "success", "message": "Rules updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/config/notifications")
def get_notification_config() -> Dict[str, Any]:
    """è·å–é€šçŸ¥é…ç½®ï¼ˆSMTPç­‰ï¼‰"""
    import json
    import os
    from pathlib import Path
    
    config_dir = Path(os.path.expanduser("~/.cloudlens"))
    config_file = config_dir / "notifications.json"
    
    default_config = {
        "email": "",  # å‘ä»¶é‚®ç®±ï¼ˆç®€åŒ–é…ç½®ï¼‰
        "auth_code": "",  # æˆæƒç /å¯†ç 
        "default_receiver_email": "",  # é»˜è®¤æ¥æ”¶é‚®ç®±ï¼ˆå‘Šè­¦é€šçŸ¥çš„ç›®æ ‡é‚®ç®±ï¼‰
        "smtp_host": "",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_password": "",
        "smtp_from": ""
    }
    
    if not config_file.exists():
        return default_config
    
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
            # åˆå¹¶é»˜è®¤å€¼ï¼Œç¡®ä¿æ‰€æœ‰å­—æ®µéƒ½å­˜åœ¨
            result = {**default_config, **config}
            # å¦‚æœå·²æœ‰æ—§æ ¼å¼é…ç½®ï¼Œè½¬æ¢ä¸ºæ–°æ ¼å¼
            if result.get("smtp_user") and not result.get("email"):
                result["email"] = result.get("smtp_user", "")
            if result.get("smtp_password") and not result.get("auth_code"):
                result["auth_code"] = result.get("smtp_password", "")
            return result
    except Exception as e:
        logger.error(f"è¯»å–é€šçŸ¥é…ç½®å¤±è´¥: {e}")
        return default_config

def _get_smtp_config_by_email(email: str) -> Dict[str, Any]:
    """æ ¹æ®é‚®ç®±åœ°å€è‡ªåŠ¨è·å–SMTPé…ç½®"""
    email_lower = email.lower().strip()
    
    # QQé‚®ç®±
    if email_lower.endswith("@qq.com"):
        return {
            "smtp_host": "smtp.qq.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }
    # Gmail
    elif email_lower.endswith("@gmail.com"):
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }
    # 163é‚®ç®±
    elif email_lower.endswith("@163.com"):
        return {
            "smtp_host": "smtp.163.com",
            "smtp_port": 465,
            "smtp_use_tls": False,
            "smtp_use_ssl": True
        }
    # 126é‚®ç®±
    elif email_lower.endswith("@126.com"):
        return {
            "smtp_host": "smtp.126.com",
            "smtp_port": 465,
            "smtp_use_tls": False,
            "smtp_use_ssl": True
        }
    # é»˜è®¤é…ç½®ï¼ˆé€šç”¨SMTPï¼‰
    else:
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }

@router.post("/config/notifications")
def set_notification_config(config: Dict[str, Any]):
    """ä¿å­˜é€šçŸ¥é…ç½®ï¼ˆSMTPç­‰ï¼‰"""
    import json
    import os
    from pathlib import Path
    
    config_dir = Path(os.path.expanduser("~/.cloudlens"))
    config_file = config_dir / "notifications.json"
    
    try:
        if not config_dir.exists():
            config_dir.mkdir(parents=True, exist_ok=True)
        
        # è·å–ç”¨æˆ·è¾“å…¥çš„é‚®ç®±å’Œæˆæƒç 
        email = config.get("email", "").strip()
        auth_code = config.get("auth_code", "").strip()
        default_receiver_email = config.get("default_receiver_email", "").strip()
        
        # æ ¹æ®é‚®ç®±è‡ªåŠ¨é…ç½®SMTP
        smtp_config = {}
        if email:
            smtp_config = _get_smtp_config_by_email(email)
            smtp_config["smtp_user"] = email
            smtp_config["smtp_from"] = email
            smtp_config["smtp_password"] = auth_code
        
        # ä¿å­˜å®Œæ•´é…ç½®ï¼ˆåŒ…å«è‡ªåŠ¨ç”Ÿæˆçš„SMTPé…ç½®ï¼‰
        full_config = {
            "email": email,
            "auth_code": auth_code,  # ä¿å­˜æˆæƒç ç”¨äºæ˜¾ç¤ºï¼ˆä½†å®é™…ä½¿ç”¨smtp_passwordï¼‰
            "default_receiver_email": default_receiver_email,
            **smtp_config
        }
        
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(full_config, f, indent=2, ensure_ascii=False)
        
        return {"status": "success", "message": "é€šçŸ¥é…ç½®å·²æ›´æ–°"}
    except Exception as e:
        logger.error(f"ä¿å­˜é€šçŸ¥é…ç½®å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=f"ä¿å­˜é€šçŸ¥é…ç½®å¤±è´¥: {str(e)}")

@router.post("/analyze/trigger")
def trigger_analysis(req: TriggerAnalysisRequest, background_tasks: BackgroundTasks):
    """Trigger idle resource analysis"""
    # For MVP, we run it synchronously to provide immediate feedback, 
    # but in production this should be a background task. 
    # To avoid timeout for long requests, we could use background_tasks.
    # But for now, let's try synchronous for simplicity as user requested "Scan Now" button.
    # Actually, user might wait 10-20s.
    try:
        data, cached = AnalysisService.analyze_idle_resources(req.account, req.days, req.force)
        
        # æ¸…é™¤ dashboard_idle å’Œ dashboard_summary ç¼“å­˜ï¼Œç¡®ä¿ä»ªè¡¨ç›˜èƒ½è·å–æœ€æ–°æ•°æ®
        cache_manager = CacheManager(ttl_seconds=86400)
        cache_manager.clear(resource_type="dashboard_idle", account_name=req.account)
        cache_manager.clear(resource_type="dashboard_summary", account_name=req.account)
        cache_manager.clear(resource_type=f"resource_list_{req.account}", account_name=req.account)
        
        # åŒæ—¶æ›´æ–° dashboard_idle ç¼“å­˜
        cache_manager.set(resource_type="dashboard_idle", account_name=req.account, data=data)
        
        return {
            "status": "success", 
            "count": len(data), 
            "cached": False,
            "data": data
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"è§¦å‘åˆ†æå¤±è´¥: {str(e)}\n{error_trace}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard/summary")
@api_error_handler
async def get_summary(account: Optional[str] = None, force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜")):
    """Get dashboard summary metricsï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    import logging
    logger = logging.getLogger(__name__)
    
    cm = ConfigManager()
    
    # Resolve account - å¿…é¡»æ˜ç¡®æŒ‡å®šè´¦å·ï¼Œä¸å…è®¸è‡ªåŠ¨é€‰æ‹©
    if not account:
        raise HTTPException(status_code=400, detail="è´¦å·å‚æ•°æ˜¯å¿…éœ€çš„ï¼Œè¯·åœ¨å‰ç«¯é€‰æ‹©è´¦å·")
    
    # è°ƒè¯•æ—¥å¿—
    logger.info(f"[get_summary] æ”¶åˆ°è´¦å·å‚æ•°: {account}, force_refresh: {force_refresh}")
    logger.debug(f"æ”¶åˆ°è´¦å·å‚æ•°: {account}, force_refresh: {force_refresh}")
    logger.debug(f"æ”¶åˆ°è´¦å·å‚æ•°: {account}, force_refresh: {force_refresh}")

    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="dashboard_summary", account_name=account)
        if cached_result:
            logger.debug(f"ä½¿ç”¨ç¼“å­˜æ•°æ®ï¼Œè´¦å·: {account}")
            # ä½†æ˜¯éœ€è¦ç¡®ä¿ idle_count æ˜¯æœ€æ–°çš„ï¼ˆä»é—²ç½®èµ„æºç¼“å­˜ä¸­é‡æ–°è·å–ï¼‰
            idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account)
            if not idle_data:
                idle_data = cache_manager.get(resource_type="idle_result", account_name=account)
            # å¦‚æœç¼“å­˜ä¸­æœ‰é—²ç½®èµ„æºæ•°æ®ï¼Œæ›´æ–° idle_count
            if idle_data:
                idle_count = len(idle_data) if idle_data else 0
                cached_result["idle_count"] = idle_count
                logger.info(f"ä»ç¼“å­˜æ›´æ–° idle_count: {idle_count} (è´¦å·: {account}, é—²ç½®èµ„æºæ•°æ®: {len(idle_data) if idle_data else 0} æ¡)")
            else:
                logger.warning(f"ç¼“å­˜ä¸­æ²¡æœ‰é—²ç½®èµ„æºæ•°æ®ï¼Œidle_count ä¿æŒåŸå€¼: {cached_result.get('idle_count', 0)} (è´¦å·: {account})")
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        return {
            **cached_result,
            "cached": True,
        }

    logger.debug(f"ç¼“å­˜æœªå‘½ä¸­ï¼Œç›´æ¥è·å–çœŸå®æ•°æ®ï¼Œè´¦å·: {account}")
    account_config = cm.get_account(account)
    if not account_config:
        logger.error(f"[get_summary] è´¦å· '{account}' æœªæ‰¾åˆ°")
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    
    # ç›´æ¥è°ƒç”¨å®Œæ•´çš„è®¡ç®—é€»è¾‘å¹¶æ›´æ–°ç¼“å­˜
    # _update_dashboard_summary_cache å·²ç»åŒ…å«äº† æˆæœ¬ã€èµ„æºã€è¶‹åŠ¿ã€å‘Šè­¦ã€æ ‡ç­¾ç­‰æ‰€æœ‰æŒ‡æ ‡çš„è®¡ç®—
    try:
        result_data = _update_dashboard_summary_cache(account, account_config, force_refresh=force_refresh)
        return {
            **result_data,
            "cached": False
        }
    except Exception as e:
        import traceback
        logger.error(f"è·å–ä»ªè¡¨ç›˜æ‘˜è¦å¤±è´¥: {str(e)}\n{traceback.format_exc()}")
        # å‘ç”Ÿé”™è¯¯æ—¶è¿”å›é»˜è®¤å€¼
        return {
            "account": account,
            "total_cost": 0.0,
            "idle_count": 0,
            "cost_trend": "è·å–å¤±è´¥",
            "trend_pct": 0.0,
            "total_resources": 0,
            "resource_breakdown": {"ecs": 0, "rds": 0, "redis": 0},
            "alert_count": 0,
            "tag_coverage": 0.0,
            "savings_potential": 0.0,
            "cached": False,
        }


def _update_dashboard_summary_cache(account: str, account_config, force_refresh: bool = False):
    """æ›´æ–° dashboard summary ç¼“å­˜ï¼ˆåå°ä»»åŠ¡ï¼‰"""
    import logging
    logger = logging.getLogger(__name__)
    
    # åˆå§‹åŒ–æ‰€æœ‰å˜é‡ï¼Œç¡®ä¿å³ä½¿æŸäº›æ­¥éª¤å¤±è´¥ä¹Ÿèƒ½è¿”å›æœ‰æ•ˆæ•°æ®
    total_cost = 0.0
    trend = "æ•°æ®ä¸è¶³"
    trend_pct = 0.0
    idle_count = 0
    total_resources = 0
    resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
    tag_coverage = 0.0
    alert_count = 0
    savings_potential = 0.0
    billing_total_cost = None
    current_totals = None
    last_totals = None

    # Get Cost Data - ä½¿ç”¨çœŸå®è´¦å•æ•°æ®æ¯”è¾ƒæœ¬æœˆå’Œä¸Šæœˆï¼ˆä¼˜åŒ–ï¼šå¹¶è¡Œè·å–ï¼‰
    from datetime import datetime, timedelta
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    
    now = datetime.now()
    current_cycle = now.strftime("%Y-%m")
    first_day_this_month = now.replace(day=1)
    last_day_last_month = first_day_this_month - timedelta(days=1)
    last_cycle = last_day_last_month.strftime("%Y-%m")
    
    try:
        # å¹¶è¡Œè·å–æœ¬æœˆå’Œä¸Šæœˆçš„è´¦å•æ•°æ®ï¼ˆä¼˜åŒ–æ€§èƒ½ï¼‰
        def get_current_totals():
            try:
                return _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False)
            except Exception as e:
                logger.error(f"è·å–æœ¬æœˆè´¦å•æ•°æ®å¤±è´¥: {str(e)}")
                return None
        
        def get_last_totals():
            try:
                totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False)
                # å¦‚æœæ•°æ®åº“æ²¡æœ‰ä¸Šæœˆæ•°æ®ï¼Œå°è¯•é€šè¿‡APIè·å–
                if totals is None or (totals.get("total_pretax", 0) == 0 and totals.get("data_source") == "local_db"):
                    logger.info(f"ä¸Šæœˆæ•°æ®ä¸å¯ç”¨ï¼Œå°è¯•é€šè¿‡APIè·å–: {last_cycle}")
                    return _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=True)
                return totals
            except Exception as e:
                logger.error(f"è·å–ä¸Šæœˆè´¦å•æ•°æ®å¤±è´¥: {str(e)}")
                return None
        
        # ä½¿ç”¨çº¿ç¨‹æ± å¹¶è¡Œæ‰§è¡Œï¼ˆé¿å…é˜»å¡ï¼‰
        with ThreadPoolExecutor(max_workers=2) as executor:
            current_future = executor.submit(get_current_totals)
            last_future = executor.submit(get_last_totals)
            
            # ç­‰å¾…ç»“æœï¼ˆè®¾ç½®è¶…æ—¶ï¼š30ç§’ï¼‰
            try:
                current_totals = current_future.result(timeout=30)
                last_totals = last_future.result(timeout=30)
            except Exception as e:
                logger.warning(f"è·å–è´¦å•æ•°æ®è¶…æ—¶æˆ–å¤±è´¥: {str(e)}")
        
        # ä½¿ç”¨æœ¬æœˆæ•°æ®ä½œä¸º billing_total_cost
        if current_totals:
            billing_total_cost = float(current_totals.get("total_pretax") or 0.0)
            if billing_total_cost <= 0:
                billing_total_cost = None
        
        current_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        last_cost = float((last_totals or {}).get("total_pretax") or 0.0)
        
        # è®¡ç®—è¶‹åŠ¿ï¼ˆæœ¬æœˆ vs ä¸Šæœˆï¼‰
        if last_cost > 0:
            trend_pct = ((current_cost - last_cost) / last_cost) * 100
            if trend_pct > 1:
                trend = "ä¸Šå‡"
            elif trend_pct < -1:
                trend = "ä¸‹é™"
            else:
                trend = "å¹³ç¨³"
        else:
            trend = "æ•°æ®ä¸è¶³"
            trend_pct = 0.0
        
        # ä½¿ç”¨æœ¬æœˆæˆæœ¬ä½œä¸ºtotal_costï¼ˆå¦‚æœè´¦å•æ•°æ®å¯ç”¨ï¼‰
        if current_cost > 0:
            total_cost = current_cost
        else:
            # å¦‚æœè´¦å•æ•°æ®ä¸å¯ç”¨ï¼Œå°è¯•ä½¿ç”¨å†å²è¶‹åŠ¿æ•°æ®
            analyzer = CostTrendAnalyzer()
            try:
                history, analysis = analyzer.get_cost_trend(account, days=30)
                if isinstance(analysis, dict) and "error" not in analysis:
                    total_cost = analysis.get("latest_cost", 0.0)
                else:
                    total_cost = None
            except Exception:
                total_cost = None
        
        logger.info(f"æˆæœ¬è¶‹åŠ¿è®¡ç®—: æœ¬æœˆ({current_cycle})={current_cost:.2f}, ä¸Šæœˆ({last_cycle})={last_cost:.2f}, è¶‹åŠ¿={trend}, å˜åŒ–={trend_pct:.2f}%")
        
    except Exception as e:
        logger.error(f"è®¡ç®—æˆæœ¬è¶‹åŠ¿å¤±è´¥: {str(e)}")
        # Fallback: ä½¿ç”¨å†å²è¶‹åŠ¿æ•°æ®
        analyzer = CostTrendAnalyzer()
        try:
            history, analysis = analyzer.get_cost_trend(account, days=30)
            if isinstance(analysis, dict) and "error" in analysis:
                total_cost = None
                trend = "N/A"
                trend_pct = 0.0
            else:
                total_cost = analysis.get("latest_cost", 0.0)
                trend = analysis.get("trend", "Unknown")
                trend_pct = analysis.get("total_change_pct", 0.0)
        except Exception:
            total_cost = None
            trend = "N/A"
            trend_pct = 0.0

    # Get Idle Data - ä¼˜åŒ–ï¼šä¼˜å…ˆä½¿ç”¨ç¼“å­˜ï¼Œé¿å…è€—æ—¶åˆ†æ
    try:
        cache_manager = CacheManager(ttl_seconds=86400)
        idle_data = None
        idle_count = 0
        
        # ä¼˜å…ˆä»ç¼“å­˜è·å–ï¼ˆé¿å…è€—æ—¶åˆ†æï¼‰
        idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account)
        if not idle_data:
            idle_data = cache_manager.get(resource_type="idle_result", account_name=account)
        
        if idle_data:
            idle_count = len(idle_data) if idle_data else 0
            logger.info(f"ä»ç¼“å­˜è·å–é—²ç½®èµ„æºæ•°é‡: {idle_count} (è´¦å·: {account})")
        else:
            # ç¼“å­˜ä¸ºç©ºï¼Œè¿”å›0ï¼ˆåå°ä»»åŠ¡ä¸­ä¸è¿›è¡Œè€—æ—¶åˆ†æï¼‰
            logger.info(f"ç¼“å­˜ä¸ºç©ºï¼Œè·³è¿‡åˆ†æ (è´¦å·: {account})")
            idle_count = 0
            
    except Exception as e:
        logger.warning(f"è·å–é—²ç½®èµ„æºæ•°æ®å¤±è´¥: {str(e)}")
        idle_count = 0

    # Get Resource Statistics (Consistent Global View)
    try:
        instances, rds_list, redis_list = _get_global_resources_consistent(account, account_config, force_refresh)
    except Exception as e:
        logger.warning(f"è·å–ä¸€è‡´æ€§å…¨å±€èµ„æºå¤±è´¥: {str(e)}")
        instances, rds_list, redis_list = [], [], []
    
    # ç¡®ä¿å˜é‡å­˜åœ¨
    instances = instances or []
    rds_list = rds_list or []
    redis_list = redis_list or []

    # è®¡ç®—èµ„æºç»Ÿè®¡ï¼ˆç¡®ä¿åœ¨æ‰€æœ‰æƒ…å†µä¸‹éƒ½èƒ½æ‰§è¡Œï¼‰
    resource_breakdown = {
        "ecs": len(instances) if instances else 0,
        "rds": len(rds_list) if rds_list else 0,
        "redis": len(redis_list) if redis_list else 0,
    }
    total_resources = sum(resource_breakdown.values())
    all_resources = instances + rds_list + redis_list
    
    # è¯¦ç»†æ—¥å¿—è¾“å‡ºï¼Œä¾¿äºè°ƒè¯•
    logger.info(f"èµ„æºç»Ÿè®¡ç»“æœ (è´¦å·: {account}):")
    logger.info(f"  ECS: {resource_breakdown['ecs']} (instancesç±»å‹: {type(instances).__name__}, é•¿åº¦: {len(instances) if instances else 0})")
    logger.info(f"  RDS: {resource_breakdown['rds']} (rds_listç±»å‹: {type(rds_list).__name__}, é•¿åº¦: {len(rds_list) if rds_list else 0})")
    logger.info(f"  Redis: {resource_breakdown['redis']} (redis_listç±»å‹: {type(redis_list).__name__}, é•¿åº¦: {len(redis_list) if redis_list else 0})")
    logger.info(f"  æ€»æ•°: {total_resources}")
    
    # Tag Coverage - ç»Ÿè®¡æ‰€æœ‰èµ„æºï¼ˆECS + RDS + Redisï¼‰çš„æ ‡ç­¾è¦†ç›–ç‡
    # è®¡ç®—æ ‡ç­¾è¦†ç›–ç‡å’Œå®‰å…¨å‘Šè­¦ (Consistent via SecurityComplianceAnalyzer)
    try:
        print("DEBUG: Starting Security Analysis in Thread...")
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        from cloudlens.core.security_scanner import PublicIPScanner
        
        analyzer = SecurityComplianceAnalyzer()
        
        # 1. æ ‡ç­¾è¦†ç›–ç‡ (ä¸å®‰å…¨é¡µé¢é€»è¾‘å®Œå…¨å¯¹é½)
        tag_coverage, no_tags = analyzer.check_missing_tags(all_resources)
        print(f"DEBUG: Tag Coverage Calculated: {tag_coverage}")
        
        # 2. å®‰å…¨å‘Šè­¦ (å…¬ç½‘æš´éœ² + åœæ­¢å®ä¾‹ + æŠ¢å å¼å®ä¾‹ + æ‰«ææ¼æ´)
        exposed = analyzer.detect_public_exposure(all_resources)
        stopped_instances = analyzer.check_stopped_instances(instances)
        # æ–°å¢ï¼šåŠ å…¥æŠ¢å å¼å®ä¾‹ç»Ÿè®¡ï¼Œå¯¹é½å®‰å…¨é¡µé¢
        preemptible_instances = analyzer.check_preemptible_instances(instances)
        
        try:
            from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
            from cloudlens.core.security_scanner import PublicIPScanner
            
            analyzer = SecurityComplianceAnalyzer()
            exposed = analyzer.detect_public_exposure(all_resources)
            stopped_instances = analyzer.check_stopped_instances(instances)
            preemptible_instances = analyzer.check_preemptible_instances(instances)
            
            # è·å–æœ€è¿‘ä¸€æ¬¡æ‰«æçš„é«˜é£é™©é¡¹
            last_scan = PublicIPScanner.load_last_results()
            high_risk_count = last_scan.get("summary", {}).get("high_risk_count", 0) if last_scan else 0
            
            # æ±‡æ€»å‘Šè­¦æ•°: å…¬ç½‘æš´éœ² + é•¿æœŸåœæ­¢ + æŠ¢å å¼å®ä¾‹ + æ‰«ææ¼æ´ (ä¸ security/overview ä¿æŒä¸€è‡´)
            alert_count = len(exposed) + len(stopped_instances) + len(preemptible_instances) + high_risk_count
            logger.info(f"å®‰å…¨å‘Šè­¦è®¡ç®—: å…¬ç½‘æš´éœ²={len(exposed)}, é•¿æœŸåœæ­¢={len(stopped_instances)}, æŠ¢å å¼={len(preemptible_instances)}, æ‰«ææ¼æ´={high_risk_count}, æ€»è®¡={alert_count}")
        except Exception as e:
            logger.warning(f"è®¡ç®—å®‰å…¨å‘Šè­¦å¤±è´¥: {str(e)}")
            alert_count = 0
            stopped_instances = []
        
        # --- [NEW] ä¸­é—´ä¿å­˜: èµ„æºç»Ÿè®¡å’Œæ ‡ç­¾è¦†ç›–ç‡å·²å®Œæˆ,å…ˆå­˜å…¥ç¼“å­˜ ---
        try:
            interim_result = {
                "account": account,
                "total_cost": total_cost or 0.0,
                "idle_count": len(idle_data) if idle_data else 0,
                "cost_trend": "è®¡ç®—ä¸­...",
                "trend_pct": 0.0,
                "total_resources": total_resources,
                "resource_breakdown": resource_breakdown,
                "alert_count": alert_count,
                "tag_coverage": tag_coverage,
                "savings_potential": 0.0, # åˆå§‹ä¸º0ï¼Œåç»­è®¡ç®—
                "loading": True  # æ ‡è®°ä¸ºè¿˜åœ¨åŠ è½½è¯¦æƒ…
            }
            cache_manager.set(resource_type="dashboard_summary", account_name=account, data=interim_result)
            logger.info(f"âœ… å·²ä¿å­˜ä¸­é—´ç¼“å­˜ (èµ„æºç»Ÿè®¡å·²å®Œæˆ)")
        except Exception as e:
            logger.warning(f"ä¿å­˜ä¸­é—´ç¼“å­˜å¤±è´¥: {str(e)}")
        # -------------------------------------------------------------
        # Savings Potential: ä¼˜å…ˆä½¿ç”¨ optimization_suggestions ç¼“å­˜çš„æ•°æ®ï¼Œä¿è¯ä¸å‰ç«¯æ˜¾ç¤ºä¸€è‡´
        savings_potential = 0.0
        
        # å…ˆå°è¯•ä» optimization_suggestions ç¼“å­˜è·å–ï¼ˆä¸ /api/optimization/suggestions ä¿æŒä¸€è‡´ï¼‰
        opt_cache = cache_manager.get(resource_type="optimization_suggestions", account_name=account)
        if opt_cache and isinstance(opt_cache, dict):
            cached_savings = opt_cache.get("summary", {}).get("total_savings_potential", 0)
            if cached_savings and cached_savings > 0:
                savings_potential = float(cached_savings)
                logger.info(f"ä¼˜åŒ–æ½œåŠ›: ä½¿ç”¨ optimization_suggestions ç¼“å­˜æ•°æ® = {savings_potential:.2f}")
        
        # å¦‚æœç¼“å­˜æ²¡æœ‰æ•°æ®ï¼Œåˆ™è‡ªè¡Œè®¡ç®—
        if savings_potential == 0.0 and idle_data and account_config:
            # Get cost map for ECS resources (idle_data typically contains ECS instances)
            cost_map = _get_cost_map("ecs", account_config)
            
            # Calculate total cost of idle resources
            for idle_item in idle_data:
                instance_id = idle_item.get("instance_id") or idle_item.get("id")
                if instance_id:
                    # Try to get real cost from cost_map
                    cost = cost_map.get(instance_id)
                    if cost is None:
                        # If not found, try to estimate from resource spec
                        spec = idle_item.get("spec", "")
                        if spec:
                            cost = _estimate_monthly_cost_from_spec(spec, "ecs")
                        else:
                            # Default fallback estimate
                            cost = 300  # Average ECS cost
                    savings_potential += cost
            
            # æ–°å¢ï¼šç´¯åŠ åœæ­¢å®ä¾‹çš„æ½œåœ¨èŠ‚çœï¼ˆå‡è®¾èŠ‚çœ70%çš„ç£ç›˜/é¢„ç•™ç­‰è´¹ç”¨ï¼‰
            if stopped_instances:
                for stop_item in stopped_instances:
                    instance_id = stop_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id) or 300
                        savings_potential += (cost * 0.7)
            
            logger.info(f"ä¼˜åŒ–æ½œåŠ›è®¡ç®—: é—²ç½®èµ„æºè´¡çŒ®={sum(item.get('savings', 0) for item in idle_data)}, åœæ­¢å®ä¾‹è´¡çŒ®={len(stopped_instances) * 210} (æ¨ä¼°), æ€»è®¡={savings_potential:.2f}")
            
            # Ensure savings potential doesn't exceed total cost
            if total_cost is not None:
                savings_potential = min(savings_potential, float(total_cost) * 0.95)  # Cap at 95% of total cost

        # å¦‚æœæˆæœ¬è¶‹åŠ¿æ²¡æœ‰å†å²æ•°æ®ï¼Œåˆ™ç”¨"å½“å‰èµ„æºæœˆåº¦æˆæœ¬ï¼ˆæŠ˜åä¼˜å…ˆï¼‰"ä½œä¸ºç»Ÿä¸€å£å¾„çš„ total_cost
        if total_cost is None and account_config:
            ecs_cost_map = _get_cost_map("ecs", account_config)
            rds_cost_map = _get_cost_map("rds", account_config)
            redis_cost_map = _get_cost_map("redis", account_config)

            estimated_total = 0.0
            for inst in instances:
                cost = ecs_cost_map.get(inst.id)
                if cost is None:
                    cost = _estimate_monthly_cost(inst)
                estimated_total += float(cost or 0)
            for rds in rds_list:
                cost = rds_cost_map.get(rds.id)
                if cost is None:
                    cost = _estimate_monthly_cost(rds)
                estimated_total += float(cost or 0)
            for r in redis_list:
                cost = redis_cost_map.get(r.id)
                if cost is None:
                    cost = _estimate_monthly_cost(r)
                estimated_total += float(cost or 0)

            total_cost = round(float(estimated_total), 2)
            # å†åšä¸€æ¬¡ savings capï¼ˆæ­¤æ—¶ total_cost å·²å¯ç”¨ï¼‰
            savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0

        # ç”¨è´¦å•å…¨é‡å£å¾„è¦†ç›– total_costï¼ˆæ›´è´´è¿‘çœŸå®è´¦å•ï¼‰
        if billing_total_cost is not None:
            total_cost = round(float(billing_total_cost), 2)
            savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0
        
        print(f"DEBUG: Alert Count Calculated: {alert_count}")
        logger.info(f"æŒ‡æ ‡å¯¹é½è®¡ç®—: æ ‡ç­¾è¦†ç›–ç‡={tag_coverage}%, å‘Šè­¦æ•°={alert_count} (æš´éœ²={len(exposed)}, åœæ­¢={len(stopped_instances)}, æŠ¢å ={len(preemptible_instances)}, æ¼æ´={high_risk_count})")
    except Exception as e:
        import traceback
        print(f"DEBUG: CRITICAL FAILURE in Analysis: {e}")
        traceback.print_exc()
        logger.warning(f"å¯¹é½è®¡ç®—å¤±è´¥: {str(e)}")
        tag_coverage = 0.0
        alert_count = 0
        stopped_instances = []
        preemptible_instances = []

    # --- [NEW] ä¸­é—´ä¿å­˜: èµ„æºç»Ÿè®¡å’Œæ ‡ç­¾è¦†ç›–ç‡å·²å®Œæˆ,å…ˆå­˜å…¥ç¼“å­˜ ---
    try:
        interim_result = {
            "account": account,
            "total_cost": total_cost or 0.0,
            "idle_count": len(idle_data) if idle_data else 0,
            "cost_trend": "è®¡ç®—ä¸­...",
            "trend_pct": 0.0,
            "total_resources": total_resources,
            "resource_breakdown": resource_breakdown,
            "alert_count": alert_count,
            "tag_coverage": tag_coverage,
            "savings_potential": 0.0, # åˆå§‹ä¸º0ï¼Œåç»­è®¡ç®—
            "loading": True  # æ ‡è®°ä¸ºè¿˜åœ¨åŠ è½½è¯¦æƒ…
        }
        cache_manager.set(resource_type="dashboard_summary", account_name=account, data=interim_result)
        logger.info(f"âœ… å·²ä¿å­˜ä¸­é—´ç¼“å­˜ (èµ„æºç»Ÿè®¡å·²å®Œæˆ)")
    except Exception as e:
        logger.warning(f"ä¿å­˜ä¸­é—´ç¼“å­˜å¤±è´¥: {str(e)}")
    # -------------------------------------------------------------
    # Savings Potential: Calculate based on actual cost of idle resources
    savings_potential = 0.0
    if idle_data and account_config:
        # Get cost map for ECS resources (idle_data typically contains ECS instances)
        cost_map = _get_cost_map("ecs", account_config)
        
        # Calculate total cost of idle resources
        for idle_item in idle_data:
            instance_id = idle_item.get("instance_id") or idle_item.get("id")
            if instance_id:
                # Try to get real cost from cost_map
                cost = cost_map.get(instance_id)
                if cost is None:
                    # If not found, try to estimate from resource spec
                    spec = idle_item.get("spec", "")
                    if spec:
                        cost = _estimate_monthly_cost_from_spec(spec, "ecs")
                    else:
                        # Default fallback estimate
                        cost = 300  # Average ECS cost
                savings_potential += cost
        
        # æ–°å¢ï¼šç´¯åŠ åœæ­¢å®ä¾‹çš„æ½œåœ¨èŠ‚çœï¼ˆå‡è®¾èŠ‚çœ70%çš„ç£ç›˜/é¢„ç•™ç­‰è´¹ç”¨ï¼‰
        if stopped_instances:
            for stop_item in stopped_instances:
                instance_id = stop_item.get("id")
                if instance_id:
                    cost = cost_map.get(instance_id) or 300
                    savings_potential += (cost * 0.7)
        
        logger.info(f"ä¼˜åŒ–æ½œåŠ›è®¡ç®—: é—²ç½®èµ„æºè´¡çŒ®={sum(item.get('savings', 0) for item in idle_data)}, åœæ­¢å®ä¾‹è´¡çŒ®={len(stopped_instances) * 210} (æ¨ä¼°), æ€»è®¡={savings_potential:.2f}")
        
        # Ensure savings potential doesn't exceed total cost
        if total_cost is not None:
            savings_potential = min(savings_potential, float(total_cost) * 0.95)  # Cap at 95% of total cost

    # å¦‚æœæˆæœ¬è¶‹åŠ¿æ²¡æœ‰å†å²æ•°æ®ï¼Œåˆ™ç”¨"å½“å‰èµ„æºæœˆåº¦æˆæœ¬ï¼ˆæŠ˜åä¼˜å…ˆï¼‰"ä½œä¸ºç»Ÿä¸€å£å¾„çš„ total_cost
    if total_cost is None and account_config:
        ecs_cost_map = _get_cost_map("ecs", account_config)
        rds_cost_map = _get_cost_map("rds", account_config)
        redis_cost_map = _get_cost_map("redis", account_config)

        estimated_total = 0.0
        for inst in instances:
            cost = ecs_cost_map.get(inst.id)
            if cost is None:
                cost = _estimate_monthly_cost(inst)
            estimated_total += float(cost or 0)
        for rds in rds_list:
            cost = rds_cost_map.get(rds.id)
            if cost is None:
                cost = _estimate_monthly_cost(rds)
            estimated_total += float(cost or 0)
        for r in redis_list:
            cost = redis_cost_map.get(r.id)
            if cost is None:
                cost = _estimate_monthly_cost(r)
            estimated_total += float(cost or 0)

        total_cost = round(float(estimated_total), 2)
        # å†åšä¸€æ¬¡ savings capï¼ˆæ­¤æ—¶ total_cost å·²å¯ç”¨ï¼‰
        savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0

    # ç”¨è´¦å•å…¨é‡å£å¾„è¦†ç›– total_costï¼ˆæ›´è´´è¿‘çœŸå®è´¦å•ï¼‰
    if billing_total_cost is not None:
        total_cost = round(float(billing_total_cost), 2)
        savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0
    
    # ç¡®ä¿æ‰€æœ‰å¿…éœ€å­—æ®µéƒ½æœ‰å€¼ï¼ˆæœ€ç»ˆæ£€æŸ¥ï¼‰
    if total_cost is None:
        total_cost = 0.0
    if trend is None or trend == "":
        trend = "æ•°æ®ä¸è¶³"
    if trend_pct is None:
        trend_pct = 0.0
    if idle_count is None:
        idle_count = 0
    if total_resources is None:
        total_resources = 0
    if resource_breakdown is None:
        resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
    if tag_coverage is None:
        tag_coverage = 0.0
    if alert_count is None:
        alert_count = 0
    if savings_potential is None:
        savings_potential = 0.0

    result_data = {
        "account": str(account),
        "total_cost": float(total_cost),
        "idle_count": int(idle_count),
        "cost_trend": str(trend),
        "trend_pct": float(trend_pct),
        "total_resources": int(total_resources),
        "resource_breakdown": dict(resource_breakdown),
        "alert_count": int(alert_count),
        "tag_coverage": round(float(tag_coverage), 2),
        "savings_potential": float(savings_potential),
    }
    
    logger.info(f"âœ… Dashboard summary æ•°æ®å‡†å¤‡å®Œæˆ: account={account}, total_cost={result_data['total_cost']}, idle_count={result_data['idle_count']}, total_resources={result_data['total_resources']}")
    
    # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
    try:
        cache_manager = CacheManager(ttl_seconds=86400)
        cache_manager.set(resource_type="dashboard_summary", account_name=account, data=result_data)
        logger.info(f"âœ… ç¼“å­˜å·²ä¿å­˜: {account}")
        return result_data
    except Exception as e:
        logger.warning(f"âš ï¸ ä¿å­˜ç¼“å­˜å¤±è´¥: {str(e)}")
        return result_data


def _get_global_resources_consistent(account_name: str, account_config, force_refresh: bool = False):
    """
    ç»Ÿä¸€çš„å…¨å±€èµ„æºè·å–å‡½æ•°ï¼Œç¡®ä¿æ‰€æœ‰ Overview æ¥å£çœ‹åˆ°çš„æ˜¯åŒä¸€ä»½æ•°æ®ã€‚
    æ”¯æŒ ECS, RDS, Redis å…¨åœ°åŸŸæ‰«æ + è´¦æˆ·çº§ç¼“å­˜ã€‚
    """
    cache_manager = CacheManager(ttl_seconds=3600)  # 1å°æ—¶ç¼“å­˜
    cache_key = f"resource_list_{account_name}"
    
    if not force_refresh:
        cached_all = cache_manager.get(resource_type=cache_key, account_name=account_name)
        if cached_all:
            logger.info(f"â™»ï¸ ä½¿ç”¨å…¨å±€ä¸€è‡´æ€§èµ„æºç¼“å­˜: {account_name}")
            return cached_all.get("instances", []), cached_all.get("rds", []), cached_all.get("redis", [])

    logger.info(f"ğŸ” å¼€å§‹å…¨çƒèµ„æºæ‰«æ (è¶Šè¿‡ç¼“å­˜: {force_refresh}) - è´¦å·: {account_name}")
    from cloudlens.core.providers.aliyun_provider import AliyunProvider
    
    # 1. æ‰«æ ECS
    all_regions = AliyunProvider.get_all_regions()
    
    def fetch_ecs():
        res = []
        for r in all_regions:
            try:
                p = AliyunProvider(account_config.name, account_config.access_key_id, account_config.access_key_secret, region=r)
                count = p.check_instances_count()
                if count > 0:
                    region_instances = p.list_instances()
                    res.extend(region_instances)
                    logger.info(f"  [Scan ECS] åœ°åŸŸ {r}: å‘ç° {len(region_instances)} ä¸ªå®ä¾‹")
            except Exception as e:
                logger.debug(f"  [Scan ECS] åœ°åŸŸ {r} å¤±è´¥: {e}")
        return res

    # 2. æ‰«æ RDS/Redis (è¿™é‡Œç®€åŒ–ä¸ºå¤šçº¿ç¨‹æˆ–å¤šåœ°åŸŸ)
    def fetch_rds():
        res = []
        for r in all_regions:
            try:
                p = AliyunProvider(account_config.name, account_config.access_key_id, account_config.access_key_secret, region=r)
                rds_list = p.list_rds()
                if rds_list:
                    res.extend(rds_list)
                    logger.info(f"  [Scan RDS] åœ°åŸŸ {r}: å‘ç° {len(rds_list)} ä¸ªå®ä¾‹")
            except: pass
        return res

    def fetch_redis():
        res = []
        for r in all_regions:
            try:
                p = AliyunProvider(account_config.name, account_config.access_key_id, account_config.access_key_secret, region=r)
                redis_list = p.list_redis()
                if redis_list:
                    res.extend(redis_list)
                    logger.info(f"  [Scan Redis] åœ°åŸŸ {r}: å‘ç° {len(redis_list)} ä¸ªå®ä¾‹")
            except: pass
        return res

    # ä¸ºäº†é€Ÿåº¦ï¼Œä½¿ç”¨çº¿ç¨‹æ± 
    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
        f_ecs = executor.submit(fetch_ecs)
        f_rds = executor.submit(fetch_rds)
        f_redis = executor.submit(fetch_redis)
        
        instances = f_ecs.result()
        rds_list = f_rds.result()
        redis_list = f_redis.result()

    logger.info(f"âœ… å…¨çƒæ‰«æå®Œæˆ: ECS={len(instances)}, RDS={len(rds_list)}, Redis={len(redis_list)}")

    # åºåˆ—åŒ–ä¸º dict ä»¥ä¾¿ç¼“å­˜ï¼ˆä½¿ç”¨å¢å¼ºåçš„ to_dictï¼‰
    data_to_cache = {
        "instances": [i.to_dict() if hasattr(i, 'to_dict') else i for i in instances],
        "rds": [i.to_dict() if hasattr(i, 'to_dict') else i for i in rds_list],
        "redis": [i.to_dict() if hasattr(i, 'to_dict') else i for i in redis_list]
    }
    cache_manager.set(resource_type=cache_key, account_name=account_name, data=data_to_cache)
    
    return instances, rds_list, redis_list


@router.get("/dashboard/trend")
@api_error_handler
async def get_trend(
    account: Optional[str] = None, 
    days: int = 30, 
    start_date: Optional[str] = Query(None, description="å¼€å§‹æ—¥æœŸ YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="ç»“æŸæ—¥æœŸ YYYY-MM-DD"),
    granularity: Optional[str] = Query("daily", description="æ•°æ®ç²’åº¦: daily(æŒ‰å¤©) æˆ– monthly(æŒ‰æœˆ)"),
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜")
):
    """Get cost trend chart dataï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰
    
    Args:
        account: è´¦å·åç§°
        days: æŸ¥è¯¢å¤©æ•°ï¼Œ0è¡¨ç¤ºè·å–æ‰€æœ‰å†å²æ•°æ®ï¼ˆå½“start_dateå’Œend_dateéƒ½æä¾›æ—¶ï¼Œæ­¤å‚æ•°è¢«å¿½ç•¥ï¼‰
        start_date: å¼€å§‹æ—¥æœŸ YYYY-MM-DDæ ¼å¼
        end_date: ç»“æŸæ—¥æœŸ YYYY-MM-DDæ ¼å¼
        granularity: æ•°æ®ç²’åº¦ï¼Œdaily(æŒ‰å¤©) æˆ– monthly(æŒ‰æœˆ)ï¼Œé»˜è®¤daily
        force_refresh: æ˜¯å¦å¼ºåˆ¶åˆ·æ–°ç¼“å­˜
    """
    if not account:
        raise HTTPException(status_code=400, detail="è´¦å·å‚æ•°æ˜¯å¿…éœ€çš„")
    
    # éªŒè¯granularityå‚æ•°
    if granularity not in ["daily", "monthly"]:
        granularity = "daily"
    
    # å¦‚æœæä¾›äº†æ—¥æœŸèŒƒå›´ï¼Œä½¿ç”¨æ—¥æœŸèŒƒå›´ï¼›å¦åˆ™ä½¿ç”¨dayså‚æ•°
    if start_date and end_date:
        logger.debug(f"æ”¶åˆ°è´¦å·å‚æ•°: {account}, æ—¥æœŸèŒƒå›´: {start_date} è‡³ {end_date}, ç²’åº¦: {granularity}")
        cache_key = f"dashboard_trend_{granularity}_{start_date}_{end_date}"
    else:
        logger.debug(f"æ”¶åˆ°è´¦å·å‚æ•°: {account}, days: {days} ({'å…¨éƒ¨å†å²' if days == 0 else f'æœ€è¿‘{days}å¤©'}), ç²’åº¦: {granularity}")
        cache_key = f"dashboard_trend_{granularity}_{days}"
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type=cache_key, account_name=account)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        return {
            **cached_result,
            "cached": True,
        }
    
    analyzer = CostTrendAnalyzer()
    try:
        # å¦‚æœæä¾›äº†æ—¥æœŸèŒƒå›´ï¼Œè®¡ç®—dayså‚æ•°ï¼›å¦åˆ™ä½¿ç”¨ä¼ å…¥çš„days
        if start_date and end_date:
            from datetime import datetime
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            calculated_days = (end - start).days
            # ä½¿ç”¨æ—¥æœŸèŒƒå›´ç”ŸæˆæŠ¥å‘Š
            report = analyzer.generate_trend_report(account, calculated_days, start_date=start_date, end_date=end_date)
        else:
            # å¦‚æœgranularityæ˜¯monthlyä¸”days=0ï¼Œç›´æ¥æŒ‰è´¦æœŸæŸ¥è¯¢å¹¶èšåˆ
            if granularity == "monthly" and days == 0:
                # ç›´æ¥ä»æ•°æ®åº“æŒ‰è´¦æœŸæŸ¥è¯¢æœˆåº¦æ•°æ®
                from cloudlens.core.bill_storage import BillStorageManager
                from cloudlens.core.config import ConfigManager
                storage = BillStorageManager()
                db = storage.db
                cm = ConfigManager()
                account_config = cm.get_account(account)
                if not account_config:
                    raise HTTPException(status_code=404, detail=f"è´¦å· '{account}' ä¸å­˜åœ¨")
                
                account_id = f"{account_config.access_key_id[:10]}-{account}"
                
                # æŸ¥è¯¢æ‰€æœ‰è´¦æœŸçš„æœˆåº¦æˆæœ¬
                rows = db.query("""
                    SELECT
                        billing_cycle,
                        SUM(pretax_amount) as monthly_cost
                    FROM bill_items
                    WHERE account_id = ?
                        AND billing_cycle IS NOT NULL
                        AND billing_cycle != ''
                        AND pretax_amount IS NOT NULL
                    GROUP BY billing_cycle
                    ORDER BY billing_cycle ASC
                """, (account_id,))
                
                if not rows:
                    return {
                        "account": account,
                        "period_days": 0,
                        "granularity": granularity,
                        "analysis": {"error": "No cost history available"},
                        "chart_data": None,
                        "cost_by_type": {},
                        "cost_by_region": {},
                        "snapshots_count": 0,
                        "summary": None,
                        "cached": False,
                    }
                
                # è½¬æ¢ä¸ºå›¾è¡¨æ•°æ®æ ¼å¼
                monthly_chart_data = []
                for row in rows:
                    billing_cycle = row['billing_cycle'] if isinstance(row, dict) else row[0]
                    monthly_cost = float(row['monthly_cost'] or 0) if isinstance(row, dict) else float(row[1] or 0)
                    # ä½¿ç”¨æœˆåˆæ—¥æœŸæ ¼å¼ YYYY-MM-01
                    year, month = map(int, billing_cycle.split('-'))
                    date_str = f"{year}-{month:02d}-01"
                    monthly_chart_data.append({
                        "date": date_str,
                        "total_cost": monthly_cost,
                        "breakdown": {}
                    })
                
                logger.info(f"âœ… ç”Ÿæˆ {len(monthly_chart_data)} æ¡æœˆåº¦æ•°æ®ï¼Œæ—¥æœŸèŒƒå›´: {monthly_chart_data[0]['date']} è‡³ {monthly_chart_data[-1]['date']}")
                
                # è®¡ç®—ç»Ÿè®¡ä¿¡æ¯
                monthly_costs = [item["total_cost"] for item in monthly_chart_data]
                summary = {
                    "total_cost": sum(monthly_costs) if monthly_costs else 0,
                    "avg_monthly_cost": sum(monthly_costs) / len(monthly_costs) if monthly_costs else 0,
                    "max_monthly_cost": max(monthly_costs) if monthly_costs else 0,
                    "min_monthly_cost": min(monthly_costs) if monthly_costs else 0,
                    "trend": "å¹³ç¨³",
                    "trend_pct": 0.0
                }
                
                # å¦‚æœæœ‰å¤šä¸ªæœˆä»½ï¼Œè®¡ç®—è¶‹åŠ¿
                if len(monthly_costs) >= 2:
                    latest = monthly_costs[-1]
                    previous = monthly_costs[-2]
                    if previous > 0:
                        trend_pct = ((latest - previous) / previous) * 100
                        summary["trend_pct"] = round(trend_pct, 2)
                        if trend_pct > 5:
                            summary["trend"] = "ä¸Šå‡"
                        elif trend_pct < -5:
                            summary["trend"] = "ä¸‹é™"
                
                result = {
                    "account": account,
                    "period_days": 0,
                    "granularity": granularity,
                    "chart_data": monthly_chart_data,
                    "summary": summary,
                    "cost_by_type": {},
                    "cost_by_region": {},
                    "snapshots_count": len(monthly_chart_data),
                    "cached": False,
                }
                
                # ä¿å­˜åˆ°ç¼“å­˜
                cache_manager.set(resource_type=cache_key, account_name=account, data=result)
                return result
            else:
                report = analyzer.generate_trend_report(account, days)
        
        if "error" in report:
            # è¶‹åŠ¿å›¾å¸¸è§çš„"æ— æ•°æ®/æ•°æ®ä¸è¶³"ä¸åº”è¯¥ä½œä¸ºæœåŠ¡ç«¯é”™è¯¯ï¼›
            # è¿”å› 200 + ç©º chart_dataï¼Œå‰ç«¯å¯è‡ªç„¶é™çº§ä¸º"ä¸å±•ç¤ºè¶‹åŠ¿å›¾"ã€‚
            err = report.get("error") or "No trend data"
            if err in ("No cost history available", "Insufficient data for trend analysis"):
                return {
                    "account": account,
                    "period_days": days,
                    "granularity": granularity,
                    "analysis": {"error": err},
                    "chart_data": None,
                    "cost_by_type": {},
                    "cost_by_region": {},
                    "snapshots_count": 0,
                    "summary": None,
                    "cached": False,
                }
            raise HTTPException(status_code=404, detail=err)
        
        # è½¬æ¢æ•°æ®æ ¼å¼ä»¥æ”¯æŒæ–°çš„å‰ç«¯éœ€æ±‚
        chart_data = report.get("chart_data", {})
        dates = chart_data.get("dates", [])
        costs = chart_data.get("costs", [])
        
        # æ ¹æ®granularityè½¬æ¢æ•°æ®æ ¼å¼
        if granularity == "monthly":
            # æŒ‰æœˆèšåˆæ•°æ®
            monthly_data = {}
            for i, date in enumerate(dates):
                # æå–æœˆä»½ï¼ˆYYYY-MMæ ¼å¼ï¼‰
                if len(date) >= 7:
                    month_key = date[:7]  # YYYY-MM
                elif len(date) == 10:
                    month_key = date[:7]  # YYYY-MM-DD -> YYYY-MM
                else:
                    # å¦‚æœæ—¥æœŸæ ¼å¼ä¸å¯¹ï¼Œå°è¯•è§£æ
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(date, "%Y-%m-%d")
                        month_key = date_obj.strftime("%Y-%m")
                    except:
                        month_key = date[:7] if len(date) >= 7 else date
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        "date": f"{month_key}-01",  # ä½¿ç”¨æœˆåˆæ—¥æœŸæ ¼å¼ YYYY-MM-01
                        "total_cost": 0,
                        "breakdown": {}
                    }
                monthly_data[month_key]["total_cost"] += costs[i] if i < len(costs) else 0
                
                # å¦‚æœæœ‰èµ„æºç±»å‹åˆ†è§£ï¼Œä¹Ÿèšåˆ
                cost_by_type = report.get("cost_by_type", {})
                if cost_by_type and isinstance(cost_by_type, dict):
                    for type_name, type_cost in cost_by_type.items():
                        if type_name not in monthly_data[month_key]["breakdown"]:
                            monthly_data[month_key]["breakdown"][type_name] = 0
                        # ç®€åŒ–ï¼šå¹³å‡åˆ†é…ï¼ˆå®é™…åº”è¯¥æŒ‰æ—¥æœŸè®¡ç®—ï¼‰
                        monthly_data[month_key]["breakdown"][type_name] += type_cost / len(dates) if dates else 0
            
            # è½¬æ¢ä¸ºåˆ—è¡¨æ ¼å¼ï¼Œå¹¶æŒ‰æ—¥æœŸæ’åºï¼ˆä»æ—©åˆ°æ™šï¼‰
            chart_data_list = sorted(monthly_data.values(), key=lambda x: x["date"])
            
            # è®¡ç®—æœˆåº¦ç»Ÿè®¡
            monthly_costs = [item["total_cost"] for item in chart_data_list]
            summary = {
                "total_cost": sum(monthly_costs) if monthly_costs else 0,
                "avg_monthly_cost": sum(monthly_costs) / len(monthly_costs) if monthly_costs else 0,
                "max_monthly_cost": max(monthly_costs) if monthly_costs else 0,
                "min_monthly_cost": min(monthly_costs) if monthly_costs else 0,
                "trend": report.get("analysis", {}).get("trend", "å¹³ç¨³"),
                "trend_pct": report.get("analysis", {}).get("total_change_pct", 0.0)
            }
        else:
            # æŒ‰å¤©æ•°æ®ï¼ˆä¿æŒåŸæœ‰æ ¼å¼ï¼Œä½†è½¬æ¢ä¸ºæ–°æ ¼å¼ï¼‰
            chart_data_list = []
            for i, date in enumerate(dates):
                item = {
                    "date": date,
                    "total_cost": costs[i] if i < len(costs) else 0,
                    "breakdown": {}
                }
                
                # å¦‚æœæœ‰èµ„æºç±»å‹åˆ†è§£ï¼Œæ·»åŠ åˆ°breakdown
                cost_by_type = report.get("cost_by_type", {})
                if cost_by_type and isinstance(cost_by_type, dict):
                    # ç®€åŒ–ï¼šå¹³å‡åˆ†é…ï¼ˆå®é™…åº”è¯¥ä»å¿«ç…§æ•°æ®ä¸­æå–ï¼‰
                    for type_name, type_cost in cost_by_type.items():
                        item["breakdown"][type_name] = type_cost / len(dates) if dates else 0
                
                chart_data_list.append(item)
            
            # è®¡ç®—æ—¥ç»Ÿè®¡
            daily_costs = [item["total_cost"] for item in chart_data_list]
            summary = {
                "total_cost": sum(daily_costs) if daily_costs else 0,
                "avg_daily_cost": sum(daily_costs) / len(daily_costs) if daily_costs else 0,
                "max_daily_cost": max(daily_costs) if daily_costs else 0,
                "min_daily_cost": min(daily_costs) if daily_costs else 0,
                "trend": report.get("analysis", {}).get("trend", "å¹³ç¨³"),
                "trend_pct": report.get("analysis", {}).get("total_change_pct", 0.0)
            }
        
        # æ„å»ºæ–°çš„å“åº”æ ¼å¼
        result = {
            "account": account,
            "period_days": days,
            "granularity": granularity,
            "chart_data": chart_data_list,
            "summary": summary,
            "cost_by_type": report.get("cost_by_type", {}),
            "cost_by_region": report.get("cost_by_region", {}),
            "snapshots_count": report.get("snapshots_count", 0),
            "cached": False,
        }
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type=cache_key, account_name=account, data=result)
        
        return result
    except HTTPException:
        # ä¸è¦æŠŠ 4xx å†åŒ…è£…æˆ 500ï¼Œå¦åˆ™å‰ç«¯åªèƒ½çœ‹åˆ° "Internal Server Error"
        raise
    except Exception as e:
        logger.exception(f"è·å–æˆæœ¬è¶‹åŠ¿å¤±è´¥: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard/idle")
def get_idle_resources(account: Optional[str] = None, force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜")):
    """Get idle resources listï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    import logging
    logger = logging.getLogger(__name__)
    
    if not account:
        raise HTTPException(status_code=400, detail="è´¦å·å‚æ•°æ˜¯å¿…éœ€çš„")
    
    logger.info(f"[get_idle_resources] æ”¶åˆ°è¯·æ±‚: account={account}, force_refresh={bool(force_refresh)}")
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®ï¼ˆä¼˜å…ˆä½¿ç”¨ç¼“å­˜ï¼Œé¿å…è€—æ—¶æ“ä½œï¼‰
    cached_result = None
    try:
        # ä¼˜å…ˆä» dashboard_idle ç¼“å­˜è·å–
        cached_result = cache_manager.get(resource_type="dashboard_idle", account_name=account)
        if cached_result:
            logger.info(f"[get_idle_resources] âœ… ä» dashboard_idle ç¼“å­˜è¿”å›: {len(cached_result) if isinstance(cached_result, list) else 'N/A'} æ¡æ•°æ®")
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
        
        # å¦‚æœ dashboard_idle ç¼“å­˜ä¸ºç©ºï¼Œå°è¯•ä» idle_result ç¼“å­˜è·å–ï¼ˆä¸ summary é€»è¾‘ä¿æŒä¸€è‡´ï¼‰
        cached_result = cache_manager.get(resource_type="idle_result", account_name=account)
        if cached_result:
            logger.info(f"[get_idle_resources] âœ… ä» idle_result ç¼“å­˜è¿”å›: {len(cached_result) if isinstance(cached_result, list) else 'N/A'} æ¡æ•°æ®")
            # åŒæ—¶æ›´æ–° dashboard_idle ç¼“å­˜ï¼Œä¿æŒä¸€è‡´æ€§
            try:
                cache_manager.set(resource_type="dashboard_idle", account_name=account, data=cached_result)
            except Exception as e:
                logger.warning(f"æ›´æ–° dashboard_idle ç¼“å­˜å¤±è´¥: {e}")
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
    except Exception as e:
        logger.warning(f"[get_idle_resources] è¯»å–ç¼“å­˜å¤±è´¥: {e}")
        import traceback
        logger.error(f"ç¼“å­˜è¯»å–å¼‚å¸¸è¯¦æƒ…: {traceback.format_exc()}")
    
    # å¦‚æœç¼“å­˜ä¸ºç©ºï¼Œç›´æ¥è¿”å›ç©ºæ•°æ®ï¼ˆä¸è¿›è¡Œè€—æ—¶åˆ†æï¼Œé¿å…é˜»å¡ï¼‰
    # ç”¨æˆ·å¿…é¡»é€šè¿‡ /api/analyze/trigger ä¸»åŠ¨è§¦å‘åˆ†æ
    logger.info(f"[get_idle_resources] ç¼“å­˜ä¸ºç©ºï¼Œè¿”å›ç©ºæ•°æ®ï¼ˆä¸è¿›è¡Œè€—æ—¶åˆ†æï¼‰")
    return {
        "success": True,
        "data": [],
        "cached": False,
        "message": "ç¼“å­˜ä¸ºç©ºï¼Œè¯·ä½¿ç”¨æ‰«ææŒ‰é’®è§¦å‘åˆ†æ"
    }


# ==================== Phase 1 Week 2: Resource Management APIs ====================


# ==================== (Migrated Helpers to api_base.py or specialized modules) ====================



def _get_billing_cycle_default() -> str:
    from datetime import datetime
    return datetime.now().strftime("%Y-%m")


def _get_billing_overview_from_db(
    account_config: CloudAccount,
    billing_cycle: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    ä»æœ¬åœ°è´¦å•æ•°æ®åº“è¯»å–æˆæœ¬æ¦‚è§ˆï¼ˆä¼˜å…ˆä½¿ç”¨ï¼Œé€Ÿåº¦å¿«ï¼‰
    
    Args:
        account_config: è´¦å·é…ç½®å¯¹è±¡
        billing_cycle: è´¦æœŸï¼Œæ ¼å¼ YYYY-MMï¼Œé»˜è®¤å½“å‰æœˆ
    
    Returns:
        æˆæœ¬æ¦‚è§ˆæ•°æ®ï¼Œå¦‚æœæ•°æ®åº“ä¸å­˜åœ¨æˆ–è¯»å–å¤±è´¥åˆ™è¿”å› None
    """
    import os
    from datetime import datetime
    from cloudlens.core.database import DatabaseFactory
    
    # ç»Ÿä¸€ä½¿ç”¨ MySQLï¼Œä¸å†æ”¯æŒ SQLite
    try:
        if billing_cycle is None:
            billing_cycle = datetime.now().strftime("%Y-%m")
        
        db = DatabaseFactory.create_adapter("mysql")
        
        # æ„é€ æ­£ç¡®çš„ account_id æ ¼å¼ï¼š{access_key_id[:10]}-{account_name}
        account_id = f"{account_config.access_key_id[:10]}-{account_config.name}"
        
        # ä¼˜å…ˆå°è¯• prefix åŒ¹é…ï¼Œå› ä¸º LTAI5tECY4-ydzn è¿™ç§æ ¼å¼å¦‚æœ AK å‰ç¼€å˜åŒ–ä¼šæ— æ³•åŒ¹é…
        # å®é™… bill_items è¡¨ä¸­ä½¿ç”¨çš„æ˜¯æŠ“å–æ—¶æ„é€ çš„ account_id
        
        # éªŒè¯ account_id æ˜¯å¦å­˜åœ¨ï¼ˆç²¾ç¡®åŒ¹é…ï¼‰
        account_result = db.query_one("""
            SELECT DISTINCT account_id 
            FROM bill_items 
            WHERE account_id = %s OR account_id LIKE %s
            ORDER BY (account_id = %s) DESC
            LIMIT 1
        """, (account_id, f"%-{account_config.name}", account_id))
        
        if not account_result:
            # å¦‚æœç²¾ç¡®åŒ¹é…å¤±è´¥ï¼Œå°è¯•æ¨¡ç³ŠåŒ¹é…ï¼ˆå…¼å®¹æ—§æ•°æ®ï¼‰
            logger.warning(f"ç²¾ç¡®åŒ¹é…å¤±è´¥ï¼Œå°è¯•æ¨¡ç³ŠåŒ¹é…: {account_id}")
            account_result = db.query_one("""
                SELECT DISTINCT account_id 
                FROM bill_items 
                WHERE account_id LIKE %s
                LIMIT 1
            """, (f"%{account_config.name}%",))
            
            if not account_result:
                logger.warning(f"æœªæ‰¾åˆ°è´¦å· '{account_config.name}' (account_id: {account_id}) çš„è´¦å•æ•°æ®")
                return None
            
            # å¤„ç†å­—å…¸æ ¼å¼çš„ç»“æœï¼ˆMySQLï¼‰
            if isinstance(account_result, dict):
                matched_account_id = account_result.get('account_id')
            else:
                matched_account_id = account_result[0] if account_result else None
            
            if matched_account_id and matched_account_id != account_id:
                logger.warning(f"ä½¿ç”¨æ¨¡ç³ŠåŒ¹é…çš„ account_id: {matched_account_id} (æœŸæœ›: {account_id})ï¼Œå¯èƒ½å­˜åœ¨æ•°æ®ä¸²å·é£é™©")
                account_id = matched_account_id
        
        # æŒ‰äº§å“èšåˆå½“æœˆæˆæœ¬ï¼ˆMySQL ä½¿ç”¨ %s å ä½ç¬¦ï¼‰
        product_results = db.query("""
            SELECT 
                product_name,
                product_code,
                subscription_type,
                SUM(pretax_amount) as total_pretax
            FROM bill_items
            WHERE account_id = %s
                AND billing_cycle = %s
                AND pretax_amount IS NOT NULL
            GROUP BY product_name, product_code, subscription_type
        """, (account_id, billing_cycle))
        
        by_product: Dict[str, float] = {}
        by_product_name: Dict[str, str] = {}
        by_product_subscription: Dict[str, Dict[str, float]] = {}
        total = 0.0
        
        for row in product_results:
            # å¤„ç†å­—å…¸æ ¼å¼çš„ç»“æœï¼ˆMySQLï¼‰
            if isinstance(row, dict):
                product_name = row.get('product_name') or "unknown"
                product_code = row.get('product_code') or "unknown"
                subscription_type = row.get('subscription_type') or "Unknown"
                pretax = float(row.get('total_pretax') or 0)
            else:
                product_name = row[0] or "unknown"
                product_code = row[1] or "unknown"
                subscription_type = row[2] or "Unknown"
                pretax = float(row[3] or 0)
            
            if pretax <= 0:
                continue
            
            if product_code not in by_product_name:
                by_product_name[product_code] = product_name
            
            by_product[product_code] = by_product.get(product_code, 0.0) + pretax
            by_product_subscription.setdefault(product_code, {})
            by_product_subscription[product_code][subscription_type] = (
                by_product_subscription[product_code].get(subscription_type, 0.0) + pretax
            )
            
            total += pretax
        
        # æ£€æŸ¥æ˜¯å¦æœ‰ä»»ä½•è®°å½•ï¼ˆå³ä½¿æ€»æˆæœ¬ä¸º0ï¼Œä¹Ÿå¯èƒ½æœ‰è®°å½•ï¼‰
        # å¦‚æœæ²¡æœ‰è®°å½•ï¼Œè¿”å›Noneè®©APIæŸ¥è¯¢ï¼›å¦‚æœæœ‰è®°å½•ä½†æ€»æˆæœ¬ä¸º0ï¼Œä¹Ÿè¿”å›æ•°æ®ï¼ˆå¯èƒ½æ˜¯çœŸå®æƒ…å†µï¼‰
        if len(by_product) == 0:
            # æ²¡æœ‰æ‰¾åˆ°ä»»ä½•è®°å½•ï¼Œè¿”å›Noneè®©APIæŸ¥è¯¢
            logger.info(f"æ•°æ®åº“ä¸­æ²¡æœ‰è´¦æœŸ {billing_cycle} çš„æ•°æ®ï¼Œå°†ä½¿ç”¨APIæŸ¥è¯¢")
            return None
        
        return {
            "billing_cycle": billing_cycle,
            "total_pretax": round(total, 2),
            "by_product": {k: round(v, 2) for k, v in by_product.items()},
            "by_product_name": by_product_name,
            "by_product_subscription": {
                code: {k: round(v, 2) for k, v in sub.items()}
                for code, sub in by_product_subscription.items()
            },
            "data_source": "local_db"
        }
        
    except Exception as e:
        logger.error(f"ä»æœ¬åœ°æ•°æ®åº“è¯»å–è´¦å•æ¦‚è§ˆå¤±è´¥: {str(e)}")
        return None



# ==================== (Migrated to api_billing.py / internal helpers) ====================


@router.get("/resources")
def list_resources(
    type: str = Query("ecs", description="èµ„æºç±»å‹"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    account: Optional[str] = None,
    sortBy: Optional[str] = None,
    sortOrder: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filter: Optional[str] = None,
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
):
    """è·å–èµ„æºåˆ—è¡¨ï¼ˆæ”¯æŒåˆ†é¡µã€æ’åºã€ç­›é€‰ï¼Œå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[list_resources] æ”¶åˆ°è´¦å·å‚æ•°: {account}, type: {type}")
    print(f"[DEBUG list_resources] æ”¶åˆ°è´¦å·å‚æ•°: {account}, type: {type}")
    
    provider, account_name = _get_provider_for_account(account)
    logger.info(f"[list_resources] ä½¿ç”¨è´¦å·: {account_name}")
    print(f"[DEBUG list_resources] ä½¿ç”¨è´¦å·: {account_name}")
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    cached_result = None
    if not force_refresh:
        try:
            cached_result = cache_manager.get(resource_type=type, account_name=account_name)
        except Exception as e:
            logger.warning(f"è·å–ç¼“å­˜å¤±è´¥ï¼Œå°†é‡æ–°æŸ¥è¯¢: {e}")
            cached_result = None
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        result = cached_result
    else:
        # ç¼“å­˜æ— æ•ˆæˆ–ä¸å­˜åœ¨ï¼Œä»provideræŸ¥è¯¢
        cm = ConfigManager()
        account_config = cm.get_account(account_name)

        # åˆå§‹åŒ– resources å˜é‡
        resources = []
        
        # æ ¹æ®ç±»å‹è·å–èµ„æºï¼ˆECS æŸ¥è¯¢æ‰€æœ‰åŒºåŸŸï¼‰
        if type == "ecs":
            # ECS æŸ¥è¯¢æ‰€æœ‰åŒºåŸŸï¼Œè€Œä¸æ˜¯åªæŸ¥è¯¢é…ç½®çš„ region
            try:
                from cloudlens.core.services.analysis_service import AnalysisService
                from cloudlens.providers.aliyun.provider import AliyunProvider
                
                logger.info(f"å¼€å§‹æŸ¥è¯¢æ‰€æœ‰åŒºåŸŸçš„ECSå®ä¾‹ï¼Œè´¦å·: {account_name}")
                
                # è·å–æ‰€æœ‰åŒºåŸŸ
                all_regions = AnalysisService._get_all_regions(
                    account_config.access_key_id,
                    account_config.access_key_secret
                )
                logger.info(f"è·å–åˆ° {len(all_regions)} ä¸ªå¯ç”¨åŒºåŸŸ")
                
                all_instances = []
                for idx, region in enumerate(all_regions):
                    try:
                        region_provider = AliyunProvider(
                            account_name=account_config.name,
                            access_key=account_config.access_key_id,
                            secret_key=account_config.access_key_secret,
                            region=region,
                        )
                        # å¿«é€Ÿæ£€æŸ¥æ˜¯å¦æœ‰èµ„æº
                        count = region_provider.check_instances_count()
                        if count > 0:
                            logger.info(f"åŒºåŸŸ {region} ({idx+1}/{len(all_regions)}): æœ‰ {count} ä¸ªå®ä¾‹ï¼Œå¼€å§‹è¯¦ç»†æŸ¥è¯¢...")
                            region_instances = region_provider.list_instances()
                            all_instances.extend(region_instances)
                            logger.info(f"åŒºåŸŸ {region}: å®é™…è·å– {len(region_instances)} ä¸ªECSå®ä¾‹")
                    except Exception as e:
                        logger.warning(f"æŸ¥è¯¢åŒºåŸŸ {region} çš„ECSå®ä¾‹å¤±è´¥: {str(e)}")
                        import traceback
                        logger.debug(f"å¼‚å¸¸è¯¦æƒ…: {traceback.format_exc()}")
                        continue
                
                logger.info(f"æ€»å…±æ‰¾åˆ° {len(all_instances)} ä¸ªECSå®ä¾‹ï¼ˆä» {len(all_regions)} ä¸ªåŒºåŸŸï¼‰")
                resources = all_instances
            except Exception as e:
                logger.error(f"æŸ¥è¯¢æ‰€æœ‰åŒºåŸŸçš„ECSå®ä¾‹å¤±è´¥ï¼Œå›é€€åˆ°å•åŒºåŸŸæŸ¥è¯¢: {str(e)}")
                import traceback
                logger.error(f"å¼‚å¸¸è¯¦æƒ…: {traceback.format_exc()}")
                # å¦‚æœæŸ¥è¯¢æ‰€æœ‰åŒºåŸŸå¤±è´¥ï¼Œå›é€€åˆ°åªæŸ¥è¯¢é…ç½®çš„ region
                try:
                    resources = provider.list_instances()
                    logger.info(f"å•åŒºåŸŸæŸ¥è¯¢ç»“æœ: {len(resources)} ä¸ªå®ä¾‹")
                except Exception as e2:
                    logger.error(f"å•åŒºåŸŸæŸ¥è¯¢ä¹Ÿå¤±è´¥: {str(e2)}")
                    resources = []
        elif type == "rds":
            resources = provider.list_rds()
        elif type == "redis":
            resources = provider.list_redis()
        elif type == "slb":
            resources = provider.list_slb() if hasattr(provider, "list_slb") else []
        elif type == "nat":
            resources = provider.list_nat_gateways() if hasattr(provider, "list_nat_gateways") else []
        elif type == "eip":
            # EIP provider è¿”å› dict åˆ—è¡¨
            resources = provider.list_eip() if hasattr(provider, "list_eip") else (provider.list_eips() if hasattr(provider, "list_eips") else [])
        elif type == "oss":
            # OSS bucket åˆ—è¡¨ï¼ˆå¦‚æœå®‰è£…äº† oss2ï¼‰
            resources = provider.list_oss() if hasattr(provider, "list_oss") else []
        elif type == "disk":
            resources = provider.list_disks() if hasattr(provider, "list_disks") else []
        elif type == "snapshot":
            resources = provider.list_snapshots() if hasattr(provider, "list_snapshots") else []
        elif type == "vpc":
            vpcs = provider.list_vpcs()
            # Convert VPC dict to list format
            resources = []
            for vpc in vpcs:
                from cloudlens.models.resource import UnifiedResource, ResourceType, ResourceStatus
                
                # Get VPC ID and name from dict - check both possible key formats
                vpc_id = vpc.get("id") or vpc.get("VpcId") or ""
                vpc_name = vpc.get("name") or vpc.get("VpcName") or ""
                
                # Log for debugging - è¯¦ç»†è®°å½•VPCæ•°æ®
                logger.info(f"Processing VPC: id={vpc_id}, name={vpc_name}, raw_vpc={vpc}, keys={list(vpc.keys())}")
                
                # If name is empty or just whitespace, use ID as name
                if not vpc_name or not vpc_name.strip():
                    vpc_name = vpc_id if vpc_id else "æœªå‘½åVPC"
                
                # Ensure we have at least an ID
                if not vpc_id:
                    logger.warning(f"VPC missing ID, skipping: {vpc}")
                    continue

                resources.append(
                    UnifiedResource(
                        id=vpc_id,
                        name=vpc_name,
                        resource_type=ResourceType.VPC,
                        status=ResourceStatus.RUNNING,
                        provider=provider.provider_name,
                        region=vpc.get("region") or vpc.get("RegionId", account_name),
                    )
                )
        else:
            raise HTTPException(status_code=400, detail=f"ä¸æ”¯æŒçš„èµ„æºç±»å‹: {type}")

        # æ‰¹é‡è·å–çœŸå®æˆæœ¬æ˜ å°„ï¼ˆæé«˜æ•ˆç‡ï¼‰
        cost_map = {}
        if account_config and type not in ("vpc",):
            cost_map = _get_cost_map(type, account_config)

        # è½¬æ¢ä¸ºç»Ÿä¸€æ ¼å¼ï¼Œä½¿ç”¨çœŸå®æˆæœ¬
        result = []

        # dict èµ„æºï¼ˆEIP/OSS ç­‰ï¼‰
        if resources and isinstance(resources[0], dict):
            # å¿«ç…§ï¼šQueryInstanceBill å¾ˆå¤šæƒ…å†µä¸‹è¿”å› RegionId è€Œä¸æ˜¯ SnapshotIdï¼Œå¯¼è‡´æ— æ³•é€å®ä¾‹å¯¹é½
            # è¿™ç§æƒ…å†µæ”¹ç”¨è´¦å•æ€»é¢æŒ‰å®¹é‡æ¯”ä¾‹åˆ†æ‘Šåˆ°æ¯ä¸ªå¿«ç…§ï¼Œä¿è¯â€œå®ä¾‹çº§ cost ä¹‹å’Œ == è´¦å•å…¨é‡â€
            if account_config and type == "snapshot":
                has_snapshot_keys = any(str(k).startswith("s-") for k in (cost_map or {}).keys())
                if not has_snapshot_keys:
                    cost_map = {}
                try:
                    totals = _get_billing_overview_totals(account_config, force_refresh=force_refresh)
                    product_total = float(((totals or {}).get("by_product") or {}).get("snapshot") or 0.0)
                except Exception:
                    product_total = 0.0

                if product_total > 0:
                    weights = []
                    for r in resources:
                        rid = r.get("id")
                        if not rid:
                            continue
                        w = r.get("size_gb") or 0
                        try:
                            w = float(w)
                        except Exception:
                            w = 0.0
                        weights.append((rid, max(0.0, w)))

                    total_w = sum(w for _, w in weights)
                    if total_w <= 0:
                        # æ²¡æœ‰å®¹é‡ä¿¡æ¯ï¼Œå‡åˆ†
                        n = len(weights)
                        if n > 0:
                            per = product_total / n
                            for rid, _ in weights:
                                cost_map[rid] = per
                    else:
                        for rid, w in weights:
                            cost_map[rid] = product_total * (w / total_w)

            for r in resources:
                rid = r.get("id") or r.get("Id") or r.get("ResourceId") or r.get("name")
                if not rid:
                    continue

                # EIPï¼šid=AllocationIdï¼›OSSï¼šid=bucket name
                cost = cost_map.get(rid, 0.0)

                name = r.get("name") or r.get("ip_address") or r.get("id") or "-"
                spec = "-"
                region_val = r.get("region") or r.get("RegionId") or getattr(provider, "region", "")
                status_val = r.get("status") or r.get("Status") or "-"

                if type == "eip":
                    spec = f"{r.get('bandwidth', '-') }Mbps"
                elif type == "disk":
                    size_gb = r.get("size_gb", "-")
                    cat = r.get("disk_category", "-")
                    dtyp = r.get("disk_type", "-")
                    spec = f"{cat} / {dtyp} / {size_gb}GB"
                elif type == "snapshot":
                    src = r.get("source_disk_id") or "-"
                    size_gb = r.get("size_gb", "-")
                    spec = f"æºç›˜: {src} / {size_gb}GB"

                result.append(
                    {
                        "id": rid,
                        "name": name,
                        "type": type,
                        "status": str(status_val),
                        "region": str(region_val),
                        "spec": str(spec),
                        "cost": float(cost or 0),
                        "tags": {},
                    "created_time": r.get("created_time") if isinstance(r.get("created_time"), str) else None,
                        "public_ips": [r.get("ip_address")] if r.get("ip_address") else [],
                        "private_ips": [],
                        "vpc_id": r.get("vpc_id") or r.get("VpcId") or None,
                    }
                )
        else:
            for r in resources:
                # ä»æˆæœ¬æ˜ å°„ä¸­è·å–çœŸå®æˆæœ¬ï¼Œå¦‚æœæ²¡æœ‰åˆ™ä½¿ç”¨ä¼°ç®—å€¼
                cost = cost_map.get(r.id)
                if cost is None:
                    cost = _estimate_monthly_cost(r)

                # For VPC resources, ensure name is not empty
                display_name = r.name or r.id or "-"
                if type == "vpc" and not r.name:
                    display_name = r.id or "-"
                
                # For VPC resources, vpc_id should be the VPC's own ID
                # For other resources, vpc_id is the associated VPC ID
                if type == "vpc":
                    # VPCèµ„æºæœ¬èº«ï¼Œvpc_idåº”è¯¥æ˜¾ç¤ºä¸ºVPCçš„IDï¼ˆVPCèµ„æºæœ¬èº«æ²¡æœ‰å…³è”çš„VPCï¼Œæ‰€ä»¥æ˜¾ç¤ºè‡ªå·±çš„IDï¼‰
                    # ç¡®ä¿å³ä½¿r.idæ˜¯ç©ºå­—ç¬¦ä¸²ä¹Ÿèƒ½æ­£ç¡®å¤„ç†
                    vpc_id_value = r.id if (hasattr(r, "id") and r.id and str(r.id).strip()) else None
                    # è°ƒè¯•æ—¥å¿—
                    if not vpc_id_value:
                        logger.warning(f"VPC resource has empty ID: id={r.id}, name={r.name}, type={type}")
                else:
                    vpc_id_value = r.vpc_id if hasattr(r, "vpc_id") and r.vpc_id else None
                
                result.append(
                    {
                        "id": r.id or "-",
                        "name": display_name,
                        "type": type,
                        "status": r.status.value if hasattr(r.status, "value") else str(r.status),
                        "region": r.region,
                        "spec": r.spec or "-",
                        "cost": float(cost or 0),
                        "tags": r.tags if hasattr(r, "tags") and r.tags else {},
                        "created_time": r.created_time.isoformat()
                        if hasattr(r, "created_time") and r.created_time
                        else None,
                        "public_ips": r.public_ips if hasattr(r, "public_ips") else [],
                        "private_ips": r.private_ips if hasattr(r, "private_ips") else [],
                        "vpc_id": vpc_id_value,
                    }
                )

        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type=type, account_name=account_name, data=result)
    
    # æ’åºï¼ˆåœ¨ç¼“å­˜æ•°æ®ä¸Šæ’åºï¼‰
    if sortBy:
        reverse = sortOrder == "desc"
        try:
            result.sort(key=lambda x: x.get(sortBy, ""), reverse=reverse)
        except:
            pass  # Ignore sort errors
    
    # ç­›é€‰ï¼ˆåœ¨ç¼“å­˜æ•°æ®ä¸Šç­›é€‰ï¼‰
    if filter:
        try:
            import json
            filter_dict = json.loads(filter)
            filtered_result = []
            for item in result:
                match = True
                for key, value in filter_dict.items():
                    if item.get(key) != value:
                        match = False
                        break
                if match:
                    filtered_result.append(item)
            result = filtered_result
        except:
            pass  # Ignore filter errors
    
    # åˆ†é¡µï¼ˆåœ¨ç¼“å­˜æ•°æ®ä¸Šåˆ†é¡µï¼‰
    total = len(result)
    start = (page - 1) * pageSize
    end = start + pageSize
    paginated_resources = result[start:end]
    
    return {
        "success": True,
        "data": paginated_resources,
        "pagination": {
            "page": page,
            "pageSize": pageSize,
            "total": total,
            "totalPages": (total + pageSize - 1) // pageSize,
        },
        "cached": cached_result is not None,  # æ ‡è¯†æ˜¯å¦æ¥è‡ªç¼“å­˜
    }


@router.get("/resources/{resource_id}")
def get_resource(
    resource_id: str,
    account: Optional[str] = None,
    resource_type: Optional[str] = Query(None, description="èµ„æºç±»å‹ï¼Œç”¨äºåŠ é€ŸæŸ¥æ‰¾")
):
    """è·å–èµ„æºè¯¦æƒ…ï¼ˆLegacy APIï¼Œå»ºè®®ä½¿ç”¨ api_resources æ¨¡å—ï¼‰"""
    # æ·»åŠ è°ƒè¯•æ—¥å¿—
    logger.info(f"[Legacy API] æ”¶åˆ°è¯·æ±‚: resource_id={resource_id}, type={resource_type}, account={account}")

    # å¦‚æœæä¾›äº† resource_typeï¼Œç›´æ¥ä½¿ç”¨æ–°çš„ API é€»è¾‘ï¼ˆä¼˜å…ˆï¼‰
    if resource_type:
        logger.info(f"[Legacy API] æ£€æµ‹åˆ° resource_type={resource_type}ï¼Œå°è¯•è½¬å‘åˆ°æ–° API")
        try:
            from web.backend.api_resources import get_resource_detail
            logger.info(f"[Legacy API] è°ƒç”¨æ–° API: get_resource_detail({resource_id}, {account}, {resource_type})")
            result = get_resource_detail(resource_id, account, resource_type)
            logger.info(f"[Legacy API] æ–° API è¿”å›: success={result.get('success')}, has_data={bool(result.get('data'))}")

            if result.get('success') and result.get('data'):
                logger.info(f"[Legacy API] âœ… è½¬å‘æˆåŠŸï¼Œè¿”å›æ–° API ç»“æœ")
                # å¯¹äº ACK èµ„æºï¼Œè®°å½•æ˜¯å¦åŒ…å« ack_details
                if resource_type == "ack" and result.get('data'):
                    has_ack_details = 'ack_details' in result.get('data', {})
                    logger.info(f"[Legacy API] ACK èµ„æºè¿”å›æ•°æ®åŒ…å« ack_details: {has_ack_details}")
                return result
            else:
                logger.warning(f"[Legacy API] æ–° API è¿”å›å¤±è´¥æˆ–æ— æ•°æ®ï¼Œä½¿ç”¨æ—§é€»è¾‘")
        except Exception as e:
            logger.error(f"[Legacy API] è½¬å‘åˆ°æ–° API å¤±è´¥: {e}, ä½¿ç”¨æ—§é€»è¾‘", exc_info=True)
    
    provider, account_name = _get_provider_for_account(account)
    
    # è·å–è´¦å·é…ç½®ç”¨äºæˆæœ¬æŸ¥è¯¢
    cm = ConfigManager()
    account_config = cm.get_account(account_name)
    
    # å°è¯•ä»å„ç§èµ„æºç±»å‹ä¸­æŸ¥æ‰¾ï¼ˆèµ„æºé‡è¾ƒå¤§æ—¶å»ºè®®æŒ‰ type æŸ¥ï¼›è¿™é‡Œä¸ºé€šç”¨è¯¦æƒ…å…¥å£åšå°½é‡è¦†ç›–ï¼‰
    # UnifiedResource ç±»èµ„æº
    resources = []
    try:
        resources.extend(provider.list_instances())
        resources.extend(provider.list_rds())
        resources.extend(provider.list_redis())
        if hasattr(provider, "list_slb"):
            resources.extend(provider.list_slb())
        if hasattr(provider, "list_nat_gateways"):
            resources.extend(provider.list_nat_gateways())
    except Exception:
        pass

    resource = next((r for r in resources if getattr(r, "id", None) == resource_id), None)

    # dict èµ„æºï¼ˆEIP/OSSï¼‰
    dict_resource = None
    if resource is None:
        try:
            if hasattr(provider, "list_eip"):
                eips_list = provider.list_eip() if hasattr(provider, "list_eip") else (provider.list_eips() if hasattr(provider, "list_eips") else [])
                for e in eips_list:
                    if e.get("id") == resource_id:
                        dict_resource = ("eip", e)
                        break
            if dict_resource is None and hasattr(provider, "list_oss"):
                for b in provider.list_oss():
                    if b.get("id") == resource_id:
                        dict_resource = ("oss", b)
                        break
            if dict_resource is None and hasattr(provider, "list_disks"):
                for d in provider.list_disks():
                    if d.get("id") == resource_id:
                        dict_resource = ("disk", d)
                        break
            if dict_resource is None and hasattr(provider, "list_snapshots"):
                for s in provider.list_snapshots():
                    if s.get("id") == resource_id:
                        dict_resource = ("snapshot", s)
                        break
        except Exception:
            dict_resource = None

    if resource is None and dict_resource is None:
        raise HTTPException(status_code=404, detail="èµ„æºä¸å­˜åœ¨")

    # ç¡®å®šèµ„æºç±»å‹
    resource_type = "ecs"
    if dict_resource is not None:
        resource_type = dict_resource[0]
    elif hasattr(resource, "resource_type"):
        rt = resource.resource_type.value if hasattr(resource.resource_type, "value") else str(resource.resource_type)
        if "rds" in rt.lower():
            resource_type = "rds"
        elif "redis" in rt.lower():
            resource_type = "redis"
        elif "vpc" in rt.lower():
            resource_type = "vpc"
        elif "slb" in rt.lower():
            resource_type = "slb"
        elif "nat" in rt.lower():
            resource_type = "nat"
    
    # è·å–çœŸå®æˆæœ¬æ˜ å°„
    cost_map = {}
    if account_config and resource_type not in ("vpc",):
        cost_map = _get_cost_map(resource_type, account_config)

    if dict_resource is not None:
        _, r = dict_resource
        cost = float(cost_map.get(resource_id) or 0.0)
        spec = "-"
        if resource_type == "eip":
            spec = f"{r.get('bandwidth', '-') }Mbps"
        elif resource_type == "disk":
            spec = f"{r.get('disk_category', '-') } / {r.get('disk_type', '-') } / {r.get('size_gb', '-') }GB"
        elif resource_type == "snapshot":
            spec = f"æºç›˜: {r.get('source_disk_id') or '-'}"
        return {
            "success": True,
            "data": {
                "id": resource_id,
                "name": r.get("name") or r.get("ip_address") or resource_id,
                "type": resource_type,
                "status": str(r.get("status") or "-"),
                "region": str(r.get("region") or getattr(provider, "region", "")),
                "spec": spec,
                "cost": cost,
                "tags": {},
                "created_time": r.get("created_time") if isinstance(r.get("created_time"), str) else None,
                "public_ips": [r.get("ip_address")] if r.get("ip_address") else [],
                "private_ips": [],
                "raw_data": r,
            },
        }

    # UnifiedResource
    cost = cost_map.get(resource_id)
    if cost is None:
        cost = _estimate_monthly_cost(resource)

    return {
        "success": True,
        "data": {
            "id": resource.id,
            "name": resource.name or "-",
            "type": resource_type,
            "status": resource.status.value if hasattr(resource.status, "value") else str(resource.status),
            "region": resource.region,
            "spec": resource.spec or "-",
            "cost": float(cost or 0),
            "tags": resource.tags if hasattr(resource, "tags") and resource.tags else {},
            "created_time": resource.created_time.isoformat()
            if hasattr(resource, "created_time") and resource.created_time
            else None,
            "public_ips": resource.public_ips if hasattr(resource, "public_ips") else [],
            "private_ips": resource.private_ips if hasattr(resource, "private_ips") else [],
            "raw_data": getattr(resource, "raw_data", {}),
        },
    }


@router.get("/resources/{resource_id}/metrics")
def get_resource_metrics(
    resource_id: str,
    days: int = Query(7, ge=1, le=30),
    account: Optional[str] = None,
):
    """è·å–èµ„æºç›‘æ§æ•°æ®"""
    provider, account_name = _get_provider_for_account(account)
    
    # è·å–èµ„æº
    resources = []
    try:
        resources.extend(provider.list_instances())
    except:
        pass
    
    resource = next((r for r in resources if r.id == resource_id), None)
    if not resource:
        raise HTTPException(status_code=404, detail="èµ„æºä¸å­˜åœ¨")
    
    # è·å–ç›‘æ§æ•°æ®
    try:
        from cloudlens.core.idle_detector import IdleDetector
        metrics = IdleDetector.fetch_ecs_metrics(provider, resource_id, days)
        
        # è½¬æ¢ä¸ºå›¾è¡¨æ•°æ®æ ¼å¼
        chart_data = {
            "cpu": [],
            "memory": [],
            "network_in": [],
            "network_out": [],
            "dates": [],
        }
        
        # ç®€åŒ–ï¼šè¿”å›å¹³å‡å€¼ï¼ˆå®é™…åº”è¯¥è¿”å›æ—¶é—´åºåˆ—æ•°æ®ï¼‰
        return {
            "success": True,
            "data": {
                "metrics": metrics,
                "chart_data": chart_data,
            }
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "metrics": {},
                "chart_data": {},
                "error": str(e),
            }
        }


# ==================== Phase 1 Week 3: Account Management APIs ====================

@router.get("/settings/accounts")
def list_accounts_settings():
    """è·å–è´¦å·åˆ—è¡¨ï¼ˆç”¨äºè®¾ç½®é¡µé¢ï¼‰"""
    cm = ConfigManager()
    accounts = cm.list_accounts()
    result = []
    for account in accounts:
        if isinstance(account, CloudAccount):
            result.append({
                "name": account.name,
                "alias": getattr(account, 'alias', None),  # åˆ«åï¼ˆå¯é€‰ï¼‰
                "region": account.region,
                "provider": account.provider,
                "access_key_id": account.access_key_id,
            })
    return {"success": True, "data": result}


@router.post("/settings/accounts")
def add_account(account_data: AccountCreateRequest):
    """æ·»åŠ è´¦å·"""
    cm = ConfigManager()
    try:
        cm.add_account(
            name=account_data.name,
            provider=account_data.provider,
            access_key_id=account_data.access_key_id,
            access_key_secret=account_data.access_key_secret,
            region=account_data.region,
            alias=account_data.alias,  # åˆ«åï¼ˆå¯é€‰ï¼‰
        )
        return {"success": True, "message": "è´¦å·æ·»åŠ æˆåŠŸ"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/settings/accounts/{account_name}")
def update_account(account_name: str, account_data: AccountUpdateRequest):
    """æ›´æ–°è´¦å·"""
    import keyring
    cm = ConfigManager()
    try:
        # æ£€æŸ¥è´¦å·æ˜¯å¦å­˜åœ¨
        existing_account = cm.get_account(account_name)
        if not existing_account:
            raise HTTPException(status_code=404, detail=f"è´¦å· '{account_name}' ä¸å­˜åœ¨")
        
        # è´¦å·åç§°ä¸å¯ä¿®æ”¹ï¼ˆç”¨äºæ•°æ®å…³è”ï¼‰ï¼Œåªå…è®¸ä¿®æ”¹åˆ«å
        # è·å–åˆ«åï¼ˆå¦‚æœæä¾›äº†ï¼‰
        alias = account_data.alias.strip() if account_data.alias else None
        
        # è·å–æ–°å¯†é’¥ï¼Œå¦‚æœæ²¡æœ‰æä¾›åˆ™ä½¿ç”¨ç°æœ‰å¯†é’¥
        new_secret = account_data.access_key_secret
        if not new_secret:
            # ä» keyring è·å–ç°æœ‰å¯†é’¥
            try:
                existing_secret = keyring.get_password("cloudlens", f"{account_name}_access_key_secret")
                if existing_secret:
                    new_secret = existing_secret
                else:
                    raise HTTPException(status_code=400, detail="æ— æ³•è·å–ç°æœ‰å¯†é’¥ï¼Œè¯·æä¾›æ–°å¯†é’¥")
            except Exception:
                raise HTTPException(status_code=400, detail="æ— æ³•è·å–ç°æœ‰å¯†é’¥ï¼Œè¯·æä¾›æ–°å¯†é’¥")
        
        # æ›´æ–°è´¦å·é…ç½®ï¼ˆä¸ä¿®æ”¹è´¦å·åç§°ï¼Œåªæ›´æ–°å…¶ä»–å­—æ®µå’Œåˆ«åï¼‰
        cm.add_account(
            name=account_name,  # ä¿æŒåŸåç§°ä¸å˜ï¼ˆç”¨äºæ•°æ®å…³è”ï¼‰
            provider=account_data.provider or existing_account.provider,
            access_key_id=account_data.access_key_id or existing_account.access_key_id,
            access_key_secret=new_secret,
            region=account_data.region or existing_account.region,
            alias=alias,  # æ›´æ–°åˆ«å
        )
        
        return {"success": True, "message": "è´¦å·æ›´æ–°æˆåŠŸ"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/settings/accounts/{account_name}")
def delete_account(account_name: str):
    """åˆ é™¤è´¦å·"""
    cm = ConfigManager()
    try:
        cm.remove_account(account_name)
        return {"success": True, "message": "è´¦å·åˆ é™¤æˆåŠŸ"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Phase 2: Cost Analysis APIs ====================

@router.get("/cost/overview")
def get_cost_overview(account: Optional[str] = None, force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜")):
    """è·å–æˆæœ¬æ¦‚è§ˆï¼ˆä¼˜å…ˆè´¦å•å£å¾„ï¼Œå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    provider, account_name = _get_provider_for_account(account)
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="cost_overview", account_name=account_name)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    # ç¼“å­˜æ— æ•ˆæˆ–ä¸å­˜åœ¨ï¼Œè®¡ç®—æ–°æ•°æ®
    cm = ConfigManager()
    account_config = cm.get_account(account_name)
    
    try:
        # è´¦å•ä¼˜å…ˆï¼šä½¿ç”¨ BSS è´¦å•æ¦‚è§ˆä½œä¸º"å…¨é‡æˆæœ¬"å£å¾„
        from datetime import datetime, timedelta
        now = datetime.now()
        current_cycle = now.strftime("%Y-%m")
        
        # è®¡ç®—æœ¬æœˆå·²è¿‡å¤©æ•°ï¼ˆç”¨äºç¯æ¯”å¯¹æ¯”ï¼‰
        current_day = now.day  # ä»Šå¤©æ˜¯å‡ å·
        first_day_this_month = now.replace(day=1)
        
        # æœ¬æœˆæˆæœ¬ï¼š1æœˆ1æ—¥åˆ°1æœˆ6æ—¥ï¼ˆå¦‚æœä»Šå¤©æ˜¯6å·ï¼‰
        current_month_start = first_day_this_month
        current_month_end = now
        
        # ä¸Šæœˆç›¸åŒå¤©æ•°ï¼š12æœˆ1æ—¥åˆ°12æœˆ6æ—¥
        last_month_end = first_day_this_month - timedelta(days=1)  # ä¸Šä¸ªæœˆæœ€åä¸€å¤©
        last_month_start = last_month_end.replace(day=1)  # ä¸Šä¸ªæœˆç¬¬ä¸€å¤©
        
        # è®¡ç®—ä¸Šæœˆç›¸åŒå¤©æ•°çš„ç»“æŸæ—¥æœŸï¼ˆä¸èƒ½è¶…è¿‡ä¸Šä¸ªæœˆçš„æœ€åä¸€å¤©ï¼‰
        last_month_comparable_end = last_month_start + timedelta(days=current_day - 1)
        if last_month_comparable_end > last_month_end:
            last_month_comparable_end = last_month_end
        
        # è®¡ç®—ä¸Šæœˆè´¦æœŸï¼ˆç”¨äºæ˜¾ç¤ºï¼‰
        last_cycle = last_month_end.strftime("%Y-%m")
        
        logger.info(f"ğŸ“Š æˆæœ¬æ¦‚è§ˆæŸ¥è¯¢: è´¦å·={account_name}")
        logger.info(f"   å½“å‰è´¦æœŸ={current_cycle}, æœ¬æœˆå·²è¿‡å¤©æ•°={current_day}")
        logger.info(f"   æœ¬æœˆæˆæœ¬èŒƒå›´: {current_month_start.strftime('%Y-%m-%d')} è‡³ {current_month_end.strftime('%Y-%m-%d')}")
        logger.info(f"   ä¸Šæœˆå¯¹æ¯”èŒƒå›´: {last_month_start.strftime('%Y-%m-%d')} è‡³ {last_month_comparable_end.strftime('%Y-%m-%d')}")

        # ä½¿ç”¨æˆæœ¬è¶‹åŠ¿åˆ†æå™¨è·å–æŒ‡å®šæ—¥æœŸèŒƒå›´çš„æˆæœ¬ï¼ˆæ›´å‡†ç¡®ï¼‰
        from cloudlens.core.cost_trend_analyzer import CostTrendAnalyzer
        analyzer = CostTrendAnalyzer()
        
        # è·å–æœ¬æœˆæˆæœ¬ï¼ˆä»1æœˆ1æ—¥åˆ°ä»Šå¤©ï¼‰
        current_month_cost = 0.0
        try:
            current_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=current_month_start.strftime("%Y-%m-%d"),
                end_date=current_month_end.strftime("%Y-%m-%d")
            )
            # ä»è¿”å›æ•°æ®ä¸­æå–æ€»æˆæœ¬ï¼ˆå¯èƒ½æ˜¯chart_data.costsçš„æ€»å’Œï¼Œæˆ–è€…analysisä¸­çš„å€¼ï¼‰
            if current_cost_data and "error" not in current_cost_data:
                # å°è¯•ä»chart_dataä¸­è®¡ç®—æ€»æˆæœ¬
                if "chart_data" in current_cost_data and "costs" in current_cost_data["chart_data"]:
                    costs = current_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        current_month_cost = float(sum(costs))
                        logger.info(f"âœ… æœ¬æœˆæˆæœ¬ï¼ˆä»chart_dataè®¡ç®—ï¼Œ{current_month_start.strftime('%Y-%m-%d')} è‡³ {current_month_end.strftime('%Y-%m-%d')}ï¼‰: {current_month_cost}")
                    else:
                        # å¦‚æœchart_dataä¸ºç©ºï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI
                        logger.warning("âš ï¸  chart_dataä¸ºç©ºï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI")
                        current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                        current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
                elif "total_cost" in current_cost_data:
                    current_month_cost = float(current_cost_data.get("total_cost", 0.0))
                    logger.info(f"âœ… æœ¬æœˆæˆæœ¬ï¼ˆ{current_month_start.strftime('%Y-%m-%d')} è‡³ {current_month_end.strftime('%Y-%m-%d')}ï¼‰: {current_month_cost}")
                else:
                    # å¦‚æœæ•°æ®æ ¼å¼ä¸ç¬¦åˆé¢„æœŸï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI
                    logger.warning("âš ï¸  æ•°æ®åº“æŸ¥è¯¢è¿”å›æ ¼å¼ä¸ç¬¦åˆé¢„æœŸï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI")
                    current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                    current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
            else:
                # å¦‚æœæ•°æ®åº“æŸ¥è¯¢å¤±è´¥ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI
                error_msg = current_cost_data.get("error", "Unknown error") if current_cost_data else "No data returned"
                logger.warning(f"âš ï¸  æ•°æ®åº“æŸ¥è¯¢å¤±è´¥: {error_msg}ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI")
                current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        except Exception as e:
            logger.warning(f"âš ï¸  è·å–æœ¬æœˆæˆæœ¬å¤±è´¥ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPI: {str(e)}")
            current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
        current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        
        # è·å–ä¸Šæœˆç›¸åŒå¤©æ•°çš„æˆæœ¬ï¼ˆä»12æœˆ1æ—¥åˆ°12æœˆ6æ—¥ï¼‰
        last_month_cost = 0.0
        try:
            last_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=last_month_start.strftime("%Y-%m-%d"),
                end_date=last_month_comparable_end.strftime("%Y-%m-%d")
            )
            # ä»è¿”å›æ•°æ®ä¸­æå–æ€»æˆæœ¬ï¼ˆå¯èƒ½æ˜¯chart_data.costsçš„æ€»å’Œï¼Œæˆ–è€…analysisä¸­çš„å€¼ï¼‰
            if last_cost_data and "error" not in last_cost_data:
                # å°è¯•ä»chart_dataä¸­è®¡ç®—æ€»æˆæœ¬
                if "chart_data" in last_cost_data and "costs" in last_cost_data["chart_data"]:
                    costs = last_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        last_month_cost = float(sum(costs))
                        logger.info(f"âœ… ä¸Šæœˆæˆæœ¬ï¼ˆä»chart_dataè®¡ç®—ï¼Œ{last_month_start.strftime('%Y-%m-%d')} è‡³ {last_month_comparable_end.strftime('%Y-%m-%d')}ï¼‰: {last_month_cost}")
                    else:
                        # å¦‚æœchart_dataä¸ºç©ºï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
                        logger.warning("âš ï¸  chart_dataä¸ºç©ºï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰")
                        last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                        if last_totals:
                            last_month_total = float(last_totals.get("total_pretax") or 0.0)
                            last_month_days = last_month_end.day
                            last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                            logger.info(f"   ä¸Šæœˆæ€»æˆæœ¬={last_month_total}, æ€»å¤©æ•°={last_month_days}, å·²è¿‡å¤©æ•°={current_day}, æŒ‰æ¯”ä¾‹è®¡ç®—={last_month_cost}")
                elif "total_cost" in last_cost_data:
                    last_month_cost = float(last_cost_data.get("total_cost", 0.0))
                    logger.info(f"âœ… ä¸Šæœˆæˆæœ¬ï¼ˆ{last_month_start.strftime('%Y-%m-%d')} è‡³ {last_month_comparable_end.strftime('%Y-%m-%d')}ï¼‰: {last_month_cost}")
                else:
                    # å¦‚æœæ•°æ®æ ¼å¼ä¸ç¬¦åˆé¢„æœŸï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
                    logger.warning("âš ï¸  æ•°æ®åº“æŸ¥è¯¢è¿”å›æ ¼å¼ä¸ç¬¦åˆé¢„æœŸï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰")
                    last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                    if last_totals:
                        last_month_total = float(last_totals.get("total_pretax") or 0.0)
                        last_month_days = last_month_end.day
                        last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                        logger.info(f"   ä¸Šæœˆæ€»æˆæœ¬={last_month_total}, æ€»å¤©æ•°={last_month_days}, å·²è¿‡å¤©æ•°={current_day}, æŒ‰æ¯”ä¾‹è®¡ç®—={last_month_cost}")
            else:
                # å¦‚æœæ•°æ®åº“æŸ¥è¯¢å¤±è´¥ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
                error_msg = last_cost_data.get("error", "Unknown error") if last_cost_data else "No data returned"
                logger.warning(f"âš ï¸  æ•°æ®åº“æŸ¥è¯¢å¤±è´¥: {error_msg}ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰")
                last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                if last_totals:
                    # æŒ‰æ¯”ä¾‹è®¡ç®—ï¼šä¸Šæœˆæ€»æˆæœ¬ * (å·²è¿‡å¤©æ•° / ä¸Šæœˆæ€»å¤©æ•°)
                    last_month_total = float(last_totals.get("total_pretax") or 0.0)
                    last_month_days = last_month_end.day  # ä¸Šä¸ªæœˆæ€»å¤©æ•°
                    last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                    logger.info(f"   ä¸Šæœˆæ€»æˆæœ¬={last_month_total}, æ€»å¤©æ•°={last_month_days}, å·²è¿‡å¤©æ•°={current_day}, æŒ‰æ¯”ä¾‹è®¡ç®—={last_month_cost}")
        except Exception as e:
            logger.warning(f"âš ï¸  è·å–ä¸Šæœˆæˆæœ¬å¤±è´¥ï¼Œå›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰: {str(e)}")
            last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
            if last_totals:
                last_month_total = float(last_totals.get("total_pretax") or 0.0)
                last_month_days = last_month_end.day
                last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
        
        logger.info(f"ğŸ’° æˆæœ¬æ•°æ®: æœ¬æœˆï¼ˆ{current_day}å¤©ï¼‰={current_month_cost}, ä¸Šæœˆï¼ˆ{current_day}å¤©ï¼‰={last_month_cost}")
        
        # å¦‚æœä¸Šæœˆæ•°æ®ä¸º0ï¼Œè®°å½•è­¦å‘Š
        if last_month_cost == 0:
            logger.warning(f"âš ï¸  ä¸Šæœˆç›¸åŒå¤©æ•°ï¼ˆ{last_month_start.strftime('%Y-%m-%d')} è‡³ {last_month_comparable_end.strftime('%Y-%m-%d')}ï¼‰æˆæœ¬ä¸º0ï¼Œå¯èƒ½è¯¥æ—¶é—´æ®µç¡®å®æ— æˆæœ¬æˆ–æ•°æ®æœªåŒæ­¥")
        
        mom = ((current_month_cost - last_month_cost) / last_month_cost * 100) if last_month_cost > 0 else 0.0
        
        # è®¡ç®—åŒæ¯”ï¼ˆå»å¹´åŒæœŸç›¸åŒå¤©æ•°ï¼š2025å¹´1æœˆ1æ—¥-1æœˆ6æ—¥ vs 2026å¹´1æœˆ1æ—¥-1æœˆ6æ—¥ï¼‰
        yoy = 0.0
        last_year_cost = 0.0
        try:
            # å»å¹´åŒæœŸèŒƒå›´
            last_year_start = datetime(now.year - 1, now.month, 1)
            last_year_end = datetime(now.year - 1, now.month, current_day)
            
            logger.info(f"ğŸ“Š è®¡ç®—åŒæ¯”: å»å¹´åŒæœŸèŒƒå›´ {last_year_start.strftime('%Y-%m-%d')} è‡³ {last_year_end.strftime('%Y-%m-%d')}")
            
            last_year_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=last_year_start.strftime("%Y-%m-%d"),
                end_date=last_year_end.strftime("%Y-%m-%d")
            )
            if last_year_cost_data and "error" not in last_year_cost_data:
                if "chart_data" in last_year_cost_data and "costs" in last_year_cost_data["chart_data"]:
                    costs = last_year_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        last_year_cost = float(sum(costs))
                        logger.info(f"âœ… å»å¹´åŒæœŸæˆæœ¬ï¼ˆä»chart_dataè®¡ç®—ï¼‰: {last_year_cost}")
                    else:
                        # å¦‚æœæ•°æ®åº“æŸ¥è¯¢å¤±è´¥ï¼Œå°è¯•ä»è´¦å•æ¦‚è§ˆAPIè·å–ï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
                        last_year_cycle = last_year_start.strftime("%Y-%m")
                        last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                        if last_year_totals:
                            last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                            # æŒ‰æ¯”ä¾‹è®¡ç®—ï¼šå»å¹´åŒæœŸæ€»æˆæœ¬ * (å·²è¿‡å¤©æ•° / è¯¥æœˆæ€»å¤©æ•°)
                            last_year_month_days = (datetime(last_year_start.year, last_year_start.month + 1, 1) - timedelta(days=1)).day if last_year_start.month < 12 else 31
                            last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
                            logger.info(f"   å»å¹´åŒæœŸæ€»æˆæœ¬={last_year_total}, æ€»å¤©æ•°={last_year_month_days}, å·²è¿‡å¤©æ•°={current_day}, æŒ‰æ¯”ä¾‹è®¡ç®—={last_year_cost}")
                elif "total_cost" in last_year_cost_data:
                    last_year_cost = float(last_year_cost_data.get("total_cost", 0.0))
                else:
                    # å›é€€åˆ°è´¦å•æ¦‚è§ˆAPIï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
                    last_year_cycle = last_year_start.strftime("%Y-%m")
                    last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                    if last_year_totals:
                        last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                        last_year_month_days = (datetime(last_year_start.year, last_year_start.month + 1, 1) - timedelta(days=1)).day if last_year_start.month < 12 else 31
                        last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
        except Exception as e:
            logger.warning(f"âš ï¸  è·å–å»å¹´åŒæœŸæˆæœ¬å¤±è´¥: {str(e)}")
            # å°è¯•ä»è´¦å•æ¦‚è§ˆAPIè·å–ï¼ˆæŒ‰æ¯”ä¾‹è®¡ç®—ï¼‰
            try:
                last_year_cycle = datetime(now.year - 1, now.month, 1).strftime("%Y-%m")
                last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                if last_year_totals:
                    last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                    last_year_month_days = (datetime(now.year - 1, now.month + 1, 1) - timedelta(days=1)).day if now.month < 12 else 31
                    last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
            except:
                last_year_cost = 0.0
        
        # è®¡ç®—åŒæ¯”
        yoy = (
            (current_month_cost - last_year_cost) / last_year_cost * 100
            if last_year_cost > 0 else 0.0
        )
        
        logger.info(f"ğŸ“Š åŒæ¯”æ•°æ®: ä»Šå¹´ï¼ˆ{current_day}å¤©ï¼‰={current_month_cost}, å»å¹´ï¼ˆ{current_day}å¤©ï¼‰={last_year_cost}, åŒæ¯”={yoy:.2f}%")
        
        result_data = {
            "current_month": round(current_month_cost, 2),
            "last_month": round(last_month_cost, 2),
            "yoy": round(yoy, 2),
            "mom": round(mom, 2),
            "current_cycle": current_cycle,
            "last_cycle": last_cycle,
            "current_days": current_day,  # æœ¬æœˆå·²è¿‡å¤©æ•°
            "comparable_days": current_day,  # å¯¹æ¯”å¤©æ•°ï¼ˆç›¸åŒï¼‰
        }
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type="cost_overview", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        logger.error(f"âŒ è·å–æˆæœ¬æ¦‚è§ˆå¤±è´¥: {str(e)}", exc_info=True)
        # è¿”å›é”™è¯¯ä¿¡æ¯ï¼Œè€Œä¸æ˜¯é™é»˜è¿”å›0
        return {
            "success": False,
            "error": str(e),
            "data": {
                "current_month": 0,
                "last_month": 0,
                "yoy": 0,
                "mom": 0,
            },
            "cached": False,
        }


@router.get("/cost/breakdown")
def get_cost_breakdown(
    account: Optional[str] = None,
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
    billing_cycle: Optional[str] = Query(None, description="è´¦æœŸ yyyy-MMï¼Œé»˜è®¤å½“æœˆ"),
):
    """è·å–æˆæœ¬æ„æˆï¼ˆä¼˜å…ˆè´¦å•å£å¾„ï¼Œå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    provider, account_name = _get_provider_for_account(account)
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="cost_breakdown", account_name=account_name)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    # ç¼“å­˜æ— æ•ˆæˆ–ä¸å­˜åœ¨ï¼Œè®¡ç®—æ–°æ•°æ®
    cm = ConfigManager()
    account_config = cm.get_account(account_name)

    try:
        # è´¦å•ä¼˜å…ˆï¼šç”¨ BSS BillOverview çš„ ProductCode èšåˆå¾—åˆ°â€œå…¨é‡æˆæœ¬æ„æˆâ€
        totals = _get_billing_overview_totals(account_config, billing_cycle=billing_cycle) if account_config else None
        by_product = (totals or {}).get("by_product") or {}
        total = float((totals or {}).get("total_pretax") or 0.0)
        by_product_name = (totals or {}).get("by_product_name") or {}
        by_product_subscription = (totals or {}).get("by_product_subscription") or {}

        # ä¾¿äºå‰ç«¯å±•ç¤ºçš„åˆ—è¡¨ç»“æ„ï¼ˆæ’åºåï¼‰
        categories = []
        for code, amount in by_product.items():
            try:
                amount_f = float(amount or 0.0)
            except Exception:
                amount_f = 0.0
            if amount_f <= 0:
                continue
            categories.append(
                {
                    "code": code,
                    "name": by_product_name.get(code) or code,
                    "amount": round(amount_f, 2),
                    "subscription": by_product_subscription.get(code) or {},
                }
            )
        categories.sort(key=lambda x: float(x.get("amount") or 0.0), reverse=True)

        result_data = {
            # å…¼å®¹æ—§å‰ç«¯å­—æ®µï¼šby_type ä»è¿”å› {code: amount}
            "by_type": by_product,
            "total": round(float(total), 2),
            "billing_cycle": (totals or {}).get("billing_cycle") or billing_cycle,
            "source": "billing_overview",
            # æ–°å­—æ®µï¼šå‰ç«¯ç›´æ¥ç”¨ categories æ¸²æŸ“æ›´å‹å¥½
            "categories": categories,
            "by_product_name": by_product_name,
        }
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type="cost_breakdown", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "by_type": {},
                "total": 0,
            },
            "cached": False,
        }


# ==================== Phase 2: Security APIs ====================

@router.get("/security/overview")
def get_security_overview(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
    locale: Optional[str] = Query("zh", description="è¯­è¨€è®¾ç½®: zh æˆ– en")
):
    """è·å–å®‰å…¨æ¦‚è§ˆï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    # è·å–è¯­è¨€è®¾ç½®
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    # æ³¨æ„ï¼šä¸ºäº†æ”¯æŒå¤šè¯­è¨€ï¼Œæˆ‘ä»¬éœ€è¦é‡æ–°ç”Ÿæˆç¿»è¯‘åçš„æ–‡æœ¬
    # æ–¹æ¡ˆï¼šç¼“å­˜åŸå§‹æ•°æ®ï¼ˆæ•°å­—ï¼‰ï¼Œè¿”å›æ—¶æ ¹æ®è¯­è¨€ç¿»è¯‘
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="security_overview", account_name=account_name)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œéœ€è¦æ ¹æ®è¯­è¨€é‡æ–°ç¿»è¯‘
    if cached_result is not None:
        # é‡æ–°ç”Ÿæˆ score_deductionsï¼ˆéœ€è¦åŸå§‹æ•°æ®ï¼‰
        # ç”±äºç¼“å­˜ä¸­å¯èƒ½å·²ç»åŒ…å«ç¿»è¯‘åçš„æ–‡æœ¬ï¼Œæˆ‘ä»¬éœ€è¦é‡æ–°è®¡ç®—
        # ä¸ºäº†ç®€åŒ–ï¼Œå¦‚æœè¯­è¨€ä¸æ˜¯ä¸­æ–‡ï¼Œæˆ‘ä»¬å¼ºåˆ¶åˆ·æ–°ä»¥é‡æ–°ç”Ÿæˆ
        # æ›´å¥½çš„æ–¹æ¡ˆæ˜¯ç¼“å­˜åŸå§‹æ•°æ®ï¼ˆä¸ç¿»è¯‘ï¼‰ï¼Œä½†éœ€è¦ä¿®æ”¹ç¼“å­˜ç»“æ„
        # æš‚æ—¶ï¼šå¦‚æœè¯­è¨€æ˜¯è‹±æ–‡ä¸”ç¼“å­˜å­˜åœ¨ï¼Œæˆ‘ä»¬é‡æ–°ç”Ÿæˆï¼ˆè·³è¿‡ç¼“å­˜ï¼‰
        if lang == "en":
            # è‹±æ–‡æ¨¡å¼ä¸‹ï¼Œè·³è¿‡ç¼“å­˜ï¼Œé‡æ–°ç”Ÿæˆä»¥ç¡®ä¿ç¿»è¯‘æ­£ç¡®
            cached_result = None
        else:
            # ä¸­æ–‡æ¨¡å¼ä¸‹ï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
    
    try:
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        cm = ConfigManager()
        account_config = cm.get_account(account_name)
        
        # ç»Ÿä¸€å…¨å±€èµ„æºè·å– (Consistent Global View)
        try:
            instances, rds_list, redis_list = _get_global_resources_consistent(account_name, account_config, force_refresh)
        except Exception as e:
            logger.warning(f"è·å–ä¸€è‡´æ€§å…¨å±€èµ„æºå¤±è´¥ï¼Œä½¿ç”¨å½“å‰åœ°åŸŸå…œåº•: {e}")
            instances = provider.list_instances()
            rds_list = provider.list_rds()
            redis_list = provider.list_redis()
            
        all_resources = instances + rds_list + redis_list
        analyzer = SecurityComplianceAnalyzer()
        
        # è·å–æ‰«ææ¼æ´ (å¯¹é½ä»ªè¡¨ç›˜)
        from cloudlens.core.security_scanner import PublicIPScanner
        last_scan = PublicIPScanner.load_last_results()
        high_risk_count = last_scan.get("summary", {}).get("high_risk_count", 0) if last_scan else 0
        
        # å…¬ç½‘æš´éœ²æ£€æµ‹
        exposed = analyzer.detect_public_exposure(all_resources)
        
        # å®‰å…¨æ£€æŸ¥
        stopped = analyzer.check_stopped_instances(instances)
        tag_coverage, no_tags = analyzer.check_missing_tags(all_resources)
        
        # ç£ç›˜åŠ å¯†æ£€æŸ¥
        encryption_info = analyzer.check_disk_encryption(instances)
        
        # æŠ¢å å¼å®ä¾‹æ£€æŸ¥
        preemptible = analyzer.check_preemptible_instances(instances)
        
        # EIPä½¿ç”¨æƒ…å†µï¼ˆå¦‚æœæœ‰EIPæ•°æ®ï¼‰
        eip_info = {"total": 0, "bound": 0, "unbound": 0, "unbound_rate": 0}
        try:
            eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
        except:
            pass
        
        # è®¡ç®—å®‰å…¨è¯„åˆ†ï¼ˆæ›´è¯¦ç»†çš„è¯„åˆ†é€»è¾‘ï¼‰
        security_score = 100
        score_deductions = []
        
        if len(exposed) > 0:
            deduction = min(len(exposed) * 5, 30)  # æœ€å¤šæ‰£30åˆ†
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.public_exposure", lang, count=len(exposed)),
                "deduction": deduction
            })
        
        if len(stopped) > 0:
            deduction = min(len(stopped) * 2, 20)  # æœ€å¤šæ‰£20åˆ†
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.stopped_instances", lang, count=len(stopped)),
                "deduction": deduction
            })
        
        if tag_coverage < 80:
            deduction = 10 if tag_coverage < 50 else 5
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.tag_coverage", lang, coverage=tag_coverage),
                "deduction": deduction
            })
        
        if encryption_info.get("encryption_rate", 100) < 50:
            deduction = 15
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.disk_encryption", lang, rate=encryption_info.get('encryption_rate', 0)),
                "deduction": deduction
            })
        
        if len(preemptible) > 0:
            deduction = min(len(preemptible) * 3, 15)  # æœ€å¤šæ‰£15åˆ†
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.preemptible_instances", lang, count=len(preemptible)),
                "deduction": deduction
            })
        
        if eip_info.get("unbound_rate", 0) > 20:
            deduction = 5
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.eip_unbound", lang, rate=eip_info.get('unbound_rate', 0)),
                "deduction": deduction
            })
        
        security_score = max(0, min(100, security_score))
        
        # ç”Ÿæˆå®‰å…¨æ”¹è¿›å»ºè®®ï¼ˆæ”¯æŒå›½é™…åŒ–ï¼‰
        security_summary = {
            "exposed_count": len(exposed),
            "stopped_count": len(stopped),
            "tag_coverage_rate": tag_coverage,
            "encryption_rate": encryption_info.get("encryption_rate", 100),
            "preemptible_count": len(preemptible),
            "unbound_eip": eip_info.get("unbound", 0),
        }
        suggestions = analyzer.suggest_security_improvements(security_summary, locale=lang)
        
        result_data = {
            "security_score": security_score,
            "exposed_count": len(exposed),
            "stopped_count": len(stopped),
            "tag_coverage": tag_coverage,
            "missing_tags_count": len(no_tags),
            # ç»Ÿä¸€å£å¾„: æš´éœ² + åœæ­¢ + æŠ¢å  + æ¼æ´ (å¯¹é½ä»ªè¡¨ç›˜)
            "alert_count": len(exposed) + len(stopped) + len(preemptible) + (high_risk_count if 'high_risk_count' in locals() else 0),
            "encryption_rate": encryption_info.get("encryption_rate", 100),
            "encrypted_count": encryption_info.get("encrypted", 0),
            "unencrypted_count": encryption_info.get("unencrypted_count", 0),
            "preemptible_count": len(preemptible),
            "eip_total": eip_info.get("total", 0),
            "eip_bound": eip_info.get("bound", 0),
            "eip_unbound": eip_info.get("unbound", 0),
            "eip_unbound_rate": eip_info.get("unbound_rate", 0),
            "score_deductions": score_deductions,
            "suggestions": suggestions,
        }
        
        # è¯¦ç»†æ—¥å¿—è¾“å‡ºï¼Œç”¨äºä¸ Dashboard å¯¹é½
        hr_count = high_risk_count if 'high_risk_count' in locals() else 0
        logger.info(f"æŒ‡æ ‡å¯¹é½è®¡ç®— (Security): æ ‡ç­¾è¦†ç›–ç‡={tag_coverage}%, å‘Šè­¦æ•°={result_data['alert_count']} (æš´éœ²={len(exposed)}, åœæ­¢={len(stopped)}, æŠ¢å ={len(preemptible)}, æ¼æ´={hr_count})")
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type="security_overview", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "security_score": 0,
                "exposed_count": 0,
                "stopped_count": 0,
                "tag_coverage": 0,
                "missing_tags_count": 0,
                "alert_count": 0,
                "encryption_rate": 0,
                "encrypted_count": 0,
                "unencrypted_count": 0,
                "preemptible_count": 0,
                "eip_total": 0,
                "eip_bound": 0,
                "eip_unbound": 0,
                "eip_unbound_rate": 0,
                "score_deductions": [],
                "suggestions": [],
            },
            "cached": False,
        }


@router.get("/security/checks")
def get_security_checks(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
    locale: Optional[str] = Query("zh", description="è¯­è¨€è®¾ç½®: zh æˆ– en")
):
    """è·å–å®‰å…¨æ£€æŸ¥ç»“æœï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    # è·å–è¯­è¨€è®¾ç½®
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®
    # æ³¨æ„ï¼šä¸ºäº†æ”¯æŒå¤šè¯­è¨€ï¼Œå¦‚æœè¯­è¨€ä¸æ˜¯ä¸­æ–‡ï¼Œæˆ‘ä»¬è·³è¿‡ç¼“å­˜é‡æ–°ç”Ÿæˆ
    cached_result = None
    if not force_refresh and lang == "zh":
        cached_result = cache_manager.get(resource_type="security_checks", account_name=account_name)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®ï¼ˆä¸­æ–‡ï¼‰
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    try:
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        redis_list = provider.list_redis()
        all_resources = instances + rds_list + redis_list
        
        analyzer = SecurityComplianceAnalyzer()
        
        exposed = analyzer.detect_public_exposure(all_resources)
        stopped = analyzer.check_stopped_instances(instances)
        tag_coverage, no_tags = analyzer.check_missing_tags(all_resources)
        encryption_info = analyzer.check_disk_encryption(instances)
        preemptible = analyzer.check_preemptible_instances(instances)
        
        # EIPä½¿ç”¨æƒ…å†µ
        eip_info = {"total": 0, "bound": 0, "unbound": 0, "unbound_eips": []}
        try:
            eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
        except:
            pass
        
        checks = []
        
        # å…¬ç½‘æš´éœ²æ£€æŸ¥
        if exposed:
            checks.append({
                "type": "public_exposure",
                "title": get_translation("security.public_exposure.title", lang),
                "description": get_translation("security.public_exposure.description_failed", lang),
                "status": "failed",
                "severity": "HIGH",
                "count": len(exposed),
                "resources": exposed[:20],
                "recommendation": get_translation("security.public_exposure.recommendation", lang),
            })
        else:
            checks.append({
                "type": "public_exposure",
                "title": get_translation("security.public_exposure.title", lang),
                "description": get_translation("security.public_exposure.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # åœæ­¢å®ä¾‹æ£€æŸ¥
        if stopped:
            checks.append({
                "type": "stopped_instances",
                "title": get_translation("security.stopped_instances.title", lang),
                "description": get_translation("security.stopped_instances.description", lang),
                "status": "warning",
                "severity": "MEDIUM",
                "count": len(stopped),
                "resources": stopped[:20],
                "recommendation": get_translation("security.stopped_instances.recommendation", lang),
            })
        else:
            checks.append({
                "type": "stopped_instances",
                "title": get_translation("security.stopped_instances.title", lang),
                "description": get_translation("security.stopped_instances.description", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # æ ‡ç­¾æ£€æŸ¥
        if tag_coverage >= 80:
            checks.append({
                "type": "tag_coverage",
                "title": get_translation("security.tag_coverage.title", lang),
                "description": get_translation("security.tag_coverage.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "coverage": tag_coverage,
                "missing_count": len(no_tags),
                "resources": [],
                "recommendation": get_translation("security.tag_coverage.recommendation", lang),
            })
        else:
            checks.append({
                "type": "tag_coverage",
                "title": get_translation("security.tag_coverage.title"),
                "description": get_translation("security.tag_coverage.description_failed"),
                "status": "warning",
                "severity": "MEDIUM",
                "coverage": tag_coverage,
                "missing_count": len(no_tags),
                "resources": no_tags[:20],
                "recommendation": get_translation("security.tag_coverage.recommendation", lang),
            })
        
        # ç£ç›˜åŠ å¯†æ£€æŸ¥
        encryption_rate = encryption_info.get("encryption_rate", 100)
        if encryption_rate < 100:
            checks.append({
                "type": "disk_encryption",
                "title": get_translation("security.disk_encryption.title"),
                "description": get_translation("security.disk_encryption.description_failed"),
                "status": "warning" if encryption_rate < 50 else "passed",
                "severity": "HIGH" if encryption_rate < 50 else "MEDIUM",
                "encryption_rate": encryption_rate,
                "encrypted_count": encryption_info.get("encrypted", 0),
                "unencrypted_count": encryption_info.get("unencrypted_count", 0),
                "resources": encryption_info.get("unencrypted_instances", [])[:20],
                "recommendation": get_translation("security.disk_encryption.recommendation", lang),
            })
        else:
            checks.append({
                "type": "disk_encryption",
                "title": get_translation("security.disk_encryption.title"),
                "description": get_translation("security.disk_encryption.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "encryption_rate": encryption_rate,
                "encrypted_count": encryption_info.get("encrypted", 0),
                "unencrypted_count": 0,
                "resources": [],
            })
        
        # æŠ¢å å¼å®ä¾‹æ£€æŸ¥
        if preemptible:
            checks.append({
                "type": "preemptible_instances",
                "title": get_translation("security.preemptible_instances.title", lang),
                "description": get_translation("security.preemptible_instances.description", lang),
                "status": "warning",
                "severity": "MEDIUM",
                "count": len(preemptible),
                "resources": preemptible[:20],
                "recommendation": get_translation("security.preemptible_instances.recommendation", lang),
            })
        else:
            checks.append({
                "type": "preemptible_instances",
                "title": get_translation("security.preemptible_instances.title", lang),
                "description": get_translation("security.preemptible_instances.description", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # EIPä½¿ç”¨æƒ…å†µæ£€æŸ¥
        if eip_info.get("total", 0) > 0:
            unbound_rate = eip_info.get("unbound_rate", 0)
            if unbound_rate > 20:
                checks.append({
                    "type": "eip_usage",
                    "title": get_translation("security.eip_usage.title", lang),
                    "description": get_translation("security.eip_usage.description_failed", lang),
                    "status": "warning",
                    "severity": "MEDIUM",
                    "total": eip_info.get("total", 0),
                    "bound": eip_info.get("bound", 0),
                    "unbound": eip_info.get("unbound", 0),
                    "unbound_rate": unbound_rate,
                    "resources": eip_info.get("unbound_eips", [])[:20],
                    "recommendation": get_translation("security.eip_usage.recommendation", lang, unbound=eip_info.get('unbound', 0)),
                })
            else:
                checks.append({
                    "type": "eip_usage",
                    "title": get_translation("security.eip_usage.title"),
                    "description": get_translation("security.eip_usage.description_passed", lang),
                    "status": "passed",
                    "severity": "INFO",
                    "total": eip_info.get("total", 0),
                    "bound": eip_info.get("bound", 0),
                    "unbound": eip_info.get("unbound", 0),
                    "unbound_rate": unbound_rate,
                    "resources": [],
                })
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type="security_checks", account_name=account_name, data=checks)
        
        return {
            "success": True,
            "data": checks,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": [],
            "cached": False,
        }


# ==================== Phase 2: Optimization APIs ====================

@router.get("/optimization/suggestions")
def get_optimization_suggestions(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
    locale: Optional[str] = Query("zh", description="è¯­è¨€è®¾ç½®: zh æˆ– en")
):
    """è·å–ä¼˜åŒ–å»ºè®®ï¼ˆå¸¦24å°æ—¶ç¼“å­˜ï¼‰"""
    # è·å–è¯­è¨€è®¾ç½®
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # åˆå§‹åŒ–ç¼“å­˜ç®¡ç†å™¨ï¼ŒTTLè®¾ç½®ä¸º24å°æ—¶ï¼ˆ86400ç§’ï¼‰
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # å°è¯•ä»ç¼“å­˜è·å–æ•°æ®ï¼ˆä¼˜å…ˆä½¿ç”¨ç¼“å­˜ï¼Œé¿å…é•¿æ—¶é—´ç­‰å¾…ï¼‰
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="optimization_suggestions", account_name=account_name)
    
    # å¦‚æœç¼“å­˜æœ‰æ•ˆï¼Œç›´æ¥ä½¿ç”¨ç¼“å­˜æ•°æ®
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    try:
        from cloudlens.core.optimization_engine import OptimizationEngine
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        from cloudlens.core.cost_analyzer import CostAnalyzer
        cm = ConfigManager()
        account_config = cm.get_account(account_name)
        
        suggestions = []
        all_opportunities = []
        
        # 1. ä½¿ç”¨ OptimizationEngine åˆ†æä¼˜åŒ–æœºä¼šï¼ˆè·³è¿‡ï¼Œå› ä¸ºå·²ç»æœ‰å…¶ä»–åˆ†æè¦†ç›–ï¼‰
        # è¿™ä¸ªæ“ä½œå¾ˆæ…¢ï¼Œè€Œä¸”å…¶ä»–åˆ†æå·²ç»è¦†ç›–äº†ä¸»è¦åœºæ™¯
        # try:
        #     engine = OptimizationEngine()
        #     opportunities = engine.analyze_optimization_opportunities(account_name)
        #     all_opportunities.extend(opportunities)
        # except Exception as e:
        #     pass  # å¦‚æœå¤±è´¥ï¼Œç»§ç»­å…¶ä»–åˆ†æ
        
        # 2. é—²ç½®èµ„æºå»ºè®®ï¼ˆåŸºäºçœŸå®æˆæœ¬ï¼Œä½¿ç”¨ç»Ÿä¸€çš„ç¼“å­˜é”®ï¼‰
        idle_data = cache_manager.get(resource_type="idle_result", account_name=account_name)
        # å¦‚æœæ²¡æœ‰ï¼Œå°è¯•ä» dashboard_idle ç¼“å­˜è·å–ï¼ˆå…¼å®¹æ—§ç¼“å­˜ï¼‰
        if not idle_data:
            idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account_name)
        if idle_data:
            # è®¡ç®—çœŸå®èŠ‚çœæ½œåŠ›
            total_savings = 0.0
            if account_config:
                cost_map = _get_cost_map("ecs", account_config)
                for idle_item in idle_data:
                    instance_id = idle_item.get("instance_id") or idle_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id)
                        if cost is None:
                            cost = _estimate_monthly_cost_from_spec(idle_item.get("spec", ""), "ecs")
                        total_savings += cost
            
            suggestions.append({
                "type": "idle_resources",
                "category": get_translation("optimization.idle_resources.category", lang),
                "priority": "high",
                "title": get_translation("optimization.idle_resources.title", lang),
                "description": get_translation("optimization.idle_resources.description", lang, count=len(idle_data)),
                "savings_potential": round(total_savings, 2),
                "resource_count": len(idle_data),
                "resources": idle_data[:10],  # è¿”å›å‰10ä¸ª
                "action": "release_or_downgrade",
                "recommendation": get_translation("optimization.idle_resources.recommendation", lang),
            })
        
        # 3. åœæ­¢å®ä¾‹å»ºè®® (Consistent Global View)
        try:
            instances, rds_list, redis_list = _get_global_resources_consistent(account_name, account_config, force_refresh)
        except Exception as e:
            logger.warning(f"è·å–ä¸€è‡´æ€§å…¨å±€èµ„æºå¤±è´¥ï¼Œä½¿ç”¨ç©ºåˆ—è¡¨: {e}")
            instances, rds_list, redis_list = [], [], []
        
        analyzer = SecurityComplianceAnalyzer()
        stopped = analyzer.check_stopped_instances(instances or [])
        if stopped:
            # è®¡ç®—åœæ­¢å®ä¾‹çš„æˆæœ¬
            stopped_savings = 0.0
            if account_config:
                cost_map = _get_cost_map("ecs", account_config)
                for stop_item in stopped:
                    instance_id = stop_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id)
                        if cost is None:
                            cost = 300  # é»˜è®¤ä¼°ç®—
                        # åœæ­¢å®ä¾‹ä»äº§ç”Ÿç£ç›˜è´¹ç”¨ï¼Œå‡è®¾å¯èŠ‚çœ70%
                        stopped_savings += cost * 0.7
            
            suggestions.append({
                "type": "stopped_instances",
                "category": get_translation("optimization.stopped_instances.category", lang),
                "priority": "medium",
                "title": get_translation("optimization.stopped_instances.title", lang),
                "description": get_translation("optimization.stopped_instances.description", lang, count=len(stopped)),
                "savings_potential": round(stopped_savings, 2),
                "resource_count": len(stopped),
                "resources": stopped[:10],
                "action": "release",
                "recommendation": get_translation("optimization.stopped_instances.recommendation", lang),
            })
        
        # 4. æœªç»‘å®šEIPå»ºè®®ï¼ˆä½¿ç”¨ç¼“å­˜ï¼Œé¿å…é‡å¤APIè°ƒç”¨ï¼‰
        try:
            # å°è¯•ä»ç¼“å­˜è·å–EIPåˆ—è¡¨
            eips_cache = cache_manager.get(resource_type="eip_list", account_name=account_name)
            if eips_cache:
                eips = eips_cache
            else:
                eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
                # ç¼“å­˜EIPåˆ—è¡¨ï¼ˆ5åˆ†é’Ÿæœ‰æ•ˆï¼‰
                if eips:
                    cache_manager.set(resource_type="eip_list", account_name=account_name, data=eips, ttl_seconds=300)
            
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
                unbound_eips = eip_info.get("unbound_eips", [])
                if unbound_eips:
                    # EIP è´¹ç”¨ä¼°ç®—ï¼šæ¯ä¸ªæœªç»‘å®šEIPçº¦20å…ƒ/æœˆ
                    eip_savings = len(unbound_eips) * 20
                    suggestions.append({
                        "type": "unbound_eips",
                        "category": get_translation("optimization.unbound_eips.category", lang),
                        "priority": "high",
                        "title": get_translation("optimization.unbound_eips.title", lang),
                        "description": get_translation("optimization.unbound_eips.description", lang, count=len(unbound_eips)),
                        "savings_potential": eip_savings,
                        "resource_count": len(unbound_eips),
                        "resources": unbound_eips[:10],
                        "action": "release",
                        "recommendation": get_translation("optimization.unbound_eips.recommendation", lang),
                    })
        except Exception as e:
            logger.warning(f"EIPåˆ†æå¤±è´¥: {e}")
            pass
        
        # 5. æ ‡ç­¾å®Œå–„å»ºè®®ï¼ˆéœ€è¦å®ä¾‹åˆ—è¡¨ï¼‰
        tag_coverage, no_tags = analyzer.check_missing_tags(instances or [])
        if len(no_tags) > 0:
            suggestions.append({
                "type": "missing_tags",
                "category": get_translation("optimization.missing_tags.category", lang),
                "priority": "medium",
                "title": get_translation("optimization.missing_tags.title", lang),
                "description": get_translation("optimization.missing_tags.description", lang, count=len(no_tags)),
                "savings_potential": 0,
                "resource_count": len(no_tags),
                "resources": no_tags[:10],
                "action": "add_tags",
                "recommendation": get_translation("optimization.missing_tags.recommendation", lang),
            })
        
        # 6. è§„æ ¼é™é…å»ºè®®ï¼ˆè·³è¿‡ï¼Œå› ä¸º OptimizationEngine å·²è¢«ç¦ç”¨ï¼‰
        # downgrade_opportunities = [opp for opp in all_opportunities if opp.get("action") == "downgrade"]
        # if downgrade_opportunities:
        #     total_downgrade_savings = sum(opp.get("estimated_savings", 0) for opp in downgrade_opportunities)
        #     suggestions.append({
        #         "type": "spec_downgrade",
        #         "category": get_translation("optimization.spec_downgrade.category", lang),
        #         "priority": "medium",
        #         "title": get_translation("optimization.spec_downgrade.title", lang),
        #         "description": get_translation("optimization.spec_downgrade.description", lang, count=len(downgrade_opportunities)),
        #         "savings_potential": round(total_downgrade_savings, 2),
        #         "resource_count": len(downgrade_opportunities),
        #         "resources": downgrade_opportunities[:10],
        #         "action": "downgrade",
        #         "recommendation": get_translation("optimization.spec_downgrade.recommendation", lang),
        #     })
        
        # 7. å…¬ç½‘æš´éœ²ä¼˜åŒ–å»ºè®®ï¼ˆéœ€è¦å®ä¾‹åˆ—è¡¨ï¼‰
        exposed = analyzer.detect_public_exposure(instances or [])
        if exposed:
            suggestions.append({
                "type": "public_exposure",
                "category": get_translation("optimization.public_exposure.category", lang),
                "priority": "high",
                "title": get_translation("optimization.public_exposure.title", lang),
                "description": get_translation("optimization.public_exposure.description", lang, count=len(exposed)),
                "savings_potential": 0,
                "resource_count": len(exposed),
                "resources": exposed[:10],
                "action": "secure",
                "recommendation": get_translation("optimization.public_exposure.recommendation", lang),
            })
        
        # 8. ç£ç›˜åŠ å¯†å»ºè®®ï¼ˆéœ€è¦å®ä¾‹åˆ—è¡¨ï¼‰
        encryption_info = analyzer.check_disk_encryption(instances or [])
        if encryption_info.get("encryption_rate", 100) < 50:
            suggestions.append({
                "type": "disk_encryption",
                "category": get_translation("optimization.disk_encryption.category", lang),
                "priority": "high",
                "title": get_translation("optimization.disk_encryption.title", lang),
                "description": get_translation("optimization.disk_encryption.description", lang, count=encryption_info.get('unencrypted_count', 0)),
                "savings_potential": 0,
                "resource_count": encryption_info.get("unencrypted_count", 0),
                "resources": encryption_info.get("unencrypted_instances", [])[:10],
                "action": "enable_encryption",
                "recommendation": get_translation("optimization.disk_encryption.recommendation", lang),
            })
        
        # æŒ‰ä¼˜å…ˆçº§å’ŒèŠ‚çœæ½œåŠ›æ’åº
        suggestions.sort(key=lambda x: (
            {"high": 0, "medium": 1, "low": 2}.get(x.get("priority", "low"), 2),
            -x.get("savings_potential", 0)
        ))
        
        # è®¡ç®—æ€»èŠ‚çœæ½œåŠ›
        total_savings = sum(s.get("savings_potential", 0) for s in suggestions)
        
        result = {
            "suggestions": suggestions,
            "summary": {
                "total_suggestions": len(suggestions),
                "total_savings_potential": round(total_savings, 2),
                "high_priority_count": sum(1 for s in suggestions if s.get("priority") == "high"),
                "medium_priority_count": sum(1 for s in suggestions if s.get("priority") == "medium"),
                "low_priority_count": sum(1 for s in suggestions if s.get("priority") == "low"),
            }
        }
        
        # ä¿å­˜åˆ°ç¼“å­˜ï¼ˆ24å°æ—¶æœ‰æ•ˆï¼‰
        cache_manager.set(resource_type="optimization_suggestions", account_name=account_name, data=result)
        
        return {
            "success": True,
            "data": result,
            "cached": False,
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"è·å–ä¼˜åŒ–å»ºè®®å¤±è´¥: {str(e)}\n{error_trace}")
        # å³ä½¿å‡ºé”™ä¹Ÿè¿”å›ç©ºç»“æœï¼Œè€Œä¸æ˜¯æŠ›å‡ºå¼‚å¸¸ï¼Œé¿å…å‰ç«¯å´©æºƒ
        return {
            "success": True,
            "data": {
                "suggestions": [],
                "summary": {
                    "total_suggestions": 0,
                    "total_savings_potential": 0,
                    "high_priority_count": 0,
                    "medium_priority_count": 0,
                    "low_priority_count": 0,
                }
            },
            "cached": False,
            "error": str(e)  # å¯é€‰ï¼šè¿”å›é”™è¯¯ä¿¡æ¯ç”¨äºè°ƒè¯•
        }


# ==================== Phase 3: Reports APIs ====================

@router.post("/reports/generate")
def generate_report(report_data: Dict[str, Any]):
    """ç”ŸæˆæŠ¥å‘Š"""
    account = report_data.get("account")
    report_type = report_data.get("type", "comprehensive")
    format_type = report_data.get("format", "excel")
    
    try:
        from cloudlens.core.report_generator import ReportGenerator
        
        provider, account_name = _get_provider_for_account(account)
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        
        # æ„å»ºæŠ¥å‘Šæ•°æ®
        data = {
            "account": account_name,
            "instances": instances,
            "rds": rds_list,
        }
        
        if format_type == "html":
            report_content = ReportGenerator.generate_html(account_name, data)
            return {
                "success": True,
                "data": {
                    "format": "html",
                    "content": report_content,
                }
            }
        elif format_type == "excel":
            # å®ç°Excelç”Ÿæˆ
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from fastapi.responses import StreamingResponse

            wb = Workbook()

            # èµ„æºæ¦‚è§ˆSheet
            ws_overview = wb.active
            ws_overview.title = "èµ„æºæ¦‚è§ˆ"

            # æ ‡é¢˜è¡Œæ ·å¼
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            # ECSæ¦‚è§ˆ
            ws_overview['A1'] = "CloudLens äº‘èµ„æºæŠ¥å‘Š"
            ws_overview['A1'].font = Font(size=16, bold=True)
            ws_overview.merge_cells('A1:E1')

            ws_overview['A3'] = "è´¦å·"
            ws_overview['B3'] = account_name

            # ECSå®ä¾‹åˆ—è¡¨
            ecs_start_row = 5
            headers = ['å®ä¾‹ID', 'åç§°', 'è§„æ ¼', 'çŠ¶æ€', 'åŒºåŸŸ']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_overview.cell(row=ecs_start_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # å¡«å……ECSæ•°æ®
            for row_idx, instance in enumerate(instances, ecs_start_row + 1):
                ws_overview.cell(row=row_idx, column=1, value=instance.id)
                ws_overview.cell(row=row_idx, column=2, value=instance.name)
                ws_overview.cell(row=row_idx, column=3, value=instance.spec)
                ws_overview.cell(row=row_idx, column=4, value=instance.status.value)
                ws_overview.cell(row=row_idx, column=5, value=instance.region)

            # RDSå®ä¾‹Sheet
            if rds_list:
                ws_rds = wb.create_sheet("RDSå®ä¾‹")
                rds_headers = ['å®ä¾‹ID', 'åç§°', 'å¼•æ“', 'ç‰ˆæœ¬', 'çŠ¶æ€', 'åŒºåŸŸ']
                for col_idx, header in enumerate(rds_headers, 1):
                    cell = ws_rds.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, rds in enumerate(rds_list, 2):
                    ws_rds.cell(row=row_idx, column=1, value=rds.id)
                    ws_rds.cell(row=row_idx, column=2, value=rds.name)
                    ws_rds.cell(row=row_idx, column=3, value=rds.spec)
                    ws_rds.cell(row=row_idx, column=4, value=getattr(rds, 'version', 'N/A'))
                    ws_rds.cell(row=row_idx, column=5, value=rds.status.value)
                    ws_rds.cell(row=row_idx, column=6, value=rds.region)

            # ä¿å­˜åˆ°BytesIO
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # è¿”å›æ–‡ä»¶æµ
            filename = f"cloudlens_report_{account_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            raise HTTPException(status_code=400, detail=f"ä¸æ”¯æŒçš„æ ¼å¼: {format_type}")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# Additional API endpoints to append to api.py

# ==================== Phase 2: Budget Management APIs ====================

@router.get("/cost/budget")
def get_budget(account: Optional[str] = None):
    """è·å–é¢„ç®—ä¿¡æ¯"""
    # TODO: å®ç°é¢„ç®—å­˜å‚¨å’ŒæŸ¥è¯¢ï¼ˆå¯ä»¥ä½¿ç”¨æ–‡ä»¶æˆ–æ•°æ®åº“ï¼‰
    return {
        "success": True,
        "data": {
            "monthly_budget": 0,
            "annual_budget": 0,
            "current_month_spent": 0,
            "usage_rate": 0,
        }
    }


# ==================== Billing APIs (BSS OpenAPI) ====================

@router.get("/billing/overview")
def get_billing_overview(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="è´¦æœŸï¼Œæ ¼å¼ yyyy-MMï¼Œé»˜è®¤å½“æœˆ"),
    product_code: Optional[str] = Query(None, description="äº§å“ä»£ç è¿‡æ»¤ï¼ˆå¯é€‰ï¼‰"),
    subscription_type: Optional[str] = Query(None, description="è®¢é˜…ç±»å‹è¿‡æ»¤ï¼šSubscription / PayAsYouGoï¼ˆå¯é€‰ï¼‰"),
):
    """
    è·å–è´¦å•æ¦‚è§ˆï¼ˆé˜¿é‡Œäº‘ BSS OpenAPIï¼‰ã€‚

    ç”¨é€”ï¼š
    - éªŒè¯å½“å‰è´¦å· AK æ˜¯å¦å…·å¤‡è´¦å•è¯»å–æƒé™
    - ä¸ºåç»­â€œæŒ‰å®ä¾‹çœŸå®æˆæœ¬ï¼ˆå«æŠ˜æ‰£ï¼‰â€æ‰“é€šæ•°æ®æº
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")

    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="å½“å‰ä»…æ”¯æŒé˜¿é‡Œäº‘è´¦å•ï¼ˆBSS OpenAPIï¼‰")

    # é»˜è®¤å½“æœˆè´¦æœŸ
    if not billing_cycle:
        from datetime import datetime
        billing_cycle = datetime.now().strftime("%Y-%m")

    # åŠ¨æ€å¯¼å…¥ï¼šé¿å…åœ¨æœªå®‰è£… SDK çš„ç¯å¢ƒä¸‹ç›´æ¥ import å¤±è´¥
    try:
        from aliyunsdkcore.client import AcsClient
        from aliyunsdkcore.request import CommonRequest
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"é˜¿é‡Œäº‘ SDK ä¸å¯ç”¨ï¼š{e}")

    try:
        import json

        # BSS OpenAPI ä¸åŒºåˆ†åœ°åŸŸï¼Œä½† SDK éœ€è¦ region å‚æ•°
        client = AcsClient(
            account_config.access_key_id,
            account_config.access_key_secret,
            "cn-hangzhou",
        )

        request = CommonRequest()
        request.set_domain("business.aliyuncs.com")
        request.set_version("2017-12-14")
        request.set_action_name("QueryBillOverview")
        request.set_method("POST")

        request.add_query_param("BillingCycle", billing_cycle)
        if product_code:
            request.add_query_param("ProductCode", product_code)
        if subscription_type:
            request.add_query_param("SubscriptionType", subscription_type)

        resp = client.do_action_with_exception(request)
        data = json.loads(resp)
        return {"success": True, "data": data}
    except Exception as e:
        # å¸¸è§ï¼šUnauthorizedOperation / Forbidden / InvalidAccessKeyId.NotFound / SignatureDoesNotMatch
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/billing/instance-bill")
def get_billing_instance_bill(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="è´¦æœŸ yyyy-MMï¼Œé»˜è®¤å½“æœˆ"),
    product_code: str = Query(..., description="äº§å“ä»£ç ï¼Œå¦‚ ecs/rds/kvstore/yundisk/snapshot/slb/eip/nat_gw"),
    subscription_type: Optional[str] = Query(None, description="Subscription / PayAsYouGoï¼ˆå¯é€‰ï¼‰"),
    limit: int = Query(50, ge=1, le=500, description="è¿”å›å‰ N æ¡ç”¨äºè°ƒè¯•"),
):
    """
    è°ƒè¯•æ¥å£ï¼šæ‹‰å– BSS QueryInstanceBill åŸå§‹æ•°æ®ï¼Œä¾¿äºç¡®è®¤ InstanceID çš„å­—æ®µä¸æ ¼å¼ã€‚
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="å½“å‰ä»…æ”¯æŒé˜¿é‡Œäº‘è´¦å•ï¼ˆBSS OpenAPIï¼‰")

    if not billing_cycle:
        billing_cycle = _get_billing_cycle_default()

    try:
        rows = _bss_query_instance_bill(
            account_config=account_config,
            billing_cycle=billing_cycle,
            product_code=product_code,
            subscription_type=subscription_type,
        )
        return {
            "success": True,
            "data": {
                "billing_cycle": billing_cycle,
                "product_code": product_code,
                "subscription_type": subscription_type,
                "count": len(rows),
                "items": rows[:limit],
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/discounts/trend")
def get_discount_trend(
    account: Optional[str] = Query(None, description="è´¦å·åç§°"),
    months: int = Query(19, ge=1, le=999, description="åˆ†ææœˆæ•°ï¼Œé»˜è®¤19ä¸ªæœˆï¼Œè®¾ç½®ä¸º99æˆ–æ›´å¤§è¡¨ç¤ºå…¨éƒ¨å†å²æ•°æ®"),
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
):
    """
    æŠ˜æ‰£è¶‹åŠ¿åˆ†æ - åŸºäºæ•°æ®åº“å…¨é‡æ•°æ®åˆ†æ
    
    æ•°æ®æ¥æºï¼šMySQLæ•°æ®åº“ï¼ˆè‡ªåŠ¨åŒæ­¥æœ€æ–°è´¦å•æ•°æ®ï¼‰
    æ”¯æŒï¼š
    - æŸ¥çœ‹é•¿æœŸæŠ˜æ‰£è¶‹åŠ¿ï¼ˆæœ€å¤š19ä¸ªæœˆå†å²ï¼‰
    - åˆ†æå•†åŠ¡åˆåŒæŠ˜æ‰£æ•ˆæœ
    - æŒ‰äº§å“/å®ä¾‹/åˆåŒç»´åº¦æŸ¥çœ‹æŠ˜æ‰£åˆ†å¸ƒ
    - å®æ—¶æ›´æ–°ï¼Œæ— éœ€æ‰‹åŠ¨ä¸‹è½½CSV
    """
    from cloudlens.core.discount_analyzer_db import DiscountAnalyzerDB
    import os
    
    try:
        # è·å–è´¦å·ä¿¡æ¯
        cm = ConfigManager()
        if not account:
            # å°è¯•è·å–å½“å‰è´¦å·
            ctx = ContextManager()
            account = ctx.get_last_account()
        if not account:
            # ä½¿ç”¨ç¬¬ä¸€ä¸ªè´¦å·
            accounts = cm.list_accounts()
            if accounts:
                account = accounts[0].name
            else:
                raise HTTPException(status_code=400, detail="æœªæ‰¾åˆ°è´¦å·é…ç½®")
        
        account_config = cm.get_account(account)
        if not account_config:
            raise HTTPException(status_code=404, detail=f"è´¦å· '{account}' ä¸å­˜åœ¨")
        
        # ç”Ÿæˆè´¦å·IDï¼ˆä¸bill_fetcherä¿æŒä¸€è‡´ï¼‰
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        # ä½¿ç”¨æ•°æ®åº“ç‰ˆæŠ˜æ‰£åˆ†æå™¨ï¼ˆé»˜è®¤ä½¿ç”¨MySQLï¼‰
        analyzer = DiscountAnalyzerDB()
        
        # åˆ†ææŠ˜æ‰£è¶‹åŠ¿
        result = analyzer.analyze_discount_trend(
            account_id=account_id,
            months=months
        )
        
        if 'error' in result:
            return {
                "success": False,
                "error": result['error']
            }
        
        # è½¬æ¢æ•°æ®æ ¼å¼ä»¥åŒ¹é…å‰ç«¯æœŸæœ›çš„ç»“æ„
        from datetime import datetime
        
        # æå–æ•°æ®
        monthly_trend = result.get('monthly_trend', [])
        product_discounts = result.get('product_discounts', [])
        instance_discounts = result.get('instance_discounts', [])
        contract_discounts = result.get('contract_discounts', [])
        summary = result.get('summary', {})
        
        # æ„å»ºå‰ç«¯æœŸæœ›çš„æ•°æ®ç»“æ„
        response_data = {
            "account_name": account,
            "analysis_periods": [m['month'] for m in monthly_trend],
            
            # trend_analysis æ ¼å¼
            "trend_analysis": {
                "timeline": [
                    {
                        "period": m['month'],
                        "official_price": m['official_price'],
                        "discount_amount": m['discount_amount'],
                        "discount_rate": m['discount_rate'],
                        "payable_amount": m['actual_amount']
                    }
                    for m in monthly_trend
                ],
                "latest_period": monthly_trend[-1]['month'] if monthly_trend else "",
                "latest_discount_rate": summary.get('latest_discount_rate', 0),
                "discount_rate_change": summary.get('trend_change_pct', 0) / 100,
                "discount_rate_change_pct": summary.get('trend_change_pct', 0),
                "discount_amount_change": 0,  # å¯ä»¥è®¡ç®—
                "trend_direction": summary.get('trend', 'å¹³ç¨³'),
                "average_discount_rate": summary.get('avg_discount_rate', 0),
                "max_discount_rate": max([m['discount_rate'] for m in monthly_trend], default=0),
                "min_discount_rate": min([m['discount_rate'] for m in monthly_trend], default=0),
                "total_savings_6m": summary.get('total_discount', 0),
            },
            
            # product_analysis æ ¼å¼
            "product_analysis": {
                p['product']: {
                    "total_discount": p['discount_amount'],
                    "avg_discount_rate": p['discount_rate'],
                    "latest_discount_rate": p['discount_rate'],
                    "rate_change": 0,
                    "trend": "å¹³ç¨³",
                    "periods": [m['month'] for m in monthly_trend],
                    "discount_rates": [p['discount_rate']] * len(monthly_trend),
                }
                for p in product_discounts
            },
            
            # contract_analysis æ ¼å¼ï¼ˆå¦‚æœæœ‰åˆåŒæ•°æ®ï¼‰
            "contract_analysis": {
                c['contract_name']: {
                    "discount_name": c['contract_name'],
                    "total_discount": c.get('total_discount', 0),
                    "avg_discount_rate": c.get('avg_discount_rate', 0),
                    "latest_discount_rate": c.get('latest_discount_rate', 0),
                    "periods": c.get('periods', []),
                    "discount_amounts": c.get('discount_amounts', []),
                }
                for c in contract_discounts
            },
            
            # top_instance_discounts æ ¼å¼
            "top_instance_discounts": [
                {
                    "instance_id": i['instance_id'],
                    "instance_name": i.get('instance_name', i['instance_id']),
                    "product_name": i['product'],
                    "official_price": i['official_price'],
                    "discount_amount": i['discount_amount'],
                    "payable_amount": i['actual_amount'],
                    "discount_rate": i['discount_rate'],
                    "discount_pct": i['discount_rate'] * 100,
                }
                for i in instance_discounts
            ],
            
            "generated_at": datetime.now().isoformat(),
        }
        
        return {
            "success": True,
            "data": response_data,
            "cached": False,
            "source": "database",
            "account": account,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/discounts/products")
def get_product_discounts(
    account: Optional[str] = Query(None, description="è´¦å·åç§°"),
    product: Optional[str] = Query(None, description="äº§å“åç§°è¿‡æ»¤"),
    months: int = Query(19, ge=1, le=999, description="åˆ†ææœˆæ•°ï¼Œè®¾ç½®ä¸º99æˆ–æ›´å¤§è¡¨ç¤ºå…¨éƒ¨å†å²æ•°æ®"),
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
):
    """
    äº§å“æŠ˜æ‰£è¯¦æƒ… - åŸºäºæ•°æ®åº“æŸ¥çœ‹ç‰¹å®šäº§å“çš„æŠ˜æ‰£æ˜ç»†
    """
    from cloudlens.core.discount_analyzer_db import DiscountAnalyzerDB
    import os
    
    try:
        # è·å–è´¦å·ä¿¡æ¯
        cm = ConfigManager()
        if not account:
            ctx = ContextManager()
            account = ctx.get_last_account()
        if not account:
            accounts = cm.list_accounts()
            if accounts:
                account = accounts[0].name
            else:
                raise HTTPException(status_code=400, detail="æœªæ‰¾åˆ°è´¦å·é…ç½®")
        
        account_config = cm.get_account(account)
        if not account_config:
            raise HTTPException(status_code=404, detail=f"è´¦å· '{account}' ä¸å­˜åœ¨")
        
        # ç”Ÿæˆè´¦å·ID
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        # ä½¿ç”¨æ•°æ®åº“ç‰ˆæŠ˜æ‰£åˆ†æå™¨ï¼ˆé»˜è®¤ä½¿ç”¨MySQLï¼‰
        analyzer = DiscountAnalyzerDB()
        
        result = analyzer.analyze_discount_trend(account_id=account_id, months=months)
        
        if 'error' in result:
            return {"success": False, "error": result['error']}
        
        # æå–äº§å“æŠ˜æ‰£æ•°æ®
        product_data = result['product_analysis']
        
        # å¦‚æœæŒ‡å®šäº†äº§å“è¿‡æ»¤
        if product:
            product_data = {k: v for k, v in product_data.items() if product.lower() in k.lower()}
        
        return {
            "success": True,
            "data": {
                "products": product_data,
                "analysis_periods": result['analysis_periods'],
            },
            "source": "database",
            "account": account,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/billing/discounts")
def get_billing_discounts(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="è´¦æœŸ yyyy-MMï¼Œé»˜è®¤å½“æœˆ"),
    force_refresh: bool = Query(False, description="å¼ºåˆ¶åˆ·æ–°ç¼“å­˜"),
):
    """
    æŠ˜æ‰£æ¢³ç†ï¼ˆæŒ‰äº§å“ + è®¡è´¹æ–¹å¼èšåˆï¼‰- åŸºäºBSS APIå®æ—¶æŸ¥è¯¢
    
    æ³¨æ„ï¼šè¿™æ˜¯å®æ—¶APIæ¥å£ï¼Œä¸ /discounts/trendï¼ˆåŸºäºCSVç¦»çº¿åˆ†æï¼‰äº’è¡¥ã€‚
    - å®æ—¶APIï¼šæŸ¥è¯¢å½“å‰æœˆæŠ˜æ‰£æƒ…å†µ
    - CSVåˆ†æï¼šæŸ¥çœ‹å†å²6ä¸ªæœˆæŠ˜æ‰£è¶‹åŠ¿

    å£å¾„è¯´æ˜ï¼š
    - PretaxGrossAmountï¼šç¨å‰åŸä»·ï¼ˆæœªæŠ˜æ‰£/æœªä¼˜æƒ æŠµæ‰£å‰ï¼‰
    - PretaxAmountï¼šç¨å‰åº”ä»˜ï¼ˆæŠ˜æ‰£/ä¼˜æƒ æŠµæ‰£åï¼‰
    - å¯¹ PayAsYouGoï¼ŒPaymentAmount å¸¸ä¸º 0ï¼ˆæœªå‡ºè´¦/æœªç»“ç®—ï¼‰ï¼Œè¯·ä»¥ PretaxAmount/OutstandingAmount ä¸ºä¸»
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="å½“å‰ä»…æ”¯æŒé˜¿é‡Œäº‘è´¦å•ï¼ˆBSS OpenAPIï¼‰")

    if not billing_cycle:
        billing_cycle = _get_billing_cycle_default()

    cache_manager = CacheManager(ttl_seconds=86400)
    cache_key = f"billing_discounts_{billing_cycle}"
    if not force_refresh:
        cached = cache_manager.get(resource_type=cache_key, account_name=account_config.name)
        if isinstance(cached, dict) and cached.get("billing_cycle") == billing_cycle:
            return {"success": True, "data": cached, "cached": True}

    def f(x: Any) -> float:
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    items = _bss_query_bill_overview(account_config, billing_cycle)

    agg: Dict[str, Dict[str, Any]] = {}
    for it in items:
        code = (it.get("ProductCode") or it.get("PipCode") or "unknown")
        sub = (it.get("SubscriptionType") or "Unknown")
        key = f"{code}::{sub}"
        if key not in agg:
            agg[key] = {
                "product_code": code,
                "product_name": it.get("ProductName") or code,
                "subscription_type": sub,
                "pretax_gross_amount": 0.0,
                "pretax_amount": 0.0,
                "payment_amount": 0.0,
                "outstanding_amount": 0.0,
                "invoice_discount": 0.0,
                "round_down_discount": 0.0,
                "deducted_by_coupons": 0.0,
                "deducted_by_cash_coupons": 0.0,
                "deducted_by_prepaid_card": 0.0,
                "adjust_amount": 0.0,
                "cash_amount": 0.0,
                "currency": it.get("Currency") or "CNY",
            }

        row = agg[key]
        row["pretax_gross_amount"] += f(it.get("PretaxGrossAmount"))
        row["pretax_amount"] += f(it.get("PretaxAmount"))
        row["payment_amount"] += f(it.get("PaymentAmount"))
        row["outstanding_amount"] += f(it.get("OutstandingAmount"))
        row["invoice_discount"] += f(it.get("InvoiceDiscount"))
        row["round_down_discount"] += f(it.get("RoundDownDiscount"))
        row["deducted_by_coupons"] += f(it.get("DeductedByCoupons"))
        row["deducted_by_cash_coupons"] += f(it.get("DeductedByCashCoupons"))
        row["deducted_by_prepaid_card"] += f(it.get("DeductedByPrepaidCard"))
        row["adjust_amount"] += f(it.get("AdjustAmount"))
        row["cash_amount"] += f(it.get("CashAmount"))

    rows = []
    total_gross = 0.0
    total_pretax = 0.0
    total_discount_amount = 0.0
    for row in agg.values():
        gross = float(row.get("pretax_gross_amount") or 0.0)
        pretax = float(row.get("pretax_amount") or 0.0)
        payment_amount = float(row.get("payment_amount") or 0.0)
        outstanding_amount = float(row.get("outstanding_amount") or 0.0)
        invoice_discount = float(row.get("invoice_discount") or 0.0)
        round_down_discount = float(row.get("round_down_discount") or 0.0)
        deducted_by_coupons = float(row.get("deducted_by_coupons") or 0.0)
        deducted_by_cash_coupons = float(row.get("deducted_by_cash_coupons") or 0.0)
        deducted_by_prepaid_card = float(row.get("deducted_by_prepaid_card") or 0.0)
        adjust_amount = float(row.get("adjust_amount") or 0.0)
        cash_amount = float(row.get("cash_amount") or 0.0)

        # â€œæ²¡æœ‰ä½¿ç”¨/æ²¡æœ‰å‘ç”Ÿè´¹ç”¨â€çš„äº§å“ä¸å±•ç¤ºï¼šæ‰€æœ‰é‡‘é¢å­—æ®µå‡ä¸º 0
        # æ³¨æ„ï¼šå¦‚æœæ˜¯â€œå…å•/å…¨é¢å‡å…â€ï¼Œä¸€èˆ¬è¡¨ç°ä¸º gross>0 ä½† pretax=0ï¼Œè¿™ç§è¦å±•ç¤ºä¸ºâ€œå…è´¹â€
        has_any_amount = any(
            abs(x) > 0
            for x in (
                gross,
                pretax,
                payment_amount,
                outstanding_amount,
                invoice_discount,
                round_down_discount,
                deducted_by_coupons,
                deducted_by_cash_coupons,
                deducted_by_prepaid_card,
                adjust_amount,
                cash_amount,
            )
        )
        if not has_any_amount:
            continue

        discount_amount = max(0.0, gross - pretax) if gross > 0 else 0.0
        # æŠ˜æ‰£å£å¾„ï¼šæŒ‰â€œå®ä»˜æ¯”ä¾‹â€è®¡ç®—æŠ˜æ‰£ï¼ˆx.xæŠ˜ï¼‰ï¼Œä¾‹å¦‚ 30/100 => 0.3 => 3.0æŠ˜
        # æ³¨æ„ï¼š0.0æŠ˜é€šå¸¸æ„å‘³ç€â€œå…¨é¢å‡å…/å®Œå…¨è¢«ä¼˜æƒ æŠµæ‰£â€ï¼Œä¸åº”å±•ç¤ºä¸º 0.0æŠ˜ï¼Œæ”¹ç”¨ free æ ‡è¯†åœ¨å‰ç«¯å±•ç¤ºâ€œå…è´¹â€
        discount_rate = (pretax / gross) if gross > 0 else None
        is_free = (gross > 0 and pretax == 0)
        discount_zhe = (float(discount_rate) * 10.0) if (discount_rate is not None and pretax > 0) else None
        row["pretax_gross_amount"] = round(gross, 2)
        row["pretax_amount"] = round(pretax, 2)
        row["discount_amount"] = round(discount_amount, 2)
        row["discount_rate"] = round(float(discount_rate), 6) if discount_rate is not None else None
        row["discount_pct"] = round((1.0 - float(discount_rate)) * 100, 2) if discount_rate is not None else None
        row["discount_zhe"] = round(float(discount_zhe), 1) if discount_zhe is not None else None
        row["free"] = bool(is_free)
        row["payment_amount"] = round(payment_amount, 2)
        row["outstanding_amount"] = round(outstanding_amount, 2)
        row["invoice_discount"] = round(invoice_discount, 2)
        row["round_down_discount"] = round(round_down_discount, 6)
        row["deducted_by_coupons"] = round(deducted_by_coupons, 2)
        row["deducted_by_cash_coupons"] = round(deducted_by_cash_coupons, 2)
        row["deducted_by_prepaid_card"] = round(deducted_by_prepaid_card, 2)
        row["adjust_amount"] = round(adjust_amount, 2)
        row["cash_amount"] = round(cash_amount, 2)

        total_gross += gross
        total_pretax += pretax
        total_discount_amount += discount_amount
        rows.append(row)

    rows.sort(key=lambda r: float(r.get("discount_amount") or 0.0), reverse=True)

    overall_rate = (total_pretax / total_gross) if total_gross > 0 else None
    overall_free = (total_gross > 0 and total_pretax == 0)
    overall_zhe = (float(overall_rate) * 10.0) if (overall_rate is not None and total_pretax > 0) else None
    result = {
        "billing_cycle": billing_cycle,
        "summary": {
            "total_pretax_gross_amount": round(total_gross, 2),
            "total_pretax_amount": round(total_pretax, 2),
            "total_discount_amount": round(total_discount_amount, 2),
            "discount_rate": round(float(overall_rate), 6) if overall_rate is not None else None,
            "discount_pct": round((1.0 - float(overall_rate)) * 100, 2) if overall_rate is not None else None,
            "discount_zhe": round(float(overall_zhe), 1) if overall_zhe is not None else None,
            "free": bool(overall_free),
        },
        "rows": rows,
    }

    cache_manager.set(resource_type=cache_key, account_name=account_config.name, data=result)
    return {"success": True, "data": result, "cached": False}


@router.post("/cost/budget")
def set_budget(budget_data: Dict[str, Any]):
    """è®¾ç½®é¢„ç®—"""
    # TODO: å®ç°é¢„ç®—ä¿å­˜
    return {
        "success": True,
        "message": "é¢„ç®—è®¾ç½®æˆåŠŸ"
    }


# ==================== Phase 2: CIS Compliance APIs ====================

@router.get("/security/cis")
def get_cis_compliance(account: Optional[str] = None):
    """è·å–CISåˆè§„æ£€æŸ¥ç»“æœ"""
    provider, account_name = _get_provider_for_account(account)
    
    try:
        from cloudlens.core.cis_compliance import CISBenchmark
        
        # è·å–èµ„æº
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        all_resources = instances + rds_list
        
        # è¿è¡ŒCISæ£€æŸ¥
        checker = CISBenchmark()
        results = checker.run_all_checks(all_resources, provider)
        
        # è®¡ç®—åˆè§„åº¦
        total_checks = len(results.get("results", []))
        passed_checks = sum(1 for check in results.get("results", []) if check.get("status") == "PASS")
        compliance_rate = (passed_checks / total_checks * 100) if total_checks > 0 else 0
        
        return {
            "success": True,
            "data": {
                "compliance_rate": round(compliance_rate, 2),
                "checks": results.get("results", []),
                "summary": results.get("summary", {}),
            }
        }
    except Exception as e:
        # å¦‚æœCISæ£€æŸ¥å™¨ä¸å­˜åœ¨æˆ–å‡ºé”™ï¼Œè¿”å›æç¤º
        return {
            "success": True,
            "data": {
                "compliance_rate": 0,
                "checks": [],
                "message": f"CISåˆè§„æ£€æŸ¥åŠŸèƒ½å¼€å‘ä¸­: {str(e)}"
            }
        }




@router.get("/reports")
def list_reports(account: Optional[str] = None, limit: int = Query(50, ge=1, le=100)):
    """è·å–æŠ¥å‘Šå†å²åˆ—è¡¨"""
    # TODO: å®ç°æŠ¥å‘Šå†å²å­˜å‚¨å’ŒæŸ¥è¯¢
    return {
        "success": True,
        "data": []
    }


# ==================== Phase 1: é«˜çº§æŠ˜æ‰£åˆ†æAPI ====================

@router.get("/discounts/quarterly")
def get_quarterly_discount_comparison(
    account: Optional[str] = Query(None, description="è´¦å·åç§°"),
    quarters: int = Query(8, ge=1, le=20, description="åˆ†æå­£åº¦æ•°"),
    start_date: Optional[str] = Query(None, description="å¼€å§‹æ—¥æœŸ (YYYY-MMæ ¼å¼)"),
    end_date: Optional[str] = Query(None, description="ç»“æŸæ—¥æœŸ (YYYY-MMæ ¼å¼)"),
):
    """
    å­£åº¦æŠ˜æ‰£å¯¹æ¯”åˆ†æ
    
    è¿”å›å­£åº¦ç»´åº¦çš„æŠ˜æ‰£ç‡ã€æ¶ˆè´¹é‡‘é¢ã€ç¯æ¯”å˜åŒ–ç­‰æ•°æ®
    """
    import os
    from cloudlens.core.discount_analyzer_advanced import AdvancedDiscountAnalyzer
    
    cm = ConfigManager()
    
    # è§£æè´¦å·
    if not account:
        accounts = cm.list_accounts()
        if not accounts:
            raise HTTPException(status_code=404, detail="æœªæ‰¾åˆ°ä»»ä½•è´¦å·é…ç½®")
        account = accounts[0].name if isinstance(accounts[0], CloudAccount) else accounts[0].get('name')
    
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"è´¦å· '{account}' æœªæ‰¾åˆ°")
    
    try:
        # ä½¿ç”¨é»˜è®¤é…ç½®ï¼ˆMySQLï¼‰ï¼Œä¸å†éœ€è¦db_path
        analyzer = AdvancedDiscountAnalyzer()
        
        # æ„é€ è´¦å·IDï¼ˆä¸bill_cmd.pyä¿æŒä¸€è‡´ï¼‰
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        result = analyzer.get_quarterly_comparison(account_id, quarters, start_date, end_date)
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', 'åˆ†æå¤±è´¥'))
        
        return {
            "success": True,
            "data": result,
            "account": account,
            "source": "database"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"å­£åº¦å¯¹æ¯”åˆ†æå¤±è´¥: {str(e)}")


@router.get("/discounts/yearly")
def get_yearly_discount_comparison(
    account: Optional[str] = Query(None, description="è´¦å·åç§°"),
    start_date: Optional[str] = Query(None, description="å¼€å§‹æ—¥æœŸ (YYYY-MMæ ¼å¼)"),
    end_date: Optional[str] = Query(None, description="ç»“æŸæ—¥æœŸ (YYYY-MMæ ¼å¼)"),
):
    """
    å¹´åº¦æŠ˜æ‰£å¯¹æ¯”åˆ†æ
    
    è¿”å›å¹´åº¦ç»´åº¦çš„æŠ˜æ‰£ç‡ã€æ¶ˆè´¹é‡‘é¢ã€åŒæ¯”å˜åŒ–ç­‰æ•°æ®
    """
    import os
    from cloudlens.core.discount_analyzer_advanced import AdvancedDiscountAnalyzer
    
    cm = ConfigManager()
    
    # è§£æè´¦å·
    if not account:
        accounts = cm.list_accounts()
        if not accounts:
            raise HTTPException(status_code=404, detail="æœªæ‰¾åˆ°ä»»ä½•è´¦å·é…ç½®")
        account = accounts[0].name if isinstance(accounts[0], CloudAccount) else accounts[0].get('name')
    
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"è´¦å· '{account}' æœªæ‰¾åˆ°")
    
    try:
        # ä½¿ç”¨é»˜è®¤é…ç½®ï¼ˆMySQLï¼‰ï¼Œä¸å†éœ€è¦db_path
        analyzer = AdvancedDiscountAnalyzer()
        
        # æ„é€ è´¦å·IDï¼ˆä¸bill_cmd.pyä¿æŒä¸€è‡´ï¼‰
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        result = analyzer.get_yearly_comparison(account_id, start_date, end_date)
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', 'åˆ†æå¤±è´¥'))
        
        return {
            "success": True,
            "data": result,
            "account": account,
            "source": "database"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"å¹´åº¦å¯¹æ¯”åˆ†æå¤±è´¥: {str(e)}")


@router.get("/discounts/product-trends")
def get_product_discount_trends(
    account: Optional[str] = Query(None, description="è´¦å·åç§°"),
    months: int = Query(19, ge=1, le=999, description="åˆ†ææœˆæ•°"),
    top_n: int = Query(20, ge=1, le=50, description="TOP Näº§å“"),
    start_date: Optional[str] = Query(None, description="å¼€å§‹æ—¥æœŸ (YYYY-MMæ ¼å¼)"),
    end_date: Optional[str] = Query(None, description="ç»“æŸæ—¥æœŸ (YYYY-MMæ ¼å¼)"),
):
    """
    äº§å“æŠ˜æ‰£è¶‹åŠ¿åˆ†æ
    
    è¿”å›æ¯ä¸ªäº§å“çš„æœˆåº¦æŠ˜æ‰£è¶‹åŠ¿ã€æ³¢åŠ¨ç‡ã€è¶‹åŠ¿å˜åŒ–ç­‰æ•°æ®
    """
    import os

# ==================== (Migrated to api_discounts.py) ====================



# ==================== è™šæ‹Ÿæ ‡ç­¾ç³»ç»Ÿ API ====================


# ==================== (Migrated to api_tags.py) ====================




# ==================== (Migrated to api_budgets.py) ====================




# ==================== (Migrated to api_dashboards.py) ====================




# ==================== (Migrated to api_dashboards.py) ====================

