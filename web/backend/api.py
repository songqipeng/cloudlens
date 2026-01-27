
from fastapi import APIRouter, HTTPException, Query, BackgroundTasks, Body, Depends, Request
from typing import List, Dict, Optional, Any
from datetime import datetime, timedelta
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import sys
import logging

logger = logging.getLogger(__name__)

# 创建限流器（使用 IP 地址作为 key）
limiter = Limiter(key_func=get_remote_address)
from cloudlens.core.config import ConfigManager, CloudAccount
from web.backend.i18n import get_translation, get_locale_from_request, Locale
from cloudlens.core.context import ContextManager
from cloudlens.core.cost_trend_analyzer import CostTrendAnalyzer
from cloudlens.core.cache import CacheManager  # MySQL缓存管理器（统一使用）
from cloudlens.core.rules_manager import RulesManager
from cloudlens.core.services.analysis_service import AnalysisService
from cloudlens.core.virtual_tags import VirtualTagStorage, VirtualTag, TagRule, TagEngine
from cloudlens.core.bill_storage import BillStorageManager
from web.backend.api_cost import _get_billing_overview_totals
from cloudlens.core.dashboard_manager import DashboardStorage, Dashboard, WidgetConfig
from web.backend.api_resources import _get_cost_map, _estimate_monthly_cost, _estimate_monthly_cost_from_spec
from web.backend.error_handler import api_error_handler
from pydantic import BaseModel

router = APIRouter(prefix="/api")

class AccountInfo(BaseModel):
    name: str
    region: str
    access_key_id: str

class AccountUpdateRequest(BaseModel):
    """账号更新请求"""
    alias: Optional[str] = None
    provider: Optional[str] = None
    region: Optional[str] = None
    access_key_id: Optional[str] = None
    access_key_secret: Optional[str] = None

class AccountCreateRequest(BaseModel):
    """账号创建请求"""
    name: str
    alias: Optional[str] = None
    provider: str = "aliyun"
    region: str = "cn-hangzhou"
    access_key_id: str
    access_key_secret: str

class DashboardSummary(BaseModel):
    account: str
    total_cost: float
    idle_count: int
    cost_trend: str
    trend_pct: float

class TriggerAnalysisRequest(BaseModel):
    account: str
    days: int = 7
    force: bool = True

@router.get("/accounts")
@limiter.limit("100/minute")
def list_accounts(request: Request) -> List[Dict]:
    """List all configured accounts (限流: 100次/分钟)"""
    cm = ConfigManager()
    accounts = cm.list_accounts()
    result = []
    for account in accounts:
        if isinstance(account, CloudAccount):
            result.append({
                "name": account.name,
                "region": account.region,
                "access_key_id": account.access_key_id,
            })
    return result

@router.get("/config/rules")
def get_rules() -> Dict[str, Any]:
    """Get current optimization rules"""
    rm = RulesManager()
    return rm.get_rules()

@router.post("/config/rules")
def set_rules(rules: Dict[str, Any]):
    """Update optimization rules"""
    rm = RulesManager()
    try:
        rm.set_rules(rules)
        return {"status": "success", "message": "Rules updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/config/notifications")
def get_notification_config() -> Dict[str, Any]:
    """获取通知配置（SMTP等）"""
    import json
    import os
    from pathlib import Path
    
    config_dir = Path(os.path.expanduser("~/.cloudlens"))
    config_file = config_dir / "notifications.json"
    
    default_config = {
        "email": "",  # 发件邮箱（简化配置）
        "auth_code": "",  # 授权码/密码
        "default_receiver_email": "",  # 默认接收邮箱（告警通知的目标邮箱）
        "smtp_host": "",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_password": "",
        "smtp_from": ""
    }
    
    if not config_file.exists():
        return default_config
    
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
            # 合并默认值，确保所有字段都存在
            result = {**default_config, **config}
            # 如果已有旧格式配置，转换为新格式
            if result.get("smtp_user") and not result.get("email"):
                result["email"] = result.get("smtp_user", "")
            if result.get("smtp_password") and not result.get("auth_code"):
                result["auth_code"] = result.get("smtp_password", "")
            return result
    except Exception as e:
        logger.error(f"读取通知配置失败: {e}")
        return default_config

def _get_smtp_config_by_email(email: str) -> Dict[str, Any]:
    """根据邮箱地址自动获取SMTP配置"""
    email_lower = email.lower().strip()
    
    # QQ邮箱
    if email_lower.endswith("@qq.com"):
        return {
            "smtp_host": "smtp.qq.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }
    # Gmail
    elif email_lower.endswith("@gmail.com"):
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }
    # 163邮箱
    elif email_lower.endswith("@163.com"):
        return {
            "smtp_host": "smtp.163.com",
            "smtp_port": 465,
            "smtp_use_tls": False,
            "smtp_use_ssl": True
        }
    # 126邮箱
    elif email_lower.endswith("@126.com"):
        return {
            "smtp_host": "smtp.126.com",
            "smtp_port": 465,
            "smtp_use_tls": False,
            "smtp_use_ssl": True
        }
    # 默认配置（通用SMTP）
    else:
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_use_tls": True
        }

@router.post("/config/notifications")
def set_notification_config(config: Dict[str, Any]):
    """保存通知配置（SMTP等）"""
    import json
    import os
    from pathlib import Path
    
    config_dir = Path(os.path.expanduser("~/.cloudlens"))
    config_file = config_dir / "notifications.json"
    
    try:
        if not config_dir.exists():
            config_dir.mkdir(parents=True, exist_ok=True)
        
        # 获取用户输入的邮箱和授权码
        email = config.get("email", "").strip()
        auth_code = config.get("auth_code", "").strip()
        default_receiver_email = config.get("default_receiver_email", "").strip()
        
        # 根据邮箱自动配置SMTP
        smtp_config = {}
        if email:
            smtp_config = _get_smtp_config_by_email(email)
            smtp_config["smtp_user"] = email
            smtp_config["smtp_from"] = email
            smtp_config["smtp_password"] = auth_code
        
        # 保存完整配置（包含自动生成的SMTP配置）
        full_config = {
            "email": email,
            "auth_code": auth_code,  # 保存授权码用于显示（但实际使用smtp_password）
            "default_receiver_email": default_receiver_email,
            **smtp_config
        }
        
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(full_config, f, indent=2, ensure_ascii=False)
        
        return {"status": "success", "message": "通知配置已更新"}
    except Exception as e:
        logger.error(f"保存通知配置失败: {e}")
        raise HTTPException(status_code=500, detail=f"保存通知配置失败: {str(e)}")

@router.post("/analyze/trigger")
def trigger_analysis(req: TriggerAnalysisRequest, background_tasks: BackgroundTasks):
    """Trigger idle resource analysis"""
    # For MVP, we run it synchronously to provide immediate feedback, 
    # but in production this should be a background task. 
    # To avoid timeout for long requests, we could use background_tasks.
    # But for now, let's try synchronous for simplicity as user requested "Scan Now" button.
    # Actually, user might wait 10-20s.
    try:
        data, cached = AnalysisService.analyze_idle_resources(req.account, req.days, req.force)
        
        # 清除 dashboard_idle 缓存，确保 dashboard 页面能获取最新数据
        cache_manager = CacheManager(ttl_seconds=86400)
        cache_manager.clear(resource_type="dashboard_idle", account_name=req.account)
        # 同时更新 dashboard_idle 缓存
        cache_manager.set(resource_type="dashboard_idle", account_name=req.account, data=data)
        
        return {
            "status": "success", 
            "count": len(data), 
            "cached": cached,
            "data": data
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"触发分析失败: {str(e)}\n{error_trace}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard/summary")
@api_error_handler
async def get_summary(account: Optional[str] = None, force_refresh: bool = Query(False, description="强制刷新缓存")):
    """Get dashboard summary metrics（带24小时缓存）"""
    import logging
    logger = logging.getLogger(__name__)
    
    cm = ConfigManager()
    
    # Resolve account - 必须明确指定账号，不允许自动选择
    if not account:
        raise HTTPException(status_code=400, detail="账号参数是必需的，请在前端选择账号")
    
    # 调试日志
    logger.info(f"[get_summary] 收到账号参数: {account}, force_refresh: {force_refresh}")
    logger.debug(f"收到账号参数: {account}, force_refresh: {force_refresh}")
    logger.debug(f"收到账号参数: {account}, force_refresh: {force_refresh}")

    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="dashboard_summary", account_name=account)
        if cached_result:
            logger.debug(f"使用缓存数据，账号: {account}")
            # 但是需要确保 idle_count 是最新的（从闲置资源缓存中重新获取）
            idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account)
            if not idle_data:
                idle_data = cache_manager.get(resource_type="idle_result", account_name=account)
            # 如果缓存中有闲置资源数据，更新 idle_count
            if idle_data:
                idle_count = len(idle_data) if idle_data else 0
                cached_result["idle_count"] = idle_count
                logger.info(f"从缓存更新 idle_count: {idle_count} (账号: {account}, 闲置资源数据: {len(idle_data) if idle_data else 0} 条)")
            else:
                logger.warning(f"缓存中没有闲置资源数据，idle_count 保持原值: {cached_result.get('idle_count', 0)} (账号: {account})")
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        return {
            **cached_result,
            "cached": True,
        }

    logger.debug(f"缓存未命中，直接获取真实数据，账号: {account}")
    account_config = cm.get_account(account)
    if not account_config:
        print(f"[DEBUG get_summary] 账号 '{account}' 未找到")
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    
    # 直接获取资源统计（不依赖复杂的缓存逻辑）
    logger.info(f"[get_summary] 直接获取资源数据: {account}")
    
    # 初始化结果
    total_cost = 0.0
    trend = "暂无数据"
    trend_pct = 0.0
    idle_count = 0
    total_resources = 0
    resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
    tag_coverage = 0.0
    savings_potential = 0.0
    
    try:
        # 1. 获取资源统计 - 直接调用list_resources函数
        from web.backend.api import list_resources as get_resources
        
        # 获取ECS数量
        try:
            ecs_result = await get_resources(type="ecs", page=1, pageSize=1000, account=account, force_refresh=False)
            if isinstance(ecs_result, dict) and "data" in ecs_result:
                ecs_count = ecs_result["data"].get("pagination", {}).get("total", 0)
            else:
                ecs_count = len(ecs_result) if isinstance(ecs_result, list) else 0
            resource_breakdown["ecs"] = ecs_count
            logger.info(f"  ECS数量: {ecs_count}")
        except Exception as e:
            logger.warning(f"  获取ECS数量失败: {str(e)}")
        
        # 获取RDS数量
        try:
            rds_result = await get_resources(type="rds", page=1, pageSize=1000, account=account, force_refresh=False)
            if isinstance(rds_result, dict) and "data" in rds_result:
                rds_count = rds_result["data"].get("pagination", {}).get("total", 0)
            else:
                rds_count = len(rds_result) if isinstance(rds_result, list) else 0
            resource_breakdown["rds"] = rds_count
            logger.info(f"  RDS数量: {rds_count}")
        except Exception as e:
            logger.warning(f"  获取RDS数量失败: {str(e)}")
        
        # 获取Redis数量
        try:
            redis_result = await get_resources(type="redis", page=1, pageSize=1000, account=account, force_refresh=False)
            if isinstance(redis_result, dict) and "data" in redis_result:
                redis_count = redis_result["data"].get("pagination", {}).get("total", 0)
            else:
                redis_count = len(redis_result) if isinstance(redis_result, list) else 0
            resource_breakdown["redis"] = redis_count
            logger.info(f"  Redis数量: {redis_count}")
        except Exception as e:
            logger.warning(f"  获取Redis数量失败: {str(e)}")
        
        total_resources = sum(resource_breakdown.values())
        logger.info(f"  资源总数: {total_resources}")
        
        # 2. 获取成本数据
        try:
            from datetime import datetime, timedelta
            now = datetime.now()
            current_cycle = now.strftime("%Y-%m")
            
            billing_data = _get_billing_overview_from_db(account_config, billing_cycle=current_cycle)
            if billing_data:
                total_cost = float(billing_data.get("total_pretax", 0))
                logger.info(f"  总成本: {total_cost}")
        except Exception as e:
            logger.warning(f"  获取成本数据失败: {str(e)}")
        
        # 3. 获取闲置资源数量
        try:
            cache_manager = CacheManager(ttl_seconds=86400)
            idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account)
            if not idle_data:
                idle_data = cache_manager.get(resource_type="idle_result", account_name=account)
            if idle_data:
                idle_count = len(idle_data)
                logger.info(f"  闲置资源: {idle_count}")
        except Exception as e:
            logger.warning(f"  获取闲置资源失败: {str(e)}")
        
        # 构建返回结果
        result = {
            "account": account,
            "total_cost": total_cost,
            "idle_count": idle_count,
            "cost_trend": trend,
            "trend_pct": trend_pct,
            "total_resources": total_resources,
            "resource_breakdown": resource_breakdown,
            "alert_count": 0,
            "tag_coverage": tag_coverage,
            "savings_potential": savings_potential,
            "cached": False,
        }
        
        logger.info(f"[get_summary] 成功返回数据: total_resources={total_resources}, total_cost={total_cost}")
        return result
        
    except Exception as e:
        import traceback
        error_msg = f"获取仪表盘数据失败: {str(e)}\n{traceback.format_exc()}"
        logger.error(error_msg)
        print(error_msg)
        
        # 发生错误时返回默认值
        return {
            "account": account,
            "total_cost": 0.0,
            "idle_count": 0,
            "cost_trend": "获取失败",
            "trend_pct": 0.0,
            "total_resources": 0,
            "resource_breakdown": {"ecs": 0, "rds": 0, "redis": 0},
            "alert_count": 0,
            "tag_coverage": 0.0,
            "savings_potential": 0.0,
            "cached": False,
        }


def _update_dashboard_summary_cache(account: str, account_config, force_refresh: bool = False):
    """更新 dashboard summary 缓存（后台任务）"""
    import logging
    logger = logging.getLogger(__name__)
    
    # 初始化所有变量，确保即使某些步骤失败也能返回有效数据
    total_cost = 0.0
    trend = "数据不足"
    trend_pct = 0.0
    idle_count = 0
    total_resources = 0
    resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
    tag_coverage = 0.0
    alert_count = 0
    savings_potential = 0.0
    billing_total_cost = None
    current_totals = None
    last_totals = None

    # Get Cost Data - 使用真实账单数据比较本月和上月（优化：并行获取）
    from datetime import datetime, timedelta
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    
    now = datetime.now()
    current_cycle = now.strftime("%Y-%m")
    first_day_this_month = now.replace(day=1)
    last_day_last_month = first_day_this_month - timedelta(days=1)
    last_cycle = last_day_last_month.strftime("%Y-%m")
    
    try:
        # 并行获取本月和上月的账单数据（优化性能）
        def get_current_totals():
            try:
                return _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False)
            except Exception as e:
                logger.error(f"获取本月账单数据失败: {str(e)}")
                return None
        
        def get_last_totals():
            try:
                totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False)
                # 如果数据库没有上月数据，尝试通过API获取
                if totals is None or (totals.get("total_pretax", 0) == 0 and totals.get("data_source") == "local_db"):
                    logger.info(f"上月数据不可用，尝试通过API获取: {last_cycle}")
                    return _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=True)
                return totals
            except Exception as e:
                logger.error(f"获取上月账单数据失败: {str(e)}")
                return None
        
        # 使用线程池并行执行（避免阻塞）
        with ThreadPoolExecutor(max_workers=2) as executor:
            current_future = executor.submit(get_current_totals)
            last_future = executor.submit(get_last_totals)
            
            # 等待结果（设置超时：30秒）
            try:
                current_totals = current_future.result(timeout=30)
                last_totals = last_future.result(timeout=30)
            except Exception as e:
                logger.warning(f"获取账单数据超时或失败: {str(e)}")
        
        # 使用本月数据作为 billing_total_cost
        if current_totals:
            billing_total_cost = float(current_totals.get("total_pretax") or 0.0)
            if billing_total_cost <= 0:
                billing_total_cost = None
        
        current_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        last_cost = float((last_totals or {}).get("total_pretax") or 0.0)
        
        # 计算趋势（本月 vs 上月）
        if last_cost > 0:
            trend_pct = ((current_cost - last_cost) / last_cost) * 100
            if trend_pct > 1:
                trend = "上升"
            elif trend_pct < -1:
                trend = "下降"
            else:
                trend = "平稳"
        else:
            trend = "数据不足"
            trend_pct = 0.0
        
        # 使用本月成本作为total_cost（如果账单数据可用）
        if current_cost > 0:
            total_cost = current_cost
        else:
            # 如果账单数据不可用，尝试使用历史趋势数据
            analyzer = CostTrendAnalyzer()
            try:
                history, analysis = analyzer.get_cost_trend(account, days=30)
                if isinstance(analysis, dict) and "error" not in analysis:
                    total_cost = analysis.get("latest_cost", 0.0)
                else:
                    total_cost = None
            except Exception:
                total_cost = None
        
        logger.info(f"成本趋势计算: 本月({current_cycle})={current_cost:.2f}, 上月({last_cycle})={last_cost:.2f}, 趋势={trend}, 变化={trend_pct:.2f}%")
        
    except Exception as e:
        logger.error(f"计算成本趋势失败: {str(e)}")
        # Fallback: 使用历史趋势数据
        analyzer = CostTrendAnalyzer()
        try:
            history, analysis = analyzer.get_cost_trend(account, days=30)
            if isinstance(analysis, dict) and "error" in analysis:
                total_cost = None
                trend = "N/A"
                trend_pct = 0.0
            else:
                total_cost = analysis.get("latest_cost", 0.0)
                trend = analysis.get("trend", "Unknown")
                trend_pct = analysis.get("total_change_pct", 0.0)
        except Exception:
            total_cost = None
            trend = "N/A"
            trend_pct = 0.0

    # Get Idle Data - 优化：优先使用缓存，避免耗时分析
    try:
        cache_manager = CacheManager(ttl_seconds=86400)
        idle_data = None
        idle_count = 0
        
        # 优先从缓存获取（避免耗时分析）
        idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account)
        if not idle_data:
            idle_data = cache_manager.get(resource_type="idle_result", account_name=account)
        
        if idle_data:
            idle_count = len(idle_data) if idle_data else 0
            logger.info(f"从缓存获取闲置资源数量: {idle_count} (账号: {account})")
        else:
            # 缓存为空，返回0（后台任务中不进行耗时分析）
            logger.info(f"缓存为空，跳过分析 (账号: {account})")
            idle_count = 0
            
    except Exception as e:
        logger.warning(f"获取闲置资源数据失败: {str(e)}")
        idle_count = 0

    # Get Resource Statistics (Task 1.1) - 优化：使用缓存或快速查询
    # 初始化资源列表变量
    instances = []
    rds_list = []
    redis_list = []
    
    try:
        from cloudlens.cli.utils import get_provider
        provider = get_provider(account_config)
        
        # 尝试从缓存获取资源列表（避免重复查询）
        # 如果force_refresh为True，跳过缓存，强制重新查询
        cache_manager = CacheManager(ttl_seconds=86400)
        resource_cache_key = f"resource_list_{account}"
        cached_resources = None
        
        if not force_refresh:
            cached_resources = cache_manager.get(resource_type=resource_cache_key, account_name=account)
        
        if cached_resources:
            instances = cached_resources.get("instances", []) or []
            rds_list = cached_resources.get("rds", []) or []
            redis_list = cached_resources.get("redis", []) or []
            logger.info(f"从缓存获取资源列表 (账号: {account}): ECS={len(instances)}, RDS={len(rds_list)}, Redis={len(redis_list)}")
        else:
            # 初始化变量
            instances = []
            rds_list = []
            redis_list = []
            
            # 如果缓存中没有，尝试从资源API的缓存中获取（避免重复查询）
            # 资源API使用不同的缓存键：resource_type=ecs/rds/redis
            logger.info(f"资源列表缓存未命中，尝试从资源API缓存获取 (账号: {account})")
            try:
                # 尝试从资源API的缓存获取
                ecs_cached = cache_manager.get(resource_type="ecs", account_name=account)
                rds_cached = cache_manager.get(resource_type="rds", account_name=account)
                redis_cached = cache_manager.get(resource_type="redis", account_name=account)
                
                if ecs_cached:
                    instances = ecs_cached if isinstance(ecs_cached, list) else []
                    logger.info(f"从资源API缓存获取ECS: {len(instances)} 个")
                if rds_cached:
                    rds_list = rds_cached if isinstance(rds_cached, list) else []
                    logger.info(f"从资源API缓存获取RDS: {len(rds_list)} 个")
                if redis_cached:
                    redis_list = redis_cached if isinstance(redis_cached, list) else []
                    logger.info(f"从资源API缓存获取Redis: {len(redis_list)} 个")
                
                if instances or rds_list or redis_list:
                    logger.info(f"从资源API缓存获取成功 (账号: {account}): ECS={len(instances)}, RDS={len(rds_list)}, Redis={len(redis_list)}")
            except Exception as e:
                logger.warning(f"从资源API缓存获取失败: {str(e)}")
            
            # 如果从缓存获取失败或数据不完整，先尝试从数据库读取
            if not instances and not rds_list and not redis_list:
                logger.info(f"缓存未命中，尝试从数据库缓存表读取 (账号: {account})")
                try:
                    from cloudlens.core.cache_manager import get_db_connection
                    import json
                    
                    conn = get_db_connection()
                    if conn:
                        cursor = conn.cursor()
                        
                        # 查询ECS缓存
                        cursor.execute("""
                            SELECT data FROM cache_data 
                            WHERE account_name = %s AND resource_type = 'ecs' 
                            AND expires_at > NOW()
                            ORDER BY created_at DESC LIMIT 1
                        """, (account,))
                        ecs_row = cursor.fetchone()
                        if ecs_row and ecs_row[0]:
                            instances = json.loads(ecs_row[0])
                            logger.info(f"从数据库获取ECS: {len(instances)} 个")
                        
                        # 查询RDS缓存
                        cursor.execute("""
                            SELECT data FROM cache_data 
                            WHERE account_name = %s AND resource_type = 'rds' 
                            AND expires_at > NOW()
                            ORDER BY created_at DESC LIMIT 1
                        """, (account,))
                        rds_row = cursor.fetchone()
                        if rds_row and rds_row[0]:
                            rds_list = json.loads(rds_row[0])
                            logger.info(f"从数据库获取RDS: {len(rds_list)} 个")
                        
                        # 查询Redis缓存
                        cursor.execute("""
                            SELECT data FROM cache_data 
                            WHERE account_name = %s AND resource_type = 'redis' 
                            AND expires_at > NOW()
                            ORDER BY created_at DESC LIMIT 1
                        """, (account,))
                        redis_row = cursor.fetchone()
                        if redis_row and redis_row[0]:
                            redis_list = json.loads(redis_row[0])
                            logger.info(f"从数据库获取Redis: {len(redis_list)} 个")
                        
                        cursor.close()
                        conn.close()
                        
                        logger.info(f"从数据库获取资源: ECS={len(instances)}, RDS={len(rds_list)}, Redis={len(redis_list)}")
                except Exception as e:
                    logger.warning(f"从数据库读取失败: {str(e)}")
            
            # 如果数据库也没有数据，才执行API查询
            if not instances and not rds_list and not redis_list:
                logger.info(f"数据库也为空，开始查询所有区域 (账号: {account})")
            # 查询资源（查询所有区域，而不是只查询配置的 region）
            def get_instances():
                try:
                    from cloudlens.core.services.analysis_service import AnalysisService
                    from cloudlens.providers.aliyun.provider import AliyunProvider
                    
                    # 获取所有区域
                    all_regions = AnalysisService._get_all_regions(
                        account_config.access_key_id,
                        account_config.access_key_secret
                    )
                    
                    all_instances = []
                    for region in all_regions:
                        try:
                            region_provider = AliyunProvider(
                                account_name=account_config.name,
                                access_key=account_config.access_key_id,
                                secret_key=account_config.access_key_secret,
                                region=region,
                            )
                            # 快速检查是否有资源
                            count = region_provider.check_instances_count()
                            if count > 0:
                                region_instances = region_provider.list_instances()
                                all_instances.extend(region_instances)
                                logger.info(f"区域 {region}: 找到 {len(region_instances)} 个ECS实例")
                        except Exception as e:
                            logger.warning(f"查询区域 {region} 的ECS实例失败: {str(e)}")
                            continue
                    
                    logger.info(f"总共找到 {len(all_instances)} 个ECS实例（从 {len(all_regions)} 个区域）")
                    return all_instances
                except Exception as e:
                    logger.warning(f"获取ECS列表失败: {str(e)}")
                    # 如果查询所有区域失败，回退到只查询配置的 region
                    try:
                        return provider.list_instances()
                    except:
                        return []
            
            def get_rds():
                try:
                        from cloudlens.core.services.analysis_service import AnalysisService
                        from cloudlens.providers.aliyun.provider import AliyunProvider
                        
                        # 获取所有区域
                        all_regions = AnalysisService._get_all_regions(
                            account_config.access_key_id,
                            account_config.access_key_secret
                        )
                        
                        all_rds = []
                        for region in all_regions:
                            try:
                                region_provider = AliyunProvider(
                                    account_name=account_config.name,
                                    access_key=account_config.access_key_id,
                                    secret_key=account_config.access_key_secret,
                                    region=region,
                                )
                                region_rds = region_provider.list_rds()
                                if region_rds and len(region_rds) > 0:
                                    all_rds.extend(region_rds)
                                    logger.info(f"区域 {region}: 找到 {len(region_rds)} 个RDS实例")
                                else:
                                    logger.debug(f"区域 {region}: 没有RDS实例")
                            except Exception as e:
                                logger.warning(f"查询区域 {region} 的RDS实例失败: {str(e)}")
                                import traceback
                                logger.debug(f"RDS查询异常详情: {traceback.format_exc()}")
                                continue
                        
                        logger.info(f"总共找到 {len(all_rds)} 个RDS实例（从 {len(all_regions)} 个区域）")
                        return all_rds
                except Exception as e:
                    logger.warning(f"获取RDS列表失败: {str(e)}")
                    # 如果查询所有区域失败，回退到只查询配置的 region
                    try:
                        return provider.list_rds()
                    except:
                        return []
            
            def get_redis():
                try:
                        from cloudlens.core.services.analysis_service import AnalysisService
                        from cloudlens.providers.aliyun.provider import AliyunProvider
                        
                        # 获取所有区域
                        all_regions = AnalysisService._get_all_regions(
                            account_config.access_key_id,
                            account_config.access_key_secret
                        )
                        
                        all_redis = []
                        for region in all_regions:
                            try:
                                region_provider = AliyunProvider(
                                    account_name=account_config.name,
                                    access_key=account_config.access_key_id,
                                    secret_key=account_config.access_key_secret,
                                    region=region,
                                )
                                region_redis = region_provider.list_redis()
                                if region_redis and len(region_redis) > 0:
                                    all_redis.extend(region_redis)
                                    logger.info(f"区域 {region}: 找到 {len(region_redis)} 个Redis实例")
                                else:
                                    logger.debug(f"区域 {region}: 没有Redis实例")
                            except Exception as e:
                                logger.warning(f"查询区域 {region} 的Redis实例失败: {str(e)}")
                                import traceback
                                logger.debug(f"Redis查询异常详情: {traceback.format_exc()}")
                                continue
                        
                        logger.info(f"总共找到 {len(all_redis)} 个Redis实例（从 {len(all_regions)} 个区域）")
                        return all_redis
                except Exception as e:
                    logger.warning(f"获取Redis列表失败: {str(e)}")
                    # 如果查询所有区域失败，回退到只查询配置的 region
                    try:
                        return provider.list_redis()
                    except:
                        return []
            
            # 并行查询资源（优化性能）
            with ThreadPoolExecutor(max_workers=5) as executor:
                instances_future = executor.submit(get_instances)
                rds_future = executor.submit(get_rds)
                redis_future = executor.submit(get_redis)
                
                try:
                    # 调大超时时间或容忍部分失败
                    instances = instances_future.result(timeout=90) or []
                    rds_list = rds_future.result(timeout=30) or []
                    redis_list = redis_future.result(timeout=30) or []
                    
                    # 如果查询结果为空，尝试从之前的缓存中恢复（避免显示0资源）
                    if not instances and not rds_list and not redis_list:
                        resource_cache_key = f"resource_list_{account}"
                        cached_resources_retry = cache_manager.get(resource_type=resource_cache_key, account_name=account)
                        if cached_resources_retry:
                            instances = cached_resources_retry.get("instances", []) or []
                            rds_list = cached_resources_retry.get("rds", []) or []
                            redis_list = cached_resources_retry.get("redis", []) or []
                            logger.info(f"⚠️ 实时查询资源为空，从缓存中恢复了 {len(instances)} 个实例")
                    else:
                        # 只有在非空时才更新长期资源列表缓存
                        cache_manager_short = CacheManager(ttl_seconds=1800) # 30分钟
                        
                        # Convert items to dict for JSON serialization
                        instances_dict = [inst.to_dict() if hasattr(inst, "to_dict") else inst for inst in instances]
                        rds_dict = [r.to_dict() if hasattr(r, "to_dict") else r for r in rds_list]
                        redis_dict = [r.to_dict() if hasattr(r, "to_dict") else r for r in redis_list]
                        
                        cache_manager_short.set(
                            resource_type=f"resource_list_{account}",
                            account_name=account,
                            data={"instances": instances_dict, "rds": rds_dict, "redis": redis_dict}
                        )
                except Exception as e:
                    logger.warning(f"查询资源列表发生异常: {str(e)}")
                    # 发生异常时也尝试从缓存恢复
                    resource_cache_key = f"resource_list_{account}"
                    cached_resources_err = cache_manager.get(resource_type=resource_cache_key, account_name=account)
                    if cached_resources_err:
                        instances = cached_resources_err.get("instances", []) or []
                        rds_list = cached_resources_err.get("rds", []) or []
                        redis_list = cached_resources_err.get("redis", []) or []
        
        # 确保变量存在（处理作用域问题）- 移到try块外，确保在所有情况下都能执行
        try:
            _ = instances
        except NameError:
            instances = []
        try:
            _ = rds_list
        except NameError:
            rds_list = []
        try:
            _ = redis_list
        except NameError:
            redis_list = []
        
        # 计算资源统计（确保在所有情况下都能执行）
        resource_breakdown = {
            "ecs": len(instances) if instances else 0,
            "rds": len(rds_list) if rds_list else 0,
            "redis": len(redis_list) if redis_list else 0,
        }
        total_resources = sum(resource_breakdown.values())
        
        # 详细日志输出，便于调试
        logger.info(f"资源统计结果 (账号: {account}):")
        logger.info(f"  ECS: {resource_breakdown['ecs']} (instances类型: {type(instances).__name__}, 长度: {len(instances) if instances else 0})")
        logger.info(f"  RDS: {resource_breakdown['rds']} (rds_list类型: {type(rds_list).__name__}, 长度: {len(rds_list) if rds_list else 0})")
        logger.info(f"  Redis: {resource_breakdown['redis']} (redis_list类型: {type(redis_list).__name__}, 长度: {len(redis_list) if redis_list else 0})")
        logger.info(f"  总数: {total_resources}")
        
        # Tag Coverage - 统计所有资源（ECS + RDS + Redis）的标签覆盖率
        all_resources = list(instances) + list(rds_list) + list(redis_list)
        tagged_count = 0
        for resource in all_resources:
            has_tags = False
            # 检查资源是否有tags属性且tags不为空
            if hasattr(resource, 'tags'):
                # UnifiedResource对象，tags是字典
                if resource.tags and isinstance(resource.tags, dict) and len(resource.tags) > 0:
                    has_tags = True
            elif isinstance(resource, dict):
                # 字典格式的资源，检查tags字段
                tags = resource.get('tags') or resource.get('Tags') or {}
                if tags and isinstance(tags, dict) and len(tags) > 0:
                    has_tags = True
            
            # 如果tags为空，尝试从raw_data中提取
            if not has_tags and hasattr(resource, 'raw_data') and resource.raw_data:
                raw_tags = resource.raw_data.get('Tags') or resource.raw_data.get('tags') or {}
                if raw_tags:
                    # 处理阿里云API返回的Tags格式: {'Tag': [{'TagKey': '...', 'TagValue': '...'}]}
                    if isinstance(raw_tags, dict) and 'Tag' in raw_tags:
                        tag_list = raw_tags['Tag']
                        if isinstance(tag_list, list) and len(tag_list) > 0:
                            has_tags = True
                    elif isinstance(raw_tags, dict) and len(raw_tags) > 0:
                        has_tags = True
            
            if has_tags:
                tagged_count += 1
        
        tag_coverage = (tagged_count / total_resources * 100) if total_resources > 0 else 0
        logger.info(f"标签覆盖率计算: 总资源数={total_resources}, 有标签资源数={tagged_count}, 覆盖率={tag_coverage:.2f}%")

        # --- [NEW] 计算安全告警数量 ---
        try:
            from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
            from cloudlens.core.security_scanner import PublicIPScanner
            
            analyzer = SecurityComplianceAnalyzer()
            exposed = analyzer.detect_public_exposure(all_resources)
            stopped_instances = analyzer.check_stopped_instances(instances)
            preemptible_instances = analyzer.check_preemptible_instances(instances)
            
            # 获取最近一次扫描的高风险项
            last_scan = PublicIPScanner.load_last_results()
            high_risk_count = last_scan.get("summary", {}).get("high_risk_count", 0) if last_scan else 0
            
            # 汇总告警数: 公网暴露 + 长期停止 + 抢占式实例 + 扫描漏洞 (与 security/overview 保持一致)
            alert_count = len(exposed) + len(stopped_instances) + len(preemptible_instances) + high_risk_count
            logger.info(f"安全告警计算: 公网暴露={len(exposed)}, 长期停止={len(stopped_instances)}, 抢占式={len(preemptible_instances)}, 扫描漏洞={high_risk_count}, 总计={alert_count}")
        except Exception as e:
            logger.warning(f"计算安全告警失败: {str(e)}")
            alert_count = 0
            stopped_instances = []
        
        # --- [NEW] 中间保存: 资源统计和标签覆盖率已完成,先存入缓存 ---
        try:
            interim_result = {
                "account": account,
                "total_cost": total_cost or 0.0,
                "idle_count": len(idle_data) if idle_data else 0,
                "cost_trend": "计算中...",
                "trend_pct": 0.0,
                "total_resources": total_resources,
                "resource_breakdown": resource_breakdown,
                "alert_count": alert_count,
                "tag_coverage": tag_coverage,
                "savings_potential": 0.0, # 初始为0，后续计算
                "loading": True  # 标记为还在加载详情
            }
            cache_manager.set(resource_type="dashboard_summary", account_name=account, data=interim_result)
            logger.info(f"✅ 已保存中间缓存 (资源统计已完成)")
        except Exception as e:
            logger.warning(f"保存中间缓存失败: {str(e)}")
        # -------------------------------------------------------------
        # Savings Potential: 优先使用 optimization_suggestions 缓存的数据，保证与前端显示一致
        savings_potential = 0.0
        
        # 先尝试从 optimization_suggestions 缓存获取（与 /api/optimization/suggestions 保持一致）
        opt_cache = cache_manager.get(resource_type="optimization_suggestions", account_name=account)
        if opt_cache and isinstance(opt_cache, dict):
            cached_savings = opt_cache.get("summary", {}).get("total_savings_potential", 0)
            if cached_savings and cached_savings > 0:
                savings_potential = float(cached_savings)
                logger.info(f"优化潜力: 使用 optimization_suggestions 缓存数据 = {savings_potential:.2f}")
        
        # 如果缓存没有数据，则自行计算
        if savings_potential == 0.0 and idle_data and account_config:
            # Get cost map for ECS resources (idle_data typically contains ECS instances)
            cost_map = _get_cost_map("ecs", account_config)
            
            # Calculate total cost of idle resources
            for idle_item in idle_data:
                instance_id = idle_item.get("instance_id") or idle_item.get("id")
                if instance_id:
                    # Try to get real cost from cost_map
                    cost = cost_map.get(instance_id)
                    if cost is None:
                        # If not found, try to estimate from resource spec
                        spec = idle_item.get("spec", "")
                        if spec:
                            cost = _estimate_monthly_cost_from_spec(spec, "ecs")
                        else:
                            # Default fallback estimate
                            cost = 300  # Average ECS cost
                    savings_potential += cost
            
            # 新增：累加停止实例的潜在节省（假设节省70%的磁盘/预留等费用）
            if stopped_instances:
                for stop_item in stopped_instances:
                    instance_id = stop_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id) or 300
                        savings_potential += (cost * 0.7)
            
            logger.info(f"优化潜力计算: 闲置资源贡献={sum(item.get('savings', 0) for item in idle_data)}, 停止实例贡献={len(stopped_instances) * 210} (推估), 总计={savings_potential:.2f}")
            
            # Ensure savings potential doesn't exceed total cost
            if total_cost is not None:
                savings_potential = min(savings_potential, float(total_cost) * 0.95)  # Cap at 95% of total cost

        # 如果成本趋势没有历史数据，则用"当前资源月度成本（折后优先）"作为统一口径的 total_cost
        if total_cost is None and account_config:
            ecs_cost_map = _get_cost_map("ecs", account_config)
            rds_cost_map = _get_cost_map("rds", account_config)
            redis_cost_map = _get_cost_map("redis", account_config)

            estimated_total = 0.0
            for inst in instances:
                cost = ecs_cost_map.get(inst.id)
                if cost is None:
                    cost = _estimate_monthly_cost(inst)
                estimated_total += float(cost or 0)
            for rds in rds_list:
                cost = rds_cost_map.get(rds.id)
                if cost is None:
                    cost = _estimate_monthly_cost(rds)
                estimated_total += float(cost or 0)
            for r in redis_list:
                cost = redis_cost_map.get(r.id)
                if cost is None:
                    cost = _estimate_monthly_cost(r)
                estimated_total += float(cost or 0)

            total_cost = round(float(estimated_total), 2)
            # 再做一次 savings cap（此时 total_cost 已可用）
            savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0

        # 用账单全量口径覆盖 total_cost（更贴近真实账单）
        if billing_total_cost is not None:
            total_cost = round(float(billing_total_cost), 2)
            savings_potential = min(float(savings_potential), float(total_cost) * 0.95) if total_cost else 0.0
        
    except Exception as e:
        # Fallback if resource query fails
        logger.warning(f"获取资源统计失败: {str(e)}")
        # 确保所有变量都有默认值（使用 try-except 处理作用域问题）
        try:
            _ = total_resources
        except NameError:
            total_resources = 0
        try:
            _ = resource_breakdown
        except NameError:
            resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
        try:
            _ = tag_coverage
        except NameError:
            tag_coverage = 0.0
        try:
            _ = alert_count
        except NameError:
            alert_count = 0
        try:
            _ = savings_potential
        except NameError:
            savings_potential = 0.0

    # 确保所有必需字段都有值（最终检查）
    if total_cost is None:
        total_cost = 0.0
    if trend is None or trend == "":
        trend = "数据不足"
    if trend_pct is None:
        trend_pct = 0.0
    if idle_count is None:
        idle_count = 0
    if total_resources is None:
        total_resources = 0
    if resource_breakdown is None:
        resource_breakdown = {"ecs": 0, "rds": 0, "redis": 0}
    if tag_coverage is None:
        tag_coverage = 0.0
    if alert_count is None:
        alert_count = 0
    if savings_potential is None:
        savings_potential = 0.0

    result_data = {
        "account": str(account),
        "total_cost": float(total_cost),
        "idle_count": int(idle_count),
        "cost_trend": str(trend),
        "trend_pct": float(trend_pct),
        "total_resources": int(total_resources),
        "resource_breakdown": dict(resource_breakdown),
        "alert_count": int(alert_count),
        "tag_coverage": round(float(tag_coverage), 2),
        "savings_potential": float(savings_potential),
    }
    
    logger.info(f"✅ Dashboard summary 数据准备完成: account={account}, total_cost={result_data['total_cost']}, idle_count={result_data['idle_count']}, total_resources={result_data['total_resources']}")
    
    # 保存到缓存（24小时有效）
    try:
        cache_manager = CacheManager(ttl_seconds=86400)
        cache_manager.set(resource_type="dashboard_summary", account_name=account, data=result_data)
        logger.info(f"✅ 缓存已保存: {account}")
        return result_data
    except Exception as e:
        logger.warning(f"⚠️ 保存缓存失败: {str(e)}")


@router.get("/dashboard/trend")
@api_error_handler
async def get_trend(
    account: Optional[str] = None, 
    days: int = 30, 
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    granularity: Optional[str] = Query("daily", description="数据粒度: daily(按天) 或 monthly(按月)"),
    force_refresh: bool = Query(False, description="强制刷新缓存")
):
    """Get cost trend chart data（带24小时缓存）
    
    Args:
        account: 账号名称
        days: 查询天数，0表示获取所有历史数据（当start_date和end_date都提供时，此参数被忽略）
        start_date: 开始日期 YYYY-MM-DD格式
        end_date: 结束日期 YYYY-MM-DD格式
        granularity: 数据粒度，daily(按天) 或 monthly(按月)，默认daily
        force_refresh: 是否强制刷新缓存
    """
    if not account:
        raise HTTPException(status_code=400, detail="账号参数是必需的")
    
    # 验证granularity参数
    if granularity not in ["daily", "monthly"]:
        granularity = "daily"
    
    # 如果提供了日期范围，使用日期范围；否则使用days参数
    if start_date and end_date:
        logger.debug(f"收到账号参数: {account}, 日期范围: {start_date} 至 {end_date}, 粒度: {granularity}")
        cache_key = f"dashboard_trend_{granularity}_{start_date}_{end_date}"
    else:
        logger.debug(f"收到账号参数: {account}, days: {days} ({'全部历史' if days == 0 else f'最近{days}天'}), 粒度: {granularity}")
        cache_key = f"dashboard_trend_{granularity}_{days}"
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type=cache_key, account_name=account)
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        return {
            **cached_result,
            "cached": True,
        }
    
    analyzer = CostTrendAnalyzer()
    try:
        # 如果提供了日期范围，计算days参数；否则使用传入的days
        if start_date and end_date:
            from datetime import datetime
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            calculated_days = (end - start).days
            # 使用日期范围生成报告
            report = analyzer.generate_trend_report(account, calculated_days, start_date=start_date, end_date=end_date)
        else:
            # 如果granularity是monthly且days=0，直接按账期查询并聚合
            if granularity == "monthly" and days == 0:
                # 直接从数据库按账期查询月度数据
                from cloudlens.core.bill_storage import BillStorageManager
                from cloudlens.core.config import ConfigManager
                storage = BillStorageManager()
                db = storage._get_db()  # 使用延迟初始化方法获取数据库连接
                cm = ConfigManager()
                account_config = cm.get_account(account)
                if not account_config:
                    raise HTTPException(status_code=404, detail=f"账号 '{account}' 不存在")
                
                account_id = f"{account_config.access_key_id[:10]}-{account}"
                
                # 查询所有账期的月度成本
                rows = db.query("""
                    SELECT
                        billing_cycle,
                        SUM(pretax_amount) as monthly_cost
                    FROM bill_items
                    WHERE account_id = ?
                        AND billing_cycle IS NOT NULL
                        AND billing_cycle != ''
                        AND pretax_amount IS NOT NULL
                    GROUP BY billing_cycle
                    ORDER BY billing_cycle ASC
                """, (account_id,))
                
                if not rows:
                    return {
                        "account": account,
                        "period_days": 0,
                        "granularity": granularity,
                        "analysis": {"error": "No cost history available"},
                        "chart_data": None,
                        "cost_by_type": {},
                        "cost_by_region": {},
                        "snapshots_count": 0,
                        "summary": None,
                        "cached": False,
                    }
                
                # 转换为图表数据格式
                monthly_chart_data = []
                for row in rows:
                    billing_cycle = row['billing_cycle'] if isinstance(row, dict) else row[0]
                    monthly_cost = float(row['monthly_cost'] or 0) if isinstance(row, dict) else float(row[1] or 0)
                    # 使用月初日期格式 YYYY-MM-01
                    year, month = map(int, billing_cycle.split('-'))
                    date_str = f"{year}-{month:02d}-01"
                    monthly_chart_data.append({
                        "date": date_str,
                        "total_cost": monthly_cost,
                        "breakdown": {}
                    })
                
                logger.info(f"✅ 生成 {len(monthly_chart_data)} 条月度数据，日期范围: {monthly_chart_data[0]['date']} 至 {monthly_chart_data[-1]['date']}")
                
                # 计算统计信息
                monthly_costs = [item["total_cost"] for item in monthly_chart_data]
                summary = {
                    "total_cost": sum(monthly_costs) if monthly_costs else 0,
                    "avg_monthly_cost": sum(monthly_costs) / len(monthly_costs) if monthly_costs else 0,
                    "max_monthly_cost": max(monthly_costs) if monthly_costs else 0,
                    "min_monthly_cost": min(monthly_costs) if monthly_costs else 0,
                    "trend": "平稳",
                    "trend_pct": 0.0
                }
                
                # 如果有多个月份，计算趋势
                if len(monthly_costs) >= 2:
                    latest = monthly_costs[-1]
                    previous = monthly_costs[-2]
                    if previous > 0:
                        trend_pct = ((latest - previous) / previous) * 100
                        summary["trend_pct"] = round(trend_pct, 2)
                        if trend_pct > 5:
                            summary["trend"] = "上升"
                        elif trend_pct < -5:
                            summary["trend"] = "下降"
                
                result = {
                    "account": account,
                    "period_days": 0,
                    "granularity": granularity,
                    "chart_data": monthly_chart_data,
                    "summary": summary,
                    "cost_by_type": {},
                    "cost_by_region": {},
                    "snapshots_count": len(monthly_chart_data),
                    "cached": False,
                }
                
                # 保存到缓存
                cache_manager.set(resource_type=cache_key, account_name=account, data=result)
                return result
            else:
                report = analyzer.generate_trend_report(account, days)
        
        if "error" in report:
            # 趋势图常见的"无数据/数据不足"不应该作为服务端错误；
            # 返回 200 + 空 chart_data，前端可自然降级为"不展示趋势图"。
            err = report.get("error") or "No trend data"
            if err in ("No cost history available", "Insufficient data for trend analysis"):
                return {
                    "account": account,
                    "period_days": days,
                    "granularity": granularity,
                    "analysis": {"error": err},
                    "chart_data": None,
                    "cost_by_type": {},
                    "cost_by_region": {},
                    "snapshots_count": 0,
                    "summary": None,
                    "cached": False,
                }
            raise HTTPException(status_code=404, detail=err)
        
        # 转换数据格式以支持新的前端需求
        chart_data = report.get("chart_data", {})
        dates = chart_data.get("dates", [])
        costs = chart_data.get("costs", [])
        
        # 根据granularity转换数据格式
        if granularity == "monthly":
            # 按月聚合数据
            monthly_data = {}
            for i, date in enumerate(dates):
                # 提取月份（YYYY-MM格式）
                if len(date) >= 7:
                    month_key = date[:7]  # YYYY-MM
                elif len(date) == 10:
                    month_key = date[:7]  # YYYY-MM-DD -> YYYY-MM
                else:
                    # 如果日期格式不对，尝试解析
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(date, "%Y-%m-%d")
                        month_key = date_obj.strftime("%Y-%m")
                    except:
                        month_key = date[:7] if len(date) >= 7 else date
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        "date": f"{month_key}-01",  # 使用月初日期格式 YYYY-MM-01
                        "total_cost": 0,
                        "breakdown": {}
                    }
                monthly_data[month_key]["total_cost"] += costs[i] if i < len(costs) else 0
                
                # 如果有资源类型分解，也聚合
                cost_by_type = report.get("cost_by_type", {})
                if cost_by_type and isinstance(cost_by_type, dict):
                    for type_name, type_cost in cost_by_type.items():
                        if type_name not in monthly_data[month_key]["breakdown"]:
                            monthly_data[month_key]["breakdown"][type_name] = 0
                        # 简化：平均分配（实际应该按日期计算）
                        monthly_data[month_key]["breakdown"][type_name] += type_cost / len(dates) if dates else 0
            
            # 转换为列表格式，并按日期排序（从早到晚）
            chart_data_list = sorted(monthly_data.values(), key=lambda x: x["date"])
            
            # 计算月度统计
            monthly_costs = [item["total_cost"] for item in chart_data_list]
            summary = {
                "total_cost": sum(monthly_costs) if monthly_costs else 0,
                "avg_monthly_cost": sum(monthly_costs) / len(monthly_costs) if monthly_costs else 0,
                "max_monthly_cost": max(monthly_costs) if monthly_costs else 0,
                "min_monthly_cost": min(monthly_costs) if monthly_costs else 0,
                "trend": report.get("analysis", {}).get("trend", "平稳"),
                "trend_pct": report.get("analysis", {}).get("total_change_pct", 0.0)
            }
        else:
            # 按天数据（保持原有格式，但转换为新格式）
            chart_data_list = []
            for i, date in enumerate(dates):
                item = {
                    "date": date,
                    "total_cost": costs[i] if i < len(costs) else 0,
                    "breakdown": {}
                }
                
                # 如果有资源类型分解，添加到breakdown
                cost_by_type = report.get("cost_by_type", {})
                if cost_by_type and isinstance(cost_by_type, dict):
                    # 简化：平均分配（实际应该从快照数据中提取）
                    for type_name, type_cost in cost_by_type.items():
                        item["breakdown"][type_name] = type_cost / len(dates) if dates else 0
                
                chart_data_list.append(item)
            
            # 计算日统计
            daily_costs = [item["total_cost"] for item in chart_data_list]
            summary = {
                "total_cost": sum(daily_costs) if daily_costs else 0,
                "avg_daily_cost": sum(daily_costs) / len(daily_costs) if daily_costs else 0,
                "max_daily_cost": max(daily_costs) if daily_costs else 0,
                "min_daily_cost": min(daily_costs) if daily_costs else 0,
                "trend": report.get("analysis", {}).get("trend", "平稳"),
                "trend_pct": report.get("analysis", {}).get("total_change_pct", 0.0)
            }
        
        # 构建新的响应格式
        result = {
            "account": account,
            "period_days": days,
            "granularity": granularity,
            "chart_data": chart_data_list,
            "summary": summary,
            "cost_by_type": report.get("cost_by_type", {}),
            "cost_by_region": report.get("cost_by_region", {}),
            "snapshots_count": report.get("snapshots_count", 0),
            "cached": False,
        }
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type=cache_key, account_name=account, data=result)
        
        return result
    except HTTPException:
        # 不要把 4xx 再包装成 500，否则前端只能看到 "Internal Server Error"
        raise
    except Exception as e:
        logger.exception(f"获取成本趋势失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dashboard/idle")
def get_idle_resources(account: Optional[str] = None, force_refresh: bool = Query(False, description="强制刷新缓存")):
    """Get idle resources list（带24小时缓存）"""
    import logging
    logger = logging.getLogger(__name__)
    
    if not account:
        raise HTTPException(status_code=400, detail="账号参数是必需的")
    
    logger.info(f"[get_idle_resources] 收到请求: account={account}, force_refresh={bool(force_refresh)}")
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据（优先使用缓存，避免耗时操作）
    cached_result = None
    try:
        # 优先从 dashboard_idle 缓存获取
        cached_result = cache_manager.get(resource_type="dashboard_idle", account_name=account)
        if cached_result:
            logger.info(f"[get_idle_resources] ✅ 从 dashboard_idle 缓存返回: {len(cached_result) if isinstance(cached_result, list) else 'N/A'} 条数据")
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
        
        # 如果 dashboard_idle 缓存为空，尝试从 idle_result 缓存获取（与 summary 逻辑保持一致）
        cached_result = cache_manager.get(resource_type="idle_result", account_name=account)
        if cached_result:
            logger.info(f"[get_idle_resources] ✅ 从 idle_result 缓存返回: {len(cached_result) if isinstance(cached_result, list) else 'N/A'} 条数据")
            # 同时更新 dashboard_idle 缓存，保持一致性
            try:
                cache_manager.set(resource_type="dashboard_idle", account_name=account, data=cached_result)
            except Exception as e:
                logger.warning(f"更新 dashboard_idle 缓存失败: {e}")
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
    except Exception as e:
        logger.warning(f"[get_idle_resources] 读取缓存失败: {e}")
        import traceback
        logger.error(f"缓存读取异常详情: {traceback.format_exc()}")
    
    # 如果缓存为空，直接返回空数据（不进行耗时分析，避免阻塞）
    # 用户必须通过 /api/analyze/trigger 主动触发分析
    logger.info(f"[get_idle_resources] 缓存为空，返回空数据（不进行耗时分析）")
    return {
        "success": True,
        "data": [],
        "cached": False,
        "message": "缓存为空，请使用扫描按钮触发分析"
    }


# ==================== Phase 1 Week 2: Resource Management APIs ====================


# ==================== (Migrated Helpers to api_base.py or specialized modules) ====================



def _get_billing_cycle_default() -> str:
    from datetime import datetime
    return datetime.now().strftime("%Y-%m")


def _get_billing_overview_from_db(
    account_config: CloudAccount,
    billing_cycle: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    从本地账单数据库读取成本概览（优先使用，速度快）
    
    Args:
        account_config: 账号配置对象
        billing_cycle: 账期，格式 YYYY-MM，默认当前月
    
    Returns:
        成本概览数据，如果数据库不存在或读取失败则返回 None
    """
    import os
    from datetime import datetime
    from cloudlens.core.database import DatabaseFactory
    
    # 统一使用 MySQL，不再支持 SQLite
    try:
        if billing_cycle is None:
            billing_cycle = datetime.now().strftime("%Y-%m")
        
        db = DatabaseFactory.create_adapter("mysql")
        
        # 构造正确的 account_id 格式：{access_key_id[:10]}-{account_name}
        account_id = f"{account_config.access_key_id[:10]}-{account_config.name}"
        
        # 验证 account_id 是否存在（精确匹配）
        account_result = db.query_one("""
            SELECT DISTINCT account_id 
            FROM bill_items 
            WHERE account_id = %s
            LIMIT 1
        """, (account_id,))
        
        if not account_result:
            # 如果精确匹配失败，尝试模糊匹配（兼容旧数据）
            logger.warning(f"精确匹配失败，尝试模糊匹配: {account_id}")
            account_result = db.query_one("""
                SELECT DISTINCT account_id 
                FROM bill_items 
                WHERE account_id LIKE %s
                LIMIT 1
            """, (f"%{account_config.name}%",))
            
            if not account_result:
                logger.warning(f"未找到账号 '{account_config.name}' (account_id: {account_id}) 的账单数据")
                return None
            
            # 处理字典格式的结果（MySQL）
            if isinstance(account_result, dict):
                matched_account_id = account_result.get('account_id')
            else:
                matched_account_id = account_result[0] if account_result else None
            
            if matched_account_id and matched_account_id != account_id:
                logger.warning(f"使用模糊匹配的 account_id: {matched_account_id} (期望: {account_id})，可能存在数据串号风险")
                account_id = matched_account_id
        
        # 按产品聚合当月成本（MySQL 使用 %s 占位符）
        product_results = db.query("""
            SELECT 
                product_name,
                product_code,
                subscription_type,
                SUM(pretax_amount) as total_pretax
            FROM bill_items
            WHERE account_id = %s
                AND billing_cycle = %s
                AND pretax_amount IS NOT NULL
            GROUP BY product_name, product_code, subscription_type
        """, (account_id, billing_cycle))
        
        by_product: Dict[str, float] = {}
        by_product_name: Dict[str, str] = {}
        by_product_subscription: Dict[str, Dict[str, float]] = {}
        total = 0.0
        
        for row in product_results:
            # 处理字典格式的结果（MySQL）
            if isinstance(row, dict):
                product_name = row.get('product_name') or "unknown"
                product_code = row.get('product_code') or "unknown"
                subscription_type = row.get('subscription_type') or "Unknown"
                pretax = float(row.get('total_pretax') or 0)
            else:
                product_name = row[0] or "unknown"
                product_code = row[1] or "unknown"
                subscription_type = row[2] or "Unknown"
                pretax = float(row[3] or 0)
            
            if pretax <= 0:
                continue
            
            if product_code not in by_product_name:
                by_product_name[product_code] = product_name
            
            by_product[product_code] = by_product.get(product_code, 0.0) + pretax
            by_product_subscription.setdefault(product_code, {})
            by_product_subscription[product_code][subscription_type] = (
                by_product_subscription[product_code].get(subscription_type, 0.0) + pretax
            )
            
            total += pretax
        
        # 检查是否有任何记录（即使总成本为0，也可能有记录）
        # 如果没有记录，返回None让API查询；如果有记录但总成本为0，也返回数据（可能是真实情况）
        if len(by_product) == 0:
            # 没有找到任何记录，返回None让API查询
            logger.info(f"数据库中没有账期 {billing_cycle} 的数据，将使用API查询")
            return None
        
        return {
            "billing_cycle": billing_cycle,
            "total_pretax": round(total, 2),
            "by_product": {k: round(v, 2) for k, v in by_product.items()},
            "by_product_name": by_product_name,
            "by_product_subscription": {
                code: {k: round(v, 2) for k, v in sub.items()}
                for code, sub in by_product_subscription.items()
            },
            "data_source": "local_db"
        }
        
    except Exception as e:
        logger.error(f"从本地数据库读取账单概览失败: {str(e)}")
        return None



# ==================== (Migrated to api_billing.py / internal helpers) ====================


@router.get("/resources")
def list_resources(
    type: str = Query("ecs", description="资源类型"),
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    account: Optional[str] = None,
    sortBy: Optional[str] = None,
    sortOrder: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filter: Optional[str] = None,
    force_refresh: bool = Query(False, description="强制刷新缓存"),
):
    """获取资源列表（支持分页、排序、筛选，带24小时缓存）"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[list_resources] 收到账号参数: {account}, type: {type}")
    print(f"[DEBUG list_resources] 收到账号参数: {account}, type: {type}")
    
    provider, account_name = _get_provider_for_account(account)
    logger.info(f"[list_resources] 使用账号: {account_name}")
    print(f"[DEBUG list_resources] 使用账号: {account_name}")
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    cached_result = None
    if not force_refresh:
        try:
            cached_result = cache_manager.get(resource_type=type, account_name=account_name)
        except Exception as e:
            logger.warning(f"获取缓存失败，将重新查询: {e}")
            cached_result = None
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        result = cached_result
    else:
        # 缓存无效或不存在，从provider查询
        cm = ConfigManager()
        account_config = cm.get_account(account_name)

        # 初始化 resources 变量
        resources = []
        
        # 根据类型获取资源（ECS 查询所有区域）
        if type == "ecs":
            # ECS 查询所有区域，而不是只查询配置的 region
            try:
                from cloudlens.core.services.analysis_service import AnalysisService
                from cloudlens.providers.aliyun.provider import AliyunProvider
                
                logger.info(f"开始查询所有区域的ECS实例，账号: {account_name}")
                
                # 获取所有区域
                all_regions = AnalysisService._get_all_regions(
                    account_config.access_key_id,
                    account_config.access_key_secret
                )
                logger.info(f"获取到 {len(all_regions)} 个可用区域")
                
                all_instances = []
                for idx, region in enumerate(all_regions):
                    try:
                        region_provider = AliyunProvider(
                            account_name=account_config.name,
                            access_key=account_config.access_key_id,
                            secret_key=account_config.access_key_secret,
                            region=region,
                        )
                        # 快速检查是否有资源
                        count = region_provider.check_instances_count()
                        if count > 0:
                            logger.info(f"区域 {region} ({idx+1}/{len(all_regions)}): 有 {count} 个实例，开始详细查询...")
                            region_instances = region_provider.list_instances()
                            all_instances.extend(region_instances)
                            logger.info(f"区域 {region}: 实际获取 {len(region_instances)} 个ECS实例")
                    except Exception as e:
                        logger.warning(f"查询区域 {region} 的ECS实例失败: {str(e)}")
                        import traceback
                        logger.debug(f"异常详情: {traceback.format_exc()}")
                        continue
                
                logger.info(f"总共找到 {len(all_instances)} 个ECS实例（从 {len(all_regions)} 个区域）")
                resources = all_instances
            except Exception as e:
                logger.error(f"查询所有区域的ECS实例失败，回退到单区域查询: {str(e)}")
                import traceback
                logger.error(f"异常详情: {traceback.format_exc()}")
                # 如果查询所有区域失败，回退到只查询配置的 region
                try:
                    resources = provider.list_instances()
                    logger.info(f"单区域查询结果: {len(resources)} 个实例")
                except Exception as e2:
                    logger.error(f"单区域查询也失败: {str(e2)}")
                    resources = []
        elif type == "rds":
            resources = provider.list_rds()
        elif type == "redis":
            resources = provider.list_redis()
        elif type == "slb":
            resources = provider.list_slb() if hasattr(provider, "list_slb") else []
        elif type == "nat":
            resources = provider.list_nat_gateways() if hasattr(provider, "list_nat_gateways") else []
        elif type == "eip":
            # EIP provider 返回 dict 列表
            resources = provider.list_eip() if hasattr(provider, "list_eip") else (provider.list_eips() if hasattr(provider, "list_eips") else [])
        elif type == "oss":
            # OSS bucket 列表（如果安装了 oss2）
            resources = provider.list_oss() if hasattr(provider, "list_oss") else []
        elif type == "disk":
            resources = provider.list_disks() if hasattr(provider, "list_disks") else []
        elif type == "snapshot":
            resources = provider.list_snapshots() if hasattr(provider, "list_snapshots") else []
        elif type == "vpc":
            vpcs = provider.list_vpcs()
            # Convert VPC dict to list format
            resources = []
            for vpc in vpcs:
                from cloudlens.models.resource import UnifiedResource, ResourceType, ResourceStatus
                
                # Get VPC ID and name from dict - check both possible key formats
                vpc_id = vpc.get("id") or vpc.get("VpcId") or ""
                vpc_name = vpc.get("name") or vpc.get("VpcName") or ""
                
                # Log for debugging - 详细记录VPC数据
                logger.info(f"Processing VPC: id={vpc_id}, name={vpc_name}, raw_vpc={vpc}, keys={list(vpc.keys())}")
                
                # If name is empty or just whitespace, use ID as name
                if not vpc_name or not vpc_name.strip():
                    vpc_name = vpc_id if vpc_id else "未命名VPC"
                
                # Ensure we have at least an ID
                if not vpc_id:
                    logger.warning(f"VPC missing ID, skipping: {vpc}")
                    continue

                resources.append(
                    UnifiedResource(
                        id=vpc_id,
                        name=vpc_name,
                        resource_type=ResourceType.VPC,
                        status=ResourceStatus.RUNNING,
                        provider=provider.provider_name,
                        region=vpc.get("region") or vpc.get("RegionId", account_name),
                    )
                )
        else:
            raise HTTPException(status_code=400, detail=f"不支持的资源类型: {type}")

        # 批量获取真实成本映射（提高效率）
        cost_map = {}
        if account_config and type not in ("vpc",):
            cost_map = _get_cost_map(type, account_config)

        # 转换为统一格式，使用真实成本
        result = []

        # dict 资源（EIP/OSS 等）
        if resources and isinstance(resources[0], dict):
            # 快照：QueryInstanceBill 很多情况下返回 RegionId 而不是 SnapshotId，导致无法逐实例对齐
            # 这种情况改用账单总额按容量比例分摊到每个快照，保证“实例级 cost 之和 == 账单全量”
            if account_config and type == "snapshot":
                has_snapshot_keys = any(str(k).startswith("s-") for k in (cost_map or {}).keys())
                if not has_snapshot_keys:
                    cost_map = {}
                try:
                    totals = _get_billing_overview_totals(account_config, force_refresh=force_refresh)
                    product_total = float(((totals or {}).get("by_product") or {}).get("snapshot") or 0.0)
                except Exception:
                    product_total = 0.0

                if product_total > 0:
                    weights = []
                    for r in resources:
                        rid = r.get("id")
                        if not rid:
                            continue
                        w = r.get("size_gb") or 0
                        try:
                            w = float(w)
                        except Exception:
                            w = 0.0
                        weights.append((rid, max(0.0, w)))

                    total_w = sum(w for _, w in weights)
                    if total_w <= 0:
                        # 没有容量信息，均分
                        n = len(weights)
                        if n > 0:
                            per = product_total / n
                            for rid, _ in weights:
                                cost_map[rid] = per
                    else:
                        for rid, w in weights:
                            cost_map[rid] = product_total * (w / total_w)

            for r in resources:
                rid = r.get("id") or r.get("Id") or r.get("ResourceId") or r.get("name")
                if not rid:
                    continue

                # EIP：id=AllocationId；OSS：id=bucket name
                cost = cost_map.get(rid, 0.0)

                name = r.get("name") or r.get("ip_address") or r.get("id") or "-"
                spec = "-"
                region_val = r.get("region") or r.get("RegionId") or getattr(provider, "region", "")
                status_val = r.get("status") or r.get("Status") or "-"

                if type == "eip":
                    spec = f"{r.get('bandwidth', '-') }Mbps"
                elif type == "disk":
                    size_gb = r.get("size_gb", "-")
                    cat = r.get("disk_category", "-")
                    dtyp = r.get("disk_type", "-")
                    spec = f"{cat} / {dtyp} / {size_gb}GB"
                elif type == "snapshot":
                    src = r.get("source_disk_id") or "-"
                    size_gb = r.get("size_gb", "-")
                    spec = f"源盘: {src} / {size_gb}GB"

                result.append(
                    {
                        "id": rid,
                        "name": name,
                        "type": type,
                        "status": str(status_val),
                        "region": str(region_val),
                        "spec": str(spec),
                        "cost": float(cost or 0),
                        "tags": {},
                    "created_time": r.get("created_time") if isinstance(r.get("created_time"), str) else None,
                        "public_ips": [r.get("ip_address")] if r.get("ip_address") else [],
                        "private_ips": [],
                        "vpc_id": r.get("vpc_id") or r.get("VpcId") or None,
                    }
                )
        else:
            for r in resources:
                # 从成本映射中获取真实成本，如果没有则使用估算值
                cost = cost_map.get(r.id)
                if cost is None:
                    cost = _estimate_monthly_cost(r)

                # For VPC resources, ensure name is not empty
                display_name = r.name or r.id or "-"
                if type == "vpc" and not r.name:
                    display_name = r.id or "-"
                
                # For VPC resources, vpc_id should be the VPC's own ID
                # For other resources, vpc_id is the associated VPC ID
                if type == "vpc":
                    # VPC资源本身，vpc_id应该显示为VPC的ID（VPC资源本身没有关联的VPC，所以显示自己的ID）
                    # 确保即使r.id是空字符串也能正确处理
                    vpc_id_value = r.id if (hasattr(r, "id") and r.id and str(r.id).strip()) else None
                    # 调试日志
                    if not vpc_id_value:
                        logger.warning(f"VPC resource has empty ID: id={r.id}, name={r.name}, type={type}")
                else:
                    vpc_id_value = r.vpc_id if hasattr(r, "vpc_id") and r.vpc_id else None
                
                result.append(
                    {
                        "id": r.id or "-",
                        "name": display_name,
                        "type": type,
                        "status": r.status.value if hasattr(r.status, "value") else str(r.status),
                        "region": r.region,
                        "spec": r.spec or "-",
                        "cost": float(cost or 0),
                        "tags": r.tags if hasattr(r, "tags") and r.tags else {},
                        "created_time": r.created_time.isoformat()
                        if hasattr(r, "created_time") and r.created_time
                        else None,
                        "public_ips": r.public_ips if hasattr(r, "public_ips") else [],
                        "private_ips": r.private_ips if hasattr(r, "private_ips") else [],
                        "vpc_id": vpc_id_value,
                    }
                )

        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type=type, account_name=account_name, data=result)
    
    # 排序（在缓存数据上排序）
    if sortBy:
        reverse = sortOrder == "desc"
        try:
            result.sort(key=lambda x: x.get(sortBy, ""), reverse=reverse)
        except:
            pass  # Ignore sort errors
    
    # 筛选（在缓存数据上筛选）
    if filter:
        try:
            import json
            filter_dict = json.loads(filter)
            filtered_result = []
            for item in result:
                match = True
                for key, value in filter_dict.items():
                    if item.get(key) != value:
                        match = False
                        break
                if match:
                    filtered_result.append(item)
            result = filtered_result
        except:
            pass  # Ignore filter errors
    
    # 分页（在缓存数据上分页）
    total = len(result)
    start = (page - 1) * pageSize
    end = start + pageSize
    paginated_resources = result[start:end]
    
    return {
        "success": True,
        "data": paginated_resources,
        "pagination": {
            "page": page,
            "pageSize": pageSize,
            "total": total,
            "totalPages": (total + pageSize - 1) // pageSize,
        },
        "cached": cached_result is not None,  # 标识是否来自缓存
    }


@router.get("/resources/{resource_id}")
def get_resource(
    resource_id: str,
    account: Optional[str] = None,
    resource_type: Optional[str] = Query(None, description="资源类型，用于加速查找")
):
    """获取资源详情（Legacy API，建议使用 api_resources 模块）"""
    # 添加调试日志
    logger.info(f"[Legacy API] 收到请求: resource_id={resource_id}, type={resource_type}, account={account}")

    # 如果提供了 resource_type，直接使用新的 API 逻辑（优先）
    if resource_type:
        logger.info(f"[Legacy API] 检测到 resource_type={resource_type}，尝试转发到新 API")
        try:
            from web.backend.api_resources import get_resource_detail
            logger.info(f"[Legacy API] 调用新 API: get_resource_detail({resource_id}, {account}, {resource_type})")
            result = get_resource_detail(resource_id, account, resource_type)
            logger.info(f"[Legacy API] 新 API 返回: success={result.get('success')}, has_data={bool(result.get('data'))}")

            if result.get('success') and result.get('data'):
                logger.info(f"[Legacy API] ✅ 转发成功，返回新 API 结果")
                # 对于 ACK 资源，记录是否包含 ack_details
                if resource_type == "ack" and result.get('data'):
                    has_ack_details = 'ack_details' in result.get('data', {})
                    logger.info(f"[Legacy API] ACK 资源返回数据包含 ack_details: {has_ack_details}")
                return result
            else:
                logger.warning(f"[Legacy API] 新 API 返回失败或无数据，使用旧逻辑")
        except Exception as e:
            logger.error(f"[Legacy API] 转发到新 API 失败: {e}, 使用旧逻辑", exc_info=True)
    
    provider, account_name = _get_provider_for_account(account)
    
    # 获取账号配置用于成本查询
    cm = ConfigManager()
    account_config = cm.get_account(account_name)
    
    # 尝试从各种资源类型中查找（资源量较大时建议按 type 查；这里为通用详情入口做尽量覆盖）
    # UnifiedResource 类资源
    resources = []
    try:
        resources.extend(provider.list_instances())
        resources.extend(provider.list_rds())
        resources.extend(provider.list_redis())
        if hasattr(provider, "list_slb"):
            resources.extend(provider.list_slb())
        if hasattr(provider, "list_nat_gateways"):
            resources.extend(provider.list_nat_gateways())
    except Exception:
        pass

    resource = next((r for r in resources if getattr(r, "id", None) == resource_id), None)

    # dict 资源（EIP/OSS）
    dict_resource = None
    if resource is None:
        try:
            if hasattr(provider, "list_eip"):
                eips_list = provider.list_eip() if hasattr(provider, "list_eip") else (provider.list_eips() if hasattr(provider, "list_eips") else [])
                for e in eips_list:
                    if e.get("id") == resource_id:
                        dict_resource = ("eip", e)
                        break
            if dict_resource is None and hasattr(provider, "list_oss"):
                for b in provider.list_oss():
                    if b.get("id") == resource_id:
                        dict_resource = ("oss", b)
                        break
            if dict_resource is None and hasattr(provider, "list_disks"):
                for d in provider.list_disks():
                    if d.get("id") == resource_id:
                        dict_resource = ("disk", d)
                        break
            if dict_resource is None and hasattr(provider, "list_snapshots"):
                for s in provider.list_snapshots():
                    if s.get("id") == resource_id:
                        dict_resource = ("snapshot", s)
                        break
        except Exception:
            dict_resource = None

    if resource is None and dict_resource is None:
        raise HTTPException(status_code=404, detail="资源不存在")

    # 确定资源类型
    resource_type = "ecs"
    if dict_resource is not None:
        resource_type = dict_resource[0]
    elif hasattr(resource, "resource_type"):
        rt = resource.resource_type.value if hasattr(resource.resource_type, "value") else str(resource.resource_type)
        if "rds" in rt.lower():
            resource_type = "rds"
        elif "redis" in rt.lower():
            resource_type = "redis"
        elif "vpc" in rt.lower():
            resource_type = "vpc"
        elif "slb" in rt.lower():
            resource_type = "slb"
        elif "nat" in rt.lower():
            resource_type = "nat"
    
    # 获取真实成本映射
    cost_map = {}
    if account_config and resource_type not in ("vpc",):
        cost_map = _get_cost_map(resource_type, account_config)

    if dict_resource is not None:
        _, r = dict_resource
        cost = float(cost_map.get(resource_id) or 0.0)
        spec = "-"
        if resource_type == "eip":
            spec = f"{r.get('bandwidth', '-') }Mbps"
        elif resource_type == "disk":
            spec = f"{r.get('disk_category', '-') } / {r.get('disk_type', '-') } / {r.get('size_gb', '-') }GB"
        elif resource_type == "snapshot":
            spec = f"源盘: {r.get('source_disk_id') or '-'}"
        return {
            "success": True,
            "data": {
                "id": resource_id,
                "name": r.get("name") or r.get("ip_address") or resource_id,
                "type": resource_type,
                "status": str(r.get("status") or "-"),
                "region": str(r.get("region") or getattr(provider, "region", "")),
                "spec": spec,
                "cost": cost,
                "tags": {},
                "created_time": r.get("created_time") if isinstance(r.get("created_time"), str) else None,
                "public_ips": [r.get("ip_address")] if r.get("ip_address") else [],
                "private_ips": [],
                "raw_data": r,
            },
        }

    # UnifiedResource
    cost = cost_map.get(resource_id)
    if cost is None:
        cost = _estimate_monthly_cost(resource)

    return {
        "success": True,
        "data": {
            "id": resource.id,
            "name": resource.name or "-",
            "type": resource_type,
            "status": resource.status.value if hasattr(resource.status, "value") else str(resource.status),
            "region": resource.region,
            "spec": resource.spec or "-",
            "cost": float(cost or 0),
            "tags": resource.tags if hasattr(resource, "tags") and resource.tags else {},
            "created_time": resource.created_time.isoformat()
            if hasattr(resource, "created_time") and resource.created_time
            else None,
            "public_ips": resource.public_ips if hasattr(resource, "public_ips") else [],
            "private_ips": resource.private_ips if hasattr(resource, "private_ips") else [],
            "raw_data": getattr(resource, "raw_data", {}),
        },
    }


@router.get("/resources/{resource_id}/metrics")
def get_resource_metrics(
    resource_id: str,
    days: int = Query(7, ge=1, le=30),
    account: Optional[str] = None,
):
    """获取资源监控数据"""
    provider, account_name = _get_provider_for_account(account)
    
    # 获取资源
    resources = []
    try:
        resources.extend(provider.list_instances())
    except:
        pass
    
    resource = next((r for r in resources if r.id == resource_id), None)
    if not resource:
        raise HTTPException(status_code=404, detail="资源不存在")
    
    # 获取监控数据
    try:
        from cloudlens.core.idle_detector import IdleDetector
        metrics = IdleDetector.fetch_ecs_metrics(provider, resource_id, days)
        
        # 转换为图表数据格式
        chart_data = {
            "cpu": [],
            "memory": [],
            "network_in": [],
            "network_out": [],
            "dates": [],
        }
        
        # 简化：返回平均值（实际应该返回时间序列数据）
        return {
            "success": True,
            "data": {
                "metrics": metrics,
                "chart_data": chart_data,
            }
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "metrics": {},
                "chart_data": {},
                "error": str(e),
            }
        }


# ==================== Phase 1 Week 3: Account Management APIs ====================

@router.get("/settings/accounts")
def list_accounts_settings():
    """获取账号列表（用于设置页面）"""
    cm = ConfigManager()
    accounts = cm.list_accounts()
    result = []
    for account in accounts:
        if isinstance(account, CloudAccount):
            result.append({
                "name": account.name,
                "alias": getattr(account, 'alias', None),  # 别名（可选）
                "region": account.region,
                "provider": account.provider,
                "access_key_id": account.access_key_id,
            })
    return {"success": True, "data": result}


@router.post("/settings/accounts")
def add_account(account_data: AccountCreateRequest):
    """添加账号"""
    cm = ConfigManager()
    try:
        cm.add_account(
            name=account_data.name,
            provider=account_data.provider,
            access_key_id=account_data.access_key_id,
            access_key_secret=account_data.access_key_secret,
            region=account_data.region,
            alias=account_data.alias,  # 别名（可选）
        )
        return {"success": True, "message": "账号添加成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/settings/accounts/{account_name}")
def update_account(account_name: str, account_data: AccountUpdateRequest):
    """更新账号"""
    import keyring
    cm = ConfigManager()
    try:
        # 检查账号是否存在
        existing_account = cm.get_account(account_name)
        if not existing_account:
            raise HTTPException(status_code=404, detail=f"账号 '{account_name}' 不存在")
        
        # 账号名称不可修改（用于数据关联），只允许修改别名
        # 获取别名（如果提供了）
        alias = account_data.alias.strip() if account_data.alias else None
        
        # 获取新密钥，如果没有提供则使用现有密钥
        new_secret = account_data.access_key_secret
        if not new_secret:
            # 从 keyring 获取现有密钥
            try:
                existing_secret = keyring.get_password("cloudlens", f"{account_name}_access_key_secret")
                if existing_secret:
                    new_secret = existing_secret
                else:
                    raise HTTPException(status_code=400, detail="无法获取现有密钥，请提供新密钥")
            except Exception:
                raise HTTPException(status_code=400, detail="无法获取现有密钥，请提供新密钥")
        
        # 更新账号配置（不修改账号名称，只更新其他字段和别名）
        cm.add_account(
            name=account_name,  # 保持原名称不变（用于数据关联）
            provider=account_data.provider or existing_account.provider,
            access_key_id=account_data.access_key_id or existing_account.access_key_id,
            access_key_secret=new_secret,
            region=account_data.region or existing_account.region,
            alias=alias,  # 更新别名
        )
        
        return {"success": True, "message": "账号更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/settings/accounts/{account_name}")
def delete_account(account_name: str):
    """删除账号"""
    cm = ConfigManager()
    try:
        cm.remove_account(account_name)
        return {"success": True, "message": "账号删除成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Phase 2: Cost Analysis APIs ====================

@router.get("/cost/overview")
def get_cost_overview(account: Optional[str] = None, force_refresh: bool = Query(False, description="强制刷新缓存")):
    """获取成本概览（优先账单口径，带24小时缓存）"""
    provider, account_name = _get_provider_for_account(account)
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="cost_overview", account_name=account_name)
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    # 缓存无效或不存在，计算新数据
    cm = ConfigManager()
    account_config = cm.get_account(account_name)
    
    try:
        # 账单优先：使用 BSS 账单概览作为"全量成本"口径
        from datetime import datetime, timedelta
        now = datetime.now()
        current_cycle = now.strftime("%Y-%m")
        
        # 计算本月已过天数（用于环比对比）
        current_day = now.day  # 今天是几号
        first_day_this_month = now.replace(day=1)
        
        # 本月成本：1月1日到1月6日（如果今天是6号）
        current_month_start = first_day_this_month
        current_month_end = now
        
        # 上月相同天数：12月1日到12月6日
        last_month_end = first_day_this_month - timedelta(days=1)  # 上个月最后一天
        last_month_start = last_month_end.replace(day=1)  # 上个月第一天
        
        # 计算上月相同天数的结束日期（不能超过上个月的最后一天）
        last_month_comparable_end = last_month_start + timedelta(days=current_day - 1)
        if last_month_comparable_end > last_month_end:
            last_month_comparable_end = last_month_end
        
        # 计算上月账期（用于显示）
        last_cycle = last_month_end.strftime("%Y-%m")
        
        logger.info(f"📊 成本概览查询: 账号={account_name}")
        logger.info(f"   当前账期={current_cycle}, 本月已过天数={current_day}")
        logger.info(f"   本月成本范围: {current_month_start.strftime('%Y-%m-%d')} 至 {current_month_end.strftime('%Y-%m-%d')}")
        logger.info(f"   上月对比范围: {last_month_start.strftime('%Y-%m-%d')} 至 {last_month_comparable_end.strftime('%Y-%m-%d')}")

        # 使用成本趋势分析器获取指定日期范围的成本（更准确）
        from cloudlens.core.cost_trend_analyzer import CostTrendAnalyzer
        analyzer = CostTrendAnalyzer()
        
        # 获取本月成本（从1月1日到今天）
        current_month_cost = 0.0
        try:
            current_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=current_month_start.strftime("%Y-%m-%d"),
                end_date=current_month_end.strftime("%Y-%m-%d")
            )
            # 从返回数据中提取总成本（可能是chart_data.costs的总和，或者analysis中的值）
            if current_cost_data and "error" not in current_cost_data:
                # 尝试从chart_data中计算总成本
                if "chart_data" in current_cost_data and "costs" in current_cost_data["chart_data"]:
                    costs = current_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        current_month_cost = float(sum(costs))
                        logger.info(f"✅ 本月成本（从chart_data计算，{current_month_start.strftime('%Y-%m-%d')} 至 {current_month_end.strftime('%Y-%m-%d')}）: {current_month_cost}")
                    else:
                        # 如果chart_data为空，回退到账单概览API
                        logger.warning("⚠️  chart_data为空，回退到账单概览API")
                        current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                        current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
                elif "total_cost" in current_cost_data:
                    current_month_cost = float(current_cost_data.get("total_cost", 0.0))
                    logger.info(f"✅ 本月成本（{current_month_start.strftime('%Y-%m-%d')} 至 {current_month_end.strftime('%Y-%m-%d')}）: {current_month_cost}")
                else:
                    # 如果数据格式不符合预期，回退到账单概览API
                    logger.warning("⚠️  数据库查询返回格式不符合预期，回退到账单概览API")
                    current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                    current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
            else:
                # 如果数据库查询失败，回退到账单概览API
                error_msg = current_cost_data.get("error", "Unknown error") if current_cost_data else "No data returned"
                logger.warning(f"⚠️  数据库查询失败: {error_msg}，回退到账单概览API")
                current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
                current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        except Exception as e:
            logger.warning(f"⚠️  获取本月成本失败，回退到账单概览API: {str(e)}")
            current_totals = _get_billing_overview_totals(account_config, billing_cycle=current_cycle, force_refresh=False) if account_config else None
        current_month_cost = float((current_totals or {}).get("total_pretax") or 0.0)
        
        # 获取上月相同天数的成本（从12月1日到12月6日）
        last_month_cost = 0.0
        try:
            last_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=last_month_start.strftime("%Y-%m-%d"),
                end_date=last_month_comparable_end.strftime("%Y-%m-%d")
            )
            # 从返回数据中提取总成本（可能是chart_data.costs的总和，或者analysis中的值）
            if last_cost_data and "error" not in last_cost_data:
                # 尝试从chart_data中计算总成本
                if "chart_data" in last_cost_data and "costs" in last_cost_data["chart_data"]:
                    costs = last_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        last_month_cost = float(sum(costs))
                        logger.info(f"✅ 上月成本（从chart_data计算，{last_month_start.strftime('%Y-%m-%d')} 至 {last_month_comparable_end.strftime('%Y-%m-%d')}）: {last_month_cost}")
                    else:
                        # 如果chart_data为空，回退到账单概览API（按比例计算）
                        logger.warning("⚠️  chart_data为空，回退到账单概览API（按比例计算）")
                        last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                        if last_totals:
                            last_month_total = float(last_totals.get("total_pretax") or 0.0)
                            last_month_days = last_month_end.day
                            last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                            logger.info(f"   上月总成本={last_month_total}, 总天数={last_month_days}, 已过天数={current_day}, 按比例计算={last_month_cost}")
                elif "total_cost" in last_cost_data:
                    last_month_cost = float(last_cost_data.get("total_cost", 0.0))
                    logger.info(f"✅ 上月成本（{last_month_start.strftime('%Y-%m-%d')} 至 {last_month_comparable_end.strftime('%Y-%m-%d')}）: {last_month_cost}")
                else:
                    # 如果数据格式不符合预期，回退到账单概览API（按比例计算）
                    logger.warning("⚠️  数据库查询返回格式不符合预期，回退到账单概览API（按比例计算）")
                    last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                    if last_totals:
                        last_month_total = float(last_totals.get("total_pretax") or 0.0)
                        last_month_days = last_month_end.day
                        last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                        logger.info(f"   上月总成本={last_month_total}, 总天数={last_month_days}, 已过天数={current_day}, 按比例计算={last_month_cost}")
            else:
                # 如果数据库查询失败，回退到账单概览API（按比例计算）
                error_msg = last_cost_data.get("error", "Unknown error") if last_cost_data else "No data returned"
                logger.warning(f"⚠️  数据库查询失败: {error_msg}，回退到账单概览API（按比例计算）")
                last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
                if last_totals:
                    # 按比例计算：上月总成本 * (已过天数 / 上月总天数)
                    last_month_total = float(last_totals.get("total_pretax") or 0.0)
                    last_month_days = last_month_end.day  # 上个月总天数
                    last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
                    logger.info(f"   上月总成本={last_month_total}, 总天数={last_month_days}, 已过天数={current_day}, 按比例计算={last_month_cost}")
        except Exception as e:
            logger.warning(f"⚠️  获取上月成本失败，回退到账单概览API（按比例计算）: {str(e)}")
            last_totals = _get_billing_overview_totals(account_config, billing_cycle=last_cycle, force_refresh=False) if account_config else None
            if last_totals:
                last_month_total = float(last_totals.get("total_pretax") or 0.0)
                last_month_days = last_month_end.day
                last_month_cost = last_month_total * (current_day / last_month_days) if last_month_days > 0 else 0.0
        
        logger.info(f"💰 成本数据: 本月（{current_day}天）={current_month_cost}, 上月（{current_day}天）={last_month_cost}")
        
        # 如果上月数据为0，记录警告
        if last_month_cost == 0:
            logger.warning(f"⚠️  上月相同天数（{last_month_start.strftime('%Y-%m-%d')} 至 {last_month_comparable_end.strftime('%Y-%m-%d')}）成本为0，可能该时间段确实无成本或数据未同步")
        
        mom = ((current_month_cost - last_month_cost) / last_month_cost * 100) if last_month_cost > 0 else 0.0
        
        # 计算同比（去年同期相同天数：2025年1月1日-1月6日 vs 2026年1月1日-1月6日）
        yoy = 0.0
        last_year_cost = 0.0
        try:
            # 去年同期范围
            last_year_start = datetime(now.year - 1, now.month, 1)
            last_year_end = datetime(now.year - 1, now.month, current_day)
            
            logger.info(f"📊 计算同比: 去年同期范围 {last_year_start.strftime('%Y-%m-%d')} 至 {last_year_end.strftime('%Y-%m-%d')}")
            
            last_year_cost_data = analyzer.get_real_cost_from_bills(
                account_name=account_name,
                start_date=last_year_start.strftime("%Y-%m-%d"),
                end_date=last_year_end.strftime("%Y-%m-%d")
            )
            if last_year_cost_data and "error" not in last_year_cost_data:
                if "chart_data" in last_year_cost_data and "costs" in last_year_cost_data["chart_data"]:
                    costs = last_year_cost_data["chart_data"]["costs"]
                    if isinstance(costs, list) and len(costs) > 0:
                        last_year_cost = float(sum(costs))
                        logger.info(f"✅ 去年同期成本（从chart_data计算）: {last_year_cost}")
                    else:
                        # 如果数据库查询失败，尝试从账单概览API获取（按比例计算）
                        last_year_cycle = last_year_start.strftime("%Y-%m")
                        last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                        if last_year_totals:
                            last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                            # 按比例计算：去年同期总成本 * (已过天数 / 该月总天数)
                            last_year_month_days = (datetime(last_year_start.year, last_year_start.month + 1, 1) - timedelta(days=1)).day if last_year_start.month < 12 else 31
                            last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
                            logger.info(f"   去年同期总成本={last_year_total}, 总天数={last_year_month_days}, 已过天数={current_day}, 按比例计算={last_year_cost}")
                elif "total_cost" in last_year_cost_data:
                    last_year_cost = float(last_year_cost_data.get("total_cost", 0.0))
                else:
                    # 回退到账单概览API（按比例计算）
                    last_year_cycle = last_year_start.strftime("%Y-%m")
                    last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                    if last_year_totals:
                        last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                        last_year_month_days = (datetime(last_year_start.year, last_year_start.month + 1, 1) - timedelta(days=1)).day if last_year_start.month < 12 else 31
                        last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
        except Exception as e:
            logger.warning(f"⚠️  获取去年同期成本失败: {str(e)}")
            # 尝试从账单概览API获取（按比例计算）
            try:
                last_year_cycle = datetime(now.year - 1, now.month, 1).strftime("%Y-%m")
                last_year_totals = _get_billing_overview_totals(account_config, billing_cycle=last_year_cycle, force_refresh=False) if account_config else None
                if last_year_totals:
                    last_year_total = float(last_year_totals.get("total_pretax") or 0.0)
                    last_year_month_days = (datetime(now.year - 1, now.month + 1, 1) - timedelta(days=1)).day if now.month < 12 else 31
                    last_year_cost = last_year_total * (current_day / last_year_month_days) if last_year_month_days > 0 else 0.0
            except:
                last_year_cost = 0.0
        
        # 计算同比
        yoy = (
            (current_month_cost - last_year_cost) / last_year_cost * 100
            if last_year_cost > 0 else 0.0
        )
        
        logger.info(f"📊 同比数据: 今年（{current_day}天）={current_month_cost}, 去年（{current_day}天）={last_year_cost}, 同比={yoy:.2f}%")
        
        result_data = {
            "current_month": round(current_month_cost, 2),
            "last_month": round(last_month_cost, 2),
            "yoy": round(yoy, 2),
            "mom": round(mom, 2),
            "current_cycle": current_cycle,
            "last_cycle": last_cycle,
            "current_days": current_day,  # 本月已过天数
            "comparable_days": current_day,  # 对比天数（相同）
        }
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type="cost_overview", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        logger.error(f"❌ 获取成本概览失败: {str(e)}", exc_info=True)
        # 返回错误信息，而不是静默返回0
        return {
            "success": False,
            "error": str(e),
            "data": {
                "current_month": 0,
                "last_month": 0,
                "yoy": 0,
                "mom": 0,
            },
            "cached": False,
        }


@router.get("/cost/breakdown")
def get_cost_breakdown(
    account: Optional[str] = None,
    force_refresh: bool = Query(False, description="强制刷新缓存"),
    billing_cycle: Optional[str] = Query(None, description="账期 yyyy-MM，默认当月"),
):
    """获取成本构成（优先账单口径，带24小时缓存）"""
    provider, account_name = _get_provider_for_account(account)
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="cost_breakdown", account_name=account_name)
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    # 缓存无效或不存在，计算新数据
    cm = ConfigManager()
    account_config = cm.get_account(account_name)

    try:
        # 账单优先：用 BSS BillOverview 的 ProductCode 聚合得到“全量成本构成”
        totals = _get_billing_overview_totals(account_config, billing_cycle=billing_cycle) if account_config else None
        by_product = (totals or {}).get("by_product") or {}
        total = float((totals or {}).get("total_pretax") or 0.0)
        by_product_name = (totals or {}).get("by_product_name") or {}
        by_product_subscription = (totals or {}).get("by_product_subscription") or {}

        # 便于前端展示的列表结构（排序后）
        categories = []
        for code, amount in by_product.items():
            try:
                amount_f = float(amount or 0.0)
            except Exception:
                amount_f = 0.0
            if amount_f <= 0:
                continue
            categories.append(
                {
                    "code": code,
                    "name": by_product_name.get(code) or code,
                    "amount": round(amount_f, 2),
                    "subscription": by_product_subscription.get(code) or {},
                }
            )
        categories.sort(key=lambda x: float(x.get("amount") or 0.0), reverse=True)

        result_data = {
            # 兼容旧前端字段：by_type 仍返回 {code: amount}
            "by_type": by_product,
            "total": round(float(total), 2),
            "billing_cycle": (totals or {}).get("billing_cycle") or billing_cycle,
            "source": "billing_overview",
            # 新字段：前端直接用 categories 渲染更友好
            "categories": categories,
            "by_product_name": by_product_name,
        }
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type="cost_breakdown", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "by_type": {},
                "total": 0,
            },
            "cached": False,
        }


# ==================== Phase 2: Security APIs ====================

@router.get("/security/overview")
def get_security_overview(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="强制刷新缓存"),
    locale: Optional[str] = Query("zh", description="语言设置: zh 或 en")
):
    """获取安全概览（带24小时缓存）"""
    # 获取语言设置
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    # 注意：为了支持多语言，我们需要重新生成翻译后的文本
    # 方案：缓存原始数据（数字），返回时根据语言翻译
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="security_overview", account_name=account_name)
    
    # 如果缓存有效，需要根据语言重新翻译
    if cached_result is not None:
        # 重新生成 score_deductions（需要原始数据）
        # 由于缓存中可能已经包含翻译后的文本，我们需要重新计算
        # 为了简化，如果语言不是中文，我们强制刷新以重新生成
        # 更好的方案是缓存原始数据（不翻译），但需要修改缓存结构
        # 暂时：如果语言是英文且缓存存在，我们重新生成（跳过缓存）
        if lang == "en":
            # 英文模式下，跳过缓存，重新生成以确保翻译正确
            cached_result = None
        else:
            # 中文模式下，直接使用缓存
            return {
                "success": True,
                "data": cached_result,
                "cached": True,
            }
    
    try:
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        redis_list = provider.list_redis()
        all_resources = instances + rds_list + redis_list
        
        analyzer = SecurityComplianceAnalyzer()
        
        # 公网暴露检测
        exposed = analyzer.detect_public_exposure(all_resources)
        
        # 安全检查
        stopped = analyzer.check_stopped_instances(instances)
        tag_coverage, no_tags = analyzer.check_missing_tags(all_resources)
        
        # 磁盘加密检查
        encryption_info = analyzer.check_disk_encryption(instances)
        
        # 抢占式实例检查
        preemptible = analyzer.check_preemptible_instances(instances)
        
        # EIP使用情况（如果有EIP数据）
        eip_info = {"total": 0, "bound": 0, "unbound": 0, "unbound_rate": 0}
        try:
            eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
        except:
            pass
        
        # 计算安全评分（更详细的评分逻辑）
        security_score = 100
        score_deductions = []
        
        if len(exposed) > 0:
            deduction = min(len(exposed) * 5, 30)  # 最多扣30分
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.public_exposure", lang, count=len(exposed)),
                "deduction": deduction
            })
        
        if len(stopped) > 0:
            deduction = min(len(stopped) * 2, 20)  # 最多扣20分
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.stopped_instances", lang, count=len(stopped)),
                "deduction": deduction
            })
        
        if tag_coverage < 80:
            deduction = 10 if tag_coverage < 50 else 5
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.tag_coverage", lang, coverage=tag_coverage),
                "deduction": deduction
            })
        
        if encryption_info.get("encryption_rate", 100) < 50:
            deduction = 15
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.disk_encryption", lang, rate=encryption_info.get('encryption_rate', 0)),
                "deduction": deduction
            })
        
        if len(preemptible) > 0:
            deduction = min(len(preemptible) * 3, 15)  # 最多扣15分
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.preemptible_instances", lang, count=len(preemptible)),
                "deduction": deduction
            })
        
        if eip_info.get("unbound_rate", 0) > 20:
            deduction = 5
            security_score -= deduction
            score_deductions.append({
                "reason": get_translation("security.score_deductions.eip_unbound", lang, rate=eip_info.get('unbound_rate', 0)),
                "deduction": deduction
            })
        
        security_score = max(0, min(100, security_score))
        
        # 生成安全改进建议（支持国际化）
        security_summary = {
            "exposed_count": len(exposed),
            "stopped_count": len(stopped),
            "tag_coverage_rate": tag_coverage,
            "encryption_rate": encryption_info.get("encryption_rate", 100),
            "preemptible_count": len(preemptible),
            "unbound_eip": eip_info.get("unbound", 0),
        }
        suggestions = analyzer.suggest_security_improvements(security_summary, locale=lang)
        
        result_data = {
            "security_score": security_score,
            "exposed_count": len(exposed),
            "stopped_count": len(stopped),
            "tag_coverage": tag_coverage,
            "missing_tags_count": len(no_tags),
            "alert_count": len(exposed) + len(stopped) + len(preemptible),
            "encryption_rate": encryption_info.get("encryption_rate", 100),
            "encrypted_count": encryption_info.get("encrypted", 0),
            "unencrypted_count": encryption_info.get("unencrypted_count", 0),
            "preemptible_count": len(preemptible),
            "eip_total": eip_info.get("total", 0),
            "eip_bound": eip_info.get("bound", 0),
            "eip_unbound": eip_info.get("unbound", 0),
            "eip_unbound_rate": eip_info.get("unbound_rate", 0),
            "score_deductions": score_deductions,
            "suggestions": suggestions,
        }
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type="security_overview", account_name=account_name, data=result_data)
        
        return {
            "success": True,
            "data": result_data,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": {
                "security_score": 0,
                "exposed_count": 0,
                "stopped_count": 0,
                "tag_coverage": 0,
                "missing_tags_count": 0,
                "alert_count": 0,
                "encryption_rate": 0,
                "encrypted_count": 0,
                "unencrypted_count": 0,
                "preemptible_count": 0,
                "eip_total": 0,
                "eip_bound": 0,
                "eip_unbound": 0,
                "eip_unbound_rate": 0,
                "score_deductions": [],
                "suggestions": [],
            },
            "cached": False,
        }


@router.get("/security/checks")
def get_security_checks(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="强制刷新缓存"),
    locale: Optional[str] = Query("zh", description="语言设置: zh 或 en")
):
    """获取安全检查结果（带24小时缓存）"""
    # 获取语言设置
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据
    # 注意：为了支持多语言，如果语言不是中文，我们跳过缓存重新生成
    cached_result = None
    if not force_refresh and lang == "zh":
        cached_result = cache_manager.get(resource_type="security_checks", account_name=account_name)
    
    # 如果缓存有效，直接使用缓存数据（中文）
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    try:
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        redis_list = provider.list_redis()
        all_resources = instances + rds_list + redis_list
        
        analyzer = SecurityComplianceAnalyzer()
        
        exposed = analyzer.detect_public_exposure(all_resources)
        stopped = analyzer.check_stopped_instances(instances)
        tag_coverage, no_tags = analyzer.check_missing_tags(all_resources)
        encryption_info = analyzer.check_disk_encryption(instances)
        preemptible = analyzer.check_preemptible_instances(instances)
        
        # EIP使用情况
        eip_info = {"total": 0, "bound": 0, "unbound": 0, "unbound_eips": []}
        try:
            eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
        except:
            pass
        
        checks = []
        
        # 公网暴露检查
        if exposed:
            checks.append({
                "type": "public_exposure",
                "title": get_translation("security.public_exposure.title", lang),
                "description": get_translation("security.public_exposure.description_failed", lang),
                "status": "failed",
                "severity": "HIGH",
                "count": len(exposed),
                "resources": exposed[:20],
                "recommendation": get_translation("security.public_exposure.recommendation", lang),
            })
        else:
            checks.append({
                "type": "public_exposure",
                "title": get_translation("security.public_exposure.title", lang),
                "description": get_translation("security.public_exposure.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # 停止实例检查
        if stopped:
            checks.append({
                "type": "stopped_instances",
                "title": get_translation("security.stopped_instances.title", lang),
                "description": get_translation("security.stopped_instances.description", lang),
                "status": "warning",
                "severity": "MEDIUM",
                "count": len(stopped),
                "resources": stopped[:20],
                "recommendation": get_translation("security.stopped_instances.recommendation", lang),
            })
        else:
            checks.append({
                "type": "stopped_instances",
                "title": get_translation("security.stopped_instances.title", lang),
                "description": get_translation("security.stopped_instances.description", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # 标签检查
        if tag_coverage >= 80:
            checks.append({
                "type": "tag_coverage",
                "title": get_translation("security.tag_coverage.title", lang),
                "description": get_translation("security.tag_coverage.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "coverage": tag_coverage,
                "missing_count": len(no_tags),
                "resources": [],
                "recommendation": get_translation("security.tag_coverage.recommendation", lang),
            })
        else:
            checks.append({
                "type": "tag_coverage",
                "title": get_translation("security.tag_coverage.title"),
                "description": get_translation("security.tag_coverage.description_failed"),
                "status": "warning",
                "severity": "MEDIUM",
                "coverage": tag_coverage,
                "missing_count": len(no_tags),
                "resources": no_tags[:20],
                "recommendation": get_translation("security.tag_coverage.recommendation", lang),
            })
        
        # 磁盘加密检查
        encryption_rate = encryption_info.get("encryption_rate", 100)
        if encryption_rate < 100:
            checks.append({
                "type": "disk_encryption",
                "title": get_translation("security.disk_encryption.title"),
                "description": get_translation("security.disk_encryption.description_failed"),
                "status": "warning" if encryption_rate < 50 else "passed",
                "severity": "HIGH" if encryption_rate < 50 else "MEDIUM",
                "encryption_rate": encryption_rate,
                "encrypted_count": encryption_info.get("encrypted", 0),
                "unencrypted_count": encryption_info.get("unencrypted_count", 0),
                "resources": encryption_info.get("unencrypted_instances", [])[:20],
                "recommendation": get_translation("security.disk_encryption.recommendation", lang),
            })
        else:
            checks.append({
                "type": "disk_encryption",
                "title": get_translation("security.disk_encryption.title"),
                "description": get_translation("security.disk_encryption.description_passed", lang),
                "status": "passed",
                "severity": "INFO",
                "encryption_rate": encryption_rate,
                "encrypted_count": encryption_info.get("encrypted", 0),
                "unencrypted_count": 0,
                "resources": [],
            })
        
        # 抢占式实例检查
        if preemptible:
            checks.append({
                "type": "preemptible_instances",
                "title": get_translation("security.preemptible_instances.title", lang),
                "description": get_translation("security.preemptible_instances.description", lang),
                "status": "warning",
                "severity": "MEDIUM",
                "count": len(preemptible),
                "resources": preemptible[:20],
                "recommendation": get_translation("security.preemptible_instances.recommendation", lang),
            })
        else:
            checks.append({
                "type": "preemptible_instances",
                "title": get_translation("security.preemptible_instances.title", lang),
                "description": get_translation("security.preemptible_instances.description", lang),
                "status": "passed",
                "severity": "INFO",
                "count": 0,
                "resources": [],
            })
        
        # EIP使用情况检查
        if eip_info.get("total", 0) > 0:
            unbound_rate = eip_info.get("unbound_rate", 0)
            if unbound_rate > 20:
                checks.append({
                    "type": "eip_usage",
                    "title": get_translation("security.eip_usage.title", lang),
                    "description": get_translation("security.eip_usage.description_failed", lang),
                    "status": "warning",
                    "severity": "MEDIUM",
                    "total": eip_info.get("total", 0),
                    "bound": eip_info.get("bound", 0),
                    "unbound": eip_info.get("unbound", 0),
                    "unbound_rate": unbound_rate,
                    "resources": eip_info.get("unbound_eips", [])[:20],
                    "recommendation": get_translation("security.eip_usage.recommendation", lang, unbound=eip_info.get('unbound', 0)),
                })
            else:
                checks.append({
                    "type": "eip_usage",
                    "title": get_translation("security.eip_usage.title"),
                    "description": get_translation("security.eip_usage.description_passed", lang),
                    "status": "passed",
                    "severity": "INFO",
                    "total": eip_info.get("total", 0),
                    "bound": eip_info.get("bound", 0),
                    "unbound": eip_info.get("unbound", 0),
                    "unbound_rate": unbound_rate,
                    "resources": [],
                })
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type="security_checks", account_name=account_name, data=checks)
        
        return {
            "success": True,
            "data": checks,
            "cached": False,
        }
    except Exception as e:
        return {
            "success": True,
            "data": [],
            "cached": False,
        }


# ==================== Phase 2: Optimization APIs ====================

@router.get("/optimization/suggestions")
def get_optimization_suggestions(
    account: Optional[str] = None, 
    force_refresh: bool = Query(False, description="强制刷新缓存"),
    locale: Optional[str] = Query("zh", description="语言设置: zh 或 en")
):
    """获取优化建议（带24小时缓存）"""
    # 获取语言设置
    lang: Locale = get_locale_from_request(
        request_headers=None,
        query_params={"locale": locale}
    )
    
    provider, account_name = _get_provider_for_account(account)
    
    # 初始化缓存管理器，TTL设置为24小时（86400秒）
    cache_manager = CacheManager(ttl_seconds=86400)
    
    # 尝试从缓存获取数据（优先使用缓存，避免长时间等待）
    cached_result = None
    if not force_refresh:
        cached_result = cache_manager.get(resource_type="optimization_suggestions", account_name=account_name)
    
    # 如果缓存有效，直接使用缓存数据
    if cached_result is not None:
        return {
            "success": True,
            "data": cached_result,
            "cached": True,
        }
    
    try:
        from cloudlens.core.optimization_engine import OptimizationEngine
        from cloudlens.core.security_compliance import SecurityComplianceAnalyzer
        from cloudlens.core.cost_analyzer import CostAnalyzer
        cm = ConfigManager()
        account_config = cm.get_account(account_name)
        
        suggestions = []
        all_opportunities = []
        
        # 1. 使用 OptimizationEngine 分析优化机会（跳过，因为已经有其他分析覆盖）
        # 这个操作很慢，而且其他分析已经覆盖了主要场景
        # try:
        #     engine = OptimizationEngine()
        #     opportunities = engine.analyze_optimization_opportunities(account_name)
        #     all_opportunities.extend(opportunities)
        # except Exception as e:
        #     pass  # 如果失败，继续其他分析
        
        # 2. 闲置资源建议（基于真实成本，使用统一的缓存键）
        idle_data = cache_manager.get(resource_type="idle_result", account_name=account_name)
        # 如果没有，尝试从 dashboard_idle 缓存获取（兼容旧缓存）
        if not idle_data:
            idle_data = cache_manager.get(resource_type="dashboard_idle", account_name=account_name)
        if idle_data:
            # 计算真实节省潜力
            total_savings = 0.0
            if account_config:
                cost_map = _get_cost_map("ecs", account_config)
                for idle_item in idle_data:
                    instance_id = idle_item.get("instance_id") or idle_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id)
                        if cost is None:
                            cost = _estimate_monthly_cost_from_spec(idle_item.get("spec", ""), "ecs")
                        total_savings += cost
            
            suggestions.append({
                "type": "idle_resources",
                "category": get_translation("optimization.idle_resources.category", lang),
                "priority": "high",
                "title": get_translation("optimization.idle_resources.title", lang),
                "description": get_translation("optimization.idle_resources.description", lang, count=len(idle_data)),
                "savings_potential": round(total_savings, 2),
                "resource_count": len(idle_data),
                "resources": idle_data[:10],  # 返回前10个
                "action": "release_or_downgrade",
                "recommendation": get_translation("optimization.idle_resources.recommendation", lang),
            })
        
        # 3. 停止实例建议（优先使用缓存，避免重复调用API）
        instances = None
        try:
            # 尝试从缓存获取实例列表
            instances_cache = cache_manager.get(resource_type="ecs_instances", account_name=account_name)
            if instances_cache:
                instances = instances_cache
            else:
                instances = provider.list_instances()
                # 缓存实例列表（5分钟有效）
                if instances:
                    cache_manager.set(resource_type="ecs_instances", account_name=account_name, data=instances, ttl_seconds=300)
        except Exception as e:
            logger.warning(f"获取实例列表失败，使用空列表: {e}")
            instances = []
        
        analyzer = SecurityComplianceAnalyzer()
        stopped = analyzer.check_stopped_instances(instances or [])
        if stopped:
            # 计算停止实例的成本
            stopped_savings = 0.0
            if account_config:
                cost_map = _get_cost_map("ecs", account_config)
                for stop_item in stopped:
                    instance_id = stop_item.get("id")
                    if instance_id:
                        cost = cost_map.get(instance_id)
                        if cost is None:
                            cost = 300  # 默认估算
                        # 停止实例仍产生磁盘费用，假设可节省70%
                        stopped_savings += cost * 0.7
            
            suggestions.append({
                "type": "stopped_instances",
                "category": get_translation("optimization.stopped_instances.category", lang),
                "priority": "medium",
                "title": get_translation("optimization.stopped_instances.title", lang),
                "description": get_translation("optimization.stopped_instances.description", lang, count=len(stopped)),
                "savings_potential": round(stopped_savings, 2),
                "resource_count": len(stopped),
                "resources": stopped[:10],
                "action": "release",
                "recommendation": get_translation("optimization.stopped_instances.recommendation", lang),
            })
        
        # 4. 未绑定EIP建议（使用缓存，避免重复API调用）
        try:
            # 尝试从缓存获取EIP列表
            eips_cache = cache_manager.get(resource_type="eip_list", account_name=account_name)
            if eips_cache:
                eips = eips_cache
            else:
                eips = provider.list_eip() if hasattr(provider, 'list_eip') else (provider.list_eips() if hasattr(provider, 'list_eips') else [])
                # 缓存EIP列表（5分钟有效）
                if eips:
                    cache_manager.set(resource_type="eip_list", account_name=account_name, data=eips, ttl_seconds=300)
            
            if eips:
                eip_info = analyzer.analyze_eip_usage(eips)
                unbound_eips = eip_info.get("unbound_eips", [])
                if unbound_eips:
                    # EIP 费用估算：每个未绑定EIP约20元/月
                    eip_savings = len(unbound_eips) * 20
                    suggestions.append({
                        "type": "unbound_eips",
                        "category": get_translation("optimization.unbound_eips.category", lang),
                        "priority": "high",
                        "title": get_translation("optimization.unbound_eips.title", lang),
                        "description": get_translation("optimization.unbound_eips.description", lang, count=len(unbound_eips)),
                        "savings_potential": eip_savings,
                        "resource_count": len(unbound_eips),
                        "resources": unbound_eips[:10],
                        "action": "release",
                        "recommendation": get_translation("optimization.unbound_eips.recommendation", lang),
                    })
        except Exception as e:
            logger.warning(f"EIP分析失败: {e}")
            pass
        
        # 5. 标签完善建议（需要实例列表）
        tag_coverage, no_tags = analyzer.check_missing_tags(instances or [])
        if len(no_tags) > 0:
            suggestions.append({
                "type": "missing_tags",
                "category": get_translation("optimization.missing_tags.category", lang),
                "priority": "medium",
                "title": get_translation("optimization.missing_tags.title", lang),
                "description": get_translation("optimization.missing_tags.description", lang, count=len(no_tags)),
                "savings_potential": 0,
                "resource_count": len(no_tags),
                "resources": no_tags[:10],
                "action": "add_tags",
                "recommendation": get_translation("optimization.missing_tags.recommendation", lang),
            })
        
        # 6. 规格降配建议（跳过，因为 OptimizationEngine 已被禁用）
        # downgrade_opportunities = [opp for opp in all_opportunities if opp.get("action") == "downgrade"]
        # if downgrade_opportunities:
        #     total_downgrade_savings = sum(opp.get("estimated_savings", 0) for opp in downgrade_opportunities)
        #     suggestions.append({
        #         "type": "spec_downgrade",
        #         "category": get_translation("optimization.spec_downgrade.category", lang),
        #         "priority": "medium",
        #         "title": get_translation("optimization.spec_downgrade.title", lang),
        #         "description": get_translation("optimization.spec_downgrade.description", lang, count=len(downgrade_opportunities)),
        #         "savings_potential": round(total_downgrade_savings, 2),
        #         "resource_count": len(downgrade_opportunities),
        #         "resources": downgrade_opportunities[:10],
        #         "action": "downgrade",
        #         "recommendation": get_translation("optimization.spec_downgrade.recommendation", lang),
        #     })
        
        # 7. 公网暴露优化建议（需要实例列表）
        exposed = analyzer.detect_public_exposure(instances or [])
        if exposed:
            suggestions.append({
                "type": "public_exposure",
                "category": get_translation("optimization.public_exposure.category", lang),
                "priority": "high",
                "title": get_translation("optimization.public_exposure.title", lang),
                "description": get_translation("optimization.public_exposure.description", lang, count=len(exposed)),
                "savings_potential": 0,
                "resource_count": len(exposed),
                "resources": exposed[:10],
                "action": "secure",
                "recommendation": get_translation("optimization.public_exposure.recommendation", lang),
            })
        
        # 8. 磁盘加密建议（需要实例列表）
        encryption_info = analyzer.check_disk_encryption(instances or [])
        if encryption_info.get("encryption_rate", 100) < 50:
            suggestions.append({
                "type": "disk_encryption",
                "category": get_translation("optimization.disk_encryption.category", lang),
                "priority": "high",
                "title": get_translation("optimization.disk_encryption.title", lang),
                "description": get_translation("optimization.disk_encryption.description", lang, count=encryption_info.get('unencrypted_count', 0)),
                "savings_potential": 0,
                "resource_count": encryption_info.get("unencrypted_count", 0),
                "resources": encryption_info.get("unencrypted_instances", [])[:10],
                "action": "enable_encryption",
                "recommendation": get_translation("optimization.disk_encryption.recommendation", lang),
            })
        
        # 按优先级和节省潜力排序
        suggestions.sort(key=lambda x: (
            {"high": 0, "medium": 1, "low": 2}.get(x.get("priority", "low"), 2),
            -x.get("savings_potential", 0)
        ))
        
        # 计算总节省潜力
        total_savings = sum(s.get("savings_potential", 0) for s in suggestions)
        
        result = {
            "suggestions": suggestions,
            "summary": {
                "total_suggestions": len(suggestions),
                "total_savings_potential": round(total_savings, 2),
                "high_priority_count": sum(1 for s in suggestions if s.get("priority") == "high"),
                "medium_priority_count": sum(1 for s in suggestions if s.get("priority") == "medium"),
                "low_priority_count": sum(1 for s in suggestions if s.get("priority") == "low"),
            }
        }
        
        # 保存到缓存（24小时有效）
        cache_manager.set(resource_type="optimization_suggestions", account_name=account_name, data=result)
        
        return {
            "success": True,
            "data": result,
            "cached": False,
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"获取优化建议失败: {str(e)}\n{error_trace}")
        # 即使出错也返回空结果，而不是抛出异常，避免前端崩溃
        return {
            "success": True,
            "data": {
                "suggestions": [],
                "summary": {
                    "total_suggestions": 0,
                    "total_savings_potential": 0,
                    "high_priority_count": 0,
                    "medium_priority_count": 0,
                    "low_priority_count": 0,
                }
            },
            "cached": False,
            "error": str(e)  # 可选：返回错误信息用于调试
        }


# ==================== Phase 3: Reports APIs ====================

@router.post("/reports/generate")
def generate_report(report_data: Dict[str, Any]):
    """生成报告"""
    account = report_data.get("account")
    report_type = report_data.get("type", "comprehensive")
    format_type = report_data.get("format", "excel")
    
    try:
        from cloudlens.core.report_generator import ReportGenerator
        
        provider, account_name = _get_provider_for_account(account)
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        
        # 构建报告数据
        data = {
            "account": account_name,
            "instances": instances,
            "rds": rds_list,
        }
        
        if format_type == "html":
            report_content = ReportGenerator.generate_html(account_name, data)
            return {
                "success": True,
                "data": {
                    "format": "html",
                    "content": report_content,
                }
            }
        elif format_type == "excel":
            # 实现Excel生成
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from fastapi.responses import StreamingResponse

            wb = Workbook()

            # 资源概览Sheet
            ws_overview = wb.active
            ws_overview.title = "资源概览"

            # 标题行样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            # ECS概览
            ws_overview['A1'] = "CloudLens 云资源报告"
            ws_overview['A1'].font = Font(size=16, bold=True)
            ws_overview.merge_cells('A1:E1')

            ws_overview['A3'] = "账号"
            ws_overview['B3'] = account_name

            # ECS实例列表
            ecs_start_row = 5
            headers = ['实例ID', '名称', '规格', '状态', '区域']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_overview.cell(row=ecs_start_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 填充ECS数据
            for row_idx, instance in enumerate(instances, ecs_start_row + 1):
                ws_overview.cell(row=row_idx, column=1, value=instance.id)
                ws_overview.cell(row=row_idx, column=2, value=instance.name)
                ws_overview.cell(row=row_idx, column=3, value=instance.spec)
                ws_overview.cell(row=row_idx, column=4, value=instance.status.value)
                ws_overview.cell(row=row_idx, column=5, value=instance.region)

            # RDS实例Sheet
            if rds_list:
                ws_rds = wb.create_sheet("RDS实例")
                rds_headers = ['实例ID', '名称', '引擎', '版本', '状态', '区域']
                for col_idx, header in enumerate(rds_headers, 1):
                    cell = ws_rds.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, rds in enumerate(rds_list, 2):
                    ws_rds.cell(row=row_idx, column=1, value=rds.id)
                    ws_rds.cell(row=row_idx, column=2, value=rds.name)
                    ws_rds.cell(row=row_idx, column=3, value=rds.spec)
                    ws_rds.cell(row=row_idx, column=4, value=getattr(rds, 'version', 'N/A'))
                    ws_rds.cell(row=row_idx, column=5, value=rds.status.value)
                    ws_rds.cell(row=row_idx, column=6, value=rds.region)

            # 保存到BytesIO
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # 返回文件流
            filename = f"cloudlens_report_{account_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format_type}")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# Additional API endpoints to append to api.py

# ==================== Phase 2: Budget Management APIs ====================

@router.get("/cost/budget")
def get_budget(account: Optional[str] = None):
    """获取预算信息"""
    # TODO: 实现预算存储和查询（可以使用文件或数据库）
    return {
        "success": True,
        "data": {
            "monthly_budget": 0,
            "annual_budget": 0,
            "current_month_spent": 0,
            "usage_rate": 0,
        }
    }


# ==================== Billing APIs (BSS OpenAPI) ====================

@router.get("/billing/overview")
def get_billing_overview(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="账期，格式 yyyy-MM，默认当月"),
    product_code: Optional[str] = Query(None, description="产品代码过滤（可选）"),
    subscription_type: Optional[str] = Query(None, description="订阅类型过滤：Subscription / PayAsYouGo（可选）"),
):
    """
    获取账单概览（阿里云 BSS OpenAPI）。

    用途：
    - 验证当前账号 AK 是否具备账单读取权限
    - 为后续“按实例真实成本（含折扣）”打通数据源
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")

    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="当前仅支持阿里云账单（BSS OpenAPI）")

    # 默认当月账期
    if not billing_cycle:
        from datetime import datetime
        billing_cycle = datetime.now().strftime("%Y-%m")

    # 动态导入：避免在未安装 SDK 的环境下直接 import 失败
    try:
        from aliyunsdkcore.client import AcsClient
        from aliyunsdkcore.request import CommonRequest
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"阿里云 SDK 不可用：{e}")

    try:
        import json

        # BSS OpenAPI 不区分地域，但 SDK 需要 region 参数
        client = AcsClient(
            account_config.access_key_id,
            account_config.access_key_secret,
            "cn-hangzhou",
        )

        request = CommonRequest()
        request.set_domain("business.aliyuncs.com")
        request.set_version("2017-12-14")
        request.set_action_name("QueryBillOverview")
        request.set_method("POST")

        request.add_query_param("BillingCycle", billing_cycle)
        if product_code:
            request.add_query_param("ProductCode", product_code)
        if subscription_type:
            request.add_query_param("SubscriptionType", subscription_type)

        resp = client.do_action_with_exception(request)
        data = json.loads(resp)
        return {"success": True, "data": data}
    except Exception as e:
        # 常见：UnauthorizedOperation / Forbidden / InvalidAccessKeyId.NotFound / SignatureDoesNotMatch
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/billing/instance-bill")
def get_billing_instance_bill(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="账期 yyyy-MM，默认当月"),
    product_code: str = Query(..., description="产品代码，如 ecs/rds/kvstore/yundisk/snapshot/slb/eip/nat_gw"),
    subscription_type: Optional[str] = Query(None, description="Subscription / PayAsYouGo（可选）"),
    limit: int = Query(50, ge=1, le=500, description="返回前 N 条用于调试"),
):
    """
    调试接口：拉取 BSS QueryInstanceBill 原始数据，便于确认 InstanceID 的字段与格式。
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="当前仅支持阿里云账单（BSS OpenAPI）")

    if not billing_cycle:
        billing_cycle = _get_billing_cycle_default()

    try:
        rows = _bss_query_instance_bill(
            account_config=account_config,
            billing_cycle=billing_cycle,
            product_code=product_code,
            subscription_type=subscription_type,
        )
        return {
            "success": True,
            "data": {
                "billing_cycle": billing_cycle,
                "product_code": product_code,
                "subscription_type": subscription_type,
                "count": len(rows),
                "items": rows[:limit],
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/discounts/trend")
def get_discount_trend(
    account: Optional[str] = Query(None, description="账号名称"),
    months: int = Query(19, ge=1, le=999, description="分析月数，默认19个月，设置为99或更大表示全部历史数据"),
    force_refresh: bool = Query(False, description="强制刷新缓存"),
):
    """
    折扣趋势分析 - 基于数据库全量数据分析
    
    数据来源：MySQL数据库（自动同步最新账单数据）
    支持：
    - 查看长期折扣趋势（最多19个月历史）
    - 分析商务合同折扣效果
    - 按产品/实例/合同维度查看折扣分布
    - 实时更新，无需手动下载CSV
    """
    from cloudlens.core.discount_analyzer_db import DiscountAnalyzerDB
    import os
    
    try:
        # 获取账号信息
        cm = ConfigManager()
        if not account:
            # 尝试获取当前账号
            ctx = ContextManager()
            account = ctx.get_last_account()
        if not account:
            # 使用第一个账号
            accounts = cm.list_accounts()
            if accounts:
                account = accounts[0].name
            else:
                raise HTTPException(status_code=400, detail="未找到账号配置")
        
        account_config = cm.get_account(account)
        if not account_config:
            raise HTTPException(status_code=404, detail=f"账号 '{account}' 不存在")
        
        # 生成账号ID（与bill_fetcher保持一致）
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        # 使用数据库版折扣分析器（默认使用MySQL）
        analyzer = DiscountAnalyzerDB()
        
        # 分析折扣趋势
        result = analyzer.analyze_discount_trend(
            account_id=account_id,
            months=months
        )
        
        if 'error' in result:
            return {
                "success": False,
                "error": result['error']
            }
        
        # 转换数据格式以匹配前端期望的结构
        from datetime import datetime
        
        # 提取数据
        monthly_trend = result.get('monthly_trend', [])
        product_discounts = result.get('product_discounts', [])
        instance_discounts = result.get('instance_discounts', [])
        contract_discounts = result.get('contract_discounts', [])
        summary = result.get('summary', {})
        
        # 构建前端期望的数据结构
        response_data = {
            "account_name": account,
            "analysis_periods": [m['month'] for m in monthly_trend],
            
            # trend_analysis 格式
            "trend_analysis": {
                "timeline": [
                    {
                        "period": m['month'],
                        "official_price": m['official_price'],
                        "discount_amount": m['discount_amount'],
                        "discount_rate": m['discount_rate'],
                        "payable_amount": m['actual_amount']
                    }
                    for m in monthly_trend
                ],
                "latest_period": monthly_trend[-1]['month'] if monthly_trend else "",
                "latest_discount_rate": summary.get('latest_discount_rate', 0),
                "discount_rate_change": summary.get('trend_change_pct', 0) / 100,
                "discount_rate_change_pct": summary.get('trend_change_pct', 0),
                "discount_amount_change": 0,  # 可以计算
                "trend_direction": summary.get('trend', '平稳'),
                "average_discount_rate": summary.get('avg_discount_rate', 0),
                "max_discount_rate": max([m['discount_rate'] for m in monthly_trend], default=0),
                "min_discount_rate": min([m['discount_rate'] for m in monthly_trend], default=0),
                "total_savings_6m": summary.get('total_discount', 0),
            },
            
            # product_analysis 格式
            "product_analysis": {
                p['product']: {
                    "total_discount": p['discount_amount'],
                    "avg_discount_rate": p['discount_rate'],
                    "latest_discount_rate": p['discount_rate'],
                    "rate_change": 0,
                    "trend": "平稳",
                    "periods": [m['month'] for m in monthly_trend],
                    "discount_rates": [p['discount_rate']] * len(monthly_trend),
                }
                for p in product_discounts
            },
            
            # contract_analysis 格式（如果有合同数据）
            "contract_analysis": {
                c['contract_name']: {
                    "discount_name": c['contract_name'],
                    "total_discount": c.get('total_discount', 0),
                    "avg_discount_rate": c.get('avg_discount_rate', 0),
                    "latest_discount_rate": c.get('latest_discount_rate', 0),
                    "periods": c.get('periods', []),
                    "discount_amounts": c.get('discount_amounts', []),
                }
                for c in contract_discounts
            },
            
            # top_instance_discounts 格式
            "top_instance_discounts": [
                {
                    "instance_id": i['instance_id'],
                    "instance_name": i.get('instance_name', i['instance_id']),
                    "product_name": i['product'],
                    "official_price": i['official_price'],
                    "discount_amount": i['discount_amount'],
                    "payable_amount": i['actual_amount'],
                    "discount_rate": i['discount_rate'],
                    "discount_pct": i['discount_rate'] * 100,
                }
                for i in instance_discounts
            ],
            
            "generated_at": datetime.now().isoformat(),
        }
        
        return {
            "success": True,
            "data": response_data,
            "cached": False,
            "source": "database",
            "account": account,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/discounts/products")
def get_product_discounts(
    account: Optional[str] = Query(None, description="账号名称"),
    product: Optional[str] = Query(None, description="产品名称过滤"),
    months: int = Query(19, ge=1, le=999, description="分析月数，设置为99或更大表示全部历史数据"),
    force_refresh: bool = Query(False, description="强制刷新缓存"),
):
    """
    产品折扣详情 - 基于数据库查看特定产品的折扣明细
    """
    from cloudlens.core.discount_analyzer_db import DiscountAnalyzerDB
    import os
    
    try:
        # 获取账号信息
        cm = ConfigManager()
        if not account:
            ctx = ContextManager()
            account = ctx.get_last_account()
        if not account:
            accounts = cm.list_accounts()
            if accounts:
                account = accounts[0].name
            else:
                raise HTTPException(status_code=400, detail="未找到账号配置")
        
        account_config = cm.get_account(account)
        if not account_config:
            raise HTTPException(status_code=404, detail=f"账号 '{account}' 不存在")
        
        # 生成账号ID
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        # 使用数据库版折扣分析器（默认使用MySQL）
        analyzer = DiscountAnalyzerDB()
        
        result = analyzer.analyze_discount_trend(account_id=account_id, months=months)
        
        if 'error' in result:
            return {"success": False, "error": result['error']}
        
        # 提取产品折扣数据
        product_data = result['product_analysis']
        
        # 如果指定了产品过滤
        if product:
            product_data = {k: v for k, v in product_data.items() if product.lower() in k.lower()}
        
        return {
            "success": True,
            "data": {
                "products": product_data,
                "analysis_periods": result['analysis_periods'],
            },
            "source": "database",
            "account": account,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/billing/discounts")
def get_billing_discounts(
    account: str,
    billing_cycle: Optional[str] = Query(None, description="账期 yyyy-MM，默认当月"),
    force_refresh: bool = Query(False, description="强制刷新缓存"),
):
    """
    折扣梳理（按产品 + 计费方式聚合）- 基于BSS API实时查询
    
    注意：这是实时API接口，与 /discounts/trend（基于CSV离线分析）互补。
    - 实时API：查询当前月折扣情况
    - CSV分析：查看历史6个月折扣趋势

    口径说明：
    - PretaxGrossAmount：税前原价（未折扣/未优惠抵扣前）
    - PretaxAmount：税前应付（折扣/优惠抵扣后）
    - 对 PayAsYouGo，PaymentAmount 常为 0（未出账/未结算），请以 PretaxAmount/OutstandingAmount 为主
    """
    cm = ConfigManager()
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"Account '{account}' not found")
    provider = getattr(account_config, "provider", None) or "aliyun"
    if provider != "aliyun":
        raise HTTPException(status_code=400, detail="当前仅支持阿里云账单（BSS OpenAPI）")

    if not billing_cycle:
        billing_cycle = _get_billing_cycle_default()

    cache_manager = CacheManager(ttl_seconds=86400)
    cache_key = f"billing_discounts_{billing_cycle}"
    if not force_refresh:
        cached = cache_manager.get(resource_type=cache_key, account_name=account_config.name)
        if isinstance(cached, dict) and cached.get("billing_cycle") == billing_cycle:
            return {"success": True, "data": cached, "cached": True}

    def f(x: Any) -> float:
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    items = _bss_query_bill_overview(account_config, billing_cycle)

    agg: Dict[str, Dict[str, Any]] = {}
    for it in items:
        code = (it.get("ProductCode") or it.get("PipCode") or "unknown")
        sub = (it.get("SubscriptionType") or "Unknown")
        key = f"{code}::{sub}"
        if key not in agg:
            agg[key] = {
                "product_code": code,
                "product_name": it.get("ProductName") or code,
                "subscription_type": sub,
                "pretax_gross_amount": 0.0,
                "pretax_amount": 0.0,
                "payment_amount": 0.0,
                "outstanding_amount": 0.0,
                "invoice_discount": 0.0,
                "round_down_discount": 0.0,
                "deducted_by_coupons": 0.0,
                "deducted_by_cash_coupons": 0.0,
                "deducted_by_prepaid_card": 0.0,
                "adjust_amount": 0.0,
                "cash_amount": 0.0,
                "currency": it.get("Currency") or "CNY",
            }

        row = agg[key]
        row["pretax_gross_amount"] += f(it.get("PretaxGrossAmount"))
        row["pretax_amount"] += f(it.get("PretaxAmount"))
        row["payment_amount"] += f(it.get("PaymentAmount"))
        row["outstanding_amount"] += f(it.get("OutstandingAmount"))
        row["invoice_discount"] += f(it.get("InvoiceDiscount"))
        row["round_down_discount"] += f(it.get("RoundDownDiscount"))
        row["deducted_by_coupons"] += f(it.get("DeductedByCoupons"))
        row["deducted_by_cash_coupons"] += f(it.get("DeductedByCashCoupons"))
        row["deducted_by_prepaid_card"] += f(it.get("DeductedByPrepaidCard"))
        row["adjust_amount"] += f(it.get("AdjustAmount"))
        row["cash_amount"] += f(it.get("CashAmount"))

    rows = []
    total_gross = 0.0
    total_pretax = 0.0
    total_discount_amount = 0.0
    for row in agg.values():
        gross = float(row.get("pretax_gross_amount") or 0.0)
        pretax = float(row.get("pretax_amount") or 0.0)
        payment_amount = float(row.get("payment_amount") or 0.0)
        outstanding_amount = float(row.get("outstanding_amount") or 0.0)
        invoice_discount = float(row.get("invoice_discount") or 0.0)
        round_down_discount = float(row.get("round_down_discount") or 0.0)
        deducted_by_coupons = float(row.get("deducted_by_coupons") or 0.0)
        deducted_by_cash_coupons = float(row.get("deducted_by_cash_coupons") or 0.0)
        deducted_by_prepaid_card = float(row.get("deducted_by_prepaid_card") or 0.0)
        adjust_amount = float(row.get("adjust_amount") or 0.0)
        cash_amount = float(row.get("cash_amount") or 0.0)

        # “没有使用/没有发生费用”的产品不展示：所有金额字段均为 0
        # 注意：如果是“免单/全额减免”，一般表现为 gross>0 但 pretax=0，这种要展示为“免费”
        has_any_amount = any(
            abs(x) > 0
            for x in (
                gross,
                pretax,
                payment_amount,
                outstanding_amount,
                invoice_discount,
                round_down_discount,
                deducted_by_coupons,
                deducted_by_cash_coupons,
                deducted_by_prepaid_card,
                adjust_amount,
                cash_amount,
            )
        )
        if not has_any_amount:
            continue

        discount_amount = max(0.0, gross - pretax) if gross > 0 else 0.0
        # 折扣口径：按“实付比例”计算折扣（x.x折），例如 30/100 => 0.3 => 3.0折
        # 注意：0.0折通常意味着“全额减免/完全被优惠抵扣”，不应展示为 0.0折，改用 free 标识在前端展示“免费”
        discount_rate = (pretax / gross) if gross > 0 else None
        is_free = (gross > 0 and pretax == 0)
        discount_zhe = (float(discount_rate) * 10.0) if (discount_rate is not None and pretax > 0) else None
        row["pretax_gross_amount"] = round(gross, 2)
        row["pretax_amount"] = round(pretax, 2)
        row["discount_amount"] = round(discount_amount, 2)
        row["discount_rate"] = round(float(discount_rate), 6) if discount_rate is not None else None
        row["discount_pct"] = round((1.0 - float(discount_rate)) * 100, 2) if discount_rate is not None else None
        row["discount_zhe"] = round(float(discount_zhe), 1) if discount_zhe is not None else None
        row["free"] = bool(is_free)
        row["payment_amount"] = round(payment_amount, 2)
        row["outstanding_amount"] = round(outstanding_amount, 2)
        row["invoice_discount"] = round(invoice_discount, 2)
        row["round_down_discount"] = round(round_down_discount, 6)
        row["deducted_by_coupons"] = round(deducted_by_coupons, 2)
        row["deducted_by_cash_coupons"] = round(deducted_by_cash_coupons, 2)
        row["deducted_by_prepaid_card"] = round(deducted_by_prepaid_card, 2)
        row["adjust_amount"] = round(adjust_amount, 2)
        row["cash_amount"] = round(cash_amount, 2)

        total_gross += gross
        total_pretax += pretax
        total_discount_amount += discount_amount
        rows.append(row)

    rows.sort(key=lambda r: float(r.get("discount_amount") or 0.0), reverse=True)

    overall_rate = (total_pretax / total_gross) if total_gross > 0 else None
    overall_free = (total_gross > 0 and total_pretax == 0)
    overall_zhe = (float(overall_rate) * 10.0) if (overall_rate is not None and total_pretax > 0) else None
    result = {
        "billing_cycle": billing_cycle,
        "summary": {
            "total_pretax_gross_amount": round(total_gross, 2),
            "total_pretax_amount": round(total_pretax, 2),
            "total_discount_amount": round(total_discount_amount, 2),
            "discount_rate": round(float(overall_rate), 6) if overall_rate is not None else None,
            "discount_pct": round((1.0 - float(overall_rate)) * 100, 2) if overall_rate is not None else None,
            "discount_zhe": round(float(overall_zhe), 1) if overall_zhe is not None else None,
            "free": bool(overall_free),
        },
        "rows": rows,
    }

    cache_manager.set(resource_type=cache_key, account_name=account_config.name, data=result)
    return {"success": True, "data": result, "cached": False}


@router.post("/cost/budget")
def set_budget(budget_data: Dict[str, Any]):
    """设置预算"""
    # TODO: 实现预算保存
    return {
        "success": True,
        "message": "预算设置成功"
    }


# ==================== Phase 2: CIS Compliance APIs ====================

@router.get("/security/cis")
def get_cis_compliance(account: Optional[str] = None):
    """获取CIS合规检查结果"""
    provider, account_name = _get_provider_for_account(account)
    
    try:
        from cloudlens.core.cis_compliance import CISBenchmark
        
        # 获取资源
        instances = provider.list_instances()
        rds_list = provider.list_rds()
        all_resources = instances + rds_list
        
        # 运行CIS检查
        checker = CISBenchmark()
        results = checker.run_all_checks(all_resources, provider)
        
        # 计算合规度
        total_checks = len(results.get("results", []))
        passed_checks = sum(1 for check in results.get("results", []) if check.get("status") == "PASS")
        compliance_rate = (passed_checks / total_checks * 100) if total_checks > 0 else 0
        
        return {
            "success": True,
            "data": {
                "compliance_rate": round(compliance_rate, 2),
                "checks": results.get("results", []),
                "summary": results.get("summary", {}),
            }
        }
    except Exception as e:
        # 如果CIS检查器不存在或出错，返回提示
        return {
            "success": True,
            "data": {
                "compliance_rate": 0,
                "checks": [],
                "message": f"CIS合规检查功能开发中: {str(e)}"
            }
        }




@router.get("/reports")
def list_reports(account: Optional[str] = None, limit: int = Query(50, ge=1, le=100)):
    """获取报告历史列表"""
    # TODO: 实现报告历史存储和查询
    return {
        "success": True,
        "data": []
    }


# ==================== Phase 1: 高级折扣分析API ====================

@router.get("/discounts/quarterly")
def get_quarterly_discount_comparison(
    account: Optional[str] = Query(None, description="账号名称"),
    quarters: int = Query(8, ge=1, le=20, description="分析季度数"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM格式)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM格式)"),
):
    """
    季度折扣对比分析
    
    返回季度维度的折扣率、消费金额、环比变化等数据
    """
    import os
    from cloudlens.core.discount_analyzer_advanced import AdvancedDiscountAnalyzer
    
    cm = ConfigManager()
    
    # 解析账号
    if not account:
        accounts = cm.list_accounts()
        if not accounts:
            raise HTTPException(status_code=404, detail="未找到任何账号配置")
        account = accounts[0].name if isinstance(accounts[0], CloudAccount) else accounts[0].get('name')
    
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"账号 '{account}' 未找到")
    
    try:
        # 使用默认配置（MySQL），不再需要db_path
        analyzer = AdvancedDiscountAnalyzer()
        
        # 构造账号ID（与bill_cmd.py保持一致）
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        result = analyzer.get_quarterly_comparison(account_id, quarters, start_date, end_date)
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', '分析失败'))
        
        return {
            "success": True,
            "data": result,
            "account": account,
            "source": "database"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"季度对比分析失败: {str(e)}")


@router.get("/discounts/yearly")
def get_yearly_discount_comparison(
    account: Optional[str] = Query(None, description="账号名称"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM格式)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM格式)"),
):
    """
    年度折扣对比分析
    
    返回年度维度的折扣率、消费金额、同比变化等数据
    """
    import os
    from cloudlens.core.discount_analyzer_advanced import AdvancedDiscountAnalyzer
    
    cm = ConfigManager()
    
    # 解析账号
    if not account:
        accounts = cm.list_accounts()
        if not accounts:
            raise HTTPException(status_code=404, detail="未找到任何账号配置")
        account = accounts[0].name if isinstance(accounts[0], CloudAccount) else accounts[0].get('name')
    
    account_config = cm.get_account(account)
    if not account_config:
        raise HTTPException(status_code=404, detail=f"账号 '{account}' 未找到")
    
    try:
        # 使用默认配置（MySQL），不再需要db_path
        analyzer = AdvancedDiscountAnalyzer()
        
        # 构造账号ID（与bill_cmd.py保持一致）
        account_id = f"{account_config.access_key_id[:10]}-{account}"
        
        result = analyzer.get_yearly_comparison(account_id, start_date, end_date)
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error', '分析失败'))
        
        return {
            "success": True,
            "data": result,
            "account": account,
            "source": "database"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"年度对比分析失败: {str(e)}")


@router.get("/discounts/product-trends")
def get_product_discount_trends(
    account: Optional[str] = Query(None, description="账号名称"),
    months: int = Query(19, ge=1, le=999, description="分析月数"),
    top_n: int = Query(20, ge=1, le=50, description="TOP N产品"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM格式)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM格式)"),
):
    """
    产品折扣趋势分析
    
    返回每个产品的月度折扣趋势、波动率、趋势变化等数据
    """
    import os

# ==================== (Migrated to api_discounts.py) ====================



# ==================== 虚拟标签系统 API ====================


# ==================== (Migrated to api_tags.py) ====================




# ==================== (Migrated to api_budgets.py) ====================




# ==================== (Migrated to api_dashboards.py) ====================




# ==================== (Migrated to api_dashboards.py) ====================

