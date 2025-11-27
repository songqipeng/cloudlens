#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
查找并列出所有未绑定的EIP
"""

import openpyxl
from datetime import datetime
import os
import glob


def read_eip_report(file_path):
    """读取EIP报告文件"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # 读取表头
        headers = [cell.value for cell in ws[1]]
        
        # 读取数据
        eips = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                eip_data = dict(zip(headers, row))
                eips.append(eip_data)
        
        wb.close()
        return eips
    except Exception as e:
        print(f"读取文件 {file_path} 失败: {e}")
        return []


def main():
    """主函数"""
    print("="*100)
    print("🔍 未绑定EIP查询")
    print(f"📅 查询时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*100)
    
    # 查找最新的EIP报告
    eip_reports = glob.glob('*eip*idle_report*.xlsx')
    
    if not eip_reports:
        print("\n❌ 未找到EIP报告文件,请先运行分析:")
        print("   python3 main.py ydzn cru eip")
        print("   python3 main.py zmyc cru eip")
        return
    
    # 按修改时间排序
    eip_reports.sort(key=os.path.getmtime, reverse=True)
    latest_report = eip_reports[0]
    
    print(f"\n📄 读取报告: {latest_report}")
    print(f"   生成时间: {datetime.fromtimestamp(os.path.getmtime(latest_report)).strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 读取EIP数据
    all_eips = read_eip_report(latest_report)
    
    # 筛选未绑定的EIP
    unbound_eips = []
    
    for eip in all_eips:
        instance_id = eip.get('绑定实例ID', '')
        instance_status = eip.get('实例状态', '')
        reason = eip.get('闲置原因', '')
        
        # 判断是否未绑定
        if (not instance_id or 
            instance_id == 'None' or 
            instance_id == '' or
            '未绑定' in str(reason) or
            '未关联' in str(reason) or
            instance_status == 'Available'):
            unbound_eips.append(eip)
    
    # 显示结果
    print("\n" + "="*100)
    if unbound_eips:
        print(f"🔴 发现 {len(unbound_eips)} 个未绑定的EIP - 建议立即释放!")
        print("="*100)
        print(f"\n{'序号':<6} {'IP地址':<18} {'分配ID':<30} {'区域':<15} {'带宽':<10} {'计费方式':<15} {'闲置原因'}")
        print("-"*100)
        
        total_cost = 0
        for i, eip in enumerate(unbound_eips, 1):
            ip = eip.get('IP地址', 'N/A')
            alloc_id = eip.get('分配ID', 'N/A')
            region = eip.get('区域', 'N/A')
            bandwidth = eip.get('带宽(Mbps)', 'N/A')
            charge_type = eip.get('计费类型', 'N/A')
            reason = str(eip.get('闲置原因', 'N/A'))[:50]
            
            # 估算每个EIP的月成本 (未绑定EIP约50元/月)
            monthly_cost = 50
            total_cost += monthly_cost
            
            print(f"{i:<6} {ip:<18} {alloc_id:<30} {region:<15} {bandwidth:<10} {charge_type:<15} {reason}")
        
        # 统计信息
        print("\n" + "="*100)
        print("📊 统计汇总:")
        print(f"  • 未绑定EIP总数: {len(unbound_eips)} 个")
        print(f"  • 预估月成本浪费: ¥{total_cost:,.0f}")
        print(f"  • 预估年成本浪费: ¥{total_cost*12:,.0f}")
        
        # 按地域统计
        regions = {}
        for eip in unbound_eips:
            region = eip.get('区域', 'unknown')
            regions[region] = regions.get(region, 0) + 1
        
        print("\n📍 按地域分布:")
        for region, count in sorted(regions.items(), key=lambda x: x[1], reverse=True):
            print(f"  • {region}: {count} 个")
        
        # 优化建议
        print("\n" + "="*100)
        print("💡 优化建议:")
        print("="*100)
        print(f"""
1. 立即释放未绑定EIP
   - 数量: {len(unbound_eips)} 个
   - 节省成本: ¥{total_cost}/月 (约¥{total_cost*12}/年)
   - 操作风险: 低 (未绑定资源,不影响业务)

2. 释放步骤:
   a. 确认EIP确实无用途
   b. 在阿里云控制台 → VPC → 弹性公网IP
   c. 选择要释放的EIP,点击"释放"
   d. 或使用CLI批量释放:
      aliyun vpc ReleaseEipAddress --AllocationId <EIP-ID>

3. 预防措施:
   - 建立EIP申请审批流程
   - 定期检查(建议每周)
   - 设置云监控告警
""")
        
        # 批量释放命令
        print("\n📝 批量释放命令 (请谨慎使用):")
        print("-"*100)
        print("# 请先确认这些EIP确实无用,再执行以下命令\n")
        
        for i, eip in enumerate(unbound_eips[:10], 1):  # 只显示前10个
            alloc_id = eip.get('分配ID', '')
            ip = eip.get('IP地址', '')
            region = eip.get('区域', 'cn-beijing')
            print(f"# {i}. {ip}")
            print(f"aliyun vpc ReleaseEipAddress --RegionId {region} --AllocationId {alloc_id}")
            print()
        
        if len(unbound_eips) > 10:
            print(f"# ... 还有 {len(unbound_eips)-10} 个EIP")
        
    else:
        print("✅ 未发现未绑定的EIP")
        print("="*100)
        print(f"\n所有EIP ({len(all_eips)} 个) 都已绑定到实例。")
        print("\n但请注意:")
        print(f"  • 部分已绑定EIP可能流量极低")
        print(f"  • 建议查看完整的闲置EIP报告")
        print(f"\n查看完整报告:")
        print(f"  python3 list_tenant_eips.py")
    
    print("\n" + "="*100)
    print("📄 详细报告文件:")
    print(f"  Excel: {latest_report}")
    print(f"  HTML:  {latest_report.replace('.xlsx', '.html')}")
    print(f"\n打开报告:")
    print(f"  open {latest_report}")
    print("="*100)


if __name__ == "__main__":
    main()
