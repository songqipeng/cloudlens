"""
Report Generator
Generates HTML and PDF reports from resource data
"""
import logging
from typing import List, Dict
from datetime import datetime
from models.resource import UnifiedResource

logger = logging.getLogger("ReportGenerator")

class ReportGenerator:
    
    @staticmethod
    def generate_html(account_name: str, data: Dict) -> str:
        """
        生成HTML报告
        
        Args:
            account_name: 账号名称
            data: 包含各类资源的字典
        
        Returns:
            HTML string
        """
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{account_name} - 云资源报告</title>
    <style>
        body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            margin: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
            border-left: 4px solid #3498db;
            padding-left: 10px;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .summary-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .summary-card h3 {{
            margin: 0;
            font-size: 14px;
            opacity: 0.9;
        }}
        .summary-card .number {{
            font-size: 36px;
            font-weight: bold;
            margin: 10px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th {{
            background: #3498db;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .status-running {{
            color: #27ae60;
            font-weight: bold;
        }}
        .status-stopped {{
            color: #e74c3c;
            font-weight: bold;
        }}
        .footer {{
            margin-top: 40px;
            text-align: center;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>☁️ {account_name} 云资源报告</h1>
        <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <h2>📊 资源概览</h2>
        <div class="summary">
"""
        
        # Summary cards
        if 'ecs' in data:
            html += f"""
            <div class="summary-card">
                <h3>ECS 实例</h3>
                <div class="number">{len(data['ecs'])}</div>
            </div>
"""
        
        if 'rds' in data:
            html += f"""
            <div class="summary-card">
                <h3>RDS 实例</h3>
                <div class="number">{len(data['rds'])}</div>
            </div>
"""
        
        if 'redis' in data:
            html += f"""
            <div class="summary-card">
                <h3>Redis 实例</h3>
                <div class="number">{len(data['redis'])}</div>
            </div>
"""
        
        if 'eip' in data:
            html += f"""
            <div class="summary-card">
                <h3>弹性公网IP</h3>
                <div class="number">{len(data['eip'])}</div>
            </div>
"""
        
        html += """
        </div>
"""
        
        # ECS Table
        if 'ecs' in data and data['ecs']:
            html += """
        <h2>💻 ECS 实例</h2>
        <table>
            <tr>
                <th>实例ID</th>
                <th>名称</th>
                <th>状态</th>
                <th>公网IP</th>
                <th>规格</th>
                <th>区域</th>
            </tr>
"""
            for inst in data['ecs']:
                status_class = "status-running" if inst.status.value == "Running" else "status-stopped"
                pub_ip = inst.public_ips[0] if inst.public_ips else "-"
                html += f"""
            <tr>
                <td>{inst.id}</td>
                <td>{inst.name}</td>
                <td class="{status_class}">{inst.status.value}</td>
                <td>{pub_ip}</td>
                <td>{inst.spec}</td>
                <td>{inst.region}</td>
            </tr>
"""
            html += """
        </table>
"""
        
        # Idle resources
        if 'idle' in data and data['idle']:
            html += """
        <h2>⚠️ 闲置资源</h2>
        <table>
            <tr>
                <th>实例ID</th>
                <th>名称</th>
                <th>状态</th>
                <th>闲置原因</th>
            </tr>
"""
            for idle_info in data['idle']:
                inst, reasons = idle_info
                html += f"""
            <tr>
                <td>{inst.id}</td>
                <td>{inst.name}</td>
                <td>{inst.status.value}</td>
                <td>{'; '.join(reasons)}</td>
            </tr>
"""
            html += """
        </table>
"""
        
        html += f"""
        <div class="footer">
            <p>CloudLens CLI - Multi-Cloud Resource Analyzer</p>
            <p>Report generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""
        return html
    
    @staticmethod
    def save_html(html_content: str, filename: str):
        """保存HTML到文件"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logger.info(f"HTML report saved to {filename}")
    
    @staticmethod
    def generate_excel(account_name: str, data: Dict, filename: str):
        """
        生成Excel报告（多Sheet）
        
        Args:
            account_name: 账号名称
            data: 包含各类资源的字典
            filename: 输出文件名
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return
        
        wb = Workbook()
        
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("概览")
        ws_summary['A1'] = f"{account_name} 资源概览"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        row = 3
        summary_data = [
            ["资源类型", "数量"],
            ["ECS实例", len(data.get('ecs', []))],
            ["RDS实例", len(data.get('rds', []))],
            ["Redis实例", len(data.get('redis', []))],
            ["弹性公网IP", len(data.get('eip', []))],
        ]
        
        if 'idle' in data:
            summary_data.append(["闲置资源", len(data['idle'])])
        
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        # 样式
        for cell in ws_summary['A4:B4'][0]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # 2. ECS Sheet
        if data.get('ecs'):
            ws_ecs = wb.create_sheet("ECS实例")
            ws_ecs.append(["实例ID", "名称", "状态", "公网IP", "规格", "区域", "计费方式"])
            
            for inst in data['ecs']:
                pub_ip = inst.public_ips[0] if inst.public_ips else "-"
                ws_ecs.append([
                    inst.id,
                    inst.name,
                    inst.status.value,
                    pub_ip,
                    inst.spec,
                    inst.region,
                    inst.charge_type
                ])
            
            # 表头样式
            for cell in ws_ecs['A1:G1'][0]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        
        # 3. RDS Sheet
        if data.get('rds'):
            ws_rds = wb.create_sheet("RDS实例")
            ws_rds.append(["实例ID", "名称", "状态", "规格", "区域"])
            
            for inst in data['rds']:
                ws_rds.append([
                    inst.id,
                    inst.name,
                    inst.status.value,
                    inst.spec,
                    inst.region
                ])
            
            for cell in ws_rds['A1:E1'][0]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # 4. Idle Resources Sheet
        if data.get('idle'):
            ws_idle = wb.create_sheet("闲置资源")
            ws_idle.append(["实例ID", "名称", "状态", "闲置原因"])
            
            for idle_info in data['idle']:
                inst, reasons = idle_info
                ws_idle.append([
                    inst.id,
                    inst.name,
                    inst.status.value,
                    "; ".join(reasons)
                ])
            
            for cell in ws_idle['A1:D1'][0]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        
        # 自动调整列宽
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 保存
        wb.save(filename)
        logger.info(f"Excel report saved to {filename}")
    
    @staticmethod
    def html_to_pdf(html_file: str, pdf_file: str):
        """将HTML转换为PDF (需要wkhtmltopdf或weasyprint)"""
        try:
            from weasyprint import HTML
            HTML(html_file).write_pdf(pdf_file)
            logger.info(f"PDF report saved to {pdf_file}")
        except ImportError:
            logger.warning("WeasyPrint not installed. Run: pip install weasyprint")
            logger.info("Alternative: Install wkhtmltopdf and use: wkhtmltopdf input.html output.pdf")

