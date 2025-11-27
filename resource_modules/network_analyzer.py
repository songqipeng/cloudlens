#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
网络资源分析模块
获取VPC、VPC Peering、VPN、专线、SLB、CDN等网络资源的详细信息
"""

import json
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional

from aliyunsdkcore.acs_exception.exceptions import ClientException, ServerException
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.request import CommonRequest

from utils.concurrent_helper import process_concurrently
from utils.error_handler import ErrorHandler
from utils.logger import get_logger


class NetworkAnalyzer:
    """网络资源分析器"""

    def __init__(self, access_key_id: str, access_key_secret: str, tenant_name: str = "default"):
        """初始化网络分析器"""
        self.access_key_id = access_key_id
        self.access_key_secret = access_key_secret
        self.tenant_name = tenant_name
        self.logger = get_logger("network_analyzer")

    def get_all_regions(self) -> List[str]:
        """获取所有可用区域"""
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, "cn-hangzhou")
            request = CommonRequest()
            request.set_domain("ecs.cn-hangzhou.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2014-05-26")
            request.set_action_name("DescribeRegions")

            response = client.do_action_with_exception(request)
            data = json.loads(response)

            regions = []
            for region in data["Regions"]["Region"]:
                regions.append(region["RegionId"])

            return regions
        except Exception as e:
            self.logger.error(f"获取区域列表失败: {e}")
            return []

    def get_vpcs(self, region: str) -> List[Dict]:
        """获取VPC列表"""
        vpcs = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            request.set_domain(f"vpc.{region}.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2016-04-28")
            request.set_action_name("DescribeVpcs")
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "Vpcs" in data and "Vpc" in data["Vpcs"]:
                    vpc_list = data["Vpcs"]["Vpc"]
                    if not isinstance(vpc_list, list):
                        vpc_list = [vpc_list]

                    if len(vpc_list) == 0:
                        break

                    for vpc in vpc_list:
                        vpc_id = vpc.get("VpcId", "")
                        # 获取VPC的详细信息
                        vpc_detail = self.get_vpc_detail(region, vpc_id)

                        vpcs.append(
                            {
                                "VpcId": vpc_id,
                                "VpcName": vpc.get("VpcName", ""),
                                "CidrBlock": vpc.get("CidrBlock", ""),
                                "Status": vpc.get("Status", ""),
                                "RegionId": region,
                                "CreationTime": vpc.get("CreationTime", ""),
                                "Description": vpc.get("Description", ""),
                                "IsDefault": vpc.get("IsDefault", False),
                                "VSwitches": vpc_detail.get("VSwitches", []),
                                "RouteTables": vpc_detail.get("RouteTables", []),
                                "NetworkAcls": vpc_detail.get("NetworkAcls", []),
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(vpcs) >= total_count or len(vpc_list) < 50:
                        break

                    page_number += 1
                else:
                    break

            return vpcs
        except Exception as e:
            self.logger.warning(f"获取VPC列表失败 {region}: {e}")
            return []

    def get_vpc_detail(self, region: str, vpc_id: str) -> Dict:
        """获取VPC详细信息（子网、路由表、网络ACL）"""
        detail = {"VSwitches": [], "RouteTables": [], "NetworkAcls": []}

        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)

            # 1. 获取子网列表
            try:
                request = CommonRequest()
                request.set_domain(f"vpc.{region}.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2016-04-28")
                request.set_action_name("DescribeVSwitches")
                request.add_query_param("VpcId", vpc_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "VSwitches" in data and "VSwitch" in data["VSwitches"]:
                        vswitch_list = data["VSwitches"]["VSwitch"]
                        if not isinstance(vswitch_list, list):
                            vswitch_list = [vswitch_list]

                        if len(vswitch_list) == 0:
                            break

                        for vswitch in vswitch_list:
                            detail["VSwitches"].append(
                                {
                                    "VSwitchId": vswitch.get("VSwitchId", ""),
                                    "VSwitchName": vswitch.get("VSwitchName", ""),
                                    "CidrBlock": vswitch.get("CidrBlock", ""),
                                    "ZoneId": vswitch.get("ZoneId", ""),
                                    "Status": vswitch.get("Status", ""),
                                    "AvailableIpAddressCount": vswitch.get(
                                        "AvailableIpAddressCount", 0
                                    ),
                                    "Description": vswitch.get("Description", ""),
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(detail["VSwitches"]) >= total_count or len(vswitch_list) < 50:
                            break

                        page_number += 1
                    else:
                        break
            except Exception as e:
                self.logger.info(f"获取子网列表失败 {vpc_id}: {e}")

            # 2. 获取路由表列表
            try:
                request = CommonRequest()
                request.set_domain(f"vpc.{region}.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2016-04-28")
                request.set_action_name("DescribeRouteTables")
                request.add_query_param("VpcId", vpc_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "RouteTables" in data and "RouteTable" in data["RouteTables"]:
                        route_table_list = data["RouteTables"]["RouteTable"]
                        if not isinstance(route_table_list, list):
                            route_table_list = [route_table_list]

                        if len(route_table_list) == 0:
                            break

                        for rt in route_table_list:
                            # 获取路由表的路由条目
                            routes = self.get_route_table_routes(region, rt.get("RouteTableId", ""))

                            detail["RouteTables"].append(
                                {
                                    "RouteTableId": rt.get("RouteTableId", ""),
                                    "RouteTableName": rt.get("RouteTableName", ""),
                                    "RouteTableType": rt.get("RouteTableType", ""),
                                    "VSwitchIds": rt.get("VSwitchIds", {}).get("VSwitchId", []),
                                    "Routes": routes,
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(detail["RouteTables"]) >= total_count or len(route_table_list) < 50:
                            break

                        page_number += 1
                    else:
                        break
            except Exception as e:
                self.logger.info(f"获取路由表列表失败 {vpc_id}: {e}")

            # 3. 获取网络ACL列表
            try:
                request = CommonRequest()
                request.set_domain(f"vpc.{region}.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2016-04-28")
                request.set_action_name("DescribeNetworkAcls")
                request.add_query_param("VpcId", vpc_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "NetworkAcls" in data and "NetworkAcl" in data["NetworkAcls"]:
                        acl_list = data["NetworkAcls"]["NetworkAcl"]
                        if not isinstance(acl_list, list):
                            acl_list = [acl_list]

                        if len(acl_list) == 0:
                            break

                        for acl in acl_list:
                            detail["NetworkAcls"].append(
                                {
                                    "NetworkAclId": acl.get("NetworkAclId", ""),
                                    "NetworkAclName": acl.get("NetworkAclName", ""),
                                    "Status": acl.get("Status", ""),
                                    "Description": acl.get("Description", ""),
                                    "VpcId": acl.get("VpcId", ""),
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(detail["NetworkAcls"]) >= total_count or len(acl_list) < 50:
                            break

                        page_number += 1
                    else:
                        break
            except Exception as e:
                self.logger.info(f"获取网络ACL列表失败 {vpc_id}: {e}")

        except Exception as e:
            self.logger.warning(f"获取VPC详情失败 {vpc_id}: {e}")

        return detail

    def get_route_table_routes(self, region: str, route_table_id: str) -> List[Dict]:
        """获取路由表的路由条目"""
        routes = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            request.set_domain(f"vpc.{region}.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2016-04-28")
            request.set_action_name("DescribeRouteTableList")
            request.add_query_param("RouteTableId", route_table_id)

            response = client.do_action_with_exception(request)
            data = json.loads(response)

            if "RouterTableList" in data and "RouterTableListType" in data["RouterTableList"]:
                rt_list = data["RouterTableList"]["RouterTableListType"]
                if not isinstance(rt_list, list):
                    rt_list = [rt_list]

                for rt in rt_list:
                    if "RouteEntrys" in rt and "RouteEntry" in rt["RouteEntrys"]:
                        route_list = rt["RouteEntrys"]["RouteEntry"]
                        if not isinstance(route_list, list):
                            route_list = [route_list]

                        for route in route_list:
                            routes.append(
                                {
                                    "DestinationCidrBlock": route.get("DestinationCidrBlock", ""),
                                    "NextHopType": route.get("NextHopType", ""),
                                    "NextHopId": route.get("NextHopId", ""),
                                    "Status": route.get("Status", ""),
                                    "Type": route.get("Type", ""),
                                }
                            )
        except Exception as e:
            self.logger.info(f"获取路由条目失败 {route_table_id}: {e}")

        return routes

    def get_vpc_peerings(self, region: str) -> List[Dict]:
        """获取VPC Peering/对等连接列表"""
        peerings = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            request.set_domain(f"vpc.{region}.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2016-04-28")
            # 尝试使用正确的API名称
            try:
                request.set_action_name("DescribeVpcPeerConnections")
            except:
                # 如果API不存在，尝试其他版本
                request.set_version("2020-04-15")
                request.set_action_name("ListVpcPeerConnections")
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if (
                    "VpcPeerConnections" in data
                    and "VpcPeerConnection" in data["VpcPeerConnections"]
                ):
                    peering_list = data["VpcPeerConnections"]["VpcPeerConnection"]
                    if not isinstance(peering_list, list):
                        peering_list = [peering_list]

                    if len(peering_list) == 0:
                        break

                    for peering in peering_list:
                        peerings.append(
                            {
                                "InstanceId": peering.get("InstanceId", ""),
                                "Name": peering.get("Name", ""),
                                "VpcId": peering.get("VpcId", ""),
                                "AcceptingVpcId": peering.get("AcceptingVpcId", ""),
                                "AcceptingRegionId": peering.get("AcceptingRegionId", ""),
                                "Status": peering.get("Status", ""),
                                "Bandwidth": peering.get("Bandwidth", 0),
                                "Description": peering.get("Description", ""),
                                "CreationTime": peering.get("CreationTime", ""),
                                "RegionId": region,
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(peerings) >= total_count or len(peering_list) < 50:
                        break

                    page_number += 1
                else:
                    break

            return peerings
        except Exception as e:
            self.logger.warning(f"获取VPC Peering列表失败 {region}: {e}")
            return []

    def get_vpn_connections(self, region: str) -> List[Dict]:
        """获取VPN连接列表"""
        vpns = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            request.set_domain(f"vpc.{region}.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2016-04-28")
            request.set_action_name("DescribeVpnConnections")
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "VpnConnections" in data and "VpnConnection" in data["VpnConnections"]:
                    vpn_list = data["VpnConnections"]["VpnConnection"]
                    if not isinstance(vpn_list, list):
                        vpn_list = [vpn_list]

                    if len(vpn_list) == 0:
                        break

                    for vpn in vpn_list:
                        vpns.append(
                            {
                                "VpnConnectionId": vpn.get("VpnConnectionId", ""),
                                "VpnConnectionName": vpn.get("VpnConnectionName", ""),
                                "VpnGatewayId": vpn.get("VpnGatewayId", ""),
                                "CustomerGatewayId": vpn.get("CustomerGatewayId", ""),
                                "Status": vpn.get("Status", ""),
                                "LocalSubnet": vpn.get("LocalSubnet", ""),
                                "RemoteSubnet": vpn.get("RemoteSubnet", ""),
                                "CreateTime": vpn.get("CreateTime", ""),
                                "RegionId": region,
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(vpns) >= total_count or len(vpn_list) < 50:
                        break

                    page_number += 1
                else:
                    break

            return vpns
        except Exception as e:
            self.logger.warning(f"获取VPN连接列表失败 {region}: {e}")
            return []

    def get_express_connect(self, region: str) -> List[Dict]:
        """获取专线配置列表"""
        connections = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            request.set_domain(f"vpc.{region}.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2016-04-28")
            request.set_action_name("DescribePhysicalConnections")
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if (
                    "PhysicalConnectionSet" in data
                    and "PhysicalConnectionType" in data["PhysicalConnectionSet"]
                ):
                    conn_list = data["PhysicalConnectionSet"]["PhysicalConnectionType"]
                    if not isinstance(conn_list, list):
                        conn_list = [conn_list]

                    if len(conn_list) == 0:
                        break

                    for conn in conn_list:
                        connections.append(
                            {
                                "PhysicalConnectionId": conn.get("PhysicalConnectionId", ""),
                                "Name": conn.get("Name", ""),
                                "AccessPointId": conn.get("AccessPointId", ""),
                                "Type": conn.get("Type", ""),
                                "Status": conn.get("Status", ""),
                                "Bandwidth": conn.get("Bandwidth", 0),
                                "BusinessStatus": conn.get("BusinessStatus", ""),
                                "Description": conn.get("Description", ""),
                                "AdLocation": conn.get("AdLocation", ""),
                                "PortType": conn.get("PortType", ""),
                                "CircuitCode": conn.get("CircuitCode", ""),
                                "RedundantPhysicalConnectionId": conn.get(
                                    "RedundantPhysicalConnectionId", ""
                                ),
                                "PeerLocation": conn.get("PeerLocation", ""),
                                "CreationTime": conn.get("CreationTime", ""),
                                "RegionId": region,
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(connections) >= total_count or len(conn_list) < 50:
                        break

                    page_number += 1
                else:
                    break

            return connections
        except Exception as e:
            self.logger.warning(f"获取专线配置列表失败 {region}: {e}")
            return []

    def get_slb_detailed_info(self, region: str) -> List[Dict]:
        """获取SLB详细配置（监听规则、后端服务器组、健康检查）"""
        from resource_modules.slb_analyzer import SLBAnalyzer

        slb_analyzer = SLBAnalyzer(self.access_key_id, self.access_key_secret)
        slb_instances = slb_analyzer.get_slb_instances(region)

        detailed_slbs = []
        for instance in slb_instances:
            instance_id = instance.get("InstanceId", "")
            lb_type = instance.get("LoadBalancerType", "clb")

            # 获取监听器详情
            listeners = self.get_slb_listeners(region, instance_id, lb_type)

            # 获取后端服务器组详情
            server_groups = self.get_slb_server_groups(region, instance_id, lb_type)

            detailed_slbs.append(
                {**instance, "Listeners": listeners, "ServerGroups": server_groups}
            )

        return detailed_slbs

    def get_slb_listeners(self, region: str, instance_id: str, lb_type: str = "clb") -> List[Dict]:
        """获取SLB监听器详情"""
        listeners = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)

            if lb_type == "clb":
                # CLB使用传统SLB API
                request = CommonRequest()
                request.set_domain(f"slb.{region}.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2014-05-15")
                request.set_action_name("DescribeLoadBalancerListeners")
                request.add_query_param("LoadBalancerId", instance_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "Listeners" in data and "Listener" in data["Listeners"]:
                        listener_list = data["Listeners"]["Listener"]
                        if not isinstance(listener_list, list):
                            listener_list = [listener_list]

                        if len(listener_list) == 0:
                            break

                        for listener in listener_list:
                            # 获取健康检查配置
                            health_check = self.get_slb_health_check(
                                region, instance_id, listener.get("ListenerPort", 0), lb_type
                            )

                            listeners.append(
                                {
                                    "ListenerPort": listener.get("ListenerPort", 0),
                                    "ListenerProtocol": listener.get("ListenerProtocol", ""),
                                    "BackendServerPort": listener.get("BackendServerPort", 0),
                                    "Bandwidth": listener.get("Bandwidth", 0),
                                    "Status": listener.get("Status", ""),
                                    "Description": listener.get("Description", ""),
                                    "HealthCheck": health_check,
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(listeners) >= total_count or len(listener_list) < 50:
                            break

                        page_number += 1
                    else:
                        break

            elif lb_type in ["alb", "nlb"]:
                # ALB和NLB使用新API
                request = CommonRequest()
                domain = f"{lb_type}.{region}.aliyuncs.com"
                version = "2020-06-16" if lb_type == "alb" else "2022-04-30"
                request.set_domain(domain)
                request.set_method("POST")
                request.set_version(version)
                request.set_action_name("ListListeners")
                request.add_query_param("LoadBalancerId", instance_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "Listeners" in data:
                        listener_list = data["Listeners"]
                        if not isinstance(listener_list, list):
                            listener_list = [listener_list] if listener_list else []

                        if len(listener_list) == 0:
                            break

                        for listener in listener_list:
                            listener_id = listener.get("ListenerId", "")
                            health_check = self.get_slb_health_check(
                                region, instance_id, listener_id, lb_type
                            )

                            listeners.append(
                                {
                                    "ListenerId": listener_id,
                                    "ListenerPort": listener.get("ListenerPort", 0),
                                    "ListenerProtocol": listener.get("ListenerProtocol", ""),
                                    "Status": listener.get("Status", ""),
                                    "Description": listener.get("Description", ""),
                                    "HealthCheck": health_check,
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(listeners) >= total_count or len(listener_list) < 50:
                            break

                        page_number += 1
                    else:
                        break

        except Exception as e:
            self.logger.info(f"获取SLB监听器失败 {instance_id}: {e}")

        return listeners

    def get_slb_health_check(
        self, region: str, instance_id: str, listener_port_or_id, lb_type: str = "clb"
    ) -> Dict:
        """获取SLB健康检查配置"""
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)

            if lb_type == "clb":
                request = CommonRequest()
                request.set_domain(f"slb.{region}.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2014-05-15")
                request.set_action_name("DescribeHealthStatus")
                request.add_query_param("LoadBalancerId", instance_id)
                request.add_query_param("ListenerPort", listener_port_or_id)

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "BackendServers" in data and "BackendServer" in data["BackendServers"]:
                    backend_list = data["BackendServers"]["BackendServer"]
                    if not isinstance(backend_list, list):
                        backend_list = [backend_list]

                    # 获取第一个后端的健康检查配置（通常所有后端使用相同配置）
                    if backend_list:
                        return {
                            "HealthCheck": backend_list[0].get("HealthCheck", ""),
                            "HealthCheckDomain": backend_list[0].get("HealthCheckDomain", ""),
                            "HealthCheckURI": backend_list[0].get("HealthCheckURI", ""),
                            "HealthCheckConnectPort": backend_list[0].get(
                                "HealthCheckConnectPort", 0
                            ),
                            "HealthCheckInterval": backend_list[0].get("HealthCheckInterval", 0),
                            "HealthCheckTimeout": backend_list[0].get("HealthCheckTimeout", 0),
                            "HealthyThreshold": backend_list[0].get("HealthyThreshold", 0),
                            "UnhealthyThreshold": backend_list[0].get("UnhealthyThreshold", 0),
                        }

            elif lb_type in ["alb", "nlb"]:
                # 对于ALB和NLB，需要从监听器详情获取
                request = CommonRequest()
                domain = f"{lb_type}.{region}.aliyuncs.com"
                version = "2020-06-16" if lb_type == "alb" else "2022-04-30"
                request.set_domain(domain)
                request.set_method("POST")
                request.set_version(version)
                request.set_action_name("GetListenerAttribute")
                request.add_query_param("ListenerId", listener_port_or_id)

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                health_check_config = data.get("HealthCheckConfig", {})
                return {
                    "HealthCheckEnabled": health_check_config.get("HealthCheckEnabled", False),
                    "HealthCheckProtocol": health_check_config.get("HealthCheckProtocol", ""),
                    "HealthCheckPath": health_check_config.get("HealthCheckPath", ""),
                    "HealthCheckPort": health_check_config.get("HealthCheckPort", 0),
                    "HealthCheckInterval": health_check_config.get("HealthCheckInterval", 0),
                    "HealthCheckTimeout": health_check_config.get("HealthCheckTimeout", 0),
                    "HealthyThreshold": health_check_config.get("HealthyThreshold", 0),
                    "UnhealthyThreshold": health_check_config.get("UnhealthyThreshold", 0),
                }

        except Exception as e:
            self.logger.info(f"获取健康检查配置失败 {instance_id}: {e}")

        return {}

    def get_slb_server_groups(
        self, region: str, instance_id: str, lb_type: str = "clb"
    ) -> List[Dict]:
        """获取SLB后端服务器组详情"""
        server_groups = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)

            if lb_type == "clb":
                # CLB的后端服务器在实例详情中
                from aliyunsdkslb.request.v20140515 import DescribeLoadBalancerAttributeRequest

                request = (
                    DescribeLoadBalancerAttributeRequest.DescribeLoadBalancerAttributeRequest()
                )
                request.set_LoadBalancerId(instance_id)

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "BackendServers" in data and "BackendServer" in data["BackendServers"]:
                    backend_list = data["BackendServers"]["BackendServer"]
                    if not isinstance(backend_list, list):
                        backend_list = [backend_list] if backend_list else []

                    for backend in backend_list:
                        server_groups.append(
                            {
                                "ServerId": backend.get("ServerId", ""),
                                "ServerIp": backend.get("ServerIp", ""),
                                "Port": backend.get("Port", 0),
                                "Weight": backend.get("Weight", 100),
                                "Type": backend.get("Type", ""),
                                "Description": backend.get("Description", ""),
                            }
                        )

            elif lb_type in ["alb", "nlb"]:
                # ALB和NLB使用服务器组API
                request = CommonRequest()
                domain = f"{lb_type}.{region}.aliyuncs.com"
                version = "2020-06-16" if lb_type == "alb" else "2022-04-30"
                request.set_domain(domain)
                request.set_method("POST")
                request.set_version(version)
                request.set_action_name("ListServerGroups")
                request.add_query_param("LoadBalancerId", instance_id)
                request.add_query_param("PageSize", 50)

                page_number = 1
                while True:
                    request.add_query_param("PageNumber", page_number)
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)

                    if "ServerGroups" in data:
                        group_list = data["ServerGroups"]
                        if not isinstance(group_list, list):
                            group_list = [group_list] if group_list else []

                        if len(group_list) == 0:
                            break

                        for group in group_list:
                            group_id = group.get("ServerGroupId", "")
                            # 获取服务器组中的服务器
                            servers = self.get_server_group_servers(region, group_id, lb_type)

                            server_groups.append(
                                {
                                    "ServerGroupId": group_id,
                                    "ServerGroupName": group.get("ServerGroupName", ""),
                                    "ServerGroupType": group.get("ServerGroupType", ""),
                                    "Servers": servers,
                                }
                            )

                        total_count = data.get("TotalCount", 0)
                        if len(server_groups) >= total_count or len(group_list) < 50:
                            break

                        page_number += 1
                    else:
                        break

        except Exception as e:
            self.logger.info(f"获取后端服务器组失败 {instance_id}: {e}")

        return server_groups

    def get_server_group_servers(
        self, region: str, server_group_id: str, lb_type: str
    ) -> List[Dict]:
        """获取服务器组中的服务器列表"""
        servers = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, region)
            request = CommonRequest()
            domain = f"{lb_type}.{region}.aliyuncs.com"
            version = "2020-06-16" if lb_type == "alb" else "2022-04-30"
            request.set_domain(domain)
            request.set_method("POST")
            request.set_version(version)
            request.set_action_name("ListServerGroupServers")
            request.add_query_param("ServerGroupId", server_group_id)
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "Servers" in data:
                    server_list = data["Servers"]
                    if not isinstance(server_list, list):
                        server_list = [server_list] if server_list else []

                    if len(server_list) == 0:
                        break

                    for server in server_list:
                        servers.append(
                            {
                                "ServerId": server.get("ServerId", ""),
                                "ServerIp": server.get("ServerIp", ""),
                                "Port": server.get("Port", 0),
                                "Weight": server.get("Weight", 100),
                                "Status": server.get("Status", ""),
                                "Description": server.get("Description", ""),
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(servers) >= total_count or len(server_list) < 50:
                        break

                    page_number += 1
                else:
                    break

        except Exception as e:
            self.logger.info(f"获取服务器列表失败 {server_group_id}: {e}")

        return servers

    def get_cdn_domains(self) -> List[Dict]:
        """获取CDN域名配置"""
        domains = []
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, "cn-hangzhou")
            request = CommonRequest()
            request.set_domain("cdn.aliyuncs.com")
            request.set_method("POST")
            request.set_version("2018-05-10")
            request.set_action_name("DescribeUserDomains")
            request.add_query_param("PageSize", 50)

            page_number = 1
            while True:
                request.add_query_param("PageNumber", page_number)
                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "Domains" in data and "PageData" in data["Domains"]:
                    domain_list = data["Domains"]["PageData"]
                    if not isinstance(domain_list, list):
                        domain_list = [domain_list]

                    if len(domain_list) == 0:
                        break

                    for domain in domain_list:
                        domain_name = domain.get("DomainName", "")
                        # 获取域名详细配置
                        domain_detail = self.get_cdn_domain_detail(domain_name)

                        domains.append(
                            {
                                "DomainName": domain_name,
                                "CdnType": domain.get("CdnType", ""),
                                "DomainStatus": domain.get("DomainStatus", ""),
                                "GmtCreated": domain.get("GmtCreated", ""),
                                "GmtModified": domain.get("GmtModified", ""),
                                "Description": domain.get("Description", ""),
                                **domain_detail,
                            }
                        )

                    total_count = data.get("TotalCount", 0)
                    if len(domains) >= total_count or len(domain_list) < 50:
                        break

                    page_number += 1
                else:
                    break

            return domains
        except Exception as e:
            self.logger.warning(f"获取CDN域名列表失败: {e}")
            return []

    def get_cdn_domain_detail(self, domain_name: str) -> Dict:
        """获取CDN域名详细配置（源站、加速区域、缓存策略、回源策略）"""
        detail = {"Sources": [], "Coverage": "", "CacheRules": [], "BackendRules": []}

        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, "cn-hangzhou")

            # 1. 获取源站配置
            try:
                request = CommonRequest()
                request.set_domain("cdn.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2018-05-10")
                request.set_action_name("DescribeDomainDetail")
                request.add_query_param("DomainName", domain_name)

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "Sources" in data and "Source" in data["Sources"]:
                    source_list = data["Sources"]["Source"]
                    if not isinstance(source_list, list):
                        source_list = [source_list]

                    for source in source_list:
                        detail["Sources"].append(
                            {
                                "Content": source.get("Content", ""),
                                "Type": source.get("Type", ""),
                                "Port": source.get("Port", 80),
                                "Priority": source.get("Priority", 20),
                                "Weight": source.get("Weight", 10),
                            }
                        )

                detail["Coverage"] = data.get("Coverage", "")
            except Exception as e:
                self.logger.info(f"获取CDN源站配置失败 {domain_name}: {e}")

            # 2. 获取缓存策略
            try:
                request = CommonRequest()
                request.set_domain("cdn.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2018-05-10")
                request.set_action_name("DescribeDomainConfigs")
                request.add_query_param("DomainName", domain_name)
                request.add_query_param(
                    "FunctionNames",
                    "set_req_host_header,filetype_based_ttl_setting,path_based_ttl_setting",
                )

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "DomainConfigs" in data and "DomainConfig" in data["DomainConfigs"]:
                    config_list = data["DomainConfigs"]["DomainConfig"]
                    if not isinstance(config_list, list):
                        config_list = [config_list]

                    for config in config_list:
                        function_name = config.get("FunctionName", "")
                        if "ttl" in function_name.lower() or "cache" in function_name.lower():
                            detail["CacheRules"].append(
                                {
                                    "FunctionName": function_name,
                                    "FunctionArgs": config.get("FunctionArgs", {}),
                                }
                            )
            except Exception as e:
                self.logger.info(f"获取CDN缓存策略失败 {domain_name}: {e}")

            # 3. 获取回源策略
            try:
                request = CommonRequest()
                request.set_domain("cdn.aliyuncs.com")
                request.set_method("POST")
                request.set_version("2018-05-10")
                request.set_action_name("DescribeDomainHttpHeaderConfigs")
                request.add_query_param("DomainName", domain_name)

                response = client.do_action_with_exception(request)
                data = json.loads(response)

                if "HttpHeaderConfigs" in data and "HttpHeaderConfig" in data["HttpHeaderConfigs"]:
                    header_list = data["HttpHeaderConfigs"]["HttpHeaderConfig"]
                    if not isinstance(header_list, list):
                        header_list = [header_list]

                    for header in header_list:
                        detail["BackendRules"].append(
                            {
                                "HeaderKey": header.get("HeaderKey", ""),
                                "HeaderValue": header.get("HeaderValue", ""),
                                "ConfigId": header.get("ConfigId", ""),
                            }
                        )
            except Exception as e:
                self.logger.info(f"获取CDN回源策略失败 {domain_name}: {e}")

        except Exception as e:
            self.logger.warning(f"获取CDN域名详情失败 {domain_name}: {e}")

        return detail

    def analyze_network_resources(self):
        """分析所有网络资源"""
        self.logger.info("开始网络资源分析...")

        regions = self.get_all_regions()
        self.logger.info(f"找到 {len(regions)} 个区域")

        all_results = {
            "VPCs": [],
            "VpcPeerings": [],
            "VpnConnections": [],
            "ExpressConnects": [],
            "SLBs": [],
            "CDNs": [],
        }

        # 并发获取各区域的网络资源
        def get_region_network_resources(region):
            """获取单个区域的网络资源"""
            try:
                vpcs = self.get_vpcs(region)
                peerings = self.get_vpc_peerings(region)
                vpns = self.get_vpn_connections(region)
                express_conns = self.get_express_connect(region)
                slbs = self.get_slb_detailed_info(region)

                return {
                    "region": region,
                    "vpcs": vpcs,
                    "peerings": peerings,
                    "vpns": vpns,
                    "express_conns": express_conns,
                    "slbs": slbs,
                }
            except Exception as e:
                self.logger.warning(f"获取区域 {region} 网络资源失败: {e}")
                return {
                    "region": region,
                    "vpcs": [],
                    "peerings": [],
                    "vpns": [],
                    "express_conns": [],
                    "slbs": [],
                }

        self.logger.info("并发获取所有区域的网络资源...")
        region_results = process_concurrently(
            regions, get_region_network_resources, max_workers=10, description="获取网络资源"
        )

        # 整理结果
        for result in region_results:
            if result:
                all_results["VPCs"].extend(result.get("vpcs", []))
                all_results["VpcPeerings"].extend(result.get("peerings", []))
                all_results["VpnConnections"].extend(result.get("vpns", []))
                all_results["ExpressConnects"].extend(result.get("express_conns", []))
                all_results["SLBs"].extend(result.get("slbs", []))

        # 获取CDN配置（CDN是全局的，不需要按区域）
        self.logger.info("获取CDN配置...")
        all_results["CDNs"] = self.get_cdn_domains()

        # 生成报告
        excel_file = self.generate_network_report(all_results)
        html_file = self.generate_html_report(all_results)

        self.logger.info("网络资源分析完成")
        self.logger.info(f"📊 Excel报告: {excel_file}")
        self.logger.info(f"🌐 HTML报告: {html_file}")
        return all_results

    def generate_html_report(self, results: Dict):
        """生成HTML格式的网络资源报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_file = f"{self.tenant_name}_network_resources_{timestamp}.html"

        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>网络资源分析报告 - {self.tenant_name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: 'Microsoft YaHei', 'PingFang SC', Arial, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{ 
            max-width: 1400px; 
            margin: 0 auto; 
            background: white; 
            border-radius: 15px; 
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        .header p {{ font-size: 1.1em; opacity: 0.9; }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        .summary-card {{
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            text-align: center;
        }}
        .summary-card h3 {{
            color: #667eea;
            font-size: 2em;
            margin-bottom: 10px;
        }}
        .summary-card p {{
            color: #666;
            font-size: 0.9em;
        }}
        .section {{
            padding: 30px;
            border-bottom: 1px solid #eee;
        }}
        .section:last-child {{ border-bottom: none; }}
        .section-title {{
            font-size: 1.8em;
            color: #333;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
            background: white;
        }}
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        tr:hover {{ background-color: #f8f9fa; }}
        tr:nth-child(even) {{ background-color: #fafafa; }}
        .badge {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: 600;
        }}
        .badge-success {{ background: #d4edda; color: #155724; }}
        .badge-warning {{ background: #fff3cd; color: #856404; }}
        .badge-danger {{ background: #f8d7da; color: #721c24; }}
        .badge-info {{ background: #d1ecf1; color: #0c5460; }}
        .code-block {{
            background: #f4f4f4;
            padding: 10px;
            border-radius: 5px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            white-space: pre-wrap;
            word-break: break-all;
        }}
        .footer {{
            text-align: center;
            padding: 30px;
            color: #666;
            background: #f8f9fa;
        }}
        .empty-state {{
            text-align: center;
            padding: 40px;
            color: #999;
        }}
        .empty-state::before {{
            content: "📭";
            font-size: 3em;
            display: block;
            margin-bottom: 10px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🌐 网络资源分析报告</h1>
            <p>租户: {self.tenant_name} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="summary">
            <div class="summary-card">
                <h3>{len(results.get('VPCs', []))}</h3>
                <p>VPC数量</p>
            </div>
            <div class="summary-card">
                <h3>{len(results.get('VpcPeerings', []))}</h3>
                <p>VPC对等连接</p>
            </div>
            <div class="summary-card">
                <h3>{len(results.get('VpnConnections', []))}</h3>
                <p>VPN连接</p>
            </div>
            <div class="summary-card">
                <h3>{len(results.get('ExpressConnects', []))}</h3>
                <p>专线连接</p>
            </div>
            <div class="summary-card">
                <h3>{len(results.get('SLBs', []))}</h3>
                <p>SLB实例</p>
            </div>
            <div class="summary-card">
                <h3>{len(results.get('CDNs', []))}</h3>
                <p>CDN域名</p>
            </div>
        </div>
"""

        # 1. VPC信息
        html_content += """
        <div class="section">
            <h2 class="section-title">📦 VPC列表</h2>
"""
        if results.get("VPCs"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>VPC ID</th>
                        <th>VPC名称</th>
                        <th>CIDR</th>
                        <th>区域</th>
                        <th>状态</th>
                        <th>子网数量</th>
                        <th>路由表数量</th>
                        <th>网络ACL数量</th>
                        <th>创建时间</th>
                    </tr>
                </thead>
                <tbody>
"""
            for vpc in results["VPCs"]:
                status_class = (
                    "badge-success" if vpc.get("Status") == "Available" else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{vpc.get('VpcId', '')}</code></td>
                        <td>{vpc.get('VpcName', '未命名')}</td>
                        <td><span class="badge badge-info">{vpc.get('CidrBlock', '')}</span></td>
                        <td>{vpc.get('RegionId', '')}</td>
                        <td><span class="badge {status_class}">{vpc.get('Status', '')}</span></td>
                        <td>{len(vpc.get('VSwitches', []))}</td>
                        <td>{len(vpc.get('RouteTables', []))}</td>
                        <td>{len(vpc.get('NetworkAcls', []))}</td>
                        <td>{vpc.get('CreationTime', '')}</td>
                    </tr>
"""
                # 添加子网详情
                if vpc.get("VSwitches"):
                    html_content += f"""
                    <tr>
                        <td colspan="9" style="background: #f8f9fa; padding-left: 30px;">
                            <strong>子网列表:</strong><br>
"""
                    for vs in vpc.get("VSwitches", []):
                        html_content += f"""
                            • {vs.get('VSwitchName', '未命名')} ({vs.get('CidrBlock', '')}) - 
                            可用IP: {vs.get('AvailableIpAddressCount', 0)} | 
                            可用区: {vs.get('ZoneId', '')}<br>
"""
                    html_content += """
                        </td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无VPC资源</div>'
        html_content += "</div>"

        # 2. VPC Peering
        html_content += """
        <div class="section">
            <h2 class="section-title">🔗 VPC对等连接</h2>
"""
        if results.get("VpcPeerings"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>对等连接ID</th>
                        <th>名称</th>
                        <th>本地VPC</th>
                        <th>对端VPC</th>
                        <th>对端区域</th>
                        <th>带宽(Mbps)</th>
                        <th>状态</th>
                        <th>创建时间</th>
                    </tr>
                </thead>
                <tbody>
"""
            for peering in results["VpcPeerings"]:
                status_class = (
                    "badge-success" if peering.get("Status") == "Active" else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{peering.get('InstanceId', '')}</code></td>
                        <td>{peering.get('Name', '未命名')}</td>
                        <td><code>{peering.get('VpcId', '')}</code></td>
                        <td><code>{peering.get('AcceptingVpcId', '')}</code></td>
                        <td>{peering.get('AcceptingRegionId', '')}</td>
                        <td>{peering.get('Bandwidth', 0)}</td>
                        <td><span class="badge {status_class}">{peering.get('Status', '')}</span></td>
                        <td>{peering.get('CreationTime', '')}</td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无VPC对等连接</div>'
        html_content += "</div>"

        # 3. VPN连接
        html_content += """
        <div class="section">
            <h2 class="section-title">🔐 VPN连接</h2>
"""
        if results.get("VpnConnections"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>VPN连接ID</th>
                        <th>VPN连接名称</th>
                        <th>VPN网关ID</th>
                        <th>用户网关ID</th>
                        <th>状态</th>
                        <th>本地网段</th>
                        <th>对端网段</th>
                        <th>区域</th>
                    </tr>
                </thead>
                <tbody>
"""
            for vpn in results["VpnConnections"]:
                status_class = (
                    "badge-success"
                    if vpn.get("Status") == "ike_sa_not_established"
                    else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{vpn.get('VpnConnectionId', '')}</code></td>
                        <td>{vpn.get('VpnConnectionName', '未命名')}</td>
                        <td><code>{vpn.get('VpnGatewayId', '')}</code></td>
                        <td><code>{vpn.get('CustomerGatewayId', '')}</code></td>
                        <td><span class="badge {status_class}">{vpn.get('Status', '')}</span></td>
                        <td><span class="code-block">{vpn.get('LocalSubnet', '')}</span></td>
                        <td><span class="code-block">{vpn.get('RemoteSubnet', '')}</span></td>
                        <td>{vpn.get('RegionId', '')}</td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无VPN连接</div>'
        html_content += "</div>"

        # 4. 专线配置
        html_content += """
        <div class="section">
            <h2 class="section-title">🌐 专线配置</h2>
"""
        if results.get("ExpressConnects"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>专线ID</th>
                        <th>名称</th>
                        <th>接入点ID</th>
                        <th>类型</th>
                        <th>状态</th>
                        <th>业务状态</th>
                        <th>带宽(Mbps)</th>
                        <th>端口类型</th>
                        <th>对端位置</th>
                        <th>区域</th>
                    </tr>
                </thead>
                <tbody>
"""
            for conn in results["ExpressConnects"]:
                status_class = (
                    "badge-success" if conn.get("Status") == "Enabled" else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{conn.get('PhysicalConnectionId', '')}</code></td>
                        <td>{conn.get('Name', '未命名')}</td>
                        <td>{conn.get('AccessPointId', '')}</td>
                        <td>{conn.get('Type', '')}</td>
                        <td><span class="badge {status_class}">{conn.get('Status', '')}</span></td>
                        <td><span class="badge badge-info">{conn.get('BusinessStatus', '')}</span></td>
                        <td>{conn.get('Bandwidth', 0)}</td>
                        <td>{conn.get('PortType', '')}</td>
                        <td>{conn.get('PeerLocation', '')}</td>
                        <td>{conn.get('RegionId', '')}</td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无专线配置</div>'
        html_content += "</div>"

        # 5. SLB详细配置
        html_content += """
        <div class="section">
            <h2 class="section-title">⚖️ SLB实例</h2>
"""
        if results.get("SLBs"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>实例ID</th>
                        <th>实例名称</th>
                        <th>负载均衡类型</th>
                        <th>区域</th>
                        <th>地址类型</th>
                        <th>规格</th>
                        <th>带宽(Mbps)</th>
                        <th>监听器数</th>
                        <th>后端服务器组数</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
"""
            for slb in results["SLBs"]:
                status_class = (
                    "badge-success" if slb.get("InstanceStatus") == "active" else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{slb.get('InstanceId', '')}</code></td>
                        <td>{slb.get('InstanceName', '未命名')}</td>
                        <td><span class="badge badge-info">{slb.get('LoadBalancerTypeName', 'CLB')}</span></td>
                        <td>{slb.get('Region', '')}</td>
                        <td>{slb.get('AddressType', '')}</td>
                        <td>{slb.get('InstanceType', '')}</td>
                        <td>{slb.get('Bandwidth', 0)}</td>
                        <td>{len(slb.get('Listeners', []))}</td>
                        <td>{len(slb.get('ServerGroups', []))}</td>
                        <td><span class="badge {status_class}">{slb.get('InstanceStatus', '')}</span></td>
                    </tr>
"""
                # 添加监听器详情
                if slb.get("Listeners"):
                    html_content += f"""
                    <tr>
                        <td colspan="10" style="background: #f8f9fa; padding-left: 30px;">
                            <strong>监听器列表:</strong><br>
"""
                    for listener in slb.get("Listeners", []):
                        html_content += f"""
                            • 端口: {listener.get('ListenerPort', 0)} | 
                            协议: {listener.get('ListenerProtocol', '')} | 
                            状态: {listener.get('Status', '')}<br>
"""
                    html_content += """
                        </td>
                    </tr>
"""
                # 添加后端服务器组详情
                if slb.get("ServerGroups"):
                    html_content += f"""
                    <tr>
                        <td colspan="10" style="background: #f8f9fa; padding-left: 30px;">
                            <strong>后端服务器组:</strong><br>
"""
                    for sg in slb.get("ServerGroups", []):
                        server_count = len(sg.get("Servers", []))
                        html_content += f"""
                            • {sg.get('ServerGroupName', '未命名')} ({server_count}个服务器)<br>
"""
                    html_content += """
                        </td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无SLB实例</div>'
        html_content += "</div>"

        # 6. CDN配置
        html_content += """
        <div class="section">
            <h2 class="section-title">🚀 CDN配置</h2>
"""
        if results.get("CDNs"):
            html_content += """
            <table>
                <thead>
                    <tr>
                        <th>域名</th>
                        <th>CDN类型</th>
                        <th>状态</th>
                        <th>加速区域</th>
                        <th>源站数量</th>
                        <th>创建时间</th>
                    </tr>
                </thead>
                <tbody>
"""
            for cdn in results["CDNs"]:
                status_class = (
                    "badge-success" if cdn.get("DomainStatus") == "online" else "badge-warning"
                )
                html_content += f"""
                    <tr>
                        <td><code>{cdn.get('DomainName', '')}</code></td>
                        <td>{cdn.get('CdnType', '')}</td>
                        <td><span class="badge {status_class}">{cdn.get('DomainStatus', '')}</span></td>
                        <td>{cdn.get('Coverage', '')}</td>
                        <td>{len(cdn.get('Sources', []))}</td>
                        <td>{cdn.get('GmtCreated', '')}</td>
                    </tr>
"""
                # 添加源站详情
                if cdn.get("Sources"):
                    html_content += f"""
                    <tr>
                        <td colspan="6" style="background: #f8f9fa; padding-left: 30px;">
                            <strong>源站列表:</strong><br>
"""
                    for source in cdn.get("Sources", []):
                        html_content += f"""
                            • {source.get('Content', '')}:{source.get('Port', 80)} 
                            ({source.get('Type', '')}) - 权重: {source.get('Weight', 10)}<br>
"""
                    html_content += """
                        </td>
                    </tr>
"""
            html_content += """
                </tbody>
            </table>
"""
        else:
            html_content += '<div class="empty-state">暂无CDN配置</div>'
        html_content += "</div>"

        # Footer
        html_content += f"""
        <div class="footer">
            <p>本报告由阿里云资源分析工具自动生成</p>
            <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

        with open(html_file, "w", encoding="utf-8") as f:
            f.write(html_content)

        self.logger.info(f"📄 HTML报告已生成: {html_file}")
        return html_file

    def generate_network_report(self, results: Dict):
        """生成网络资源报告"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"{self.tenant_name}_network_resources_{timestamp}.xlsx"

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            wb.remove(wb.active)  # 删除默认sheet

            # 1. VPC信息
            if results["VPCs"]:
                vpc_data = []
                for vpc in results["VPCs"]:
                    vswitches_str = "\n".join(
                        [
                            f"{vs.get('VSwitchName', '')}({vs.get('CidrBlock', '')})"
                            for vs in vpc.get("VSwitches", [])
                        ]
                    )
                    route_tables_str = "\n".join(
                        [f"{rt.get('RouteTableName', '')}" for rt in vpc.get("RouteTables", [])]
                    )
                    acls_str = "\n".join(
                        [f"{acl.get('NetworkAclName', '')}" for acl in vpc.get("NetworkAcls", [])]
                    )

                    vpc_data.append(
                        {
                            "VPC ID": vpc.get("VpcId", ""),
                            "VPC名称": vpc.get("VpcName", ""),
                            "CIDR": vpc.get("CidrBlock", ""),
                            "区域": vpc.get("RegionId", ""),
                            "状态": vpc.get("Status", ""),
                            "是否默认": "是" if vpc.get("IsDefault") else "否",
                            "子网数量": len(vpc.get("VSwitches", [])),
                            "子网列表": vswitches_str,
                            "路由表数量": len(vpc.get("RouteTables", [])),
                            "路由表列表": route_tables_str,
                            "网络ACL数量": len(vpc.get("NetworkAcls", [])),
                            "网络ACL列表": acls_str,
                            "创建时间": vpc.get("CreationTime", ""),
                            "描述": vpc.get("Description", ""),
                        }
                    )

                df_vpc = pd.DataFrame(vpc_data)
                ws = wb.create_sheet("VPC列表")
                # 写入表头
                ws.append(list(df_vpc.columns))
                # 写入数据
                for _, row in df_vpc.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            # 2. VPC Peering
            if results["VpcPeerings"]:
                peering_data = []
                for peering in results["VpcPeerings"]:
                    peering_data.append(
                        {
                            "对等连接ID": peering.get("InstanceId", ""),
                            "名称": peering.get("Name", ""),
                            "本地VPC": peering.get("VpcId", ""),
                            "对端VPC": peering.get("AcceptingVpcId", ""),
                            "对端区域": peering.get("AcceptingRegionId", ""),
                            "带宽(Mbps)": peering.get("Bandwidth", 0),
                            "状态": peering.get("Status", ""),
                            "创建时间": peering.get("CreationTime", ""),
                            "描述": peering.get("Description", ""),
                        }
                    )

                df_peering = pd.DataFrame(peering_data)
                ws = wb.create_sheet("VPC对等连接")
                ws.append(list(df_peering.columns))
                for _, row in df_peering.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            # 3. VPN连接
            if results["VpnConnections"]:
                vpn_data = []
                for vpn in results["VpnConnections"]:
                    vpn_data.append(
                        {
                            "VPN连接ID": vpn.get("VpnConnectionId", ""),
                            "VPN连接名称": vpn.get("VpnConnectionName", ""),
                            "VPN网关ID": vpn.get("VpnGatewayId", ""),
                            "用户网关ID": vpn.get("CustomerGatewayId", ""),
                            "状态": vpn.get("Status", ""),
                            "本地网段": vpn.get("LocalSubnet", ""),
                            "对端网段": vpn.get("RemoteSubnet", ""),
                            "创建时间": vpn.get("CreateTime", ""),
                            "区域": vpn.get("RegionId", ""),
                        }
                    )

                df_vpn = pd.DataFrame(vpn_data)
                ws = wb.create_sheet("VPN连接")
                ws.append(list(df_vpn.columns))
                for _, row in df_vpn.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            # 4. 专线配置
            if results["ExpressConnects"]:
                express_data = []
                for conn in results["ExpressConnects"]:
                    express_data.append(
                        {
                            "专线ID": conn.get("PhysicalConnectionId", ""),
                            "名称": conn.get("Name", ""),
                            "接入点ID": conn.get("AccessPointId", ""),
                            "类型": conn.get("Type", ""),
                            "状态": conn.get("Status", ""),
                            "业务状态": conn.get("BusinessStatus", ""),
                            "带宽(Mbps)": conn.get("Bandwidth", 0),
                            "端口类型": conn.get("PortType", ""),
                            "电路编码": conn.get("CircuitCode", ""),
                            "对端位置": conn.get("PeerLocation", ""),
                            "创建时间": conn.get("CreationTime", ""),
                            "区域": conn.get("RegionId", ""),
                        }
                    )

                df_express = pd.DataFrame(express_data)
                ws = wb.create_sheet("专线配置")
                ws.append(list(df_express.columns))
                for _, row in df_express.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            # 5. SLB详细配置
            if results["SLBs"]:
                slb_data = []
                for slb in results["SLBs"]:
                    listeners_str = "\n".join(
                        [
                            f"端口:{l.get('ListenerPort', 0)} 协议:{l.get('ListenerProtocol', '')} 状态:{l.get('Status', '')}"
                            for l in slb.get("Listeners", [])
                        ]
                    )
                    server_groups_str = "\n".join(
                        [
                            f"{sg.get('ServerGroupName', '')}({len(sg.get('Servers', []))}个服务器)"
                            for sg in slb.get("ServerGroups", [])
                        ]
                    )

                    slb_data.append(
                        {
                            "实例ID": slb.get("InstanceId", ""),
                            "实例名称": slb.get("InstanceName", ""),
                            "负载均衡类型": slb.get("LoadBalancerTypeName", "CLB"),
                            "区域": slb.get("Region", ""),
                            "地址类型": slb.get("AddressType", ""),
                            "规格": slb.get("InstanceType", ""),
                            "带宽(Mbps)": slb.get("Bandwidth", 0),
                            "状态": slb.get("InstanceStatus", ""),
                            "监听器数量": len(slb.get("Listeners", [])),
                            "监听器列表": listeners_str,
                            "后端服务器组数量": len(slb.get("ServerGroups", [])),
                            "后端服务器组列表": server_groups_str,
                            "创建时间": slb.get("CreateTime", ""),
                        }
                    )

                df_slb = pd.DataFrame(slb_data)
                ws = wb.create_sheet("SLB实例")
                ws.append(list(df_slb.columns))
                for _, row in df_slb.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            # 6. CDN配置
            if results["CDNs"]:
                cdn_data = []
                for cdn in results["CDNs"]:
                    sources_str = "\n".join(
                        [
                            f"{s.get('Content', '')}:{s.get('Port', 80)} ({s.get('Type', '')})"
                            for s in cdn.get("Sources", [])
                        ]
                    )

                    cdn_data.append(
                        {
                            "域名": cdn.get("DomainName", ""),
                            "CDN类型": cdn.get("CdnType", ""),
                            "状态": cdn.get("DomainStatus", ""),
                            "加速区域": cdn.get("Coverage", ""),
                            "源站数量": len(cdn.get("Sources", [])),
                            "源站列表": sources_str,
                            "创建时间": cdn.get("GmtCreated", ""),
                            "修改时间": cdn.get("GmtModified", ""),
                            "描述": cdn.get("Description", ""),
                        }
                    )

                df_cdn = pd.DataFrame(cdn_data)
                ws = wb.create_sheet("CDN配置")
                ws.append(list(df_cdn.columns))
                for _, row in df_cdn.iterrows():
                    ws.append(list(row))
                format_sheet(ws)

            wb.save(output_file)
            self.logger.info(f"📄 Excel报告已生成: {output_file}")
            return output_file
        except ImportError:
            self.logger.warning("pandas或openpyxl未安装，无法生成Excel报告")
            return None
        except Exception as e:
            self.logger.error(f"生成Excel报告失败: {e}")
            import traceback

            traceback.print_exc()
            return None


def format_sheet(ws):
    """格式化工作表"""
    from openpyxl.styles import Alignment, Font, PatternFill

    # 格式化标题行
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def main(access_key_id=None, access_key_secret=None, tenant_name=None):
    """主函数"""
    import json

    # 如果没有传入参数，尝试从配置文件读取
    if access_key_id is None or access_key_secret is None:
        try:
            with open("config.json", "r") as f:
                config = json.load(f)

            if tenant_name is None:
                tenant_name = config.get("default_tenant", "default")

            tenant_config = config.get("tenants", {}).get(tenant_name, {})
            access_key_id = access_key_id or tenant_config.get("access_key_id")
            access_key_secret = access_key_secret or tenant_config.get("access_key_secret")
        except FileNotFoundError:
            raise ValueError("必须提供access_key_id和access_key_secret，或配置文件config.json")

    analyzer = NetworkAnalyzer(access_key_id, access_key_secret, tenant_name or "default")
    analyzer.analyze_network_resources()


if __name__ == "__main__":
    main()
