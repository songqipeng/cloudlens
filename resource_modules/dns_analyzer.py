#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DNS（域名解析）资源分析模块
查询所有域名及其解析记录
"""

import json
import time
import sqlite3
import pandas as pd
import sys
import os
import subprocess
import ipaddress
from datetime import datetime
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.request import CommonRequest
from utils.concurrent_helper import process_concurrently
from utils.logger import get_logger
from utils.error_handler import ErrorHandler


class DNSAnalyzer:
    """DNS资源分析器"""
    
    def __init__(self, access_key_id, access_key_secret, tenant_name=None):
        self.access_key_id = access_key_id
        self.access_key_secret = access_key_secret
        self.tenant_name = tenant_name or "default"
        self.db_name = 'dns_monitoring_data.db'
        self.logger = get_logger('dns_analyzer')
        self.init_database()
        
    def init_database(self):
        """初始化DNS数据库"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # 创建域名表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS dns_domains (
            domain_name TEXT PRIMARY KEY,
            domain_id TEXT,
            registrant_email TEXT,
            group_id TEXT,
            group_name TEXT,
            record_count INTEGER DEFAULT 0,
            creation_time TEXT,
            expiration_time TEXT,
            update_time TEXT
        )
        ''')
        
        # 创建解析记录表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS dns_records (
            record_id TEXT PRIMARY KEY,
            domain_name TEXT,
            rr TEXT,
            type TEXT,
            value TEXT,
            ttl INTEGER,
            priority INTEGER,
            line TEXT,
            status TEXT,
            locked BOOLEAN,
            update_time TEXT,
            FOREIGN KEY (domain_name) REFERENCES dns_domains (domain_name)
        )
        ''')
        
        conn.commit()
        conn.close()
        self.logger.info("DNS数据库初始化完成")
    
    def get_all_domains(self):
        """获取所有域名"""
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, 'cn-hangzhou')
            request = CommonRequest()
            request.set_domain('alidns.aliyuncs.com')
            request.set_method('POST')
            request.set_version('2015-01-09')
            request.set_action_name('DescribeDomains')
            request.add_query_param('PageSize', 100)
            
            all_domains = []
            page_number = 1
            
            while True:
                request.add_query_param('PageNumber', page_number)
                try:
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)
                    
                    if 'Domains' in data and 'Domain' in data['Domains']:
                        domains = data['Domains']['Domain']
                        if not isinstance(domains, list):
                            domains = [domains]
                        
                        if not domains:
                            break
                        
                        for domain in domains:
                            all_domains.append({
                                'DomainName': domain.get('DomainName', ''),
                                'DomainId': domain.get('DomainId', ''),
                                'RegistrantEmail': domain.get('RegistrantEmail', ''),
                                'GroupId': domain.get('GroupId', ''),
                                'GroupName': domain.get('GroupName', ''),
                                'RecordCount': int(domain.get('RecordCount', 0)),
                                'CreateTime': domain.get('CreateTime', ''),
                                'ExpirationDate': domain.get('ExpirationDate', ''),
                                'UpdateTime': domain.get('UpdateTime', '')
                            })
                        
                        total_count = data.get('TotalCount', 0)
                        if len(all_domains) >= total_count or len(domains) < 100:
                            break
                        
                        page_number += 1
                    else:
                        break
                except Exception as e:
                    self.logger.error(f"获取域名列表失败 (页码 {page_number}): {str(e)}")
                    break
            
            self.logger.info(f"共获取到 {len(all_domains)} 个域名")
            return all_domains
        except Exception as e:
            self.logger.error(f"获取域名列表失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    
    def get_domain_records(self, domain_name):
        """获取指定域名的所有解析记录"""
        try:
            client = AcsClient(self.access_key_id, self.access_key_secret, 'cn-hangzhou')
            request = CommonRequest()
            request.set_domain('alidns.aliyuncs.com')
            request.set_method('POST')
            request.set_version('2015-01-09')
            request.set_action_name('DescribeDomainRecords')
            request.add_query_param('DomainName', domain_name)
            request.add_query_param('PageSize', 500)
            
            all_records = []
            page_number = 1
            
            while True:
                request.add_query_param('PageNumber', page_number)
                try:
                    response = client.do_action_with_exception(request)
                    data = json.loads(response)
                    
                    if 'DomainRecords' in data and 'Record' in data['DomainRecords']:
                        records = data['DomainRecords']['Record']
                        if not isinstance(records, list):
                            records = [records]
                        
                        if not records:
                            break
                        
                        for record in records:
                            all_records.append({
                                'RecordId': record.get('RecordId', ''),
                                'DomainName': domain_name,
                                'RR': record.get('RR', ''),
                                'Type': record.get('Type', ''),
                                'Value': record.get('Value', ''),
                                'TTL': int(record.get('TTL', 600)),
                                'Priority': int(record.get('Priority', 0)) if record.get('Priority') else 0,
                                'Line': record.get('Line', ''),
                                'Status': record.get('Status', ''),
                                'Locked': record.get('Locked', False),
                                'UpdateTime': record.get('UpdateTime', '')
                            })
                        
                        total_count = data.get('TotalCount', 0)
                        if len(all_records) >= total_count or len(records) < 500:
                            break
                        
                        page_number += 1
                    else:
                        break
                except Exception as e:
                    self.logger.error(f"获取域名 {domain_name} 解析记录失败 (页码 {page_number}): {str(e)}")
                    break
            
            return all_records
        except Exception as e:
            self.logger.error(f"获取域名 {domain_name} 解析记录失败: {str(e)}")
            return []
    
    def save_domain_to_db(self, domain_info):
        """保存域名信息到数据库"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            cursor.execute('''
            INSERT OR REPLACE INTO dns_domains 
            (domain_name, domain_id, registrant_email, group_id, group_name, 
             record_count, creation_time, expiration_time, update_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                domain_info['DomainName'],
                domain_info.get('DomainId', ''),
                domain_info.get('RegistrantEmail', ''),
                domain_info.get('GroupId', ''),
                domain_info.get('GroupName', ''),
                domain_info.get('RecordCount', 0),
                domain_info.get('CreateTime', ''),
                domain_info.get('ExpirationDate', ''),
                domain_info.get('UpdateTime', '')
            ))
            
            conn.commit()
            conn.close()
        except Exception as e:
            self.logger.error(f"保存域名 {domain_info.get('DomainName', '')} 到数据库失败: {str(e)}")
    
    def save_records_to_db(self, records):
        """批量保存解析记录到数据库"""
        if not records:
            return
        
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            for record in records:
                cursor.execute('''
                INSERT OR REPLACE INTO dns_records
                (record_id, domain_name, rr, type, value, ttl, priority, line, status, locked, update_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    record['RecordId'],
                    record['DomainName'],
                    record['RR'],
                    record['Type'],
                    record['Value'],
                    record['TTL'],
                    record['Priority'],
                    record['Line'],
                    record['Status'],
                    record['Locked'],
                    record['UpdateTime']
                ))
            
            conn.commit()
            conn.close()
        except Exception as e:
            self.logger.error(f"保存解析记录到数据库失败: {str(e)}")
    
    def analyze_dns_resources(self):
        """分析DNS资源"""
        self.logger.info("=" * 80)
        self.logger.info("开始DNS资源分析")
        self.logger.info("=" * 80)
        
        # 获取所有域名
        self.logger.info("正在获取所有域名...")
        domains = self.get_all_domains()
        
        if not domains:
            self.logger.warning("未获取到任何域名")
            return []
        
        # 保存域名信息
        for domain in domains:
            self.save_domain_to_db(domain)
        
        # 获取每个域名的解析记录
        self.logger.info(f"正在获取 {len(domains)} 个域名的解析记录...")
        
        all_records = []
        total_records = 0
        
        def process_domain(domain_info):
            domain_name = domain_info['DomainName']
            records = self.get_domain_records(domain_name)
            return domain_name, records
        
        # 并发处理域名
        results = process_concurrently(domains, process_domain, max_workers=10)
        
        for result in results:
            if result is None:
                continue
            domain_name, records = result
            if records:
                self.save_records_to_db(records)
                all_records.extend(records)
                total_records += len(records)
                self.logger.info(f"  ✓ {domain_name}: {len(records)} 条解析记录")
            else:
                self.logger.info(f"  - {domain_name}: 无解析记录")
        
        self.logger.info("=" * 80)
        self.logger.info(f"DNS资源分析完成")
        self.logger.info(f"  域名总数: {len(domains)}")
        self.logger.info(f"  解析记录总数: {total_records}")
        self.logger.info("=" * 80)
        
        result = {
            'domains': domains,
            'records': all_records
        }
        
        print(f"\n✅ DNS分析完成:")
        print(f"   📋 域名总数: {len(domains)}")
        print(f"   📝 解析记录总数: {total_records}")
        print(f"   📊 报告已生成")
        
        return result
    
    def generate_dns_report(self, data, output_dir="."):
        """生成DNS分析报告"""
        domains = data.get('domains', [])
        records = data.get('records', [])
        
        if not domains:
            self.logger.warning("没有域名数据，跳过报告生成")
            return
        
        # 生成HTML报告
        self.generate_html_report(domains, records, output_dir)
        
        # 生成Excel报告
        self.generate_excel_report(domains, records, output_dir)
        
        # 生成公网IP解析记录报告（HTML和PDF）
        self.generate_public_ip_report(domains, records, output_dir)
    
    def generate_html_report(self, domains, records, output_dir="."):
        """生成HTML报告"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            tenant_prefix = f"{self.tenant_name}_" if self.tenant_name != "default" else ""
            filename = f"{output_dir}/{tenant_prefix}dns_report_{timestamp}.html"
            
            # 统计信息
            total_domains = len(domains)
            total_records = len(records)
            
            # 按类型统计解析记录
            record_type_stats = {}
            for record in records:
                record_type = record.get('Type', 'UNKNOWN')
                record_type_stats[record_type] = record_type_stats.get(record_type, 0) + 1
            
            # 按域名分组统计
            domain_record_stats = {}
            for record in records:
                domain_name = record.get('DomainName', '')
                if domain_name not in domain_record_stats:
                    domain_record_stats[domain_name] = {
                        'total': 0,
                        'by_type': {}
                    }
                domain_record_stats[domain_name]['total'] += 1
                record_type = record.get('Type', 'UNKNOWN')
                domain_record_stats[domain_name]['by_type'][record_type] = \
                    domain_record_stats[domain_name]['by_type'].get(record_type, 0) + 1
            
            html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DNS资源分析报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4CAF50;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
            border-left: 4px solid #4CAF50;
            padding-left: 10px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-card h3 {{
            margin: 0;
            font-size: 14px;
            opacity: 0.9;
        }}
        .stat-card .value {{
            font-size: 32px;
            font-weight: bold;
            margin: 10px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .domain-section {{
            margin: 30px 0;
            padding: 20px;
            background-color: #f9f9f9;
            border-radius: 8px;
            border-left: 4px solid #4CAF50;
        }}
        .domain-header {{
            font-size: 18px;
            font-weight: bold;
            color: #333;
            margin-bottom: 10px;
        }}
        .record-type-badge {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: bold;
            margin-right: 5px;
        }}
        .type-A {{ background-color: #4CAF50; color: white; }}
        .type-AAAA {{ background-color: #2196F3; color: white; }}
        .type-CNAME {{ background-color: #FF9800; color: white; }}
        .type-MX {{ background-color: #9C27B0; color: white; }}
        .type-TXT {{ background-color: #607D8B; color: white; }}
        .type-NS {{ background-color: #F44336; color: white; }}
        .type-SOA {{ background-color: #795548; color: white; }}
        .type-OTHER {{ background-color: #9E9E9E; color: white; }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🌐 DNS资源分析报告</h1>
        <p><strong>生成时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>租户:</strong> {self.tenant_name}</p>
        
        <div class="stats">
            <div class="stat-card">
                <h3>域名总数</h3>
                <div class="value">{total_domains}</div>
            </div>
            <div class="stat-card">
                <h3>解析记录总数</h3>
                <div class="value">{total_records}</div>
            </div>
            <div class="stat-card">
                <h3>记录类型数</h3>
                <div class="value">{len(record_type_stats)}</div>
            </div>
        </div>
        
        <h2>📊 解析记录类型统计</h2>
        <table>
            <thead>
                <tr>
                    <th>记录类型</th>
                    <th>数量</th>
                    <th>占比</th>
                </tr>
            </thead>
            <tbody>
"""
            
            # 按数量排序
            sorted_types = sorted(record_type_stats.items(), key=lambda x: x[1], reverse=True)
            for record_type, count in sorted_types:
                percentage = (count / total_records * 100) if total_records > 0 else 0
                html_content += f"""
                <tr>
                    <td><span class="record-type-badge type-{record_type}">{record_type}</span></td>
                    <td>{count}</td>
                    <td>{percentage:.2f}%</td>
                </tr>
"""
            
            html_content += """
            </tbody>
        </table>
        
        <h2>📋 域名列表</h2>
"""
            
            # 按域名分组显示
            for domain_info in sorted(domains, key=lambda x: x.get('DomainName', '')):
                domain_name = domain_info.get('DomainName', '')
                domain_records = [r for r in records if r.get('DomainName') == domain_name]
                
                html_content += f"""
        <div class="domain-section">
            <div class="domain-header">🌍 {domain_name}</div>
            <p><strong>域名ID:</strong> {domain_info.get('DomainId', 'N/A')}</p>
            <p><strong>解析记录数:</strong> {len(domain_records)}</p>
            <p><strong>注册邮箱:</strong> {domain_info.get('RegistrantEmail', 'N/A')}</p>
            <p><strong>分组:</strong> {domain_info.get('GroupName', '默认分组')}</p>
            <p><strong>创建时间:</strong> {domain_info.get('CreateTime', 'N/A')}</p>
            <p><strong>到期时间:</strong> {domain_info.get('ExpirationDate', 'N/A')}</p>
            
            <table>
                <thead>
                    <tr>
                        <th>主机记录</th>
                        <th>记录类型</th>
                        <th>记录值</th>
                        <th>TTL</th>
                        <th>线路</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
"""
                
                for record in sorted(domain_records, key=lambda x: x.get('RR', '')):
                    record_type = record.get('Type', 'UNKNOWN')
                    html_content += f"""
                    <tr>
                        <td>{record.get('RR', '')}</td>
                        <td><span class="record-type-badge type-{record_type}">{record_type}</span></td>
                        <td>{record.get('Value', '')}</td>
                        <td>{record.get('TTL', 600)}</td>
                        <td>{record.get('Line', '默认')}</td>
                        <td>{record.get('Status', 'N/A')}</td>
                    </tr>
"""
                
                html_content += """
                </tbody>
            </table>
        </div>
"""
            
            html_content += f"""
        <div class="footer">
            <p>报告由阿里云资源分析工具自动生成</p>
        </div>
    </div>
</body>
</html>
"""
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"HTML报告已生成: {filename}")
        except Exception as e:
            self.logger.error(f"生成HTML报告失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def generate_excel_report(self, domains, records, output_dir="."):
        """生成Excel报告"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            tenant_prefix = f"{self.tenant_name}_" if self.tenant_name != "default" else ""
            filename = f"{output_dir}/{tenant_prefix}dns_report_{timestamp}.xlsx"
            
            # 创建Excel写入器
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. 域名汇总表
                domains_df = pd.DataFrame(domains)
                if not domains_df.empty:
                    domains_df.columns = ['域名', '域名ID', '注册邮箱', '分组ID', '分组名称', 
                                         '解析记录数', '创建时间', '到期时间', '更新时间']
                    domains_df.to_excel(writer, sheet_name='域名列表', index=False)
                
                # 2. 解析记录详情表
                records_df = pd.DataFrame(records)
                if not records_df.empty:
                    records_df = records_df[['DomainName', 'RR', 'Type', 'Value', 'TTL', 
                                            'Priority', 'Line', 'Status', 'Locked', 'UpdateTime']]
                    records_df.columns = ['域名', '主机记录', '记录类型', '记录值', 'TTL', 
                                         '优先级', '线路', '状态', '锁定', '更新时间']
                    records_df.to_excel(writer, sheet_name='解析记录', index=False)
                
                # 3. 统计汇总表
                stats_data = []
                # 按类型统计
                record_type_stats = {}
                for record in records:
                    record_type = record.get('Type', 'UNKNOWN')
                    record_type_stats[record_type] = record_type_stats.get(record_type, 0) + 1
                
                for record_type, count in sorted(record_type_stats.items(), key=lambda x: x[1], reverse=True):
                    stats_data.append({
                        '统计项': '记录类型',
                        '名称': record_type,
                        '数量': count
                    })
                
                # 按域名统计
                for domain_info in domains:
                    domain_name = domain_info.get('DomainName', '')
                    domain_records = [r for r in records if r.get('DomainName') == domain_name]
                    stats_data.append({
                        '统计项': '域名解析记录数',
                        '名称': domain_name,
                        '数量': len(domain_records)
                    })
                
                if stats_data:
                    stats_df = pd.DataFrame(stats_data)
                    stats_df.to_excel(writer, sheet_name='统计汇总', index=False)
            
            # 格式化Excel
            wb = load_workbook(filename)
            
            # 格式化域名列表表
            if '域名列表' in wb.sheetnames:
                ws = wb['域名列表']
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 格式化解析记录表
            if '解析记录' in wb.sheetnames:
                ws = wb['解析记录']
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            self.logger.info(f"Excel报告已生成: {filename}")
        except ImportError:
            self.logger.warning("pandas或openpyxl未安装，跳过Excel报告生成")
        except Exception as e:
            self.logger.error(f"生成Excel报告失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def is_public_ip(self, ip_str):
        """判断IP是否为公网IP"""
        try:
            # 尝试解析IPv4
            try:
                ip = ipaddress.IPv4Address(ip_str)
                # 排除私有IP地址范围
                return not ip.is_private and not ip.is_loopback and not ip.is_link_local and not ip.is_multicast and not ip.is_reserved
            except:
                pass
            
            # 尝试解析IPv6
            try:
                ip = ipaddress.IPv6Address(ip_str)
                # 排除私有IPv6地址范围
                return not ip.is_private and not ip.is_loopback and not ip.is_link_local and not ip.is_multicast
            except:
                pass
            
            return False
        except:
            return False
    
    def filter_public_ip_records(self, records):
        """过滤出解析到公网IP的记录（A和AAAA类型）"""
        public_ip_records = []
        for record in records:
            record_type = record.get('Type', '').upper()
            record_value = record.get('Value', '').strip()
            
            # 只处理A和AAAA类型的记录
            if record_type in ['A', 'AAAA']:
                # 检查是否为公网IP
                if self.is_public_ip(record_value):
                    public_ip_records.append(record)
        
        return public_ip_records
    
    def generate_public_ip_report(self, domains, records, output_dir="."):
        """生成公网IP解析记录报告（HTML和PDF）"""
        # 过滤出公网IP记录
        public_ip_records = self.filter_public_ip_records(records)
        
        if not public_ip_records:
            self.logger.info("未发现解析到公网IP的记录，跳过报告生成")
            return
        
        self.logger.info(f"发现 {len(public_ip_records)} 条解析到公网IP的记录，生成报告...")
        
        # 获取相关的域名信息
        domain_map = {d['DomainName']: d for d in domains}
        related_domains = []
        domain_names = set()
        for record in public_ip_records:
            domain_name = record.get('DomainName', '')
            if domain_name and domain_name not in domain_names:
                domain_names.add(domain_name)
                if domain_name in domain_map:
                    related_domains.append(domain_map[domain_name])
        
        # 生成HTML报告
        html_file = self.generate_public_ip_html_report(related_domains, public_ip_records, output_dir)
        
        # 生成PDF报告
        if html_file:
            pdf_file = self.generate_pdf(html_file)
            if pdf_file:
                self.logger.info(f"PDF报告已生成: {pdf_file}")
    
    def generate_public_ip_html_report(self, domains, records, output_dir="."):
        """生成公网IP解析记录HTML报告"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            tenant_prefix = f"{self.tenant_name}_" if self.tenant_name != "default" else ""
            filename = f"{output_dir}/{tenant_prefix}dns_public_ip_report_{timestamp}.html"
            
            # 统计信息
            total_domains = len(set(r.get('DomainName', '') for r in records))
            total_records = len(records)
            
            # 按IP地址统计
            ip_stats = {}
            for record in records:
                ip = record.get('Value', '')
                if ip not in ip_stats:
                    ip_stats[ip] = {
                        'count': 0,
                        'domains': set(),
                        'records': []
                    }
                ip_stats[ip]['count'] += 1
                ip_stats[ip]['domains'].add(record.get('DomainName', ''))
                ip_stats[ip]['records'].append(record)
            
            # 按域名分组
            domain_records = {}
            for record in records:
                domain_name = record.get('DomainName', '')
                if domain_name not in domain_records:
                    domain_records[domain_name] = []
                domain_records[domain_name].append(record)
            
            html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DNS公网IP解析记录报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #FF6B6B;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
            border-left: 4px solid #FF6B6B;
            padding-left: 10px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #FF6B6B 0%, #FF8E53 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-card h3 {{
            margin: 0;
            font-size: 14px;
            opacity: 0.9;
        }}
        .stat-card .value {{
            font-size: 32px;
            font-weight: bold;
            margin: 10px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #FF6B6B;
            color: white;
            font-weight: bold;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .ip-address {{
            font-family: 'Courier New', monospace;
            font-weight: bold;
            color: #FF6B6B;
        }}
        .domain-section {{
            margin: 30px 0;
            padding: 20px;
            background-color: #f9f9f9;
            border-radius: 8px;
            border-left: 4px solid #FF6B6B;
        }}
        .domain-header {{
            font-size: 18px;
            font-weight: bold;
            color: #333;
            margin-bottom: 10px;
        }}
        .ip-section {{
            margin: 30px 0;
            padding: 20px;
            background-color: #fff5f5;
            border-radius: 8px;
            border-left: 4px solid #FF6B6B;
        }}
        .ip-header {{
            font-size: 18px;
            font-weight: bold;
            color: #FF6B6B;
            margin-bottom: 10px;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
        .warning {{
            background-color: #fff3cd;
            border: 1px solid #ffc107;
            color: #856404;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🌐 DNS公网IP解析记录报告</h1>
        <p><strong>生成时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>租户:</strong> {self.tenant_name}</p>
        
        <div class="warning">
            <strong>⚠️ 注意：</strong>本报告仅包含解析到公网IP地址的DNS记录（A和AAAA类型），这些记录可能暴露在互联网上。
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <h3>涉及域名数</h3>
                <div class="value">{total_domains}</div>
            </div>
            <div class="stat-card">
                <h3>公网IP记录数</h3>
                <div class="value">{total_records}</div>
            </div>
            <div class="stat-card">
                <h3>唯一IP地址数</h3>
                <div class="value">{len(ip_stats)}</div>
            </div>
        </div>
        
        <h2>📊 IP地址统计</h2>
        <table>
            <thead>
                <tr>
                    <th>IP地址</th>
                    <th>解析记录数</th>
                    <th>涉及域名数</th>
                </tr>
            </thead>
            <tbody>
"""
            
            # 按记录数排序
            sorted_ips = sorted(ip_stats.items(), key=lambda x: x[1]['count'], reverse=True)
            for ip, stats in sorted_ips:
                html_content += f"""
                <tr>
                    <td class="ip-address">{ip}</td>
                    <td>{stats['count']}</td>
                    <td>{len(stats['domains'])}</td>
                </tr>
"""
            
            html_content += """
            </tbody>
        </table>
        
        <h2>📋 按域名分组的解析记录</h2>
"""
            
            # 按域名分组显示
            for domain_name in sorted(domain_records.keys()):
                domain_recs = domain_records[domain_name]
                domain_info = next((d for d in domains if d.get('DomainName') == domain_name), {})
                
                html_content += f"""
        <div class="domain-section">
            <div class="domain-header">🌍 {domain_name}</div>
            <p><strong>解析记录数:</strong> {len(domain_recs)}</p>
            {f'<p><strong>注册邮箱:</strong> {domain_info.get("RegistrantEmail", "N/A")}</p>' if domain_info.get("RegistrantEmail") else ''}
            {f'<p><strong>分组:</strong> {domain_info.get("GroupName", "默认分组")}</p>' if domain_info.get("GroupName") else ''}
            
            <table>
                <thead>
                    <tr>
                        <th>主机记录</th>
                        <th>记录类型</th>
                        <th>IP地址</th>
                        <th>TTL</th>
                        <th>线路</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
"""
                
                for record in sorted(domain_recs, key=lambda x: x.get('RR', '')):
                    html_content += f"""
                    <tr>
                        <td>{record.get('RR', '')}</td>
                        <td>{record.get('Type', '')}</td>
                        <td class="ip-address">{record.get('Value', '')}</td>
                        <td>{record.get('TTL', 600)}</td>
                        <td>{record.get('Line', '默认')}</td>
                        <td>{record.get('Status', 'N/A')}</td>
                    </tr>
"""
                
                html_content += """
                </tbody>
            </table>
        </div>
"""
            
            html_content += f"""
        <h2>📊 按IP地址分组的解析记录</h2>
"""
            
            # 按IP地址分组显示
            for ip, stats in sorted_ips:
                html_content += f"""
        <div class="ip-section">
            <div class="ip-header">🔗 IP地址: <span class="ip-address">{ip}</span></div>
            <p><strong>解析记录数:</strong> {stats['count']}</p>
            <p><strong>涉及域名:</strong> {', '.join(sorted(stats['domains']))}</p>
            
            <table>
                <thead>
                    <tr>
                        <th>域名</th>
                        <th>主机记录</th>
                        <th>记录类型</th>
                        <th>TTL</th>
                        <th>线路</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
"""
                
                for record in sorted(stats['records'], key=lambda x: (x.get('DomainName', ''), x.get('RR', ''))):
                    html_content += f"""
                    <tr>
                        <td>{record.get('DomainName', '')}</td>
                        <td>{record.get('RR', '')}</td>
                        <td>{record.get('Type', '')}</td>
                        <td>{record.get('TTL', 600)}</td>
                        <td>{record.get('Line', '默认')}</td>
                        <td>{record.get('Status', 'N/A')}</td>
                    </tr>
"""
                
                html_content += """
                </tbody>
            </table>
        </div>
"""
            
            html_content += f"""
        <div class="footer">
            <p>报告由阿里云资源分析工具自动生成</p>
            <p>⚠️ 本报告包含解析到公网IP的DNS记录，请注意安全防护</p>
        </div>
    </div>
</body>
</html>
"""
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.logger.info(f"公网IP解析记录HTML报告已生成: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"生成公网IP解析记录HTML报告失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_pdf(self, html_file):
        """生成PDF文件（使用Chrome headless模式）"""
        pdf_file = html_file.replace('.html', '.pdf')
        
        chrome_paths = [
            '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome',
            '/Applications/Chromium.app/Contents/MacOS/Chromium',
            'google-chrome',
            'chromium',
            'chromium-browser'
        ]
        
        chrome_cmd = None
        for path in chrome_paths:
            if os.path.exists(path) or subprocess.run(['which', path.split('/')[-1]], 
                                                      capture_output=True).returncode == 0:
                chrome_cmd = path
                break
        
        if chrome_cmd:
            html_path = os.path.abspath(html_file)
            cmd = [
                chrome_cmd,
                '--headless',
                '--disable-gpu',
                '--no-pdf-header-footer',
                '--print-to-pdf=' + pdf_file,
                'file://' + html_path
            ]
            
            try:
                subprocess.run(cmd, capture_output=True, timeout=30)
                if os.path.exists(pdf_file):
                    self.logger.info(f"PDF报告已生成: {pdf_file}")
                    return pdf_file
            except Exception as e:
                self.logger.warning(f"生成PDF失败: {str(e)}")
        
        return None


def main(access_key_id=None, access_key_secret=None, tenant_name=None):
    """主函数"""
    # 如果没有传入参数，尝试从配置文件读取
    if access_key_id is None or access_key_secret is None:
        try:
            with open('config.json', 'r') as f:
                config = json.load(f)
                default_tenant = config.get('default_tenant', 'default')
                tenants = config.get('tenants', {})
                
                if tenant_name and tenant_name in tenants:
                    tenant_config = tenants[tenant_name]
                elif default_tenant in tenants:
                    tenant_config = tenants[default_tenant]
                else:
                    tenant_config = config
                
                access_key_id = access_key_id or tenant_config.get('access_key_id')
                access_key_secret = access_key_secret or tenant_config.get('access_key_secret')
                tenant_name = tenant_name or default_tenant
        except FileNotFoundError:
            raise ValueError("必须提供access_key_id和access_key_secret，或配置文件config.json")
    
    analyzer = DNSAnalyzer(access_key_id, access_key_secret, tenant_name)
    data = analyzer.analyze_dns_resources()
    analyzer.generate_dns_report(data, output_dir=".")


if __name__ == "__main__":
    main()

