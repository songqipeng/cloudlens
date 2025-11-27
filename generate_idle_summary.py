#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成所有租户闲置资源详细汇总报告
"""

import json
import os
import sqlite3
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def load_config():
    """加载配置"""
    try:
        with open("config.json", "r") as f:
            return json.load(f)
    except:
        return {"tenants": {"ydzn": {}, "zmyc": {}, "cf": {}}}


def get_ecs_idle_stats(tenant_name="all"):
    """获取ECS闲置统计"""
    db_file = "ecs_monitoring_data_fixed.db"
    if not os.path.exists(db_file):
        return []
    
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # 获取最近的监控数据,按实例分组计算平均CPU使用率
        query = """
        SELECT 
            i.instance_id,
            i.instance_name,
            i.tenant_name,
            i.instance_type,
            i.region_id,
            i.status,
            AVG(m.cpu_percent) as avg_cpu,
            AVG(m.memory_percent) as avg_memory,
            i.creation_time
        FROM instances i
        LEFT JOIN monitoring_data m ON i.instance_id = m.instance_id
        WHERE m.timestamp >= datetime('now', '-7 days')
        """
        
        if tenant_name != "all":
            query += f" AND i.tenant_name = '{tenant_name}'"
        
        query += " GROUP BY i.instance_id HAVING avg_cpu < 10 OR avg_memory < 10"
        
        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()
        
        idle_instances = []
        for row in results:
            idle_instances.append({
                "instance_id": row[0],
                "instance_name": row[1] or "未命名",
                "tenant": row[2],
                "instance_type": row[3],
                "region": row[4],
                "status": row[5],
                "avg_cpu": round(row[6], 2) if row[6] else 0,
                "avg_memory": round(row[7], 2) if row[7] else 0,
                "creation_time": row[8]
            })
        
        return idle_instances
    except Exception as e:
        print(f"❌ 查询ECS数据出错: {e}")
        return []


def get_rds_idle_stats(tenant_name="all"):
    """获取RDS闲置统计"""
    db_file = "rds_monitoring_data.db"
    if not os.path.exists(db_file):
        return []
    
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        query = """
        SELECT 
            i.instance_id,
            i.instance_name,
            i.tenant_name,
            i.engine,
            i.engine_version,
            i.region_id,
            i.status,
            AVG(m.cpu_usage) as avg_cpu,
            AVG(m.memory_usage) as avg_memory,
            AVG(m.connections) as avg_connections
        FROM rds_instances i
        LEFT JOIN rds_monitoring_data m ON i.instance_id = m.instance_id
        WHERE m.timestamp >= datetime('now', '-7 days')
        """
        
        if tenant_name != "all":
            query += f" AND i.tenant_name = '{tenant_name}'"
        
        query += " GROUP BY i.instance_id HAVING avg_cpu < 10 OR avg_connections < 5"
        
        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()
        
        idle_instances = []
        for row in results:
            idle_instances.append({
                "instance_id": row[0],
                "instance_name": row[1] or "未命名",
                "tenant": row[2],
                "engine": f"{row[3]} {row[4]}",
                "region": row[5],
                "status": row[6],
                "avg_cpu": round(row[7], 2) if row[7] else 0,
                "avg_memory": round(row[8], 2) if row[8] else 0,
                "avg_connections": round(row[9], 2) if row[9] else 0
            })
        
        return idle_instances
    except Exception as e:
        print(f"❌ 查询RDS数据出错: {e}")
        return []


def get_slb_idle_stats(tenant_name="all"):
    """获取SLB闲置统计"""
    db_file = "slb_monitoring_data.db"
    if not os.path.exists(db_file):
        return []
    
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        query = """
        SELECT 
            i.load_balancer_id,
            i.load_balancer_name,
            i.tenant_name,
            i.address_type,
            i.load_balancer_status,
            i.region_id,
            AVG(m.active_connections) as avg_connections,
            AVG(m.new_connections) as avg_new_connections
        FROM slb_instances i
        LEFT JOIN slb_monitoring_data m ON i.load_balancer_id = m.load_balancer_id
        WHERE m.timestamp >= datetime('now', '-7 days')
        """
        
        if tenant_name != "all":
            query += f" AND i.tenant_name = '{tenant_name}'"
        
        query += " GROUP BY i.load_balancer_id HAVING avg_connections < 10"
        
        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()
        
        idle_instances = []
        for row in results:
            idle_instances.append({
                "lb_id": row[0],
                "lb_name": row[1] or "未命名",
                "tenant": row[2],
                "address_type": row[3],
                "status": row[4],
                "region": row[5],
                "avg_connections": round(row[6], 2) if row[6] else 0,
                "avg_new_connections": round(row[7], 2) if row[7] else 0
            })
        
        return idle_instances
    except Exception as e:
        print(f"❌ 查询SLB数据出错: {e}")
        return []


def get_dns_unused_records():
    """获取DNS未使用记录"""
    db_file = "dns_monitoring_data.db"
    if not os.path.exists(db_file):
        return []
    
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # 查找可能未使用的DNS记录 (这里简单示例)
        query = """
        SELECT 
            domain_name,
            tenant_name,
            COUNT(*) as record_count
        FROM dns_records
        GROUP BY domain_name, tenant_name
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()
        
        dns_data = []
        for row in results:
            dns_data.append({
                "domain": row[0],
                "tenant": row[1],
                "record_count": row[2]
            })
        
        return dns_data
    except Exception as e:
        print(f"❌ 查询DNS数据出错: {e}")
        return []


def generate_summary_report():
    """生成汇总报告"""
    print("="*80)
    print("📊 所有租户闲置资源详细汇总")
    print(f"📅 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)
    
    config = load_config()
    tenants = list(config.get("tenants", {}).keys())
    
    # 按租户统计
    tenant_stats = defaultdict(lambda: {
        "ecs_idle": 0,
        "rds_idle": 0,
        "slb_idle": 0
    })
    
    # 1. ECS闲置实例
    print("\n🖥️  ECS闲置实例 (过去7天平均CPU<10% 或 内存<10%):")
    print("-"*80)
    ecs_idle = get_ecs_idle_stats()
    if ecs_idle:
        for instance in ecs_idle:
            tenant_stats[instance["tenant"]]["ecs_idle"] += 1
            print(f"租户: {instance['tenant']:8s} | "
                  f"实例: {instance['instance_name'][:30]:30s} | "
                  f"类型: {instance['instance_type']:15s} | "
                  f"CPU: {instance['avg_cpu']:5.2f}% | "
                  f"内存: {instance['avg_memory']:5.2f}% | "
                  f"地域: {instance['region']}")
        print(f"\n总计: {len(ecs_idle)} 台闲置ECS实例")
    else:
        print("未发现闲置ECS实例")
    
    # 2. RDS闲置实例
    print("\n" + "="*80)
    print("🗄️  RDS闲置实例 (过去7天平均CPU<10% 或 连接数<5):")
    print("-"*80)
    rds_idle = get_rds_idle_stats()
    if rds_idle:
        for instance in rds_idle:
            tenant_stats[instance["tenant"]]["rds_idle"] += 1
            print(f"租户: {instance['tenant']:8s} | "
                  f"实例: {instance['instance_name'][:30]:30s} | "
                  f"引擎: {instance['engine']:15s} | "
                  f"CPU: {instance['avg_cpu']:5.2f}% | "
                  f"连接: {instance['avg_connections']:5.0f} | "
                  f"地域: {instance['region']}")
        print(f"\n总计: {len(rds_idle)} 个闲置RDS实例")
    else:
        print("未发现闲置RDS实例")
    
    # 3. SLB闲置实例
    print("\n" + "="*80)
    print("⚖️  SLB闲置实例 (过去7天平均连接数<10):")
    print("-"*80)
    slb_idle = get_slb_idle_stats()
    if slb_idle:
        for instance in slb_idle:
            tenant_stats[instance["tenant"]]["slb_idle"] += 1
            print(f"租户: {instance['tenant']:8s} | "
                  f"名称: {instance['lb_name'][:30]:30s} | "
                  f"类型: {instance['address_type']:10s} | "
                  f"连接: {instance['avg_connections']:5.0f} | "
                  f"地域: {instance['region']}")
        print(f"\n总计: {len(slb_idle)} 个闲置SLB实例")
    else:
        print("未发现闲置SLB实例")
    
    # 4. DNS域名统计
    print("\n" + "="*80)
    print("🌐 DNS域名统计:")
    print("-"*80)
    dns_data = get_dns_unused_records()
    if dns_data:
        for item in dns_data:
            print(f"租户: {item['tenant']:8s} | 域名: {item['domain']:40s} | 记录数: {item['record_count']}")
        print(f"\n总计: {len(dns_data)} 个域名")
    else:
        print("未发现DNS记录")
    
    # 显示按租户汇总
    print("\n" + "="*80)
    print("📊 按租户汇总:")
    print("="*80)
    print(f"{'租户':<10} {'ECS闲置':<10} {'RDS闲置':<10} {'SLB闲置':<10}")
    print("-"*80)
    
    total_ecs = 0
    total_rds = 0
    total_slb = 0
    
    for tenant in sorted(tenant_stats.keys()):
        stats = tenant_stats[tenant]
        print(f"{tenant:<10} {stats['ecs_idle']:<10} {stats['rds_idle']:<10} {stats['slb_idle']:<10}")
        total_ecs += stats['ecs_idle']
        total_rds += stats['rds_idle']
        total_slb += stats['slb_idle']
    
    print("-"*80)
    print(f"{'总计':<10} {total_ecs:<10} {total_rds:<10} {total_slb:<10}")
    print("="*80)
    
    # 生成Excel报告
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"all_tenants_idle_summary_{timestamp}.xlsx"
    
    try:
        wb = openpyxl.Workbook()
        
        # ECS工作表
        if ecs_idle:
            ws_ecs = wb.active
            ws_ecs.title = "ECS闲置实例"
            headers = ["租户", "实例名称", "实例ID", "实例类型", "地域", "状态", "平均CPU%", "平均内存%"]
            ws_ecs.append(headers)
            
            for instance in ecs_idle:
                ws_ecs.append([
                    instance["tenant"],
                    instance["instance_name"],
                    instance["instance_id"],
                    instance["instance_type"],
                    instance["region"],
                    instance["status"],
                    instance["avg_cpu"],
                    instance["avg_memory"]
                ])
        
        # RDS工作表
        if rds_idle:
            ws_rds = wb.create_sheet("RDS闲置实例")
            headers = ["租户", "实例名称", "实例ID", "引擎", "地域", "状态", "平均CPU%", "平均连接数"]
            ws_rds.append(headers)
            
            for instance in rds_idle:
                ws_rds.append([
                    instance["tenant"],
                    instance["instance_name"],
                    instance["instance_id"],
                    instance["engine"],
                    instance["region"],
                    instance["status"],
                    instance["avg_cpu"],
                    instance["avg_connections"]
                ])
        
        # SLB工作表
        if slb_idle:
            ws_slb = wb.create_sheet("SLB闲置实例")
            headers = ["租户", "负载均衡名称", "实例ID", "类型", "地域", "状态", "平均连接数"]
            ws_slb.append(headers)
            
            for instance in slb_idle:
                ws_slb.append([
                    instance["tenant"],
                    instance["lb_name"],
                    instance["lb_id"],
                    instance["address_type"],
                    instance["region"],
                    instance["status"],
                    instance["avg_connections"]
                ])
        
        wb.save(excel_file)
        print(f"\n✅ Excel报告已生成: {excel_file}")
        
    except Exception as e:
        print(f"\n❌ 生成Excel报告失败: {e}")
    
    print("\n" + "="*80)
    print("✅ 汇总报告生成完成!")
    print("="*80)


if __name__ == "__main__":
    generate_summary_report()
