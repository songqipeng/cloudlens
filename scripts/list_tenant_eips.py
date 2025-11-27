#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
提取YDZN和ZMYC租户的闲置EIP信息
"""

import openpyxl
from datetime import datetime
import os
import glob


def read_eip_report(file_path):
    """读取EIP报告文件"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # 读取表头
        headers = [cell.value for cell in ws[1]]
        
        # 读取数据
        eips = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                eip_data = dict(zip(headers, row))
                eips.append(eip_data)
        
        wb.close()
        return eips
    except Exception as e:
        print(f"读取文件 {file_path} 失败: {e}")
        return []


def display_eip_list(eips, tenant_name, tenant_display):
    """显示EIP列表"""
    print(f"\n{'='*100}")
    print(f"🏢 {tenant_display} ({tenant_name.upper()}) - 闲置EIP分析")
    print(f"{'='*100}")
    
    if not eips:
        print("✅ 未发现闲置EIP")
        return
    
    print(f"\n📊 总计: {len(eips)} 个闲置EIP\n")
    
    # 按闲置原因分类
    unbound_eips = []  # 未绑定
    low_traffic_eips = []  # 低流量
    
    for eip in eips:
        reason = eip.get('闲置原因', '')
        if '未绑定' in str(reason):
            unbound_eips.append(eip)
        elif '流量' in str(reason) or '带宽使用率' in str(reason):
            low_traffic_eips.append(eip)
        else:
            low_traffic_eips.append(eip)
    
    # 显示未绑定的EIP
    if unbound_eips:
        print(f"\n🔴 未绑定的EIP ({len(unbound_eips)}个) - 建议立即释放")
        print("-"*100)
        print(f"{'序号':<6} {'IP地址':<18} {'分配ID':<30} {'区域':<15} {'带宽':<10} {'闲置原因'}")
        print("-"*100)
        
        for i, eip in enumerate(unbound_eips, 1):
            ip = eip.get('IP地址', 'N/A')
            alloc_id = eip.get('分配ID', 'N/A')
            region = eip.get('区域', 'N/A')
            bandwidth = eip.get('带宽(Mbps)', 'N/A')
            reason = eip.get('闲置原因', 'N/A')
            
            print(f"{i:<6} {ip:<18} {alloc_id:<30} {region:<15} {bandwidth:<10} {reason[:40]}")
    
    # 显示低流量的EIP
    if low_traffic_eips:
        print(f"\n🟡 低流量EIP ({len(low_traffic_eips)}个) - 建议评估或降低带宽")
        print("-"*100)
        print(f"{'序号':<6} {'IP地址':<18} {'绑定实例':<25} {'带宽':<10} {'流量(MB)':<12} {'使用率%':<10} {'优化建议'}")
        print("-"*100)
        
        for i, eip in enumerate(low_traffic_eips, 1):
            ip = eip.get('IP地址', 'N/A')
            instance = eip.get('绑定实例ID', 'N/A')
            bandwidth = eip.get('带宽(Mbps)', 'N/A')
            traffic = eip.get('14天总流量(MB)', 0)
            usage = eip.get('带宽使用率(%)', 0)
            suggestion = eip.get('优化建议', '')[:50]
            
            instance_short = instance[:24] if len(str(instance)) > 24 else instance
            
            print(f"{i:<6} {ip:<18} {instance_short:<25} {bandwidth:<10} {traffic:<12.2f} {usage:<10.1f} {suggestion}")
    
    # 显示统计信息
    print(f"\n{'='*100}")
    print("📈 统计汇总:")
    print(f"  • 总闲置EIP: {len(eips)} 个")
    print(f"  • 未绑定: {len(unbound_eips)} 个")
    print(f"  • 低流量: {len(low_traffic_eips)} 个")
    
    # 估算成本
    total_bandwidth = sum([eip.get('带宽(Mbps)', 0) for eip in eips if isinstance(eip.get('带宽(Mbps)'), (int, float))])
    estimated_cost = len(unbound_eips) * 50 + len(low_traffic_eips) * 30  # 粗略估算
    
    print(f"  • 总带宽配置: {total_bandwidth} Mbps")
    print(f"  • 预估月成本浪费: ¥{estimated_cost:,.0f}")
    
    print(f"\n💡 优化建议:")
    if unbound_eips:
        print(f"  1. 立即释放 {len(unbound_eips)} 个未绑定EIP,可节省约 ¥{len(unbound_eips)*50}/月")
    if low_traffic_eips:
        print(f"  2. 评估 {len(low_traffic_eips)} 个低流量EIP,考虑降低带宽或释放")
    print(f"{'='*100}")


def main():
    """主函数"""
    print("="*100)
    print("🌐 YDZN & ZMYC 租户闲置EIP详细报告")
    print(f"📅 报告时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*100)
    
    # 查找最新的EIP报告
    eip_reports = glob.glob('*eip*idle_report*.xlsx')
    
    if not eip_reports:
        print("\n❌ 未找到EIP报告文件,请先运行分析:")
        print("   python3 main.py ydzn cru eip")
        print("   python3 main.py zmyc cru eip")
        return
    
    # 按修改时间排序,获取最新的
    eip_reports.sort(key=os.path.getmtime, reverse=True)
    
    # 找到YDZN和ZMYC的报告
    ydzn_report = None
    zmyc_report = None
    
    for report in eip_reports[:10]:  # 只检查最近10个
        if 'ydzn' in report.lower():
            ydzn_report = report
        elif 'zmyc' in report.lower():
            zmyc_report = report
    
    # 如果没有找到特定租户的报告,使用最新的两个
    if not ydzn_report or not zmyc_report:
        print(f"\n⚠️  未找到明确标注租户的报告,使用最新的报告文件:")
        print(f"   最新报告: {eip_reports[0]}")
        
        # 读取最新报告
        all_eips = read_eip_report(eip_reports[0])
        
        # 尝试从绑定实例或其他字段推断租户
        ydzn_eips = []
        zmyc_eips = []
        other_eips = []
        
        for eip in all_eips:
            # 从绑定实例ID、区域等信息推断租户
            instance_id = str(eip.get('绑定实例ID', ''))
            region = str(eip.get('区域', ''))
            
            # 这里需要根据实际情况判断
            # 暂时无法准确区分,显示所有
            other_eips.append(eip)
        
        # 显示所有闲置EIP
        print(f"\n⚠️  无法区分租户,显示全部闲置EIP:")
        display_eip_list(all_eips, 'all', '所有租户')
        
    else:
        # 显示YDZN租户的EIP
        if ydzn_report:
            ydzn_eips = read_eip_report(ydzn_report)
            display_eip_list(ydzn_eips, 'ydzn', '羊小咩数科')
        
        # 显示ZMYC租户的EIP
        if zmyc_report:
            zmyc_eips = read_eip_report(zmyc_report)
            display_eip_list(zmyc_eips, 'zmyc', 'ZMYC')
    
    # 显示报告文件信息
    print(f"\n\n{'='*100}")
    print("📄 详细报告文件:")
    print("="*100)
    for i, report in enumerate(eip_reports[:5], 1):
        size = os.path.getsize(report) / 1024
        mtime = datetime.fromtimestamp(os.path.getmtime(report)).strftime('%Y-%m-%d %H:%M:%S')
        print(f"{i}. {report}")
        print(f"   生成时间: {mtime} | 大小: {size:.2f} KB")
    
    print("\n💡 查看详细报告:")
    print(f"   Excel: open {eip_reports[0]}")
    print(f"   HTML:  open {eip_reports[0].replace('.xlsx', '.html')}")
    print("="*100)


if __name__ == "__main__":
    main()
