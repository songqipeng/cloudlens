#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
汇总所有已生成的闲置资源报告
"""

import os
import glob
import openpyxl
from datetime import datetime
from collections import defaultdict


def read_excel_summary(excel_file):
    """读取Excel报告的摘要信息"""
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        
        # 计算数据行数(排除标题行)
        row_count = ws.max_row - 1 if ws.max_row > 1 else 0
        
        # 读取一些关键信息
        headers = [cell.value for cell in ws[1]]
        
        wb.close()
        
        return {
            'row_count': row_count,
            'headers': headers,
            'sheet_name': ws.title
        }
    except Exception as e:
        return {'error': str(e)}


def analyze_reports():
    """分析所有报告文件"""
    print("="*100)
    print("📊 所有租户闲置资源报告汇总")
    print(f"📅 汇总时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*100)
    
    # 查找所有报告
    xlsx_reports = glob.glob("*_idle_report_*.xlsx")
    
    if not xlsx_reports:
        print("\n❌ 未找到任何报告文件")
        return
    
    # 按租户和资源类型分组
    reports_by_tenant = defaultdict(list)
    reports_by_type = defaultdict(list)
    
    for report_file in xlsx_reports:
        stat = os.stat(report_file)
        
        # 解析文件名
        parts = report_file.replace('.xlsx', '').split('_')
        
        # 判断租户和资源类型
        if 'ydzn' in report_file:
            tenant = 'ydzn'
            resource_type = parts[1] if len(parts) > 1 else 'unknown'
        elif 'zmyc' in report_file:
            tenant = 'zmyc'
            resource_type = parts[1] if len(parts) > 1 else 'unknown'
        elif 'cf' in report_file:
            tenant = 'cf'
            resource_type = parts[1] if len(parts) > 1 else 'unknown'
        else:
            tenant = 'all'
            resource_type = parts[0]
        
        # 读取Excel内容
        summary = read_excel_summary(report_file)
        
        info = {
            'file': report_file,
            'tenant': tenant,
            'resource_type': resource_type,
            'mtime': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'size_kb': stat.st_size / 1024,
            'idle_count': summary.get('row_count', 0),
            'summary': summary
        }
        
        reports_by_tenant[tenant].append(info)
        reports_by_type[resource_type].append(info)
    
    # 显示按租户汇总
    print("\n" + "="*100)
    print("📦 按租户汇总:")
    print("="*100)
    
    tenant_names = {
        'ydzn': '羊小咩数科 (YDZN)',
        'zmyc': 'ZMYC',
        'cf': 'CF租户',
        'all': '全部租户'
    }
    
    total_idle_resources = defaultdict(int)
    
    for tenant in sorted(reports_by_tenant.keys()):
        tenant_display = tenant_names.get(tenant, tenant)
        reports = reports_by_tenant[tenant]
        
        print(f"\n🏢 {tenant_display}")
        print("-"*100)
        
        if reports:
            print(f"{'资源类型':<15s} {'闲置数量':<10s} {'报告文件':<50s} {'生成时间':<20s}")
            print("-"*100)
            
            for report in sorted(reports, key=lambda x: x['resource_type']):
                resource_display = {
                    'ecs': 'ECS实例',
                    'rds': 'RDS数据库',
                    'redis': 'Redis缓存',
                    'oss': 'OSS存储',
                    'slb': '负载均衡',
                    'eip': 'EIP地址',
                    'disk': '云盘',
                    'nas': 'NAS存储',
                    'dns': 'DNS',
                    'mongodb': 'MongoDB',
                    'polardb': 'PolarDB',
                    'clickhouse': 'ClickHouse'
                }.get(report['resource_type'], report['resource_type'].upper())
                
                idle_count = report['idle_count']
                total_idle_resources[report['resource_type']] += idle_count
                
                status_emoji = "🔴" if idle_count > 0 else "✅"
                
                print(f"{status_emoji} {resource_display:<13s} {idle_count:<10d} {report['file']:<50s} {report['mtime']}")
        else:
            print("  无报告")
    
    # 显示按资源类型汇总
    print("\n" + "="*100)
    print("📊 按资源类型汇总:")
    print("="*100)
    print(f"\n{'资源类型':<20s} {'闲置总数':<15s} {'报告数量':<10s}")
    print("-"*100)
    
    for resource_type in sorted(total_idle_resources.keys()):
        resource_display = {
            'ecs': 'ECS实例',
            'rds': 'RDS数据库',
            'redis': 'Redis缓存',
            'oss': 'OSS存储',
            'slb': '负载均衡',
            'eip': 'EIP地址',
            'disk': '云盘',
            'nas': 'NAS存储',
            'dns': 'DNS',
            'mongodb': 'MongoDB',
            'polardb': 'PolarDB',
            'clickhouse': 'ClickHouse'
        }.get(resource_type, resource_type.upper())
        
        total_count = total_idle_resources[resource_type]
        report_count = len(reports_by_type[resource_type])
        
        status_emoji = "🔴" if total_count > 10 else "🟡" if total_count > 0 else "✅"
        
        print(f"{status_emoji} {resource_display:<18s} {total_count:<15d} {report_count:<10d}")
    
    # 总计
    total_idle = sum(total_idle_resources.values())
    print("-"*100)
    print(f"{'总计':<20s} {total_idle:<15d} {len(xlsx_reports):<10d}")
    
    # 关键发现
    print("\n" + "="*100)
    print("🔍 关键发现:")
    print("="*100)
    
    # 找出闲置资源最多的类型
    sorted_resources = sorted(total_idle_resources.items(), key=lambda x: x[1], reverse=True)
    
    if sorted_resources and sorted_resources[0][1] > 0:
        print("\n🔴 闲置资源最多的类型:")
        for resource_type, count in sorted_resources[:5]:
            if count > 0:
                resource_display = {
                    'ecs': 'ECS实例',
                    'rds': 'RDS数据库',
                    'redis': 'Redis缓存',
                    'oss': 'OSS存储',
                    'slb': '负载均衡',
                    'eip': 'EIP地址',
                    'disk': '云盘',
                    'nas': 'NAS存储'
                }.get(resource_type, resource_type.upper())
                print(f"  • {resource_display}: {count} 个")
    else:
        print("\n✅ 未发现明显闲置资源")
    
    # 优化建议
    print("\n" + "="*100)
    print("💡 优化建议:")
    print("="*100)
    
    recommendations = []
    
    if total_idle_resources.get('ecs', 0) > 0:
        recommendations.append(f"• ECS: 发现 {total_idle_resources['ecs']} 个闲置实例,建议评估是否可以降配或释放")
    
    if total_idle_resources.get('rds', 0) > 0:
        recommendations.append(f"• RDS: 发现 {total_idle_resources['rds']} 个低使用率数据库,建议评估是否可以缩容或合并")
    
    if total_idle_resources.get('disk', 0) > 0:
        recommendations.append(f"• 云盘: 发现 {total_idle_resources['disk']} 个未挂载云盘,建议删除或挂载使用")
    
    if total_idle_resources.get('eip', 0) > 0:
        recommendations.append(f"• EIP: 发现 {total_idle_resources['eip']} 个未绑定IP,建议释放以节省成本")
    
    if total_idle_resources.get('slb', 0) > 0:
        recommendations.append(f"• SLB: 发现 {total_idle_resources['slb']} 个低流量负载均衡,建议评估是否可以下线")
    
    if recommendations:
        for rec in recommendations:
            print(f"\n  {rec}")
    else:
        print("\n  ✅ 当前未发现需要立即优化的闲置资源")
    
    print("\n" + "="*100)
    print("📄 查看详细报告:")
    print("="*100)
    print("""
所有报告已生成在当前目录下:
- Excel格式: *_idle_report_*.xlsx
- HTML格式: *_idle_report_*.html

使用以下命令打开报告:
- 在Finder中打开: open .
- 打开Excel报告: open ydzn_ecs_idle_report_*.xlsx
- 打开HTML报告: open ydzn_ecs_idle_report_*.html
""")
    
    print("="*100)


if __name__ == "__main__":
    analyze_reports()
