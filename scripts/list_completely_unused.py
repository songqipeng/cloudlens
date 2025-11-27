#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
汇总YDZN和ZMYC租户下完全未使用的资源
"""

import openpyxl
from datetime import datetime
import os
import glob
from collections import defaultdict


def read_excel_data(file_path):
    """读取Excel报告数据"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append(dict(zip(headers, row)))
        
        wb.close()
        return data
    except Exception as e:
        return []


def analyze_unused_resources():
    """分析完全未使用的资源"""
    
    print("="*120)
    print("🔍 YDZN & ZMYC 租户 - 完全未使用资源分析")
    print(f"📅 分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*120)
    
    unused_resources = defaultdict(lambda: {'ydzn': [], 'zmyc': [], 'unknown': []})
    
    # 1. 查找未挂载云盘
    print("\n🔍 正在检查云盘...")
    disk_reports = glob.glob('*disk*idle_report*.xlsx')
    disk_reports.sort(key=os.path.getmtime, reverse=True)
    
    if disk_reports:
        for report in disk_reports[:2]:
            disks = read_excel_data(report)
            tenant = 'zmyc' if 'zmyc' in report.lower() else 'ydzn' if 'ydzn' in report.lower() else 'unknown'
            
            for disk in disks:
                status = str(disk.get('状态', '') or disk.get('挂载状态', ''))
                if '未挂载' in status or 'Available' in status:
                    unused_resources['未挂载云盘'][tenant].append(disk)
    
    # 2. 查找未绑定EIP
    print("🔍 正在检查EIP...")
    eip_reports = glob.glob('*eip*idle_report*.xlsx')
    eip_reports.sort(key=os.path.getmtime, reverse=True)
    
    if eip_reports:
        eips = read_excel_data(eip_reports[0])
        for eip in eips:
            instance_id = eip.get('绑定实例ID', '')
            reason = str(eip.get('闲置原因', ''))
            
            if not instance_id or '未绑定' in reason:
                # 尝试识别租户
                unused_resources['未绑定EIP']['unknown'].append(eip)
    
    # 3. 查找停机ECS实例
    print("🔍 正在检查ECS实例...")
    ecs_reports = glob.glob('*ecs*idle_report*.xlsx')
    ecs_reports.sort(key=os.path.getmtime, reverse=True)
    
    if ecs_reports:
        for report in ecs_reports[:2]:
            instances = read_excel_data(report)
            tenant = 'zmyc' if 'zmyc' in report.lower() else 'ydzn' if 'ydzn' in report.lower() else 'unknown'
            
            for inst in instances:
                status = str(inst.get('状态', '') or inst.get('实例状态', ''))
                cpu = inst.get('平均CPU使用率(%)', inst.get('CPU使用率(%)', 100))
                
                # 停机或CPU极低
                if 'Stopped' in status or (cpu is not None and cpu < 1):
                    unused_resources['停机/极低使用ECS'][tenant].append(inst)
    
    # 4. 查找零连接RDS
    print("🔍 正在检查RDS数据库...")
    rds_reports = glob.glob('*rds*idle_report*.xlsx')
    rds_reports.sort(key=os.path.getmtime, reverse=True)
    
    if rds_reports:
        for report in rds_reports[:2]:
            instances = read_excel_data(report)
            tenant = 'zmyc' if 'zmyc' in report.lower() else 'ydzn' if 'ydzn' in report.lower() else 'unknown'
            
            for inst in instances:
                connections = inst.get('平均连接数', inst.get('连接数', 100))
                if connections is not None and connections < 1:
                    unused_resources['零连接RDS'][tenant].append(inst)
    
    # 5. 查找零流量SLB
    print("🔍 正在检查SLB负载均衡...")
    slb_reports = glob.glob('*slb*idle_report*.xlsx')
    slb_reports.sort(key=os.path.getmtime, reverse=True)
    
    if slb_reports:
        instances = read_excel_data(slb_reports[0])
        for inst in instances:
            connections = inst.get('平均连接数', inst.get('活跃连接数', 100))
            if connections is not None and connections < 1:
                unused_resources['零流量SLB']['unknown'].append(inst)
    
    # 6. 查找未使用NAS
    print("🔍 正在检查NAS存储...")
    nas_reports = glob.glob('*nas*idle_report*.xlsx')
    nas_reports.sort(key=os.path.getmtime, reverse=True)
    
    if nas_reports:
        for report in nas_reports[:2]:
            instances = read_excel_data(report)
            tenant = 'zmyc' if 'zmyc' in report.lower() else 'ydzn' if 'ydzn' in report.lower() else 'unknown'
            
            for inst in instances:
                mount_count = inst.get('挂载点数量', 1)
                if mount_count == 0:
                    unused_resources['零挂载点NAS'][tenant].append(inst)
    
    # 显示结果
    print("\n" + "="*120)
    print("📊 完全未使用资源汇总")
    print("="*120)
    
    total_unused = 0
    ydzn_total = 0
    zmyc_total = 0
    
    for resource_type, tenants in unused_resources.items():
        ydzn_count = len(tenants['ydzn'])
        zmyc_count = len(tenants['zmyc'])
        unknown_count = len(tenants['unknown'])
        total = ydzn_count + zmyc_count + unknown_count
        
        if total > 0:
            total_unused += total
            ydzn_total += ydzn_count
            zmyc_total += zmyc_count
            
            print(f"\n{'='*120}")
            print(f"📦 {resource_type}")
            print(f"{'='*120}")
            print(f"  总计: {total} 个 | YDZN: {ydzn_count} 个 | ZMYC: {zmyc_count} 个 | 其他: {unknown_count} 个")
            
            # 显示YDZN资源
            if ydzn_count > 0:
                print(f"\n  🏢 YDZN租户 ({ydzn_count}个):")
                print("  " + "-"*116)
                
                for i, resource in enumerate(tenants['ydzn'][:10], 1):
                    if resource_type == '未挂载云盘':
                        disk_id = resource.get('云盘ID', 'N/A')
                        name = resource.get('云盘名称', '未命名')
                        size = resource.get('容量(GB)', 'N/A')
                        region = resource.get('地域', 'N/A')
                        print(f"  {i}. {name} | ID: {disk_id} | 容量: {size}GB | 地域: {region}")
                    
                    elif resource_type == '停机/极低使用ECS':
                        inst_id = resource.get('实例ID', 'N/A')
                        name = resource.get('实例名称', '未命名')
                        status = resource.get('状态', 'N/A')
                        cpu = resource.get('平均CPU使用率(%)', 'N/A')
                        print(f"  {i}. {name} | ID: {inst_id} | 状态: {status} | CPU: {cpu}%")
                    
                    elif resource_type == '零连接RDS':
                        inst_id = resource.get('实例ID', 'N/A')
                        name = resource.get('实例名称', '未命名')
                        engine = resource.get('引擎', 'N/A')
                        conn = resource.get('平均连接数', 0)
                        print(f"  {i}. {name} | ID: {inst_id} | 引擎: {engine} | 连接数: {conn}")
                    
                    elif resource_type == '零挂载点NAS':
                        nas_id = resource.get('文件系统ID', 'N/A')
                        storage_type = resource.get('存储类型', 'N/A')
                        region = resource.get('地域', 'N/A')
                        print(f"  {i}. ID: {nas_id} | 类型: {storage_type} | 地域: {region}")
                
                if ydzn_count > 10:
                    print(f"  ... 还有 {ydzn_count - 10} 个")
            
            # 显示ZMYC资源
            if zmyc_count > 0:
                print(f"\n  🏢 ZMYC租户 ({zmyc_count}个):")
                print("  " + "-"*116)
                
                for i, resource in enumerate(tenants['zmyc'][:10], 1):
                    if resource_type == '未挂载云盘':
                        disk_id = resource.get('云盘ID', 'N/A')
                        name = resource.get('云盘名称', '未命名')
                        size = resource.get('容量(GB)', 'N/A')
                        region = resource.get('地域', 'N/A')
                        print(f"  {i}. {name} | ID: {disk_id} | 容量: {size}GB | 地域: {region}")
                    
                    elif resource_type == '停机/极低使用ECS':
                        inst_id = resource.get('实例ID', 'N/A')
                        name = resource.get('实例名称', '未命名')
                        status = resource.get('状态', 'N/A')
                        cpu = resource.get('平均CPU使用率(%)', 'N/A')
                        print(f"  {i}. {name} | ID: {inst_id} | 状态: {status} | CPU: {cpu}%")
                
                if zmyc_count > 10:
                    print(f"  ... 还有 {zmyc_count - 10} 个")
            
            # 显示其他资源
            if unknown_count > 0 and unknown_count <= 5:
                print(f"\n  🌐 其他/未分类 ({unknown_count}个):")
                print("  " + "-"*116)
                for i, resource in enumerate(tenants['unknown'], 1):
                    print(f"  {i}. {resource}")
    
    # 总结
    print("\n" + "="*120)
    print("📈 总体统计")
    print("="*120)
    print(f"\n  完全未使用资源总数: {total_unused} 个")
    print(f"  • YDZN租户: {ydzn_total} 个")
    print(f"  • ZMYC租户: {zmyc_total} 个")
    
    # 优化建议
    print("\n" + "="*120)
    print("💡 立即行动建议")
    print("="*120)
    
    if unused_resources['未挂载云盘']['ydzn'] or unused_resources['未挂载云盘']['zmyc']:
        ydzn_disks = len(unused_resources['未挂载云盘']['ydzn'])
        zmyc_disks = len(unused_resources['未挂载云盘']['zmyc'])
        total_disks = ydzn_disks + zmyc_disks
        print(f"\n  🔴 删除未挂载云盘 ({total_disks}个)")
        print(f"     YDZN: {ydzn_disks}个 | ZMYC: {zmyc_disks}个")
        print(f"     预计节省: ¥{total_disks * 80}/月")
        print(f"     风险等级: 低 (先确认数据已备份)")
    
    if unused_resources['停机/极低使用ECS']['ydzn'] or unused_resources['停机/极低使用ECS']['zmyc']:
        ydzn_ecs = len(unused_resources['停机/极低使用ECS']['ydzn'])
        zmyc_ecs = len(unused_resources['停机/极低使用ECS']['zmyc'])
        total_ecs = ydzn_ecs + zmyc_ecs
        print(f"\n  🔴 释放停机/极低使用ECS ({total_ecs}个)")
        print(f"     YDZN: {ydzn_ecs}个 | ZMYC: {zmyc_ecs}个")
        print(f"     预计节省: ¥{total_ecs * 500}/月")
        print(f"     风险等级: 中 (需与业务方确认)")
    
    if unused_resources['零连接RDS']['ydzn'] or unused_resources['零连接RDS']['zmyc']:
        ydzn_rds = len(unused_resources['零连接RDS']['ydzn'])
        zmyc_rds = len(unused_resources['零连接RDS']['zmyc'])
        total_rds = ydzn_rds + zmyc_rds
        print(f"\n  🔴 释放零连接RDS ({total_rds}个)")
        print(f"     YDZN: {ydzn_rds}个 | ZMYC: {zmyc_rds}个")
        print(f"     预计节省: ¥{total_rds * 800}/月")
        print(f"     风险等级: 高 (需仔细确认,备份数据)")
    
    if total_unused == 0:
        print("\n  ✅ 未发现完全未使用的资源!")
        print("     但请注意,可能存在低使用率资源,建议查看详细报告。")
    
    print("\n" + "="*120)


if __name__ == "__main__":
    analyze_unused_resources()
